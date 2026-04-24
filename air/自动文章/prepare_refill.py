"""
补发准备工具 - 从定时发布失败记录准备批量补发

流程：
  1. 找最新失败记录.xlsx，按账号统计失败次数
  2. 清空 账号配置.xlsx 的"本轮已发"
  3. 重写"白名单"：A列账号 + B列失败次数（=发文份数）
  4. 剪切定时发布的素材到本目录的素材/

之后双击 go.command 即可按配额补发。

用法：
  python3 prepare_refill.py                     # 自动用最新失败记录
  python3 prepare_refill.py --fail <xlsx路径>   # 指定失败记录
  python3 prepare_refill.py --dry-run           # 预览不执行
"""

import argparse
import os
import shutil
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
CFG_XLSX = BASE_DIR / "账号配置.xlsx"
DST_MAT  = BASE_DIR / "素材"

# 定时发布路径（两个候选，哪个存在用哪个）
TIMER_DIR_CANDIDATES = [
    Path.home() / "code/头条自动发布/Mac文章定时自动发布",
    Path.home() / "Desktop/Mac文章定时自动发布",
]


def find_timer_dir():
    for p in TIMER_DIR_CANDIDATES:
        if p.exists():
            return p
    raise SystemExit(f"找不到定时发布目录：{TIMER_DIR_CANDIDATES}")


def find_latest_fail_xlsx(timer_dir: Path):
    reports = timer_dir / "运行报告"
    if not reports.exists():
        raise SystemExit(f"定时报告目录不存在：{reports}")
    candidates = []
    for d in reports.iterdir():
        f = d / "失败记录.xlsx"
        if f.exists():
            candidates.append((d.name, f))
    if not candidates:
        raise SystemExit(f"没找到失败记录.xlsx，看过：{reports}")
    candidates.sort(reverse=True)  # 按日期字符串降序
    return candidates[0][1]


def read_fail_counts(fail_xlsx: Path):
    wb = openpyxl.load_workbook(fail_xlsx, read_only=True)
    ws = wb["失败记录"]
    cnt = Counter()
    reasons = Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]:
            continue
        cnt[str(r[1]).strip()] += 1
        if len(r) > 4 and r[4]:
            reasons[str(r[4])] += 1
    wb.close()
    return cnt, reasons


def write_whitelist(cfg_path: Path, fail_cnt: Counter):
    # 备份
    bak = cfg_path.with_name(
        f"{cfg_path.name}.bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    shutil.copy2(cfg_path, bak)

    wb = openpyxl.load_workbook(cfg_path)

    # 补齐新sheet
    fail_headers    = ["账号名", "失败原因", "文稿名", "失败时间", "轮次"]
    summary_headers = ["账号名", "轮次", "发文时间", "失败时间", "补发成功时间"]
    if "失败列表" not in wb.sheetnames:
        wb.create_sheet("失败列表").append(fail_headers)
    if "发文汇总" not in wb.sheetnames:
        wb.create_sheet("发文汇总").append(summary_headers)
    if "本轮已发" not in wb.sheetnames:
        wb.create_sheet("本轮已发").append(["账号名"])
    if "白名单" not in wb.sheetnames:
        wb.create_sheet("白名单")

    # 清空 本轮已发
    ws_sent = wb["本轮已发"]
    if ws_sent.max_row >= 2:
        ws_sent.delete_rows(2, ws_sent.max_row)
    ws_sent.cell(1, 1, "账号名")

    # 重写 白名单
    ws_wl = wb["白名单"]
    if ws_wl.max_row >= 2:
        ws_wl.delete_rows(2, ws_wl.max_row)
    ws_wl.cell(1, 1, "账号名")
    ws_wl.cell(1, 2, "发文份数")
    for c in (ws_wl.cell(1, 1), ws_wl.cell(1, 2)):
        c.font = Font(bold=True)
    for name, cnt in sorted(fail_cnt.items(), key=lambda x: (-x[1], x[0])):
        ws_wl.append([name, cnt])

    wb.save(cfg_path)
    return bak


def move_materials(src_dir: Path, dst_dir: Path):
    dst_dir.mkdir(parents=True, exist_ok=True)
    moved = 0
    for f in src_dir.glob("*.docx"):
        shutil.move(str(f), str(dst_dir / f.name))
        moved += 1
    return moved


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fail", help="指定失败记录.xlsx路径")
    ap.add_argument("--timer-dir", help="指定定时发布根目录")
    ap.add_argument("--dry-run", action="store_true", help="只预览不执行")
    args = ap.parse_args()

    timer_dir = Path(args.timer_dir) if args.timer_dir else find_timer_dir()
    fail_xlsx = Path(args.fail) if args.fail else find_latest_fail_xlsx(timer_dir)
    src_mat   = timer_dir / "素材"

    print(f"定时发布目录：{timer_dir}")
    print(f"失败记录    ：{fail_xlsx}")
    print(f"源素材目录  ：{src_mat}")
    print(f"自动发布素材：{DST_MAT}")
    print(f"账号配置    ：{CFG_XLSX}")
    print()

    fail_cnt, reasons = read_fail_counts(fail_xlsx)
    total_quota = sum(fail_cnt.values())
    mat_count   = len(list(src_mat.glob("*.docx")))
    dist = Counter(fail_cnt.values())

    print(f"失败账号(去重): {len(fail_cnt)}")
    print(f"配额合计      : {total_quota} 篇")
    print(f"定时素材剩余  : {mat_count} 份")
    gap = total_quota - mat_count
    if gap > 0:
        print(f"⚠ 素材缺口    : {gap} 篇（需从台机/其他来源补）")
    elif gap < 0:
        print(f"ℹ 素材多出    : {-gap} 篇")
    print()
    print("配额分布：")
    for q, c in sorted(dist.items(), reverse=True):
        print(f"  {q}篇 × {c}账号 = {q*c}")
    print()
    print("失败原因分布（TOP5）：")
    for reason, n in reasons.most_common(5):
        print(f"  {n:4d}  {reason[:70]}")
    print()

    if args.dry_run:
        print("--dry-run：不执行写入和剪切。")
        return

    if not CFG_XLSX.exists():
        raise SystemExit(f"账号配置.xlsx 不存在：{CFG_XLSX}")

    bak = write_whitelist(CFG_XLSX, fail_cnt)
    print(f"✓ 账号配置已备份 → {bak.name}")
    print(f"✓ 白名单写入 {len(fail_cnt)} 账号；本轮已发已清空")

    moved = move_materials(src_mat, DST_MAT)
    dst_total = len(list(DST_MAT.glob("*.docx")))
    print(f"✓ 素材剪切 {moved} 份 → {DST_MAT.name}/（目前共 {dst_total} 份）")
    print()
    if dst_total < total_quota:
        print(f"⚠ 素材 {dst_total} < 配额 {total_quota}，差 {total_quota - dst_total} 篇")
        print("  解决方案：从台机 台机DS创作新文章/<批次目录>/ 剪切 docx 过来。")
    else:
        print(f"✓ 素材 {dst_total} ≥ 配额 {total_quota}，可直接双击 go.command 开跑。")


if __name__ == "__main__":
    main()
