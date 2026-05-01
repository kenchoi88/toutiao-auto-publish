#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 用子目录手工下载的 04-* 各日 xlsx,填「30号阅读/推荐」列 + 加「4月30号」行
# 缺哥规则:子目录 xlsx 阅读量 = 该批截至下载时刻(05-01 凌晨 ≈ 截至 04-30)的累计快照

import os, glob
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from data_pull import (parse_export_xlsx, sc, parse_date_label, date_to_label,
                       NEW_FILL, WHT_FONT, BLD_FONT, _find_input_xlsx, _make_output_xlsx)

STAT_DATE = date(2026, 4, 30)
HERE = os.path.dirname(os.path.abspath(__file__))

SHEET_TO_SUBDIR = {
    '小馆微头条': ('小馆头条', 3),
    '迦境微头条': ('迦境头条', 3),
    '小馆文章':   ('小馆文章', 1),
    '迦境文章':   ('迦境文章', 1),
}

def read_day_local(subdir: str, day: date, ctype: int):
    pat = os.path.join(HERE, subdir, f'*{day.strftime("%Y%m%d")}-{day.strftime("%Y%m%d")}*.xlsx')
    files = sorted(glob.glob(pat))
    if not files:
        return None
    total = recommend = read = 0
    for f in files:
        rows = parse_export_xlsx(open(f, 'rb').read())
        data = rows[1:] if len(rows) > 1 else []
        total += len(data)
        for row in data:
            if ctype == 3:
                try: read += int(float(row[2])) if len(row) > 2 and row[2] else 0
                except: pass
            else:
                try: recommend += int(float(row[2])) if len(row) > 2 and row[2] else 0
                except: pass
                try: read += int(float(row[3])) if len(row) > 3 and row[3] else 0
                except: pass
    return {'total': total, 'recommend': recommend, 'read': read}

def main():
    inp = _find_input_xlsx()
    out = _make_output_xlsx(STAT_DATE)
    print(f'读取: {os.path.basename(inp)}')
    print(f'写入: {os.path.basename(out)}')
    wb = load_workbook(inp)
    stat_day = f'{STAT_DATE.day}号'
    new_label = date_to_label(STAT_DATE)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_SUBDIR:
            print(f'跳过 {sheet_name}(无本地数据源)')
            continue
        subdir, ctype = SHEET_TO_SUBDIR[sheet_name]
        ws = wb[sheet_name]
        is_micro = ctype == 3
        print(f'\n=== {sheet_name} ===')

        existing = {}
        for r in range(2, ws.max_row + 1):
            lbl = ws.cell(r, 1).value
            if lbl: existing[str(lbl)] = r

        if new_label not in existing:
            new_row = ws.max_row + 1
            sc(ws, new_row, 1, new_label, font=BLD_FONT)
            existing[new_label] = new_row
            print(f'  新增行: {new_label} @ R{new_row}')

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if is_micro:
            read_name = f'{stat_day}阅读'
            if read_name in headers:
                read_col = headers.index(read_name) + 1
            else:
                read_col = ws.max_column + 1
                sc(ws, 1, read_col, read_name, NEW_FILL, WHT_FONT)
                ws.column_dimensions[get_column_letter(read_col)].width = 11
            rec_col = None
        else:
            rec_name  = f'{stat_day}推荐'
            read_name = f'{stat_day}阅读'
            if rec_name in headers:
                rec_col = headers.index(rec_name) + 1
            else:
                last_rec = max((c for c in range(1, ws.max_column + 1)
                                if ws.cell(1, c).value and '推荐' in str(ws.cell(1, c).value)), default=0)
                if 0 < last_rec < ws.max_column:
                    ws.insert_cols(last_rec + 1)
                    rec_col = last_rec + 1
                else:
                    rec_col = ws.max_column + 1
                sc(ws, 1, rec_col, rec_name, NEW_FILL, WHT_FONT)
                ws.column_dimensions[get_column_letter(rec_col)].width = 11
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if read_name in headers:
                read_col = headers.index(read_name) + 1
            else:
                read_col = ws.max_column + 1
                sc(ws, 1, read_col, read_name, NEW_FILL, WHT_FONT)
                ws.column_dimensions[get_column_letter(read_col)].width = 11

        for pub_lbl, row_idx in sorted(existing.items()):
            d = parse_date_label(pub_lbl)
            if d is None or d > STAT_DATE: continue
            data = read_day_local(subdir, d, ctype)
            if data is None:
                print(f'    {pub_lbl}: 无本地 xlsx,跳过')
                continue
            print(f'    {pub_lbl}: 发文{data["total"]} 推荐{data["recommend"]} 阅读{data["read"]}')
            val_rec  = data['recommend'] or '-'
            val_read = data['read'] or '-'
            if is_micro:
                sc(ws, row_idx, read_col, val_read)
            else:
                sc(ws, row_idx, rec_col,  val_rec)
                sc(ws, row_idx, read_col, val_read)
            if d == STAT_DATE and data['total'] > 0:
                sc(ws, row_idx, 2, data['total'])

    wb.save(out)
    print(f'\n已保存: {out}')

if __name__ == '__main__':
    main()
