#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# MCN流量对比 - 每天运行一次，追加今日统计列到流量对比汇总.xlsx
# 【Win 台机版】路径自适应:Mac 在 ~/Library/Application Support/创作罐头/...,
#   Win 在 %APPDATA%\创作罐头\Partitions\<id>\Network\Cookies(多一层 Network)。

import sqlite3, os, sys, json, requests, urllib.request, websocket
import zipfile, io, xml.etree.ElementTree as ET, time, re, warnings, platform, glob
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── 路径自适应(Mac vs Win)─────────────────────────────────
IS_WIN = platform.system() == 'Windows'
if IS_WIN:
    GTH_BASE = os.path.join(os.environ.get('APPDATA', ''), '创作罐头')
    COOKIE_SUFFIX = os.path.join('Network', 'Cookies')   # Win:Partitions\<id>\Network\Cookies
    MAIN_COOKIE   = os.path.join(GTH_BASE, 'Network', 'Cookies')
else:
    GTH_BASE = os.path.expanduser('~/Library/Application Support/创作罐头')
    COOKIE_SUFFIX = 'Cookies'                            # Mac:Partitions/<id>/Cookies
    MAIN_COOKIE   = os.path.join(GTH_BASE, 'Cookies')

PARTITION_DIR    = os.path.join(GTH_BASE, 'Partitions')
MOTHER_PARTITION = "7477169161966321683"   # 小馆矩阵
SISTER_PARTITION = "7601367523329638450"   # 迦境矩阵

def _find_input_xlsx():
    """读取用:找最新版本(按修改时间)。支持「流量对比汇总.xlsx」「流量对比汇总4.25.xlsx」等命名。"""
    here = os.path.dirname(os.path.abspath(__file__))
    cands = [p for p in glob.glob(os.path.join(here, '流量对比汇总*.xlsx'))
             if not os.path.basename(p).startswith('~$')]
    if not cands:
        return os.path.join(here, '流量对比汇总.xlsx')
    return max(cands, key=os.path.getmtime)

def _make_output_xlsx(stat_date):
    """写入用:按 stat_date 生成「流量对比汇总<月>月<日>.xlsx」,每天一个新文件不覆盖历史。"""
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(here, f'流量对比汇总{stat_date.month}月{stat_date.day}.xlsx')

INPUT_XLSX = _find_input_xlsx()
# OUTPUT_XLSX 在 main() 里根据 stat_date 计算

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36' if IS_WIN
                  else 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Referer': 'https://mp.toutiao.com/profile_v4/lab/matrix_manage/content'
}

# ── Cookie ────────────────────────────────────────────────
def load_cookies_sqlite(partition_id=None):
    """从创作罐头 sqlite 读 cookie。Win 上路径是 Partitions/<id>/Network/Cookies。"""
    cookies = {}
    if partition_id is None:
        paths = [MAIN_COOKIE, os.path.join(PARTITION_DIR, MOTHER_PARTITION, *COOKIE_SUFFIX.split(os.sep))]
    else:
        paths = [os.path.join(PARTITION_DIR, partition_id, *COOKIE_SUFFIX.split(os.sep))]
    for f in paths:
        if not os.path.exists(f): continue
        try:
            conn = sqlite3.connect(f"file:{f}?mode=ro", uri=True)
            cur  = conn.cursor()
            cur.execute("SELECT name, value FROM cookies WHERE host_key LIKE '%toutiao%' OR host_key LIKE '%bytedance%'")
            for name, val in cur.fetchall(): cookies[name] = val
            conn.close()
        except Exception as e:
            print(f"  读 {f} 失败: {e}")
    return cookies

def load_cookies_cdp():
    port_file = os.path.join(GTH_BASE, 'DevToolsActivePort')
    if not os.path.exists(port_file):
        return {}
    port = int(open(port_file).readline().strip())
    try:
        pages = json.loads(urllib.request.urlopen(f'http://localhost:{port}/json', timeout=5).read())
    except Exception as e:
        print(f"  连不上罐头CDP: {e}"); return {}
    target = next((p for p in pages if 'matrix_manage' in p.get('url','') and p.get('type')=='webview'), None)
    if not target:
        target = next((p for p in pages if 'toutiao' in p.get('url','') and p.get('type')=='webview'), None)
    if not target: return {}
    try:
        ws = websocket.create_connection(target['webSocketDebuggerUrl'], timeout=10, suppress_origin=True)
        ws.send(json.dumps({'id':1,'method':'Network.getCookies','params':{'urls':['https://mp.toutiao.com']}}))
        deadline = time.time() + 8
        while time.time() < deadline:
            msg = json.loads(ws.recv())
            if msg.get('id') == 1:
                cookies = {c['name']:c['value'] for c in msg['result']['cookies']}
                ws.close(); return cookies
    except: pass
    return {}

# ── 解析头条导出xlsx ──────────────────────────────────────
def parse_export_xlsx(content):
    if not content or content[:2] != b'PK': return []
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    zf = zipfile.ZipFile(io.BytesIO(content))
    shared = []
    if 'xl/sharedStrings.xml' in zf.namelist():
        tree = ET.parse(zf.open('xl/sharedStrings.xml'))
        for si in tree.getroot().findall('.//s:si', ns):
            shared.append(''.join(t.text or '' for t in si.findall('.//s:t', ns)))
    sheets = sorted(n for n in zf.namelist() if n.startswith('xl/worksheets/sheet'))
    if not sheets: return []
    tree = ET.parse(zf.open(sheets[0]))
    rows = []
    for row_el in tree.getroot().findall('.//s:row', ns):
        vals = []
        for c in row_el.findall('s:c', ns):
            v = c.find('s:v', ns); t = c.get('t', '')
            if v is None: vals.append('')
            elif t == 's': vals.append(shared[int(v.text)] if v.text else '')
            else: vals.append(v.text or '')
        rows.append(vals)
    zf.close()
    return rows

def fetch_day(cookies, pub_date: date, content_type: int):
    day_str = pub_date.strftime('%Y-%m-%d')
    all_rows, page = [], 1
    while True:
        url = (f"https://mp.toutiao.com/mp/agw/media_matrix/export"
               f"?from={day_str}&to={day_str}&type={content_type}&page_num={page}&size=1000")
        try:
            r = requests.get(url, cookies=cookies, headers=HEADERS, timeout=30)
            rows = parse_export_xlsx(r.content)
            data = rows[1:] if len(rows) > 1 else []
            all_rows.extend(data)
            if len(data) < 1000: break
            page += 1
        except Exception as e:
            print(f"    {day_str} 失败: {e}"); break
    recommend = read = 0
    for row in all_rows:
        if content_type == 3:  # 微头条:阅读量在 index 2,无推荐量列
            try: read += int(float(row[2])) if len(row) > 2 and row[2] else 0
            except: pass
        else:  # 文章:推荐量 index 2,阅读量 index 3
            try: recommend += int(float(row[2])) if len(row) > 2 and row[2] else 0
            except: pass
            try: read += int(float(row[3])) if len(row) > 3 and row[3] else 0
            except: pass
    return {'total': len(all_rows), 'recommend': recommend, 'read': read}

def fetch_income_day(cookies, day: date):
    day_str = day.strftime('%Y%m%d')
    all_rows, page = [], 1
    while True:
        url = (f"https://mp.toutiao.com/mp/agw/statistic/matrix/matrix_media_daily_stat_export"
               f"?start_date={day_str}&end_date={day_str}&pagenum={page}&pagesize=50")
        try:
            r = requests.get(url, cookies=cookies, headers=HEADERS, timeout=30)
            rows = parse_export_xlsx(r.content)
            data = rows[1:] if len(rows) > 1 else []
            all_rows.extend(data)
            if len(data) < 50: break
            page += 1
        except Exception as e:
            print(f"    {day_str} 收益失败: {e}"); break
    fawen = tj = yd = 0
    liuzhuan = 0.0
    for row in all_rows:
        def g(i, cast=int):
            try: return cast(float(row[i])) if len(row) > i and row[i] else 0
            except: return 0
        fawen    += g(3)
        tj       += g(4)
        yd       += g(5)
        liuzhuan += g(11, float)
    return {'fawen': fawen, 'tj': tj, 'yd': yd, 'liuzhuan': liuzhuan}

# ── 样式 ─────────────────────────────────────────────────
NEW_FILL  = PatternFill("solid", fgColor="1F4E79")
BLUE_ROW  = PatternFill("solid", fgColor="D6E4F0")
WHT_ROW   = PatternFill("solid", fgColor="FFFFFF")
WHT_FONT  = Font(color="FFFFFF", bold=True, size=11, name="微软雅黑")
BLD_FONT  = Font(name="微软雅黑", size=11, bold=True)
BODY_FONT = Font(name="微软雅黑", size=11)
CENTER    = Alignment(horizontal='center', vertical='center')
THIN      = Side(style='thin', color='CCCCCC')
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def sc(ws, row, col, val, fill=None, font=None):
    c = ws.cell(row=row, column=col, value=val)
    c.alignment = CENTER; c.border = BORDER
    if isinstance(val, (int, float)):
        c.number_format = '#,##0'
    if row > 1 and font is None:
        c.font = BODY_FONT
    if font: c.font = font
    if fill:
        c.fill = fill
    elif row > 1:
        a_rgb = None
        if col != 1:
            af = ws.cell(row, 1).fill
            a_rgb = af.fgColor.rgb if af and af.fgColor else None
        if a_rgb and a_rgb not in ('00000000',) and str(a_rgb) != '00000000':
            c.fill = PatternFill("solid", fgColor=str(a_rgb)[-6:])
        else:
            c.fill = BLUE_ROW if row % 2 == 0 else WHT_ROW

def parse_date_label(label, year=2026):
    m = re.match(r'(\d+)月(\d+)号', str(label or ''))
    if m: return date(year, int(m.group(1)), int(m.group(2)))
    return None

def date_to_label(d: date):
    return f"{d.month}月{d.day}号"

# ── 主逻辑 ───────────────────────────────────────────────
def append_to_xlsx(stat_date: date, xg_cookies, jj_cookies, today_mode=False):
    if not os.path.exists(INPUT_XLSX):
        print(f"找不到 {INPUT_XLSX}")
        print(f"请把 air 桌面的「流量对比汇总.xlsx」拷到此目录:{os.path.dirname(INPUT_XLSX)}")
        return

    output_xlsx = _make_output_xlsx(stat_date)
    print(f"读取: {os.path.basename(INPUT_XLSX)}")
    print(f"写入: {os.path.basename(output_xlsx)}{'(覆盖同日已有)' if os.path.exists(output_xlsx) else '(新建)'}")
    wb = load_workbook(INPUT_XLSX)
    stat_day   = f"{stat_date.day}号"
    pub_date   = stat_date - timedelta(days=1)
    pub_label  = date_to_label(pub_date)
    stat_label = date_to_label(stat_date)

    # ── 重复跑拦截 ──(today_mode 跳过:用户主动刷新今日,允许覆盖)
    if not today_mode:
        read_col_name = f"{stat_day}阅读"
        for sn in wb.sheetnames:
            if '收益' in sn: continue
            ws_ck = wb[sn]
            hdr = [ws_ck.cell(1, c).value for c in range(1, ws_ck.max_column+1)]
            if read_col_name not in hdr: continue
            col_idx = hdr.index(read_col_name) + 1
            for r in range(2, ws_ck.max_row + 1):
                v = ws_ck.cell(r, col_idx).value
                if v not in (None, '', '-', 0):
                    lbl = ws_ck.cell(r, 1).value
                    print(f"\n⛔ 中止:{sn} 的「{read_col_name}」列已有数据({lbl} = {v})")
                    print(f"   stat_date={stat_date} 之前已跑过。再跑会覆盖定格快照。")
                    print(f"   真要重跑请先手动清空该列那些格子,或加 --today 强制刷新。")
                    sys.exit(1)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cookies  = xg_cookies if '小馆' in sheet_name else jj_cookies

        if '收益' in sheet_name:
            existing = {}
            for r in range(2, ws.max_row + 1):
                lbl = ws.cell(r, 1).value
                if lbl: existing[str(lbl)] = r
            d = date(stat_date.year, stat_date.month, 1)
            while d < stat_date:
                lbl = date_to_label(d)
                row_idx = existing.get(lbl)
                if row_idx and ws.cell(row_idx, 5).value not in (None, '', 0, '-'):
                    d += timedelta(days=1); continue
                data = fetch_income_day(cookies, d)
                if data['fawen'] == 0 and data['liuzhuan'] == 0:
                    d += timedelta(days=1); continue
                if row_idx is None:
                    row_idx = ws.max_row + 1
                    sc(ws, row_idx, 1, lbl, font=BLD_FONT)
                    existing[lbl] = row_idx
                sc(ws, row_idx, 2, data['fawen'])
                sc(ws, row_idx, 3, data['tj'])
                sc(ws, row_idx, 4, data['yd'])
                c = ws.cell(row=row_idx, column=5, value=data['liuzhuan'])
                c.alignment = CENTER; c.border = BORDER
                c.number_format = '¥#,##0.00'
                print(f"  [{sheet_name}] {lbl}: 发文{data['fawen']} 推荐{data['tj']} 阅读{data['yd']} 流转¥{data['liuzhuan']:.2f}")
                d += timedelta(days=1)
            continue

        is_micro = '微头条' in sheet_name
        ctype    = 3 if is_micro else 1

        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and isinstance(h, str) and '批' in h:
                ws.cell(1, c).value = h.replace('批', '号')

        existing = {}
        for r in range(2, ws.max_row + 1):
            lbl = ws.cell(r, 1).value
            if lbl: existing[str(lbl)] = r

        if today_mode and stat_label not in existing:
            new_row = ws.max_row + 1
            sc(ws, new_row, 1, stat_label, font=BLD_FONT)
            existing[stat_label] = new_row
            print(f"  [{sheet_name}] 新增今天行: {stat_label}")

        if pub_label not in existing:
            new_row = ws.max_row + 1
            sc(ws, new_row, 1, pub_label, font=BLD_FONT)
            existing[pub_label] = new_row
            print(f"  [{sheet_name}] 新增行: {pub_label}")

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if is_micro:
            read_name = f"{stat_day}阅读"
            if read_name in headers:
                col = headers.index(read_name) + 1
            else:
                col = ws.max_column + 1
                sc(ws, 1, col, read_name, NEW_FILL, WHT_FONT)
                ws.column_dimensions[get_column_letter(col)].width = 11
        else:
            rec_name = f"{stat_day}推荐"
            read_name = f"{stat_day}阅读"
            if rec_name in headers:
                rec_col = headers.index(rec_name) + 1
            else:
                last_rec = 0
                for c in range(1, ws.max_column + 1):
                    h = ws.cell(1, c).value
                    if h and '推荐' in str(h):
                        last_rec = c
                if last_rec > 0 and last_rec < ws.max_column:
                    ws.insert_cols(last_rec + 1)
                    rec_col = last_rec + 1
                else:
                    rec_col = ws.max_column + 1
                sc(ws, 1, rec_col, rec_name, NEW_FILL, WHT_FONT)
                ws.column_dimensions[get_column_letter(rec_col)].width = 11
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if read_name in headers:
                read_col = headers.index(read_name) + 1
            else:
                read_col = ws.max_column + 1
                sc(ws, 1, read_col, read_name, NEW_FILL, WHT_FONT)
                ws.column_dimensions[get_column_letter(read_col)].width = 11

        for pub_lbl, row_idx in sorted(existing.items(), key=lambda x: x[0]):
            d = parse_date_label(pub_lbl)
            if d is None or d > stat_date: continue
            data = fetch_day(cookies, d, ctype)
            print(f"    {pub_lbl}: 发文{data['total']} 推荐{data['recommend']} 阅读{data['read']}")
            val_rec  = data['recommend'] or '-'
            val_read = data['read'] or '-'
            if is_micro:
                sc(ws, row_idx, col, val_read)
            else:
                sc(ws, row_idx, rec_col,  val_rec)
                sc(ws, row_idx, read_col, val_read)
            if data['total'] > 0:
                sc(ws, row_idx, 2, data['total'])

    wb.save(output_xlsx)
    print(f"\n已保存: {output_xlsx}")

def main():
    args = sys.argv[1:]
    # 默认 today_mode=True(用户期望每天看当天) — 加 --no-today 关闭
    today_mode = True
    today = date.today()
    if args:
        if args[0] == '--no-today':
            today_mode = False
        elif args[0] == '--today':
            today_mode = True
        else:
            try:
                today = date.fromisoformat(args[0])
                today_mode = False  # 指定历史日期时不开 today_mode
            except ValueError:
                pass
    print(f"统计日: {today}  ({today.day}号列){' [today_mode]' if today_mode else ''}")
    print(f"平台: {'Win' if IS_WIN else 'Mac'} | 罐头根: {GTH_BASE}\n")

    print("读取小馆 cookie (Partition)...")
    xg = load_cookies_sqlite(MOTHER_PARTITION)
    print(f"  {'OK' if xg.get('sessionid') else '失败'}  sid={xg.get('sessionid','')[:16]}")
    print("读取迦境 cookie (Partition)...")
    jj = load_cookies_sqlite(SISTER_PARTITION)
    print(f"  {'OK' if jj.get('sessionid') else '失败'}  sid={jj.get('sessionid','')[:16]}\n")

    if not xg.get('sessionid') or not jj.get('sessionid'):
        print("⛔ 至少一个矩阵 cookie 没读到。先在罐头里登录小馆/迦境账号,或检查罐头有没有在跑。")
        sys.exit(1)

    append_to_xlsx(today, xg, jj, today_mode)
    print("完成!")

if __name__ == '__main__':
    main()
