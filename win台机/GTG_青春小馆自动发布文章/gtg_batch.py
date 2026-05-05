"""
创作罐头批量发布脚本 v3
项目目录结构:
  GTG_XXX/
  ├── 启动.bat
  ├── gtg_batch.py
  ├── 素材/           (放 .docx 文件)
  │   └── 已发送/     (发完自动移入)
  ├── gtg_log.txt     (运行日志)
  ├── 失败记录.xlsx   (发布失败的账号)
  ├── 高阅读量提醒.txt (阅读量超标时写入)
  └── 系统通知提醒.txt (检测到处罚消息时写入)
"""

import requests
import json
import websocket
import time
import os
import shutil
import glob
import sys
import random
from datetime import datetime, timedelta
import docx as docxlib
import ctypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== 配置 =====================
def _find_cdp_port():
    return 9223  # 固定端口，与 debug_launch.bat 一致

CDP_URL        = f"http://127.0.0.1:{_find_cdp_port()}"   # 自动检测罐头CDP端口
DEFAULT_TAG    = "原机构老号"               # 启动时未选择则使用此默认标签
ACCOUNT_CLASS  = "account-RALrbJ"
WAIT_LOAD      = 4
ALERT_THRESHOLD = 5000                     # 阅读量超过此值时写入提醒文件
EXCLUDE_ACCOUNTS = ["青春小馆"]            # 永不发文的账号（母账号等）
NOFIRST_ACCOUNTS = set()                   # 不选头条首发的账号（从账号配置.xlsx加载）

NO_PROXY       = {"http": "", "https": ""}
WS_OPTS        = {"suppress_origin": True}

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DOCS_FOLDER = os.path.join(BASE_DIR, "素材")
SENT_FOLDER = os.path.join(BASE_DIR, "素材", "已发送")

# 运行报告路径（main()开始时动态初始化）
RUN_REPORT_DIR = None
LOG_FILE       = None
FAIL_FILE      = None
NOTICE_FILE    = None
ALERT_FILE     = None
VIOLATION_FILE = None

VIOLATION_KEYWORDS = {
    "违规/扣分": ["违规", "扣分", "处罚", "警告"],
    "禁言封号": ["禁言", "发言受限", "封禁", "封号"],
    "原创侵权": ["原创违规", "侵权", "重复内容"],
}

# 绕过系统代理（避免 127.0.0.1 被代理拦截）
os.environ["NO_PROXY"] = "127.0.0.1,localhost"
# ================================================

# ========== 运行报告目录 ==========

def _init_run_dir():
    global LOG_FILE, FAIL_FILE, NOTICE_FILE, ALERT_FILE, VIOLATION_FILE, RUN_REPORT_DIR
    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y%m%d")
    RUN_REPORT_DIR = os.path.join(BASE_DIR, "运行报告", ts)
    os.makedirs(RUN_REPORT_DIR, exist_ok=True)
    LOG_FILE       = os.path.join(RUN_REPORT_DIR, "运行日志.txt")
    FAIL_FILE      = os.path.join(RUN_REPORT_DIR, "失败记录.xlsx")
    NOTICE_FILE    = os.path.join(RUN_REPORT_DIR, "系统通知.txt")
    ALERT_FILE     = os.path.join(RUN_REPORT_DIR, "高阅读提醒.txt")
    VIOLATION_FILE = os.path.join(RUN_REPORT_DIR, "违规提醒.txt")

# ========== 账号配置 Excel ==========

CONFIG_EXCEL  = os.path.join(BASE_DIR, "账号配置.xlsx")
SUMMARY_EXCEL = os.path.join(BASE_DIR, "发文汇总.xlsx")
_EXCEL_SHEETS = ["不首发", "永久跳过", "本轮已发", "白名单", "失败列表", "硬终止账号"]
_FAIL_HEADERS    = ["账号名", "失败原因", "文稿名", "失败时间", "轮次"]
_SUMMARY_HEADERS = ["账号名", "轮次", "发文时间", "失败时间", "补发成功时间"]
_HARD_TERMINATE_HEADERS = ["账号名", "终止原因", "终止时间", "本次已发篇数"]

# 4 类硬终止 reason — 命中即永久放弃,不再尝试
HARD_TERMINATE_REASONS = {"封号", "禁言", "侧边栏未找到", "信用分过低"}  # [v1101.1] 失登移出, 加信用分过低

def _ensure_config_excel():
    if not os.path.exists(CONFIG_EXCEL):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sh in _EXCEL_SHEETS:
            ws_new = wb.create_sheet(sh)
            if sh == "失败列表":
                ws_new.append(_FAIL_HEADERS)
            elif sh == "硬终止账号":
                ws_new.append(_HARD_TERMINATE_HEADERS)
            elif sh == "白名单":
                ws_new.append(["账号名", "发文份数"])
            else:
                ws_new.append(["账号名"])
        wb.save(CONFIG_EXCEL)
        return
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        dirty = False
        if "失败列表" not in wb.sheetnames:
            ws_n = wb.create_sheet("失败列表")
            ws_n.append(_FAIL_HEADERS); dirty = True
        if "硬终止账号" not in wb.sheetnames:
            ws_n = wb.create_sheet("硬终止账号")
            ws_n.append(_HARD_TERMINATE_HEADERS); dirty = True
        if dirty:
            wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _ensure_summary_excel():
    if os.path.exists(SUMMARY_EXCEL):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发文汇总"
        ws.append(_SUMMARY_HEADERS)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass

def _read_excel_sheet(sheet_name):
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb[sheet_name]
        result = []
        for row in ws_r.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip() and not str(val).strip().startswith('#'):
                result.append(str(val).strip())
        wb.close()
        return result
    except Exception:
        return []

def _append_sent_excel(name):
    """[v1102] 写「本轮已发」sheet:行存在 count+1,不存在 append (账号, 1)"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "本轮已发" not in wb.sheetnames:
            ws_s = wb.create_sheet("本轮已发")
            ws_s.append(["账号名", "已发次数"])
            ws_s.append([name, 1])
        else:
            ws_s = wb["本轮已发"]
            found_row = None
            for row_idx, row in enumerate(ws_s.iter_rows(min_row=2, max_col=2, values_only=False), start=2):
                if row[0].value and str(row[0].value).strip() == name:
                    found_row = row_idx
                    break
            if found_row:
                cur = ws_s.cell(row=found_row, column=2).value or 0
                try: cur = int(cur)
                except: cur = 0
                ws_s.cell(row=found_row, column=2).value = cur + 1
            else:
                ws_s.append([name, 1])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass


def _read_sent_with_count():
    """[v1102] 读「本轮已发」sheet → {账号: 已发次数}"""
    if not os.path.exists(CONFIG_EXCEL):
        return {}
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "本轮已发" not in wb.sheetnames:
            wb.close(); return {}
        ws_r = wb["本轮已发"]
        result = {}
        for row in ws_r.iter_rows(min_row=2, max_col=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            cnt = 1
            if len(row) > 1 and row[1] is not None:
                try: cnt = int(row[1])
                except: cnt = 1
            result[name] = cnt
        wb.close()
        return result
    except Exception:
        return {}

def _append_fail_list(name, reason, doc_name, round_num):
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "失败列表" not in wb.sheetnames:
            ws_f = wb.create_sheet("失败列表")
            ws_f.append(_FAIL_HEADERS)
        else:
            ws_f = wb["失败列表"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_f.append([name, reason or "", doc_name or "", ts, round_num])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _read_fail_list():
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "失败列表" not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb["失败列表"]
        result = []
        for row in ws_r.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            reason  = str(row[1]).strip() if row[1] is not None else ""
            docname = str(row[2]).strip() if row[2] is not None else ""
            ts      = str(row[3]).strip() if row[3] is not None else ""
            rnd     = int(row[4]) if (row[4] is not None and str(row[4]).strip().isdigit()) else 0
            result.append((name, reason, docname, ts, rnd))
        wb.close()
        return result
    except Exception:
        return []

def _remove_from_fail_list(name):
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "失败列表" not in wb.sheetnames:
            wb.close(); return
        ws_f = wb["失败列表"]
        rows_to_delete = []
        for idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] and str(row[0]).strip() == name:
                rows_to_delete.append(idx)
        for r_idx in reversed(rows_to_delete):
            ws_f.delete_rows(r_idx)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _append_summary(name, round_num, event_type, ts=None):
    if ts is None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    col_map = {"发文时间": 3, "失败时间": 4, "补发成功时间": 5}
    col = col_map.get(event_type)
    if col is None:
        return
    try:
        _ensure_summary_excel()
        wb = openpyxl.load_workbook(SUMMARY_EXCEL)
        ws_s = wb["发文汇总"] if "发文汇总" in wb.sheetnames else wb.create_sheet("发文汇总")
        if ws_s.max_row == 0:
            ws_s.append(_SUMMARY_HEADERS)
        target_row = None
        for r_idx in range(2, ws_s.max_row + 1):
            a = ws_s.cell(r_idx, 1).value
            b = ws_s.cell(r_idx, 2).value
            if a and str(a).strip() == name and b is not None and int(b) == round_num:
                target_row = r_idx
                break
        if target_row is None:
            ws_s.append([name, round_num, "", "", ""])
            target_row = ws_s.max_row
        ws_s.cell(target_row, col, ts)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass

def _clear_round_sheets():
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        for sn in ("本轮已发", "失败列表"):
            if sn in wb.sheetnames:
                ws_c = wb[sn]
                if ws_c.max_row > 1:
                    ws_c.delete_rows(2, ws_c.max_row - 1)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _sort_summary_by_account():
    try:
        if not os.path.exists(SUMMARY_EXCEL):
            return
        wb = openpyxl.load_workbook(SUMMARY_EXCEL)
        if "发文汇总" not in wb.sheetnames:
            wb.close(); return
        ws_s = wb["发文汇总"]
        data = []
        for r_idx in range(2, ws_s.max_row + 1):
            row = [ws_s.cell(r_idx, c).value for c in range(1, 6)]
            if row[0]:
                data.append(row)
        data.sort(key=lambda r: (str(r[0]), int(r[1]) if r[1] is not None else 0))
        if ws_s.max_row > 1:
            ws_s.delete_rows(2, ws_s.max_row - 1)
        for row in data:
            ws_s.append(row)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass


# ========== 硬终止账号 sheet ==========

def _append_hard_terminate(name, reason, count_so_far):
    """命中 4 类硬终止时,写入"硬终止账号"sheet。永久放弃,不再尝试。"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "硬终止账号" not in wb.sheetnames:
            ws_h = wb.create_sheet("硬终止账号")
            ws_h.append(_HARD_TERMINATE_HEADERS)
        else:
            ws_h = wb["硬终止账号"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_h.append([name, reason, ts, count_so_far])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _read_whitelist_with_quota():
    """读白名单 A 列+B 列。返回 [(name, quota), ...]，B 列空默认 1"""
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "白名单" not in wb.sheetnames:
            wb.close(); return []
        ws_r = wb["白名单"]
        result = []
        for row in ws_r.iter_rows(min_row=2, max_col=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            q = 1
            if len(row) > 1 and row[1] is not None:
                try:
                    q = max(1, int(row[1]))
                except Exception:
                    q = 1
            result.append((name, q))
        wb.close()
        return result
    except Exception:
        return []

# ========== 日志 ==========

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    sys.stdout.buffer.write((line + "\n").encode("utf-8", errors="replace"))
    sys.stdout.buffer.flush()
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ========== 状态文件 (实时刷新) ==========

STATUS_FILE = os.path.join(BASE_DIR, "当前状态.txt")

def write_status(big_round=None, sub_round=None, phase=None,
                 total_done=0, total_target=0, doc_pool_size=0,
                 sub_failed=0, dead_terminated=None, extra=""):
    """实时刷新状态文件,你随时打开就能看到当前进度"""
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        lines = [f"[更新于 {ts}]"]
        if big_round is not None:
            lines.append(f"当前: 第 {big_round} 大循环 / 第 {sub_round} 小轮 / {phase}")
        lines.append(f"累计已发: {total_done} / {total_target} 篇")
        lines.append(f"文档池剩余: {doc_pool_size} 篇")
        if sub_failed:
            lines.append(f"本小轮失败队列: {sub_failed} 个账号")
        dead_terminated = dead_terminated or {}
        lines.append(f"硬终止: {len(dead_terminated)} 个" +
                     (f" {list(dead_terminated.keys())}" if dead_terminated else ""))
        if extra:
            lines.append(extra)
        with open(STATUS_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
    except Exception:
        pass


# ========== 失败记录 Excel ==========

def init_fail_excel():
    if not os.path.exists(FAIL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "失败记录"
        headers = ["时间", "账号名", "失败原因"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C0392B")
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        wb.save(FAIL_FILE)


def log_failure(account_name, reason):
    try:
        wb = openpyxl.load_workbook(FAIL_FILE)
        ws = wb.active
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws.append([ts, account_name, reason])
        wb.save(FAIL_FILE)
    except Exception as e:
        log(f"  写入失败记录出错: {e}")


# ========== CDP 基础 ==========

def get_tabs():
    return requests.get(f"{CDP_URL}/json", timeout=5, proxies=NO_PROXY).json()


def get_main_ws_url():
    tabs = get_tabs()
    for t in tabs:
        if "czgts.cn" in t.get("url", "") and "webSocketDebuggerUrl" in t:
            return t["webSocketDebuggerUrl"]
    raise RuntimeError(f"找不到主窗口，请确认创作罐头已用 --remote-debugging-port={CDP_URL.split(':')[-1]} 启动")


def ws_connect(url, timeout=10):
    return websocket.create_connection(url, timeout=timeout, **WS_OPTS)


def cdp(ws, method, params=None, mid=1):
    ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
    deadline = time.time() + 15
    while time.time() < deadline:
        try:
            d = json.loads(ws.recv())
            if d.get("id") == mid:
                return d.get("result", {})
        except:
            pass
    return {}


def js(ws, expr, mid=99):
    r = cdp(ws, "Runtime.evaluate", {"expression": expr}, mid)
    return r.get("result", {}).get("value")


def click(ws, x, y, mid):
    p = {"button": "left", "clickCount": 1, "x": x, "y": y,
         "modifiers": 0, "timestamp": time.time() * 1000}
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mousePressed", **p}, mid)
    time.sleep(0.12)
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mouseReleased", **p}, mid + 1)


def click_text(ws, text, mid, area_top=None, area_bottom=None):
    text_json = json.dumps(text)
    area_check = ""
    if area_top is not None:
        area_check += f" && r.top >= {area_top}"
    if area_bottom is not None:
        area_check += f" && r.top < {area_bottom}"
    v = js(ws, f"""
    (function(){{
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){{
            var t = all[i].textContent.trim();
            if(t === {text_json} && all[i].children.length <= 2){{
                var r = all[i].getBoundingClientRect();
                if(r.width > 0{area_check})
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, mid)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], mid + 1)
        return True
    return False


# ========== 打开筛选面板 ==========

def _open_filter_panel(main_ws):
    """点击筛选图标，返回是否成功打开"""
    v = js(main_ws, """
    (function(){
        var candidates = document.querySelectorAll('[class*="filter"],[class*="Filter"],[class*="funnel"],[class*="screen"]');
        for(var i=0;i<candidates.length;i++){
            var r = candidates[i].getBoundingClientRect();
            if(r.width > 0 && r.width < 60 && r.top > 0 && r.top < 200)
                return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
        }
        var search = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(search){
            var r = search.getBoundingClientRect();
            return JSON.stringify({x:Math.round(r.right+20), y:Math.round(r.top+r.height/2)});
        }
        return null;
    })()
    """, 102)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 103)
        time.sleep(1.2)
        return True
    return False


def _close_filter_panel(main_ws):
    cdp(main_ws, "Input.dispatchKeyEvent", {
        "type": "keyDown", "key": "Escape", "code": "Escape", "windowsVirtualKeyCode": 27
    }, 109)
    time.sleep(0.5)


# ========== 交互式标签选择 ==========

def get_available_tags(main_ws):
    """
    打开筛选面板，读取所有可用标签列表。
    返回标签名列表，如读取失败返回空列表。
    """
    log("读取可用标签列表...")

    click_text(main_ws, "账号管理", 101, area_top=0, area_bottom=800)
    time.sleep(1.5)

    opened = _open_filter_panel(main_ws)
    if not opened:
        log("  警告: 未能打开筛选面板，跳过标签读取")
        return []

    tags_raw = js(main_ws, r"""
    (function(){
        var SKIP = ['全部','确定','取消','清除','清除筛选','确认','搜索','标签','分组',
                    '账号类型','平台','状态','按名称','排序','筛选'];
        // 尝试找筛选弹出容器
        var panel = document.querySelector(
            '[class*="popover"],[class*="Popover"],[class*="dropdown"],[class*="Dropdown"],' +
            '[class*="filter-panel"],[class*="filterPanel"],[class*="FilterPanel"]'
        );
        var root = panel || document.body;
        var items = root.querySelectorAll('span,label,li');
        var tags = [];
        for(var i=0;i<items.length;i++){
            var t = items[i].textContent.trim();
            if(!t || t.length < 1 || t.length > 20) continue;
            if(items[i].children.length > 0) continue;
            if(SKIP.indexOf(t) !== -1) continue;
            if(/^\d+$/.test(t)) continue;
            var r = items[i].getBoundingClientRect();
            if(r.width > 0 && r.top > 100) tags.push(t);
        }
        return JSON.stringify([...new Set(tags)]);
    })()
    """, 108)

    _close_filter_panel(main_ws)

    try:
        tags = json.loads(tags_raw) if tags_raw else []
    except:
        tags = []

    log(f"  读取到 {len(tags)} 个标签: {tags}")
    return tags


def select_tag_interactively(main_ws):
    """
    从创作罐头读取标签列表，让用户交互选择。
    返回选定的标签名。
    """
    tags = get_available_tags(main_ws)

    print("\n" + "=" * 45)
    if tags:
        print("检测到以下标签：")
        for i, tag in enumerate(tags, 1):
            print(f"  {i}. {tag}")
        print(f"\n默认标签: {DEFAULT_TAG}")
        print("请输入标签序号 (直接回车使用默认):")
        try:
            choice = input("> ").strip()
        except EOFError:
            choice = ""
        if choice.isdigit() and 1 <= int(choice) <= len(tags):
            selected = tags[int(choice) - 1]
        else:
            selected = DEFAULT_TAG
    else:
        print(f"未能读取标签列表，使用默认标签: {DEFAULT_TAG}")
        selected = DEFAULT_TAG

    print(f"已选标签: {selected}")
    print("=" * 45 + "\n")
    log(f"本次使用标签: {selected}")
    return selected


# ========== 标签筛选 ==========

def setup_tag_filter(main_ws, tag_name):
    log(f"设置标签筛选: {tag_name}")

    ok = click_text(main_ws, "账号管理", 101, area_top=0, area_bottom=800)
    if not ok:
        log("  警告: 未找到账号管理导航，尝试继续")
    time.sleep(1.5)

    _open_filter_panel(main_ws)

    ok = click_text(main_ws, "清除筛选", 104)
    if ok:
        time.sleep(0.8)
        log("  已清除筛选")

    tag_json = json.dumps(tag_name)
    v = js(main_ws, f"""
    (function(){{
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){{
            var t = all[i].textContent.trim();
            if(t === {tag_json}){{
                var r = all[i].getBoundingClientRect();
                if(r.width > 0)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 105)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 106)
        time.sleep(1)
        log(f"  已勾选标签: {tag_name}")
    else:
        log(f"  警告: 未找到标签 [{tag_name}]，将处理所有可见账号")
        _close_filter_panel(main_ws)
        return False

    _close_filter_panel(main_ws)
    return True


# ========== 收集账号列表 ==========

def collect_accounts(main_ws):
    log("开始收集账号列表...")

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 110)
    time.sleep(0.8)

    accounts = []
    seen = set()
    same_count = 0
    has_scrolled = False  # 曾经成功滚动过才开始计退出条件

    for _ in range(1000):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            var names = [];
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t) names.push(t);
            }}
            return JSON.stringify(names);
        }})()
        """, 111)

        if v:
            names = json.loads(v)
            for n in names:
                if n not in seen:
                    seen.add(n)
                    accounts.append(n)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no';
            var before = c.scrollTop;
            c.scrollTop += 200;
            return before + '->' + c.scrollTop;
        })()
        """, 112)
        time.sleep(0.3)

        if result and result != "no":
            parts = result.split("->")
            if len(parts) == 2:
                before_top = parts[0].strip()
                after_top = parts[1].strip()
                if after_top != '0' and after_top != before_top:
                    has_scrolled = True
                    same_count = 0
                elif before_top == after_top and has_scrolled:
                    same_count += 1
                    if same_count >= 4:
                        break

    # 兜底:虚拟滚动列表,最后几个账号常因 lazy render 漏读
    # 强制滚到容器最底部,等 1.5s DOM 稳定再扫一次
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = c.scrollHeight;
    })()
    """, 113)
    time.sleep(1.5)
    v = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        var names = [];
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t) names.push(t);
        }}
        return JSON.stringify(names);
    }})()
    """, 114)
    if v:
        names = json.loads(v)
        added = 0
        for n in names:
            if n not in seen:
                seen.add(n)
                accounts.append(n)
                added += 1
        if added:
            log(f"  兜底:滚到底再扫一次,补收 {added} 个账号(虚拟滚动 lazy render 漏发)")

    log(f"共收集到 {len(accounts)} 个账号")
    return accounts


# ========== 滚动查找并点击账号 ==========

def scroll_find_account(main_ws, name):
    name_json = json.dumps(name)
    same_count = 0

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 10)
    time.sleep(0.5)

    for _ in range(500):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t === {name_json} || t.startsWith({name_json})){{
                    items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                    return 'found';
                }}
            }}
            return null;
        }})()
        """, 11)

        if v == 'found':
            time.sleep(0.3)
            pos = js(main_ws, f"""
            (function(){{
                var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    if(t === {name_json} || t.startsWith({name_json})){{
                        var r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        // 坐标不在视口内,再滚一次居中
                        items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                        r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                    }}
                }}
                return null;
            }})()
            """, 13)
            if pos:
                return json.loads(pos)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no-container';
            var before = c.scrollTop;
            c.scrollTop += 250;
            return before + '->' + c.scrollTop;
        })()
        """, 12)
        time.sleep(0.25)

        if result and result != "no-container":
            parts = result.split("->")
            if len(parts) == 2 and parts[0] == parts[1]:
                same_count += 1
                if same_count >= 5:
                    break
            else:
                same_count = 0

    # 兜底 1:搜索框过滤(参考 Mac 版,用 JS nativeInputValueSetter 触发 React 友好的 input)
    # 滚动找不到时,把账号名键入侧边栏搜索框,列表自动过滤后再定位
    log(f"  滚动未找到,尝试搜索框过滤: {name}")
    search_filled = js(main_ws, f"""
    (function(){{
        var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(!s) return null;
        var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        setter.call(s, {name_json});
        s.dispatchEvent(new Event('input', {{ bubbles: true }}));
        s.dispatchEvent(new Event('change', {{ bubbles: true }}));
        return 'filled';
    }})()
    """, 14)
    if search_filled == 'filled':
        time.sleep(1.5)
        pos = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t === {name_json} || t.startsWith({name_json})){{
                    items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                    var r = items[i].getBoundingClientRect();
                    if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                        return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                }}
            }}
            return null;
        }})()
        """, 16)
        # 不管成败,清空搜索框恢复完整列表
        js(main_ws, """
        (function(){
            var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
            if(!s) return;
            var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(s, '');
            s.dispatchEvent(new Event('input', { bubbles: true }));
            s.dispatchEvent(new Event('change', { bubbles: true }));
        })()
        """, 17)
        time.sleep(0.5)
        if pos:
            log(f"  兜底1命中:搜索框过滤定位到 {name}")
            return json.loads(pos)

    # 兜底 2:虚拟滚动 + lazy render → 强制滚到容器底部再找
    # 等 1.5s DOM 稳定,scrollIntoView block:'nearest';y 钳制在视口内防止点击落空
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = c.scrollHeight;
    })()
    """, 18)
    time.sleep(1.5)
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'nearest', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0){{
                    var cx = Math.round(r.left + r.width/2);
                    var cy = Math.round(Math.max(5, Math.min(r.top + r.height/2, window.innerHeight - 5)));
                    return JSON.stringify({{x:cx, y:cy}});
                }}
            }}
        }}
        return null;
    }})()
    """, 19)
    if pos:
        log(f"  兜底2命中:滚到底再找,定位到 {name}")
        time.sleep(0.3)
        return json.loads(pos)
    return None


# ========== webview 精确匹配（含3次重试） ==========

def _find_webview_once(main_ws, name):
    partition = js(main_ws, """
    (function(){
        var webviews = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<webviews.length;i++){
            var p = webviews[i].getAttribute('partition');
            if(!p || !p.startsWith('persist:')) continue;
            var r = webviews[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = p; }
        }
        return best;
    })()
    """, 15)

    if not partition:
        log("  警告: 未找到 partition")
        return None

    marker = f"_mk{random.randint(100000, 999999)}"
    r = cdp(main_ws, "Runtime.evaluate", {
        "expression": f"""
        new Promise(function(resolve){{
            var wv = document.querySelector('webview[partition="{partition}"]');
            if(!wv){{ resolve('no_wv'); return; }}
            wv.executeJavaScript('window.{marker}=1').then(function(){{
                resolve('ok');
            }}).catch(function(){{
                resolve('err');
            }});
        }})
        """,
        "awaitPromise": True,
        "timeout": 8000
    }, 200)

    inject_result = r.get("result", {}).get("value")
    if inject_result != "ok":
        log(f"  警告: 标记注入结果={inject_result}")
        return None

    tabs = get_tabs()
    for t in tabs:
        if "webSocketDebuggerUrl" not in t:
            continue
        try:
            wsc = ws_connect(t["webSocketDebuggerUrl"], timeout=3)
            wsc.send(json.dumps({"id": 1, "method": "Runtime.evaluate",
                                 "params": {"expression": f"window.{marker}===1"}}))
            deadline = time.time() + 3
            found = False
            while time.time() < deadline:
                d = json.loads(wsc.recv())
                if d.get("id") == 1:
                    found = d.get("result", {}).get("result", {}).get("value") is True
                    break
            wsc.close()
            if found:
                log("  webview 匹配成功")
                return t["webSocketDebuggerUrl"]
        except:
            pass

    return None


def find_account_webview(main_ws, name):
    for retry in range(3):
        result = _find_webview_once(main_ws, name)
        if result:
            return result
        if retry < 2:
            log(f"  重试 webview ({retry+2}/3)...")
            time.sleep(2)
    return None


def _search_box_set(main_ws, value):
    """侧边栏搜索框设置值(空字符串=清空)。React-friendly: 用原型 setter 触发 input/change 事件。"""
    val_json = json.dumps(value)
    return js(main_ws, f"""
    (function(){{
        var inputs = document.querySelectorAll('input');
        for(var i=0;i<inputs.length;i++){{
            var ph = inputs[i].getAttribute('placeholder') || '';
            if(ph.indexOf('账号') !== -1 || ph.indexOf('手机号') !== -1){{
                var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                setter.call(inputs[i], '');
                inputs[i].dispatchEvent(new Event('input', {{bubbles:true}}));
                if({val_json}.length > 0){{
                    setter.call(inputs[i], {val_json});
                    inputs[i].dispatchEvent(new Event('input', {{bubbles:true}}));
                    inputs[i].dispatchEvent(new Event('change', {{bubbles:true}}));
                }}
                return 'ok';
            }}
        }}
        return 'no_input';
    }})()
    """, 50)


def _locate_filtered_account(main_ws, name):
    """搜索框过滤后,直接抓第一个匹配账号的中心坐标(viewport 相对)。"""
    name_json = json.dumps(name)
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 51)
    if not pos:
        return None
    return json.loads(pos)


def find_or_reopen_webview(main_ws, name, reopen_attempts=2):
    """虚拟滚动底部账号 webview partition 不渲染时,通过侧边栏搜索框定位账号让 webview 重建。

    流程:正常 find_account_webview 失败 → 关 tab → 搜索框输入账号名 → 等过滤 →
    抓过滤后的账号坐标 → click → 等更长时间 → 再尝试 find_account_webview。
    搜索框过滤后只剩匹配项渲染在 DOM 顶部,不再受虚拟滚动影响。
    """
    ws_url = find_account_webview(main_ws, name)
    if ws_url:
        return ws_url

    for attempt in range(reopen_attempts):
        log(f"  webview partition 失败,搜索框重建 {attempt+1}/{reopen_attempts}: 输入 \"{name}\"")
        try:
            close_current_tab(main_ws)
        except Exception as e:
            log(f"  关 tab 异常(忽略): {e}")
        time.sleep(1.0)

        # 用搜索框过滤
        rs = _search_box_set(main_ws, name)
        if rs != 'ok':
            log("  搜索框定位失败,降级用 scroll_find_account")
            pos = scroll_find_account(main_ws, name)
        else:
            time.sleep(1.5)  # 等过滤结果 React 渲染完
            pos = _locate_filtered_account(main_ws, name)
            if not pos:
                log("  搜索过滤后仍找不到,降级用 scroll_find_account")
                _search_box_set(main_ws, "")  # 清空恢复列表
                time.sleep(0.5)
                pos = scroll_find_account(main_ws, name)

        if not pos:
            log(f"  搜索/滚动都找不到 {name},重建中止")
            _search_box_set(main_ws, "")
            return None

        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD + 2)  # 比首次多等 2 秒,给虚拟滚动 lazy render 留余地

        ws_url = find_account_webview(main_ws, name)

        # 不论成败,清空搜索框还原列表(避免影响后续账号)
        _search_box_set(main_ws, "")
        time.sleep(0.3)

        if ws_url:
            log(f"  webview 重建成功(尝试 {attempt+1})")
            return ws_url

    return None


def get_url_from_ws(ws_url):
    try:
        wsc = ws_connect(ws_url, timeout=4)
        v = js(wsc, "location.href", 1)
        wsc.close()
        return v or ""
    except:
        return ""


# ========== 系统通知检测 ==========

def check_system_notice(ws_url, account_name):
    """
    导航到消息中心 → 分别点击系统通知、审核通知频道
    → 读取当天+昨天的消息原文，识别违规关键词
    → 通知写入 NOTICE_FILE；违规写入 VIOLATION_FILE
    → 返回 (notice_count, violation_count)
    """
    try:
        today = datetime.now()
        today_full  = today.strftime("%Y-%m-%d")
        today_short = today.strftime("%m-%d")

        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/personal/message?type=message_letter'", 300)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        notices = []

        for channel in ["系统通知", "审核通知"]:
            channel_json = channel.replace('"', '\\"')
            clicked = js(wsc, f"""
            (function(){{
                var spans = document.querySelectorAll('span.name');
                for(var i=0;i<spans.length;i++){{
                    if(spans[i].textContent.trim() === "{channel_json}"){{
                        var box = spans[i].closest('.conversation-box-primary') || spans[i].parentElement;
                        if(box){{ box.click(); return 'ok'; }}
                    }}
                }}
                return null;
            }})()
            """, 301)

            if not clicked:
                log(f"  未找到频道: {channel}")
                continue

            time.sleep(2.5)

            result = js(wsc, f"""
            (function(){{
                var keys = ["{today_full}","{today_short}"];
                var list = document.querySelector('.chat-container-list');
                if(!list) return JSON.stringify([]);
                var items = list.children;
                var results = [];
                var matched = false;
                for(var i=0;i<items.length;i++){{
                    var cls = items[i].className || '';
                    if(cls.indexOf('time-stamp') !== -1){{
                        var t = items[i].textContent.trim();
                        matched = false;
                        for(var k=0;k<keys.length;k++){{
                            if(t.indexOf(keys[k]) !== -1){{ matched = true; break; }}
                        }}
                    }} else if(matched && cls.indexOf('chat-row') !== -1){{
                        var txt = items[i].textContent.trim();
                        if(txt) results.push(txt);
                    }}
                }}
                return JSON.stringify(results);
            }})()
            """, 302)

            if result:
                try:
                    msgs = json.loads(result)
                    for msg in msgs:
                        notices.append((channel, msg))
                except Exception:
                    pass

        wsc.close()

        violations = []
        for ch, msg in notices:
            hit_cats = []
            for cat, kws in VIOLATION_KEYWORDS.items():
                for kw in kws:
                    if kw in msg:
                        hit_cats.append(cat)
                        break
            if hit_cats:
                violations.append((ch, msg, hit_cats))

        if notices:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            content_str = f"[{ts}] 账号 {account_name} 当天通知 {len(notices)} 条:\n"
            for ch, msg in notices:
                content_str += f"  【{ch}】{msg}\n"
            content_str += "\n"
            with open(NOTICE_FILE, "a", encoding="utf-8") as f:
                f.write(content_str)
            log(f"  ⚠ 当天通知 {len(notices)} 条 → 已写入系统通知.txt")
        else:
            log("  系统/审核通知: 当天无新通知")

        if violations:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            vstr = f"[{ts}] 账号 {account_name} 命中违规 {len(violations)} 条:\n"
            for ch, msg, cats in violations:
                vstr += f"  【{ch}】[{','.join(cats)}] {msg}\n"
            vstr += "\n"
            with open(VIOLATION_FILE, "a", encoding="utf-8") as f:
                f.write(vstr)
            log(f"  ★★ 违规命中 {len(violations)} 条 → 已写入违规提醒.txt")

        return len(notices), len(violations)
    except Exception as e:
        log(f"  系统通知检测出错: {e}")
        return 0, 0

# ========== 阅读量检测 ==========

def check_reading_stats(ws_url, account_name):
    """
    导航到数据概览页 → 读取三篇文章的阅读量。
    任一超过 ALERT_THRESHOLD 则写入 ALERT_FILE。
    """
    try:
        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/index'", 310)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        reads_raw = js(wsc, """
        (function(){
            var items = document.querySelectorAll('.recent-works-item');
            var results = [];
            for(var i=0;i<items.length;i++){
                var timeEl = items[i].querySelector('span.time');
                var timeText = timeEl ? timeEl.textContent.trim() : '未知时间';
                var labels = items[i].querySelectorAll('span.label');
                for(var j=0;j<labels.length;j++){
                    if(labels[j].textContent.trim() === '阅读量'){
                        var prev = labels[j].previousElementSibling;
                        if(prev){
                            var num = parseInt(prev.textContent.trim().replace(/,/g,''), 10);
                            if(!isNaN(num)) results.push({count: num, time: timeText});
                        }
                    }
                }
            }
            return JSON.stringify(results);
        })()
        """, 311)
        wsc.close()

        reads = json.loads(reads_raw) if reads_raw else []
        if not reads:
            log("  阅读量: 未读取到数据")
            return 0

        counts = [r['count'] for r in reads]
        log(f"  阅读量: {counts}")
        high = [r for r in reads if r['count'] >= ALERT_THRESHOLD]
        if high:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            lines = [f"[{ts}] 账号 {account_name} 高阅读量:"]
            for r in high:
                lines.append(f"  {r['time']} — {r['count']} 阅读量")
            msg = '\n'.join(lines) + '\n\n'
            with open(ALERT_FILE, "a", encoding="utf-8") as f:
                f.write(msg)
            log(f"  ★ 高阅读量提醒: {[r['count'] for r in high]} → 已写入高阅读提醒.txt")
        return len(high)
    except Exception as e:
        log(f"  阅读量检测出错: {e}")
        return 0


# ========== 账号状态检测 ==========

ERROR_KEYWORDS = {
    "失登": ["请登录", "登录已失效", "账号已下线", "重新登录"],
    "封号": ["账号已被封禁", "账号异常", "账号被封", "账号不可用"],
    "禁言": ["账号被禁言", "发言受限", "无法发布"],
}


def detect_account_error(wsc):
    page_text = js(wsc, "document.body.innerText || ''", 70) or ""
    for reason, keywords in ERROR_KEYWORDS.items():
        for kw in keywords:
            if kw in page_text:
                return reason
    return None


# ========== 弹窗 & Tab ==========

def close_popup(ws):
    v = js(ws, """
    (function(){
        var b = document.querySelector('.close-btn,[class*="close-btn"]');
        if(b){var r=b.getBoundingClientRect();
        if(r.width>0) return JSON.stringify({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});}
        return null;
    })()
    """, 50)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], 51)
        time.sleep(0.5)


def close_current_tab(main_ws):
    v = js(main_ws, """
    (function(){
        var closes = document.querySelectorAll('.chrome-tab-close');
        if(closes.length === 0) return null;
        for(var i=0;i<closes.length;i++){
            var tab = closes[i].closest('.chrome-tab');
            if(tab && tab.classList.contains('active')){
                var r = closes[i].getBoundingClientRect();
                if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
        }
        var last = closes[closes.length-1];
        var r = last.getBoundingClientRect();
        return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
    })()
    """, 55)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 56)
        time.sleep(0.5)
        log("  已关闭tab")
    else:
        log("  未找到tab关闭按钮")


# ========== 读取文档 ==========

def read_docx_text(doc_path):
    doc = docxlib.Document(doc_path)
    lines = [p.text for p in doc.paragraphs if p.text.strip()]
    return '\n'.join(lines)


# ========== 发布流程 ==========

def publish_article(ws_url, doc_path, main_ws, name="", _credit_out=None):

    try:
        wsc = ws_connect(ws_url, timeout=10)
    except Exception as e:
        return False, f"连接失败: {e}"

    close_popup(wsc)
    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    # 导航到图文发布页
    js(wsc, "location.href='https://mp.toutiao.com/profile_v4/graphic/publish'", 60)
    wsc.close()
    time.sleep(4)

    # 导航后 target id 会变，从 tabs 里重新找 graphic/publish 页面
    new_ws_url = None
    for _ in range(12):
        tabs = get_tabs()
        for t in tabs:
            if "graphic/publish" in t.get("url","") and "webSocketDebuggerUrl" in t:
                new_ws_url = t["webSocketDebuggerUrl"]
                break
        if new_ws_url:
            break
        time.sleep(0.5)

    if not new_ws_url:
        return False, "导航到发文页失败"

    try:
        wsc = ws_connect(new_ws_url, timeout=10)
    except Exception as e:
        return False, f"重连失败: {e}"

    close_popup(wsc)
    time.sleep(1)

    current_url = js(wsc, "location.href", 59) or ""
    if "login" in current_url:
        wsc.close()
        return False, "失登"

    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    import win32gui, win32con, win32api, threading

    # 取 webview 真实屏幕坐标（从主窗口查）
    wv_s = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 61)
    if not wv_s:
        wsc.close()
        return False, "无法获取 webview 屏幕坐标"
    wv0 = json.loads(wv_s)

    # 找文档导入按钮并真实鼠标点击
    v_import = None
    for _ in range(20):
        v_import = js(wsc, """
        (function(){
            var btns = Array.from(document.querySelectorAll('.syl-toolbar-button')).filter(function(b){
                return b.getBoundingClientRect().width > 0;
            });
            if(btns.length > 0){
                var last = btns[btns.length - 1];
                var r = last.getBoundingClientRect();
                return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
            return null;
        })()
        """, 62)
        if v_import:
            break
        time.sleep(0.5)
    if not v_import:
        wsc.close()
        return False, "找不到文档导入按鈕"
    pi = json.loads(v_import)
    import_x = wv0['sx'] + pi['x']
    import_y = wv0['sy'] + pi['y']
    log(f"  文档导入按钮CSS坐标:({pi['x']},{pi['y']}) webview原点:({wv0['sx']},{wv0['sy']}) => 屏幕:({import_x},{import_y})")

    # 置顶窗口，点编辑区聚焦，再点文档导入
    import ctypes as _ct
    _u32 = _ct.windll.user32
    import psutil
    cdp_port_num = int(CDP_URL.split(":")[-1])
    electron_pid = None
    for conn in psutil.net_connections(kind='tcp'):
        if conn.laddr.port == cdp_port_num and conn.status == 'LISTEN':
            electron_pid = conn.pid
            break
    if electron_pid:
        def _enum_top_root(h, _):
            pid_buf = _ct.c_ulong(0)
            _u32.GetWindowThreadProcessId(h, _ct.byref(pid_buf))
            if pid_buf.value == electron_pid and win32gui.IsWindowVisible(h):
                _u32.SetForegroundWindow(h)
            return True
        win32gui.EnumWindows(_enum_top_root, None)
    time.sleep(0.5)

    # [v1101 P5] verify frontmost 是罐头, 不是就重试 3 次
    if electron_pid:
        for _retry in range(3):
            fg_hwnd = _u32.GetForegroundWindow()
            fg_pid_buf = _ct.c_ulong(0)
            _u32.GetWindowThreadProcessId(fg_hwnd, _ct.byref(fg_pid_buf))
            if fg_pid_buf.value == electron_pid:
                break
            log(f"  [P5] frontmost 不是罐头(pid={fg_pid_buf.value}), 重试 {_retry+1}/3")
            win32gui.EnumWindows(_enum_top_root, None)
            time.sleep(0.3)

    # [v1101 P6] 飘屏外兜底:webview 原点 < 0 → SetWindowPos 拉回 + 重读
    if wv0.get('sx', 0) < 0 or wv0.get('sy', 0) < 0:
        log(f"  ⚠ webview 飘屏外 ({wv0['sx']},{wv0['sy']}),拉回 (100,100)")
        if electron_pid:
            def _setpos(h, _):
                pid_buf = _ct.c_ulong(0)
                _u32.GetWindowThreadProcessId(h, _ct.byref(pid_buf))
                if pid_buf.value == electron_pid and win32gui.IsWindowVisible(h):
                    _u32.SetWindowPos(h, 0, 100, 100, 1600, 1000, 0)
                return True
            win32gui.EnumWindows(_setpos, None)
        time.sleep(1)
        wv_re = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxArea = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var area = r.width * r.height;
                if(area > maxArea){ maxArea = area; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
        })()
        """, 99)
        if wv_re:
            wv0 = json.loads(wv_re)
            import_x = wv0['sx'] + pi['x']
            import_y = wv0['sy'] + pi['y']
            log(f"  拉回后 webview 原点: ({wv0['sx']},{wv0['sy']}) → 导入坐标 ({import_x},{import_y})")

    # [v1101 P7] click 文档导入 + 等弹窗,失败重试 3 次
    sel = None
    for click_attempt in range(3):
        attempt_str = f" [第{click_attempt+1}次]" if click_attempt > 0 else ""
        win32api.SetCursorPos((import_x, import_y))
        time.sleep(0.3)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        time.sleep(0.05)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
        log(f"  cliclick 点击文档导入 ({import_x},{import_y}){attempt_str}")
        time.sleep(1.5)
        for _ in range(60):
            sel = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '选择文档'){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({bx: Math.round(r.left+r.width/2), by: Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 65)
            if sel:
                break
            time.sleep(0.5)
        if sel:
            if click_attempt > 0:
                log(f"  P7 第{click_attempt+1}次成功唤出弹窗")
            break
        if click_attempt < 2:
            log(f"  弹窗未出,重 SetForegroundWindow + 重读坐标后重试")
            if electron_pid:
                win32gui.EnumWindows(_enum_top_root, None)
            time.sleep(0.6)
            wv_re2 = js(main_ws, """
            (function(){
                var wvs = document.querySelectorAll('webview');
                var maxArea = 0, best = null;
                for(var i=0;i<wvs.length;i++){
                    var r = wvs[i].getBoundingClientRect();
                    var area = r.width * r.height;
                    if(area > maxArea){ maxArea = area; best = r; }
                }
                if(!best) return null;
                return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
            })()
            """, 99)
            if wv_re2:
                wv0 = json.loads(wv_re2)
                import_x = wv0['sx'] + pi['x']
                import_y = wv0['sy'] + pi['y']
                log(f"  重试前坐标更新: ({import_x},{import_y})")
    if not sel:
        wsc.close()
        return False, "文档导入弹窗未出现(3 次 click 重试均失败)"

    sb = json.loads(sel)

    # 重新取webview坐标
    wv_s3 = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 66)
    wv1 = json.loads(wv_s3) if wv_s3 else wv0
    screen_x = wv1['sx'] + sb['bx']
    screen_y = wv1['sy'] + sb['by']
    log(f"  选择文档屏幕坐标: ({screen_x},{screen_y})")

    # 点选择文档，等对话框拿到焦点，直接发路径（参照Mac版逻辑：不检测窗口，固定等待后发键盘）
    win32api.SetCursorPos((screen_x, screen_y))
    time.sleep(0.3)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    time.sleep(0.05)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    log(f"  真实鼠标点击选择文档 ({screen_x},{screen_y})")

    # 等文件对话框打开并拿到焦点
    time.sleep(1.5)

    # 剪贴板准备路径
    import subprocess as _sp
    _safe_path = doc_path.replace("'", "''")
    _sp.run(['powershell', '-Command', f"Set-Clipboard -Value '{_safe_path}'"], capture_output=True)
    time.sleep(0.3)

    # Ctrl+A 全选文件名框 → Ctrl+V 粘贴路径 → Enter 确认 → Enter 再确认（同Mac版回车×2）
    for vk, down in [(0x11,True),(0x41,True),(0x41,False),(0x11,False)]:
        win32api.keybd_event(vk, 0, 0 if down else win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(0.1)
    for vk, down in [(0x11,True),(0x56,True),(0x56,False),(0x11,False)]:
        win32api.keybd_event(vk, 0, 0 if down else win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(0.3)
    win32api.keybd_event(0x0D, 0, 0, 0)
    win32api.keybd_event(0x0D, 0, win32con.KEYEVENTF_KEYUP, 0)
    log(f"  路径已填入: {doc_path}")

    time.sleep(3)

    # 等内容加载
    time.sleep(5)
    char_count = 0
    for _ in range(15):
        v = js(wsc, """
        (function(){
            // [v1101 P3] 取最长 ProseMirror(避免命中标题 placeholder 5 字)
            var els = document.querySelectorAll('.ProseMirror');
            var max = 0;
            for (var i = 0; i < els.length; i++) {
                var l = els[i].textContent.trim().length;
                if (l > max) max = l;
            }
            return max;
        })()
        """, 75)
        char_count = int(v) if v else 0
        if char_count >= 50:
            break
        time.sleep(0.8)

    log(f"  文章字数: {char_count}")
    # [v1101 P2] 字数<50 重试 fill_dialog 一次
    if char_count < 50:
        log(f"  对话框已关但字数仅 {char_count}（文档未真导入），重试 fill_dialog")
        try:
            result_holder[0] = None
            t = threading.Thread(target=fill_dialog, daemon=True)
            t.start()
            time.sleep(0.2)
            subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
            t.join(timeout=30)
        except NameError:
            pass  # Win 用 win32 file dialog 路径不同
        time.sleep(5)
        char_count = 0
        for _ in range(15):
            v = js(wsc, """
            (function(){
                var els = document.querySelectorAll('.ProseMirror');
                var max = 0;
                for (var i = 0; i < els.length; i++) {
                    var l = els[i].textContent.trim().length;
                    if (l > max) max = l;
                }
                return max;
            })()
            """, 75)
            char_count = int(v) if v else 0
            if char_count >= 50:
                break
            time.sleep(0.8)
        log(f"  重试后字数: {char_count}")
        if char_count < 50:
            wsc.close()
            return False, "文档导入后内容为空(重试1次仍空)"

    # 滚动到底部，等图片全部加载
    js(wsc, """
    (function(){
        var el = document.querySelector('.ProseMirror') || document.body;
        el.scrollTop = el.scrollHeight;
        window.scrollTo(0, document.body.scrollHeight);
    })()
    """, 76)
    time.sleep(2)
    for _ in range(20):
        loading = js(wsc, """
        (function(){
            var imgs = document.querySelectorAll('img');
            for(var i=0;i<imgs.length;i++){
                if(!imgs[i].complete || imgs[i].naturalWidth === 0) return true;
            }
            return false;
        })()
        """, 77)
        if not loading:
            break
        time.sleep(0.5)
    log("  图片加载完成")

    # 读取信用分，决定是否勾选头条首发
    credit_raw = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t.indexOf('\u4fe1\u7528\u5206') !== -1 && t.length < 20){
                var m = t.match(/(\\d{1,3})\u5206/);
                if(m){
                    var n = parseInt(m[1], 10);
                    if(n % 5 === 0 && n >= 5 && n <= 100) return n;
                }
            }
        }
        return null;
    })()
    """, 78)
    credit_score = int(credit_raw) if credit_raw is not None else None
    # [v1101.1] 信用分 < 60 → 硬终止
    if credit_score is not None and credit_score < 60:
        log(f"  ★ 信用分 {credit_score} < 60,硬终止")
        wsc.close()
        return False, "信用分过低"
    if credit_score is not None:
        alert = ""
        if credit_score < 80:
            alert = " ★★ 危险"
        elif credit_score < 90:
            alert = " ★ 警告"
        log(f"  信用分: {credit_score}{alert}")
    else:
        log("  信用分: 未读取到")
    if _credit_out is not None:
        try:
            _credit_out[name] = credit_score
        except Exception:
            pass
    should_first = (credit_score is not None and credit_score >= 95) and (name not in NOFIRST_ACCOUNTS)

    # [v1101.3] 头条首发复选框: 探测 cb 真坐标 + 点击 + 回读校验 + 三轮兜底 + 硬保护
    _PROBE_JS = r"""
    (function(){
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '头条首发'){
                var p = all[i].parentElement;
                var cb = null, isChecked = false, src = '';
                while(p && p.tagName !== 'BODY'){
                    var c = p.querySelector('input[type="checkbox"]');
                    if(c){ cb = c; isChecked = c.checked; src = 'input'; break; }
                    if(p.getAttribute('role') === 'checkbox'){ cb = p; isChecked = p.getAttribute('aria-checked')==='true'; src = 'role'; break; }
                    if(p.classList.contains('checked')){ cb = p; isChecked = true; src = 'class'; break; }
                    p = p.parentElement;
                }
                if(cb){
                    var r = cb.getBoundingClientRect();
                    var px = Math.round(r.left + r.width/2);
                    var py = Math.round(r.top + r.height/2);
                    if(r.width < 4 || r.height < 4){
                        var pp = cb.parentElement;
                        while(pp && pp.tagName !== 'BODY'){
                            var pr = pp.getBoundingClientRect();
                            if(pr.width >= 8 && pr.height >= 8 && pr.height < 60){
                                px = Math.round(pr.left + pr.width/2);
                                py = Math.round(pr.top + pr.height/2);
                                break;
                            }
                            pp = pp.parentElement;
                        }
                    }
                    return JSON.stringify({found:true, checked:isChecked, cb_x:px, cb_y:py, src:src});
                }
                var r2 = all[i].getBoundingClientRect();
                return JSON.stringify({found:false, checked:false, cb_x:Math.round(r2.left+r2.width/2), cb_y:Math.round(r2.top+r2.height/2), src:'text'});
            }
        }
        return null;
    })()
    """
    _CLICK_JS = r"""
    (function(){
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '头条首发'){
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){
                    var c = p.querySelector('input[type="checkbox"]');
                    if(c){ c.click(); try{ c.dispatchEvent(new Event('change',{bubbles:true})); }catch(e){} return 'input'; }
                    if(p.getAttribute('role') === 'checkbox'){ p.click(); return 'role'; }
                    p = p.parentElement;
                }
            }
        }
        return null;
    })()
    """

    def _probe_first():
        raw = js(wsc, _PROBE_JS, 79)
        if not raw:
            return None
        try:
            return json.loads(raw)
        except Exception:
            return None

    fr = _probe_first()
    if fr is None:
        log("  头条首发: 未找到复选框")
    elif bool(fr.get('checked')) == should_first:
        log(f"  头条首发: 已是{'勾选' if fr['checked'] else '未勾选'}，无需操作")
    else:
        attempts = []
        verified = False
        for attempt in range(1, 4):
            if attempt < 3 and fr.get('found'):
                sx = wv0['sx'] + fr['cb_x']
                sy = wv0['sy'] + fr['cb_y']
                win32api.SetCursorPos((sx, sy))
                time.sleep(0.2)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                time.sleep(0.05)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
                time.sleep(0.4)
                attempts.append(f"real#{attempt}@({sx},{sy})")
            else:
                rc = js(wsc, _CLICK_JS, 30)
                time.sleep(0.4)
                attempts.append(f"js#{attempt}={rc}")
            fr2 = _probe_first()
            if fr2 and bool(fr2.get('checked')) == should_first:
                log(f"  头条首发: {'勾选' if should_first else '取消勾选'} 已校验 [{'/'.join(attempts)}]")
                verified = True
                break
            fr = fr2 if fr2 else fr
        if not verified:
            actual = fr.get('checked') if fr else None
            log(f"  ✗ 头条首发校准失败: 目标={should_first} 实际={actual} 尝试=[{'/'.join(attempts)}]")
            if (not should_first) and actual is True:
                log(f"  ★ 硬保护: 应取消首发但仍勾选,跳过该篇避免扣 5 分")
                wsc.close()
                return False, "首发取消失败(硬保护)"


    # 点"预览并发布" + 等"确认发布"，最多重试5次
    confirm_clicked = False
    for attempt in range(36):
        wv_cur_r = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxArea = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var area = r.width * r.height;
                if(area > maxArea){ maxArea = area; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
        })()
        """, 79)
        wv_cur = json.loads(wv_cur_r) if wv_cur_r else wv0

        v = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim() === '\u9884\u89c8\u5e76\u53d1\u5e03' && !btns[i].disabled){
                    var r = btns[i].getBoundingClientRect();
                    if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                }
            }
            return null;
        })()
        """, 80)

        if not v:
            if attempt == 0:
                wsc.close()
                return False, "找不到预览并发布按钮"
        else:
            p = json.loads(v)
            preview_x = wv_cur['sx'] + p['x']
            preview_y = wv_cur['sy'] + p['y']
            attempt_str = f" [第{attempt+1}次]" if attempt > 0 else ""
            log(f"  cliclick 点击预览并发布 ({preview_x},{preview_y}){attempt_str}")
            win32api.SetCursorPos((preview_x, preview_y))
            time.sleep(0.5)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
            time.sleep(0.05)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
            log("  已点击预览并发布")
            time.sleep(4)

        for i in range(20):
            time.sleep(0.5)
            v2 = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    var t = btns[i].textContent.trim();
                    if(t === '\u786e\u8ba4\u53d1\u5e03' || t.indexOf('\u786e\u8ba4\u53d1\u5e03') !== -1){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 83)
            if v2:
                p2 = json.loads(v2)
                confirm_x = wv_cur['sx'] + p2['x']
                confirm_y = wv_cur['sy'] + p2['y']
                log(f"  cliclick 点击确认发布 ({confirm_x},{confirm_y})")
                win32api.SetCursorPos((confirm_x, confirm_y))
                time.sleep(0.3)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                time.sleep(0.05)
                win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
                confirm_clicked = True
                break

        if confirm_clicked:
            break
        if attempt < 4:
            log(f"  确认发布未出现，封面图可能未加载完，准备第{attempt+2}次点击预览并发布...")

    if not confirm_clicked:
        wsc.close()
        return False, "未出现确认发布按钮"

    # 检测发布成功
    published = False
    for _ in range(20):
        time.sleep(0.5)
        t2 = js(wsc, """
        (function(){
            var all = document.querySelectorAll('*');
            for(var i=0;i<all.length;i++){
                var t = all[i].textContent.trim();
                if((t==='发布成功！'||t==='发布成功'||t==='提交成功！'||t==='提交成功') && all[i].children.length<=1)
                    return t;
            }
            return null;
        })()
        """, 96)
        if t2:
            published = True
            break
        cur_url = js(wsc, "location.href", 97) or ""
        if "graphic" in cur_url and "publish" not in cur_url:
            published = True
            break

    if not published:
        err_after = detect_account_error(wsc)
        wsc.close()
        if err_after:
            return False, err_after
        return False, "未检测到发布成功"

    wsc.close()
    log("  OK 文章发布成功")
    return True, "成功"


# ========== 文档管理 ==========

def get_docs():
    docs = []
    for p in ["*.docx", "*.doc"]:
        docs.extend(glob.glob(os.path.join(DOCS_FOLDER, p)))
    return sorted([d for d in docs if "已发送" not in d])


# [v1101.4] doc_pool 实时校验 + 重扫工具,救"分发完源必删"导致罐头找不到文件
def _pick_doc(doc_pool):
    """从 doc_pool 抽一篇实存的 docx,失效引用就地清理。返回 None 表示池已空。"""
    while doc_pool:
        doc = random.choice(doc_pool)
        if os.path.exists(doc):
            return doc
        log(f"  ! 源已删除(可能被外部分发),从池剔除: {os.path.basename(doc)}")
        doc_pool.remove(doc)
    return None


def _resync_pool(doc_pool):
    """大循环开始前重扫素材池:剔除幽灵引用 + 加入新到的素材。返回 (剔除数, 新增数)。"""
    cur = set(get_docs())
    before = len(doc_pool)
    doc_pool[:] = [d for d in doc_pool if d in cur]
    removed = before - len(doc_pool)
    pool_set = set(doc_pool)
    new_docs = [d for d in cur if d not in pool_set]
    if new_docs:
        doc_pool.extend(new_docs)
    return removed, len(new_docs)


def move_to_sent(doc_path):
    os.makedirs(SENT_FOLDER, exist_ok=True)
    dest = os.path.join(SENT_FOLDER, os.path.basename(doc_path))
    if os.path.exists(dest):
        name, ext = os.path.splitext(os.path.basename(doc_path))
        dest = os.path.join(SENT_FOLDER, f"{name}_{int(time.time())}{ext}")
    shutil.move(doc_path, dest)
    log(f"  已移至已发送: {os.path.basename(dest)}")


# ========== 死磕主循环 (公共函数,timer.py 也调) ==========

def run_death_grip(
    accounts,                  # list[str] 账号顺序
    per_account_quota,         # dict[str, int]  每个账号本次目标 quota
    doc_pool,                  # list[str]       文档路径(会被 mutate,跑完是剩余)
    main_ws,                   # 主窗口 ws_connection
    sub_rounds=3,              # 1 大循环 = N 小轮(微头条 5,文章/定时 3)
    max_fail_per_sub=3,        # 本小轮内允许的失败次数
    sent_accounts_set=None,    # 中断恢复用,本轮已发
    credit_records=None,       # 信用分累积(dict,引用)
    initial_acc_count=None,    # 起步已发数(timer Stage 1 残余)
    initial_dead=None,         # 起步硬终止(timer Stage 1 已识别)
    log_label="",              # 日志前缀(用于区分 timer Stage 2 vs 普通 batch)
):
    """死磕循环:大循环 N 小轮 + 外层无限磕。
    - 4 类硬终止(失登/封号/禁言/侧边栏未找到) → 即刻永久放弃,写"硬终止账号"sheet
    - 其他失败 → 本小轮内累计 max_fail_per_sub 次跳过本小轮,下小轮恢复
    - N 小轮跑完 = 1 大循环;还有 quota 没满且 doc 没空 → 继续下一大循环(无限磕)
    返回: dict { acc_count, dead_terminated, doc_pool, ok_count, fail_count, total_notices, total_violations, total_alerts }
    """
    sent_accounts_set = sent_accounts_set if sent_accounts_set is not None else set()
    credit_records   = credit_records if credit_records is not None else {}
    acc_count        = dict(initial_acc_count) if initial_acc_count else {}
    for a in accounts:
        acc_count.setdefault(a, 0)
    dead_terminated  = dict(initial_dead) if initial_dead else {}

    ok_count = fail_count = 0
    total_notices = total_violations = total_alerts = 0
    big_round = 0
    notice_checked_set = set()  # [v1101.6] 每账号每天只读 1 次审核/系统通知

    def _do_publish(name, doc, round_label, is_retry=False):
        nonlocal ok_count, fail_count, total_notices, total_violations, total_alerts
        log(f"\n  {log_label}[剩余 {len(doc_pool)} 篇] {round_label} {'[补发]' if is_retry else ''} {name}  ->  {os.path.basename(doc)}")

        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未找到账号: {name}")
            log_failure(name, "侧边栏未找到")
            fail_count += 1
            return False, "侧边栏未找到"

        log(f"  点击坐标({pos['x']},{pos['y']})")
        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD)

        ws_url = find_or_reopen_webview(main_ws, name)
        if not ws_url:
            log("  X 找不到 webview")
            log_failure(name, "webview匹配失败")
            fail_count += 1
            close_current_tab(main_ws)
            return False, "webview匹配失败"

        page_url = get_url_from_ws(ws_url)
        if "login" in page_url:
            log("  X 账号失登")
            log_failure(name, "失登")
            fail_count += 1
            close_current_tab(main_ws)
            return False, "失登"

        # [v1101.6] 每账号每天只读 1 次审核/系统通知 — 后续发文跳过
        if name not in notice_checked_set:
            _nc, _vc = check_system_notice(ws_url, name)
            total_notices += _nc
            total_violations += _vc
            notice_checked_set.add(name)
            time.sleep(2)
        else:
            log(f"  系统/审核通知:{name} 本轮已读过,跳过")

        _d_wait = random.randint(8, 20)
        try:
            success, reason = publish_article(ws_url, doc, main_ws, name=name, _credit_out=credit_records)
            if success:
                move_to_sent(doc)
                if doc in doc_pool:
                    doc_pool.remove(doc)
                acc_count[name] = acc_count.get(name, 0) + 1
                ok_count += 1
                _a = check_reading_stats(ws_url, name)
                total_alerts += (_a or 0)
                return True, ""
            else:
                log(f"  X 发布失败: {reason}")
                log_failure(name, reason)
                fail_count += 1
                return False, reason
        except Exception as e:
            log(f"  X 异常: {e}")
            log_failure(name, f"异常: {e}")
            fail_count += 1
            return False, f"异常: {e}"
        finally:
            close_current_tab(main_ws)
            log(f"  篇间等待 {_d_wait} 秒...")
            time.sleep(_d_wait)

    def _is_eligible(name, sub_skipped):
        return (name not in dead_terminated and
                name not in sub_skipped and
                name not in sent_accounts_set and
                acc_count.get(name, 0) < per_account_quota.get(name, 0))

    def _handle_failure(name, reason, sub_fail_count, sub_skipped):
        """处理 _do_publish 返回的失败。命中 4 类硬终止即永久放弃。"""
        if reason in HARD_TERMINATE_REASONS:
            cnt_so_far = acc_count.get(name, 0)
            dead_terminated[name] = (reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cnt_so_far)
            _append_hard_terminate(name, reason, cnt_so_far)
            log(f"  ★ 4 类硬终止: {name} -> {reason} (本次已发 {cnt_so_far} 篇,永久放弃)")
            return True
        sub_fail_count[name] = sub_fail_count.get(name, 0) + 1
        if sub_fail_count[name] >= max_fail_per_sub:
            sub_skipped.add(name)
            log(f"  - {name} 本小轮 {max_fail_per_sub} 次失败,跳过本小轮(下小轮恢复)")
        return False

    def _total_target():
        return sum(per_account_quota.get(a, 0) for a in accounts if a not in dead_terminated)

    def _total_done():
        return sum(acc_count.get(a, 0) for a in accounts)

    while True:
        # 进大循环前先检查
        if not doc_pool:
            log(f"\n{log_label}文档池已空,死磕结束")
            break
        active = [a for a in accounts
                  if a not in dead_terminated
                  and acc_count.get(a, 0) < per_account_quota.get(a, 0)]
        if not active:
            log(f"\n{log_label}所有账号配额暂时满,等 60s 重扫素材池(只 doc_pool 空 / Ctrl+C 才停)...")
            time.sleep(60)
            cur_set = set(doc_pool)
            new_docs = [d for d in get_docs() if d not in cur_set]
            if new_docs:
                log(f"  + 发现 {len(new_docs)} 篇新素材,加入文档池")
                doc_pool.extend(new_docs)
            continue

        # [v1101.4] 大循环开头重扫池, 实时同步外部 mutate (scp+rm 等)
        _removed, _added = _resync_pool(doc_pool)
        if _removed or _added:
            log(f"  [v1101.4] doc_pool 重扫: 剔除 {_removed} 条幽灵引用, 新增 {_added} 篇素材")
        if not doc_pool:
            log(f"\n{log_label}重扫后文档池已空,死磕结束")
            break

        big_round += 1
        log(f"\n{'='*20} {log_label}第 {big_round} 大循环 开始 (active {len(active)} 账号,文档池 {len(doc_pool)} 篇) {'='*20}")

        for sub_idx in range(1, sub_rounds + 1):
            log(f"\n----- {log_label}第 {big_round} 大循环 / 第 {sub_idx} 小轮 (Phase A) -----")
            sub_fail_count = {}
            sub_skipped = set()
            sub_round_id = big_round * 100 + sub_idx  # 用于 sheet 标记

            # ----- Phase A: 每个未达 quota 的账号发 1 篇 -----
            for name in list(accounts):
                if not doc_pool:
                    log(f"  Phase A 中文档池已空,提前结束")
                    break
                if not _is_eligible(name, sub_skipped):
                    continue
                # [v1101.4] _pick_doc 替代 random.choice: 校验存在 + 失效就地剔除
                doc = _pick_doc(doc_pool)
                if doc is None:
                    log(f"  Phase A 中文档池被 _pick_doc 清空(全失效),提前结束")
                    break
                round_label = f"[大{big_round}/小{sub_idx}/A]"
                write_status(big_round, sub_idx, "Phase A",
                             total_done=_total_done(), total_target=_total_target(),
                             doc_pool_size=len(doc_pool),
                             sub_failed=len(sub_skipped), dead_terminated=dead_terminated)
                success, reason = _do_publish(name, doc, round_label, is_retry=False)
                if success:
                    sent_accounts_set.add(name)
                    _append_sent_excel(name)
                    _append_summary(name, sub_round_id, "发文时间")
                else:
                    _append_fail_list(name, reason, os.path.basename(doc), sub_round_id)
                    _append_summary(name, sub_round_id, "失败时间")
                    _handle_failure(name, reason, sub_fail_count, sub_skipped)

            # ----- Phase B: 重试本小轮 fail_list 内未硬终止的 -----
            phase_b_round = 0
            while phase_b_round < max_fail_per_sub and doc_pool:
                phase_b_round += 1
                pending = [f for f in _read_fail_list()
                           if f[4] == sub_round_id
                           and f[0] not in sent_accounts_set
                           and f[0] not in dead_terminated
                           and f[0] not in sub_skipped]
                if not pending:
                    log(f"  Phase B: 失败列表已清空,本小轮齐活")
                    break
                log(f"\n  {log_label}Phase B 第 {phase_b_round}/{max_fail_per_sub} 次, 待补 {len(pending)} 个")
                for f_name, f_reason, f_docname, f_ts, f_rnd in pending:
                    if not _is_eligible(f_name, sub_skipped):
                        continue
                    # [v1101.4] 先找原失败 doc, 失效就剔除; 找不到 fallback 用 _pick_doc
                    doc_path = None
                    for d in list(doc_pool):
                        if os.path.basename(d) == f_docname:
                            if os.path.exists(d):
                                doc_path = d
                            else:
                                log(f"  ! 原失败 doc 已删除: {f_docname}, 改抽随机")
                                doc_pool.remove(d)
                            break
                    if doc_path is None:
                        doc_path = _pick_doc(doc_pool)
                        if doc_path is None:
                            log(f"  Phase B 文档池清空, 提前结束")
                            break
                    round_label = f"[大{big_round}/小{sub_idx}/B-{phase_b_round}]"
                    write_status(big_round, sub_idx, f"Phase B-{phase_b_round}",
                                 total_done=_total_done(), total_target=_total_target(),
                                 doc_pool_size=len(doc_pool),
                                 sub_failed=len(sub_skipped), dead_terminated=dead_terminated)
                    success, reason = _do_publish(f_name, doc_path, round_label, is_retry=True)
                    if success:
                        sent_accounts_set.add(f_name)
                        _append_sent_excel(f_name)
                        _remove_from_fail_list(f_name)
                        _append_summary(f_name, sub_round_id, "补发成功时间")
                        log(f"  ✓ 补发成功: {f_name}")
                    else:
                        _handle_failure(f_name, reason, sub_fail_count, sub_skipped)

            # ----- 小轮收尾:本小轮的临时跳过名单清空,下小轮全员可再尝试 -----
            log(f"\n{log_label}第 {big_round} 大循环 / 第 {sub_idx} 小轮 结束。本小轮跳过 {len(sub_skipped)} 个(下小轮恢复)。硬终止累计 {len(dead_terminated)} 个")
            sent_accounts_set.clear()
            # [v1102] sheet 不再小轮末 clear,累积到大循环末才 clear

        log(f"\n{'='*20} {log_label}第 {big_round} 大循环 结束 {'='*20}")
        # [v1102] 全员齐活才 clear 「本轮已发」 sheet
        active_left = [a for a in accounts if a not in dead_terminated and acc_count.get(a, 0) < per_account_quota.get(a, 0)]
        if not active_left:
            _clear_round_sheets()
            log(f"  [v1102] 大循环全员齐活 → 「本轮已发」 sheet 已清空")

    write_status(big_round, sub_rounds, "结束",
                 total_done=_total_done(), total_target=_total_target(),
                 doc_pool_size=len(doc_pool),
                 sub_failed=0, dead_terminated=dead_terminated,
                 extra=f"{log_label}{big_round} 大循环跑完。最终硬终止 {len(dead_terminated)} 个")

    return {
        "acc_count": acc_count,
        "dead_terminated": dead_terminated,
        "doc_pool": doc_pool,
        "ok_count": ok_count,
        "fail_count": fail_count,
        "total_notices": total_notices,
        "total_violations": total_violations,
        "total_alerts": total_alerts,
        "big_rounds": big_round,
    }


# ========== 主流程 ==========

def main():
    _init_run_dir()
    log("=" * 50)
    log("创作罐头批量发布 v3 启动")
    log(f"报告目录: {RUN_REPORT_DIR}")
    log("=" * 50)

    os.makedirs(DOCS_FOLDER, exist_ok=True)
    os.makedirs(SENT_FOLDER, exist_ok=True)
    init_fail_excel()

    # 读取账号配置.xlsx - 不首发sheet
    _ensure_config_excel()
    try:
        _nofirst_list = _read_excel_sheet("不首发")
        NOFIRST_ACCOUNTS.update(_nofirst_list)
        if _nofirst_list:
            log(f"账号配置.xlsx[不首发]已加载，不首发账号: {len(NOFIRST_ACCOUNTS)} 个")
    except Exception as _ne:
        log(f"读取账号配置.xlsx[不首发]失败: {_ne}")

    # 读取账号配置.xlsx - 永久跳过sheet
    try:
        _skip_list = _read_excel_sheet("永久跳过")
        for _sv in _skip_list:
            if _sv not in EXCLUDE_ACCOUNTS:
                EXCLUDE_ACCOUNTS.append(_sv)
        if _skip_list:
            log(f"账号配置.xlsx[永久跳过]已加载，永久跳过: {EXCLUDE_ACCOUNTS}")
    except Exception as _se:
        log(f"读取账号配置.xlsx[永久跳过]失败: {_se}")

    # 读取账号配置.xlsx - 本轮已发sheet（中断恢复时跳过已发账号）
    sent_accounts_set = set()
    try:
        _sent_list = _read_excel_sheet("本轮已发")
        for _sv in _sent_list:
            sent_accounts_set.add(_sv)
        if sent_accounts_set:
            log(f"账号配置.xlsx[本轮已发]已加载，本轮已发: {len(sent_accounts_set)} 个账号（本轮将跳过）")
    except Exception as _se:
        log(f"读取账号配置.xlsx[本轮已发]失败: {_se}")

    # 读取账号配置.xlsx - 失败列表sheet（中断恢复时继续补发）
    try:
        _fail_pre = _read_fail_list()
        if _fail_pre:
            log(f"账号配置.xlsx[失败列表]已加载 {len(_fail_pre)} 条（本轮末将补发）")
    except Exception as _fe:
        log(f"读取账号配置.xlsx[失败列表]失败: {_fe}")

    docs = get_docs()
    if not docs:
        log("错误: 素材文件夹中没有 docx 文件，请放入文档后重试")
        os.system("pause")
        return

    log(f"待发文档: {len(docs)} 份")

    try:
        main_ws_url = get_main_ws_url()
    except RuntimeError as e:
        log(f"错误: {e}")
        os.system("pause")
        return

    main_ws = ws_connect(main_ws_url, timeout=10)
    log("已连接主窗口")

    # 收集账号列表
    accounts = collect_accounts(main_ws)
    if EXCLUDE_ACCOUNTS:
        before = len(accounts)
        accounts = [a for a in accounts if not any(ex in a or a in ex for ex in EXCLUDE_ACCOUNTS)]
        skipped = before - len(accounts)
        if skipped:
            log(f"已排除 {skipped} 个永不发文账号: {EXCLUDE_ACCOUNTS}")

    # per_account_quota: 本次每个账号要发几篇
    per_account_quota = {}

    # [v1102] quota 动态算 = (素材池 + 已发累计) // 账号数
    sent_count_map = _read_sent_with_count()
    sent_total = sum(sent_count_map.values())

    # 正常模式：读"白名单"A 列（账号名）+ B 列（可选配额）
    whitelist_with_q = _read_whitelist_with_quota()
    if whitelist_with_q:
        wl_map = {n: q for n, q in whitelist_with_q}
        accounts = [a for a in accounts if any(wn in a or a in wn for wn in wl_map)]
        log(f"账号配置.xlsx[白名单]已加载，白名单 {len(wl_map)} 个，过滤后剩 {len(accounts)} 个账号")
        default_q = max(1, (len(docs) + sent_total) // len(accounts)) if accounts else 1
        for a in accounts:
            matched_q = None
            for wn, wq in wl_map.items():
                if wn in a or a in wn:
                    matched_q = wq
                    break
            per_account_quota[a] = matched_q if matched_q is not None else default_q
    else:
        default_q = max(1, (len(docs) + sent_total) // len(accounts)) if accounts else 1
        for a in accounts:
            per_account_quota[a] = default_q
    total_target = sum(per_account_quota.values())
    log(f"本次发布: {len(accounts)} 个账号，{len(docs)} 篇文档，配额合计 {total_target}")

    if not accounts:
        log("错误: 未找到任何账号，请检查创作罐头是否正常登录")
        main_ws.close()
        os.system("pause")
        return

    doc_pool = list(docs)
    credit_records = {}

    # 调用死磕主循环 (3 小轮 / 大循环 + 外层无限磕)
    result = run_death_grip(
        accounts=accounts,
        per_account_quota=per_account_quota,
        doc_pool=doc_pool,
        main_ws=main_ws,
        sub_rounds=3,
        max_fail_per_sub=3,
        sent_accounts_set=sent_accounts_set,
        credit_records=credit_records,
        initial_acc_count=sent_count_map,  # [v1102] 传入已发次数
    )

    acc_count        = result["acc_count"]
    dead_terminated  = result["dead_terminated"]
    doc_pool         = result["doc_pool"]
    ok_count         = result["ok_count"]
    fail_count       = result["fail_count"]
    total_notices    = result["total_notices"]
    total_violations = result["total_violations"]
    total_alerts     = result["total_alerts"]

    if doc_pool:
        log(f"\n! 文档池还剩 {len(doc_pool)} 篇未发完(死磕已无可用账号)")
        for d in doc_pool:
            log(f"  未发: {os.path.basename(d)}")

    main_ws.close()

    # ========== 收尾:硬终止账号汇报 ==========
    if dead_terminated:
        log(f"\n★ 4 类硬终止账号 {len(dead_terminated)} 个 (已写入'硬终止账号'sheet,需人工处理):")
        for name, (reason, ts, cnt) in dead_terminated.items():
            log(f"  - {name}\t{reason}\t本次已发 {cnt} 篇\t{ts}")
    else:
        log("\n★ 无硬终止账号,本次全部账号正常完成或仍在 active")

    # ========== 信用分记录 ==========
    if credit_records:
        credit_file = os.path.join(RUN_REPORT_DIR, "信用分记录.txt")
        def _ck(item):
            n, s = item
            return (s if s is not None else -1, n)
        sorted_items = sorted(credit_records.items(), key=_ck)
        danger = warn = 0
        with open(credit_file, "w", encoding="utf-8") as f:
            f.write(f"信用分记录 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n")
            for n, s in sorted_items:
                if s is None:
                    f.write(f"{n}\t未读取\n")
                elif s < 80:
                    f.write(f"{n}\t{s}\t★★ 危险\n"); danger += 1
                elif s < 90:
                    f.write(f"{n}\t{s}\t★ 警告\n"); warn += 1
                else:
                    f.write(f"{n}\t{s}\n")
        log(f"信用分记录: {len(credit_records)} 条（危险 {danger}，警告 {warn}）")

    # 汇总报告
    summary = [
        f"{'='*60}",
        f"运行汇总 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"{'='*60}",
        f"类型    : 文章发布(死磕模式)",
        f"报告目录: {RUN_REPORT_DIR}",
        f"",
        f"成功发布: {ok_count} 篇",
        f"中途失败: {fail_count} 次(已重试到底)",
        f"硬终止号: {len(dead_terminated)} 个 (失登/封号/禁言/找不到)",
        f"文档剩余: {len(doc_pool)} 篇",
        f"",
        f"系统通知: {total_notices} 条（近两天）",
        f"违规命中: {total_violations} 条",
        f"高阅读量: {total_alerts} 条",
        f"{'='*60}",
    ]
    summary_text = "\n".join(summary)
    log(f"\n{summary_text}")
    summary_file = os.path.join(RUN_REPORT_DIR, "汇总报告.txt")
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write(summary_text + "\n")

    # 排序"发文汇总"：同账号的多轮放一起，按轮次升序
    _sort_summary_by_account()

    os.system("pause")


if __name__ == "__main__":
    main()
