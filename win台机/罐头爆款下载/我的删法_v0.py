"""第一版标题分类器(冷启动 v0)

基于:
  - 4月2号 368 删 / 1666 留 的归纳
  - 用户说的 6 类问题(标题党/夸张/不实/资质/不雅/过时效)
  - 7 天禁言案例的高频词
  - 极短标题强信号(删 22% vs 留 3%)

用法:
  python 我的删法_v0.py <原始 xlsx 路径>
  → 输出同目录 罐头爆款_xxx_我的删法.xlsx + diff 报告

注:这是 v0 — 第一次跑准度大概 60-70%,通过用户审改 + diff 学习,后续版本会越来越准
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import sys, os, re, shutil
from datetime import datetime

# === 关键词词典(从 4月2 删除集 + 禁言案例归纳) ===

# 一类:夸张诱导词 — 标题党(从 6 篇禁言实证 + 4月2 删除归纳)
EXAGG_WORDS = [
    '硬气', '彻底', '谁也没想到', '震惊', '解气', '没人敢', '出乎意料',
    '重磅炸弹', '震碎', '撂得明明白白', '不留情面', '动真格',
    '太离谱', '挺让人无语', '没白花', '真服了', '真不愧',
    '太扯', '太牛', '太可怕', '太狠', '太硬', '太逆天',
    '居然', '竟然', '万万没想到', '惊呆',
    '又动真格', '可真是', '硬刚', '正面刚',
    '这谁能想得到', '谁能想到', '想都没想到', '谁也不曾想到', '不曾想到',
    '白嫖成功', '白嫖',
]

# 二类:暗示/擦边/不雅 + 婚恋撕逼 + 性暗示 + 暴力(传播不良价值观)
SUGGEST_WORDS = [
    # 性暗示
    '出轨', '偷情', '情人', '小三', '婚外', '私生',
    '按摩', '足疗', '足浴', '按摩店', '内裤', '隐私部位',
    '床上', '同房', '一夜情', '约炮', '裸泳', '裸',
    '穿这样', '走光', '穿太少', '辣眼', '尺度', '风骚',
    '抱起', '甩出去', '腰', '胯',
    '发生了关系', '发生关系', '上了床', '睡了',
    '婚前一夜', '婚前夜', '新婚之夜', '洞房',
    # 婚恋撕逼 / 离婚纠纷
    '前妻', '前夫', '撕破脸', '离婚后第',
    # 暴力暗示
    '暴怒', '暴打', '动了手', '打了一顿', '失手', '砍人', '砍了',
    '挥刀', '一刀', '捅了', '掐死',
    '扣动扳机', '扳机', '枪打烂', '打烂头', '打烂脑', '与尸体', '尸体',
    # 性犯罪 / 猥亵 / 未成年
    '强奸', '强暴', '性侵', '猥亵', '骚扰',
    '白色液体', '体液', '精液',
    '陪酒女', '陪酒', '未成年女', '15岁陪', '14岁陪', '16岁陪',
    '滴滴原液', '迷魂药', '迷药', '下药',
    # 恶心猎奇
    '恶心坏', '恶心到', '反胃', '作呕',
    # 赌博类
    '棋牌室', '牌桌', '输光', '又输了', '骰子', '赌输', '老虎机',
]

# 三类:政治/军事/外交敏感(禁言主轴 + 用户提示)
POLI_WORDS = [
    '台独', '台湾海峡', '海峡', '台海',
    '中美', '中俄', '中日', '中印',
    '伊朗', '叙利亚', '以色列', '巴勒斯坦', '加沙',
    '北约', '联合国', '安理会', '联合国副秘书长',
    '泽连斯基', '普京', '特朗普', '拜登', '托卡耶夫',
    '军演', '舰艇', '蹭海峡', '不动手',
]

# 四类:明星/网红八卦(已洗烂 / 易翻车)
CELEB_WORDS = [
    '张雪峰', '武亮',  # 张雪峰系列出现 ~10 次
    '王石', '刘畊宏', '张丰毅', '李若彤', '丁俊晖', '赵心童',
    '何润东', '范冰冰', '甄子丹',
    '钱塘江道长', '浙大博士孟伟', '京东副总裁刘辰', '京东集团副总裁',
    '海军副司令员', '杨利伟', '钟南山',
    '罗永浩', '雷军', '马云', '刘强东', '东哥',
    '上海交大校庆',
]

# 五类:广告/资质类领域常见词
QUALI_WORDS = [
    '股票', '基金', '理财', '炒股', '财富', '收益率',
    '中央储备', '冻猪肉', '猪肉储备',  # 政策投资类
    '癌症', '高血压', '糖尿病', '老中医', '偏方', '秘方', '能治', '治愈',
    '减肥神器', '神药', '特效药',
]

# 六类:猎奇极端事件(易翻车)
EXTREME_WORDS = [
    '惨死', '惨剧', '悲剧', '溺亡', '坠亡', '抢手机',
    '装病乞讨', '社死', '丢人现眼',
    '十里河', '装病', '诈骗',
    '裸贷', '裸聊',
]

# 七类:罪犯案件黑话 / 死刑暗语 / 不良价值传播
# 用户 2026-04-26 强调:这些是网络黑话,字面看不出但出现 = 该案题材必删
CRIMINAL_NICKNAMES = [
    '金毛',          # 杀妻案罪犯绰号(不是金毛犬!)
    '三角眼',        # 罪犯亲属代号(那个案件里的妈妈)
    '吃花生米',      # 死刑/枪决暗语
    '小四毛',        # 山西黑老大代号
    '陈丹蕾',        # 2005 年清华女硕士杀夫案
    # 后续逐个补充:遇到一个用户告诉我加一个
    # 特别注意上下文:这些词单独出现不一定是案件,但跟"亲妈/笼子/走了/承受打击/泪崩"等组合就铁定是
]

# 八类:旧年代回顾 / 黑社会 / 腐败 (信息真伪难考,AI 洗稿易加油加醋)
HISTORICAL_CRIME_WORDS = [
    '黑老大', '黑社会', '老大被', '黑帮',
    '非法敛财', '受贿', '贪污', '亿元', '过亿',
    '狱中', '监狱', '越狱', '假释', '判处死刑', '判刑',
]

# 九类:军事历史(用户 2026-04-26 强调:军事领域涉及远久年代几乎全删)
# 抗日、抗美、越战、二战 + 叛逃/间谍 等几乎都不要
HISTORICAL_MILITARY_WORDS = [
    '抗日', '抗战', '抗美援朝', '抗美', '越战', '越南战争',
    '二战', '一战', '世界大战', '朝鲜战争', '解放战争', '内战',
    '老山', '对越自卫反击', '中印边境',
    '苏联', '苏军', '日军', '美军飞行员', '美苏', '冷战',
    '叛逃', '投敌', '间谍', '特工',
    '红军', '八路军', '新四军', '志愿军', '解放军老兵',
]


def score_title(text):
    """对标题打分,返回(score, reasons列表)。score 越高越倾向删。"""
    text_str = str(text)
    score = 0
    reasons = []

    # 长度规则:短文字 = 视频内容 = 洗不出来 = 必删
    # (用户 2026-04-26 确认:短文字一般是视频)
    L = len(text_str)
    if L < 25:
        score += 10
        reasons.append(f'极短/视频({L}字)')
    elif L < 40:
        score += 4
        reasons.append(f'偏短可能视频({L}字)')
    elif L < 60:
        score += 1
        reasons.append(f'较短({L}字)')

    # 词典匹配
    for kw in EXAGG_WORDS:
        if kw in text_str:
            score += 2
            reasons.append(f'夸张/{kw}')
    for kw in SUGGEST_WORDS:
        if kw in text_str:
            score += 4
            reasons.append(f'擦边/{kw}')
    for kw in POLI_WORDS:
        if kw in text_str:
            score += 3
            reasons.append(f'政治军事/{kw}')
    for kw in CELEB_WORDS:
        if kw in text_str:
            score += 3
            reasons.append(f'明星八卦/{kw}')
    for kw in QUALI_WORDS:
        if kw in text_str:
            score += 3
            reasons.append(f'资质类/{kw}')
    for kw in EXTREME_WORDS:
        if kw in text_str:
            score += 2
            reasons.append(f'猎奇/{kw}')
    for kw in CRIMINAL_NICKNAMES:
        if kw in text_str:
            score += 8
            reasons.append(f'罪犯绰号/{kw}')
    for kw in HISTORICAL_CRIME_WORDS:
        if kw in text_str:
            score += 4
            reasons.append(f'黑社/腐败/{kw}')
    for kw in HISTORICAL_MILITARY_WORDS:
        if kw in text_str:
            score += 5
            reasons.append(f'军事历史/{kw}')

    # 旧年代年份检测:标题/正文开头出现"20XX 年"且距今 ≥ 3 年 → 历史挖掘
    # 现在 2026,3 年前是 2023,所以 200X-2023 年都算旧
    m = re.search(r'(20\d\d)\s*年', text_str[:30])  # 只看开头 30 字
    if m:
        year = int(m.group(1))
        if year <= 2023:
            score += 5
            reasons.append(f'旧年代/{year}年')

    return score, reasons


THRESHOLD = 4  # score >= 4 判定为删


def classify_xlsx(input_xlsx):
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    title_idx = header.index('标题')

    rows = list(ws.iter_rows(min_row=2))
    keep, delete_with_reason = [], []
    for row in rows:
        title = row[title_idx].value
        if not title:
            continue
        sc, reasons = score_title(title)
        if sc >= THRESHOLD:
            delete_with_reason.append((row, sc, reasons))
        else:
            keep.append(row)

    return wb, ws, header, rows, keep, delete_with_reason


def emit_my_pick(input_xlsx):
    """跑分类器,直接把工作版 xlsx 改成"我筛后的剩下"(删除 Claude 判定为删的行)。
    备份(_原始.xlsx) 保持 2000 原始 9 列,做 diff 学习用。"""
    wb, ws, header, all_rows, keep, deleted = classify_xlsx(input_xlsx)

    # 收集要保留的 URL
    url_idx = header.index('链接')
    keep_urls = set(row[url_idx].value for row in keep)

    # 直接在工作版上删掉 Claude 判定为删的行(从下往上删避免索引错位)
    rows_to_delete = []
    for r_idx in range(2, ws.max_row + 1):
        url_v = ws.cell(r_idx, url_idx + 1).value
        if url_v not in keep_urls:
            rows_to_delete.append(r_idx)

    for r_idx in reversed(rows_to_delete):
        ws.delete_rows(r_idx)

    wb.save(input_xlsx)
    print(f'★ 我已筛过一轮: {os.path.basename(input_xlsx)}: 剩 {len(keep)} 条 (删 {len(deleted)},删除率 {100*len(deleted)/len(all_rows):.1f}%)')

    return input_xlsx, None


def self_test_on_apr2():
    """用 4月2 数据自检准度"""
    orig = r'C:\Users\kench\Downloads\低粉爆款4月2号.xlsx'
    filt = r'C:\Users\kench\Downloads\低粉爆款4月2号改好.xlsx'

    wb_o = openpyxl.load_workbook(orig, read_only=True)
    wb_f = openpyxl.load_workbook(filt, read_only=True)
    rows_o = list(wb_o.active.iter_rows(values_only=True))
    rows_f = list(wb_f.active.iter_rows(values_only=True))
    h = rows_o[0]; rows_o = rows_o[1:]; rows_f = rows_f[1:]
    url_idx = h.index('链接'); title_idx = h.index('标题')

    user_kept = set(r[url_idx] for r in rows_f)
    # ground truth: True = should delete, False = should keep
    gt = {r[url_idx]: r[url_idx] not in user_kept for r in rows_o}

    # 我的预测
    my_pred = {}
    for r in rows_o:
        sc, _ = score_title(r[title_idx])
        my_pred[r[url_idx]] = sc >= THRESHOLD

    # 混淆矩阵
    tp = sum(1 for u in gt if gt[u] and my_pred[u])  # 该删,我说删
    fn = sum(1 for u in gt if gt[u] and not my_pred[u])  # 该删,我说留(漏删)
    fp = sum(1 for u in gt if not gt[u] and my_pred[u])  # 该留,我说删(误删)
    tn = sum(1 for u in gt if not gt[u] and not my_pred[u])  # 该留,我说留

    total = len(gt)
    acc = (tp + tn) / total
    prec = tp / (tp + fp) if (tp + fp) else 0
    rec = tp / (tp + fn) if (tp + fn) else 0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0

    print(f'\n=== 4月2 自检 (Day 1 baseline) ===')
    print(f'  准度: {acc*100:.1f}% ({tp+tn}/{total})')
    print(f'  Precision(我说删的真该删): {prec*100:.1f}%')
    print(f'  Recall(该删的我能找出): {rec*100:.1f}%')
    print(f'  F1: {f1:.3f}')
    print(f'  TP={tp} FP={fp} FN={fn} TN={tn}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python 我的删法_v0.py <xlsx 路径> [--selftest]')
        # 默认行为:先自检,然后处理最新下载
        self_test_on_apr2()
        # 找最新的 罐头爆款_*.xlsx
        import glob
        DL = os.path.expanduser('~/Downloads')
        cand = sorted(glob.glob(os.path.join(DL, '罐头爆款_*.xlsx')),
                      key=os.path.getmtime, reverse=True)
        cand = [c for c in cand if '_我的删法' not in c and '_改好' not in c]
        if cand:
            print(f'\n处理最新: {cand[0]}')
            emit_my_pick(cand[0])
    elif sys.argv[1] == '--selftest':
        self_test_on_apr2()
    else:
        if '--selftest' in sys.argv:
            self_test_on_apr2()
        emit_my_pick(sys.argv[1])
