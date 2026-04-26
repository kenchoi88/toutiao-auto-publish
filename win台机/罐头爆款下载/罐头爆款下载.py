"""罐头低粉爆款一键下载 + URL 提取

前提:
  1. 罐头已通过 debug_launch.bat 启动到 CDP 9223
  2. 罐头已登录(显示 home 页 https://www.czgts.cn/v1/home)

行为:
  1. 连 CDP 9223
  2. 跳到低粉爆款页 /v1/hots/popular
  3. 设筛选: 默认 2 天内 (用户偏好)
  4. CDP 真实鼠标事件点 "下载数据" — JS click 不触发,必须 dispatchMouseEvent
  5. 监听 ~/Downloads 等 .tmp / article-*.xlsx 出现
  6. .tmp 自动改名为 罐头爆款_YYYYMMDD_HHMMSS.xlsx
  7. 提取"链接"列输出同名 .txt

输出位置: ~/Downloads/
"""
import requests, websocket, json, time, os, glob, shutil, sys
from datetime import datetime

CDP = 'http://127.0.0.1:9223'
DL = os.path.expanduser('~/Downloads')


def main():
    # 1. 连 CDP
    try:
        tabs = requests.get(f'{CDP}/json', timeout=3, proxies={'http': '', 'https': ''}).json()
    except Exception as e:
        print(f'X 罐头 CDP 9223 不在线: {e}')
        print('  先双击 win台机/GTG_*/debug_launch.bat 启动')
        return 1

    home = next((t for t in tabs if 'czgts.cn' in t.get('url', '')), None)
    if not home:
        print('X 没找到罐头主界面 tab')
        return 1
    print(f'连上罐头: {home["url"]}')

    ws = websocket.create_connection(home['webSocketDebuggerUrl'], suppress_origin=True)
    mid = [0]

    def cdp(method, params=None):
        mid[0] += 1
        ws.send(json.dumps({'id': mid[0], 'method': method, 'params': params or {}}))
        while True:
            d = json.loads(ws.recv())
            if d.get('id') == mid[0]:
                return d.get('result')

    def js(expr):
        return cdp('Runtime.evaluate', {'expression': expr, 'returnByValue': True}).get('result', {}).get('value')

    # 2. 跳到低粉爆款
    cur = js('location.href')
    if 'hots/popular' not in (cur or ''):
        # 如果在 home,点击"低粉爆款"快捷入口;否则直接 navigate
        result = js('''
        (function(){
          var els = document.querySelectorAll('.quickAccessItemTitle-O__czI');
          for(var i=0;i<els.length;i++){
            if(els[i].textContent.trim() === '低粉爆款'){ els[i].click(); return 'clicked'; }
          }
          location.href = 'https://www.czgts.cn/v1/hots/popular';
          return 'navigated';
        })()
        ''')
        print(f'切换到低粉爆款页: {result}')
        time.sleep(2.5)

    # 3. 选 2 天内
    js('''
    (function(){
      var all = document.querySelectorAll('*');
      for(var i=0;i<all.length;i++){
        var t = (all[i].textContent || '').trim();
        if(t === '2天内' && all[i].children.length === 0){
          var r = all[i].getBoundingClientRect();
          if(r.width > 0 && r.width < 150){ all[i].click(); return; }
        }
      }
    })()
    ''')
    time.sleep(1.0)

    # 4. 找下载按钮坐标(每次重新取,布局可能变)
    pos = js('''
    (function(){
      var all = document.querySelectorAll('*');
      for(var i=0;i<all.length;i++){
        if((all[i].textContent || '').trim() === '下载数据' && all[i].children.length === 0){
          var r = all[i].getBoundingClientRect();
          if(r.width > 0)
            return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
        }
      }
      return null;
    })()
    ''')
    if not pos:
        print('X 找不到下载数据按钮')
        ws.close()
        return 1
    p = json.loads(pos)
    print(f'下载按钮坐标: ({p["x"]}, {p["y"]})')

    # 5. CDP 真实鼠标点击
    before = set(os.listdir(DL))
    cdp('Input.dispatchMouseEvent', {'type': 'mouseMoved', 'x': p['x'], 'y': p['y']})
    time.sleep(0.1)
    cdp('Input.dispatchMouseEvent', {'type': 'mousePressed', 'x': p['x'], 'y': p['y'], 'button': 'left', 'clickCount': 1})
    time.sleep(0.05)
    cdp('Input.dispatchMouseEvent', {'type': 'mouseReleased', 'x': p['x'], 'y': p['y'], 'button': 'left', 'clickCount': 1})
    print('已触发下载,等文件落地...')

    # 6. 等新文件
    new_file = None
    for i in range(60):
        time.sleep(1)
        after = set(os.listdir(DL))
        new_names = [n for n in after - before if n.endswith('.tmp') or n.startswith('article-')]
        if new_names:
            new_file = os.path.join(DL, new_names[0])
            # 等文件大小稳定
            last_size = -1
            for j in range(15):
                cur = os.path.getsize(new_file)
                if cur == last_size and cur > 0:
                    break
                last_size = cur
                time.sleep(0.5)
            break

    if not new_file:
        print('X 60 秒未见新文件')
        ws.close()
        return 1

    # 7. 改名为 罐头爆款_YYYYMMDD_HHMMSS.xlsx
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    final_xlsx = os.path.join(DL, f'罐头爆款_{ts}.xlsx')
    shutil.move(new_file, final_xlsx)
    print(f'★ 工作版: {final_xlsx} ({os.path.getsize(final_xlsx)} 字节)')

    # 7b. 自动备份原始版(用户删行不影响,Claude 用这份做 diff 学习)
    backup_dir = os.path.join(DL, '罐头爆款_原始')
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f'罐头爆款_{ts}.xlsx')
    shutil.copy(final_xlsx, backup_path)
    print(f'★ 原始备份: {backup_path}')

    # 8. 提取 URL 列输出 .txt
    import openpyxl
    wb = openpyxl.load_workbook(final_xlsx, read_only=True)
    ws_x = wb.active
    header = [c.value for c in next(ws_x.iter_rows(max_row=1))]
    try:
        url_idx = header.index('链接')
    except ValueError:
        print('X xlsx 没有"链接"列,跳过 URL 提取')
        ws.close()
        return 0
    urls = []
    for row in ws_x.iter_rows(min_row=2, values_only=True):
        if row[url_idx]:
            urls.append(str(row[url_idx]).strip())

    final_txt = final_xlsx.replace('.xlsx', '_urls.txt')
    with open(final_txt, 'w', encoding='utf-8') as f:
        f.write('\n'.join(urls))
    print(f'★ URL 列表: {final_txt} ({len(urls)} 条)')

    ws.close()
    return 0


if __name__ == '__main__':
    sys.exit(main())
