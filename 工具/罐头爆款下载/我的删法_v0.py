"""第一版标题分类器(冷启动 v0)

基于:
  - 4月2号 368 删 / 1666 留 的归纳
  - 用户说的 6 类问题(标题党/夸张/不实/资质/不雅/过时效)
  - 7 天禁言案例的高频词
  - 极短标题强信号(删 22% vs 留 3%)

用法:
  python 我的删法_v0.py <原始 xlsx 路径>
  → 输出同目录 罐头爆款_xxx_我的删法.xlsx + diff 报告

注:这是 v0 — 第一次跑准度大概 60-70%,通过用户审改 + diff 学习,后续版本会越来越准
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import sys, os, re, shutil
from datetime import datetime

# === 关键词词典(从 4月2 删除集 + 禁言案例归纳) ===

# 一类:夸张诱导词 — 标题党(从 6 篇禁言实证 + 4月2 删除归纳)
EXAGG_WORDS = [
    '硬气', '彻底', '谁也没想到', '震惊', '解气', '没人敢', '出乎意料',
    '重磅炸弹', '震碎', '撂得明明白白', '不留情面', '动真格',
    '太离谱', '挺让人无语', '没白花', '真服了', '真不愧',
    '太扯', '太牛', '太可怕', '太狠', '太硬', '太逆天',
    '居然', '竟然', '万万没想到', '惊呆',
    '又动真格', '可真是', '硬刚', '正面刚',
]

# 二类:暗示/擦边/不雅词
SUGGEST_WORDS = [
    '出轨', '偷情', '情人', '小三', '婚外', '私生',
    '按摩', '足疗', '足浴', '按摩店', '内裤', '隐私部位',
    '床上', '同房', '一夜情', '约炮', '裸泳', '裸',
    '穿这样', '走光', '穿太少', '辣眼', '尺度', '风骚',
    '抱起', '甩出去', '腰', '胯',
]

# 三类:政治/军事/外交敏感(禁言主轴 + 用户提示)
POLI_WORDS = [
    '台独', '台湾海峡', '海峡', '台海',
    '中美', '中俄', '中日', '中印',
    '伊朗', '叙利亚', '以色列', '巴勒斯坦', '加沙',
    '北约', '联合国', '安理会', '联合国副秘书长',
    '泽连斯基', '普京', '特朗普', '拜登', '托卡耶夫',
    '军演', '舰艇', '蹭海峡', '不动手',
]

# 四类:明星/网红八卦(已洗烂 / 易翻车)
CELEB_WORDS = [
    '张雪峰', '武亮',  # 张雪峰系列出现 ~10 次
    '王石', '刘畊宏', '张丰毅', '李若彤', '丁俊晖', '赵心童',
    '何润东', '范冰冰', '甄子丹',
    '钱塘江道长', '浙大博士孟伟', '京东副总裁刘辰', '京东集团副总裁',
    '海军副司令员', '杨利伟', '钟南山',
    '罗永浩', '雷军', '马云', '刘强东', '东哥',
    '上海交大校庆',
]

# 五类:广告/资质类领域常见词
QUALI_WORDS = [
    '股票', '基金', '理财', '炒股', '财富', '收益率',
    '中央储备', '冻猪肉', '猪肉储备',  # 政策投资类
    '癌症', '高血压', '糖尿病', '老中医', '偏方', '秘方', '能治', '治愈',
    '减肥神器', '神药', '特效药',
]

# 六类:猎奇极端事件(易翻车)
EXTREME_WORDS = [
    '惨死', '惨剧', '悲剧', '溺亡', '坠亡', '抢手机',
    '装病乞讨', '社死', '丢人现眼',
    '十里河', '装病', '诈骗',
    '裸贷', '裸聊',
]


def score_title(text):
    """对标题打分,返回(score, reasons列表)。score 越高越倾向删。"""
    text_str = str(text)
    score = 0
    reasons = []

    # 长度 < 25 → 强烈倾向删 (4月2 数据 22% vs 3%)
    L = len(text_str)
    if L < 15:
        score += 5
        reasons.append(f'极短({L}字)')
    elif L < 25:
        score += 3
        reasons.append(f'偏短({L}字)')
    elif L < 40:
        score += 1
        reasons.append(f'较短({L}字)')

    # 词典匹配
    for kw in EXAGG_WORDS:
        if kw in text_str:
            score += 2
            reasons.append(f'夸张/{kw}')
    for kw in SUGGEST_WORDS:
        if kw in text_str:
            score += 4
            reasons.append(f'擦边/{kw}')
    for kw in POLI_WORDS:
        if kw in text_str:
            score += 3
            reasons.append(f'政治军事/{kw}')
    for kw in CELEB_WORDS:
        if kw in text_str:
            score += 3
            reasons.append(f'明星八卦/{kw}')
    for kw in QUALI_WORDS:
        if kw in text_str:
            score += 3
            reasons.append(f'资质类/{kw}')
    for kw in EXTREME_WORDS:
        if kw in text_str:
            score += 2
            reasons.append(f'猎奇/{kw}')

    return score, reasons


THRESHOLD = 4  # score >= 4 判定为删


def classify_xlsx(input_xlsx):
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    title_idx = header.index('标题')

    rows = list(ws.iter_rows(min_row=2))
    keep, delete_with_reason = [], []
    for row in rows:
        title = row[title_idx].value
        if not title:
            continue
        sc, reasons = score_title(title)
        if sc >= THRESHOLD:
            delete_with_reason.append((row, sc, reasons))
        else:
            keep.append(row)

    return wb, ws, header, rows, keep, delete_with_reason


def emit_my_pick(input_xlsx):
    """跑分类器,输出 我的删法.xlsx (只保留 keep 行) + 删除报告"""
    wb, ws, header, all_rows, keep, deleted = classify_xlsx(input_xlsx)

    # 输出 我的删法.xlsx — 只留下 keep
    base = os.path.splitext(input_xlsx)[0]
    out_xlsx = f'{base}_我的删法.xlsx'
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title
    new_ws.append(header)
    for row in keep:
        new_ws.append([c.value for c in row])
    new_wb.save(out_xlsx)
    print(f'★ 我的删法.xlsx: {out_xlsx} ({len(keep)} 留 / {len(all_rows)} 总,{len(deleted)} 删)')

    # 输出 删除原因报告
    report_path = f'{base}_删除原因.txt'
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(f'分类器 v0 删除报告\n')
        f.write(f'总: {len(all_rows)},留: {len(keep)},删: {len(deleted)} ({100*len(deleted)/len(all_rows):.1f}%)\n')
        f.write(f'阈值 score >= {THRESHOLD}\n')
        f.write('=' * 80 + '\n\n')
        for row, sc, reasons in sorted(deleted, key=lambda x: -x[1])[:200]:
            title = str(row[header.index('标题')].value)[:80].replace('\n', ' ')
            field = row[header.index('领域')].value
            f.write(f'[score={sc}] [{field}] {title}\n  → {", ".join(reasons[:5])}\n\n')
    print(f'★ 删除原因报告(前 200): {report_path}')

    return out_xlsx, report_path


def self_test_on_apr2():
    """用 4月2 数据自检准度"""
    orig = r'C:\Users\kench\Downloads\低粉爆款4月2号.xlsx'
    filt = r'C:\Users\kench\Downloads\低粉爆款4月2号改好.xlsx'

    wb_o = openpyxl.load_workbook(orig, read_only=True)
    wb_f = openpyxl.load_workbook(filt, read_only=True)
    rows_o = list(wb_o.active.iter_rows(values_only=True))
    rows_f = list(wb_f.active.iter_rows(values_only=True))
    h = rows_o[0]; rows_o = rows_o[1:]; rows_f = rows_f[1:]
    url_idx = h.index('链接'); title_idx = h.index('标题')

    user_kept = set(r[url_idx] for r in rows_f)
    # ground truth: True = should delete, False = should keep
    gt = {r[url_idx]: r[url_idx] not in user_kept for r in rows_o}

    # 我的预测
    my_pred = {}
    for r in rows_o:
        sc, _ = score_title(r[title_idx])
        my_pred[r[url_idx]] = sc >= THRESHOLD

    # 混淆矩阵
    tp = sum(1 for u in gt if gt[u] and my_pred[u])  # 该删,我说删
    fn = sum(1 for u in gt if gt[u] and not my_pred[u])  # 该删,我说留(漏删)
    fp = sum(1 for u in gt if not gt[u] and my_pred[u])  # 该留,我说删(误删)
    tn = sum(1 for u in gt if not gt[u] and not my_pred[u])  # 该留,我说留

    total = len(gt)
    acc = (tp + tn) / total
    prec = tp / (tp + fp) if (tp + fp) else 0
    rec = tp / (tp + fn) if (tp + fn) else 0
    f1 = 2 * prec * rec / (prec + rec) if (prec + rec) else 0

    print(f'\n=== 4月2 自检 (Day 1 baseline) ===')
    print(f'  准度: {acc*100:.1f}% ({tp+tn}/{total})')
    print(f'  Precision(我说删的真该删): {prec*100:.1f}%')
    print(f'  Recall(该删的我能找出): {rec*100:.1f}%')
    print(f'  F1: {f1:.3f}')
    print(f'  TP={tp} FP={fp} FN={fn} TN={tn}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python 我的删法_v0.py <xlsx 路径> [--selftest]')
        # 默认行为:先自检,然后处理最新下载
        self_test_on_apr2()
        # 找最新的 罐头爆款_*.xlsx
        import glob
        DL = os.path.expanduser('~/Downloads')
        cand = sorted(glob.glob(os.path.join(DL, '罐头爆款_*.xlsx')),
                      key=os.path.getmtime, reverse=True)
        cand = [c for c in cand if '_我的删法' not in c and '_改好' not in c]
        if cand:
            print(f'\n处理最新: {cand[0]}')
            emit_my_pick(cand[0])
    elif sys.argv[1] == '--selftest':
        self_test_on_apr2()
    else:
        if '--selftest' in sys.argv:
            self_test_on_apr2()
        emit_my_pick(sys.argv[1])
