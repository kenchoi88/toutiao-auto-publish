"""罐头低粉爆款一键下载 + URL 提取

前提:
  1. 罐头已通过 debug_launch.bat 启动到 CDP 9223
  2. 罐头已登录(显示 home 页 https://www.czgts.cn/v1/home)

行为:
  1. 连 CDP 9223
  2. 跳到低粉爆款页 /v1/hots/popular
  3. 设筛选: 默认 2 天内 (用户偏好)
  4. CDP 真实鼠标事件点 "下载数据" — JS click 不触发,必须 dispatchMouseEvent
  5. 监听 ~/Downloads 等 .tmp / article-*.xlsx 出现
  6. .tmp 自动改名为 罐头爆款_YYYYMMDD_HHMMSS.xlsx
  7. 提取"链接"列输出同名 .txt

输出位置: ~/Downloads/
"""
import requests, websocket, json, time, os, glob, shutil, sys
from datetime import datetime

CDP = 'http://127.0.0.1:9223'

# 用户半年来稳定使用的 20 个领域(2026-04-26 确认)
TARGET_DOMAINS = [
    '国际', '职业职场', '军事', '教育', '科学科技',
    '健康', '养老', '美食', '三农', '法律',
    '育儿', '旅游', '音乐', '运动健身', '动物宠物',
    '房产', '科普', '游戏', '动漫', '家居家装',
]
# WATCH_DIR = Chromium 实际下载到的位置(默认 ~/Downloads,即使 CDP 设了 path 也常被忽略)
# OUTPUT_DIR = 我们整理后的位置 = 脚本所在目录(双击哪个 .bat 启动,落在那)
WATCH_DIR = os.path.expanduser('~/Downloads')
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
DL = OUTPUT_DIR  # 主要输出位置(主版+备份+urls 都落这)


def main():
    # 1. 连 CDP
    try:
        tabs = requests.get(f'{CDP}/json', timeout=3, proxies={'http': '', 'https': ''}).json()
    except Exception as e:
        print(f'X 罐头 CDP 9223 不在线: {e}')
        print('  先双击 win台机/GTG_*/debug_launch.bat 启动')
        return 1

    home = next((t for t in tabs if 'czgts.cn' in t.get('url', '')), None)
    if not home:
        print('X 没找到罐头主界面 tab')
        return 1
    print(f'连上罐头: {home["url"]}')

    ws = websocket.create_connection(home['webSocketDebuggerUrl'], suppress_origin=True)
    mid = [0]

    def cdp(method, params=None):
        mid[0] += 1
        ws.send(json.dumps({'id': mid[0], 'method': method, 'params': params or {}}))
        while True:
            d = json.loads(ws.recv())
            if d.get('id') == mid[0]:
                return d.get('result')

    def js(expr):
        return cdp('Runtime.evaluate', {'expression': expr, 'returnByValue': True}).get('result', {}).get('value')

    # 关键:让 Chromium 自动下载到 Downloads,不弹"另存为"对话框
    # 必须用 Browser.setDownloadBehavior(browser-level 连接) + behavior='allowAndName'
    # behavior='allow' 会让 Chromium 等 CDP client 接管下载,我们没接管 → 文件丢失
    # behavior='allowAndName' 让 Chromium 自动保存到 downloadPath,文件名为 GUID
    try:
        ver = requests.get(f'{CDP}/json/version', timeout=3, proxies={'http': '', 'https': ''}).json()
        browser_ws_url = ver.get('webSocketDebuggerUrl')
        if browser_ws_url:
            bws = websocket.create_connection(browser_ws_url, suppress_origin=True)
            bws.send(json.dumps({'id': 1, 'method': 'Browser.setDownloadBehavior',
                                 'params': {'behavior': 'allowAndName', 'downloadPath': DL}}))
            for _ in range(5):
                d = json.loads(bws.recv())
                if d.get('id') == 1:
                    if 'error' in d:
                        print(f'  Browser.setDownloadBehavior 失败: {d["error"]}')
                    break
            bws.close()
    except Exception as e:
        print(f'  设 download behavior 异常 (可忽略): {e}')

    # 2. 跳到低粉爆款
    cur = js('location.href')
    if 'hots/popular' not in (cur or ''):
        # 如果在 home,点击"低粉爆款"快捷入口;否则直接 navigate
        result = js('''
        (function(){
          var els = document.querySelectorAll('.quickAccessItemTitle-O__czI');
          for(var i=0;i<els.length;i++){
            if(els[i].textContent.trim() === '低粉爆款'){ els[i].click(); return 'clicked'; }
          }
          location.href = 'https://www.czgts.cn/v1/hots/popular';
          return 'navigated';
        })()
        ''')
        print(f'切换到低粉爆款页: {result}')
        time.sleep(2.5)

    # 3a. 自动勾选 20 个目标领域(取消"全部",再逐个勾)
    targets_json = json.dumps(TARGET_DOMAINS)
    js(f'''
    (function(){{
      var targets = {targets_json};
      var all = document.querySelectorAll('*');

      // 先找"全部"按钮,如果是 checked 就 click 它取消(所有领域变 notChecked)
      for(var i=0;i<all.length;i++){{
        if((all[i].textContent || '').trim() === '全部' && all[i].children.length === 0){{
          var r = all[i].getBoundingClientRect();
          if(r.width > 0 && r.width < 100 && r.top < 300){{
            if(all[i].className.toString().indexOf('notChecked') < 0){{
              all[i].click();  // checked → 取消
            }}
            break;
          }}
        }}
      }}
      return 'done';
    }})()
    ''')
    time.sleep(0.5)
    # 重新查找元素并 click 每个 target
    js(f'''
    (function(){{
      var targets = {targets_json};
      var all = document.querySelectorAll('*');
      var clicked = [];
      for(var ti=0; ti<targets.length; ti++){{
        var name = targets[ti];
        for(var i=0;i<all.length;i++){{
          if((all[i].textContent || '').trim() === name && all[i].children.length === 0){{
            var r = all[i].getBoundingClientRect();
            if(r.width > 0 && r.width < 100 && r.top < 300){{
              if(all[i].className.toString().indexOf('notChecked') >= 0){{
                all[i].click();
                clicked.push(name);
              }}
              break;
            }}
          }}
        }}
      }}
      return clicked.length;
    }})()
    ''')
    time.sleep(1.5)
    print(f'★ 已勾选 {len(TARGET_DOMAINS)} 个目标领域')

    # 3b. 选 2 天内
    js('''
    (function(){
      var all = document.querySelectorAll('*');
      for(var i=0;i<all.length;i++){
        var t = (all[i].textContent || '').trim();
        if(t === '2天内' && all[i].children.length === 0){
          var r = all[i].getBoundingClientRect();
          if(r.width > 0 && r.width < 150){ all[i].click(); return; }
        }
      }
    })()
    ''')
    time.sleep(1.0)

    # 4. 找下载按钮坐标(每次重新取,布局可能变)
    pos = js('''
    (function(){
      var all = document.querySelectorAll('*');
      for(var i=0;i<all.length;i++){
        if((all[i].textContent || '').trim() === '下载数据' && all[i].children.length === 0){
          var r = all[i].getBoundingClientRect();
          if(r.width > 0)
            return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
        }
      }
      return null;
    })()
    ''')
    if not pos:
        print('X 找不到下载数据按钮')
        ws.close()
        return 1
    p = json.loads(pos)
    print(f'下载按钮坐标: ({p["x"]}, {p["y"]})')

    # 5. CDP 真实鼠标点击
    before = set(os.listdir(WATCH_DIR))   # 监听 Chromium 默认下载位置
    cdp('Input.dispatchMouseEvent', {'type': 'mouseMoved', 'x': p['x'], 'y': p['y']})
    time.sleep(0.1)
    cdp('Input.dispatchMouseEvent', {'type': 'mousePressed', 'x': p['x'], 'y': p['y'], 'button': 'left', 'clickCount': 1})
    time.sleep(0.05)
    cdp('Input.dispatchMouseEvent', {'type': 'mouseReleased', 'x': p['x'], 'y': p['y'], 'button': 'left', 'clickCount': 1})
    print('已触发下载,等文件落地...')

    # 5b. Win API 自动关闭"另存为"对话框(Chromium CDP 已接管下载,对话框是冗余的)
    if sys.platform == 'win32':
        import ctypes, threading
        user32 = ctypes.windll.user32
        WM_CLOSE = 0x0010
        def _dismiss_savedialog():
            for _ in range(40):  # 持续 8 秒,每 0.2s 扫一次
                for title in ['另存为', 'Save As', '另存为...', '保存为']:
                    hwnd = user32.FindWindowW(None, title)
                    if hwnd:
                        user32.PostMessageW(hwnd, WM_CLOSE, 0, 0)
                        return
                time.sleep(0.2)
        threading.Thread(target=_dismiss_savedialog, daemon=True).start()

    # 6. 等新文件 (allowAndName 模式下文件名是 GUID,可能带或不带 .xlsx 扩展名)
    new_file = None
    for i in range(60):
        time.sleep(1)
        after = set(os.listdir(WATCH_DIR))
        # 任何新出现的非目录文件,优先 .tmp / .xlsx / 无扩展名(GUID 形式)
        candidates = []
        for n in after - before:
            full = os.path.join(WATCH_DIR, n)
            if os.path.isfile(full):
                candidates.append(n)
        if candidates:
            # 优先取最像下载的(.tmp / 无扩展 GUID / .xlsx)
            candidates.sort(key=lambda n: (
                0 if n.endswith('.tmp') else
                1 if '.' not in n else
                2 if n.startswith('article-') else
                3 if n.endswith('.xlsx') else 9
            ))
            new_file = os.path.join(WATCH_DIR, candidates[0])
            # 等文件大小稳定
            last_size = -1
            for j in range(15):
                cur = os.path.getsize(new_file)
                if cur == last_size and cur > 0:
                    break
                last_size = cur
                time.sleep(0.5)
            break

    if not new_file:
        print('X 60 秒未见新文件')
        ws.close()
        return 1

    # 7. 改名为 罐头爆款_YYYYMMDD_HHMMSS.xlsx,放到 当天日期子目录里
    now = datetime.now()
    ts = now.strftime('%Y%m%d_%H%M%S')
    day = now.strftime('%Y%m%d')
    day_dir = os.path.join(DL, f'罐头爆款_{day}')
    os.makedirs(day_dir, exist_ok=True)

    final_xlsx = os.path.join(day_dir, f'罐头爆款_{ts}.xlsx')
    shutil.move(new_file, final_xlsx)
    print(f'★ 工作版: {final_xlsx} ({os.path.getsize(final_xlsx)} 字节)')

    # 7b. 自动备份原始版(同目录 + _原始 后缀)
    backup_path = os.path.join(day_dir, f'罐头爆款_{ts}_原始.xlsx')
    shutil.copy(final_xlsx, backup_path)
    print(f'★ 原始备份: {backup_path}')

    # 8. 提取 URL 列输出 .txt
    import openpyxl
    wb = openpyxl.load_workbook(final_xlsx, read_only=True)
    ws_x = wb.active
    header = [c.value for c in next(ws_x.iter_rows(max_row=1))]
    try:
        url_idx = header.index('链接')
    except ValueError:
        print('X xlsx 没有"链接"列,跳过 URL 提取')
        ws.close()
        return 0
    urls = []
    for row in ws_x.iter_rows(min_row=2, values_only=True):
        if row[url_idx]:
            urls.append(str(row[url_idx]).strip())

    final_txt = final_xlsx.replace('.xlsx', '_urls.txt')
    with open(final_txt, 'w', encoding='utf-8') as f:
        f.write('\n'.join(urls))
    print(f'★ URL 列表: {final_txt} ({len(urls)} 条)')

    ws.close()

    # 9. 自动跑分类器 (我的删法_v0)
    classifier = os.path.join(os.path.dirname(os.path.abspath(__file__)), '我的删法_v0.py')
    if os.path.exists(classifier):
        print('\n=== 自动跑分类器(我的删法 v0)===')
        import subprocess
        subprocess.run([sys.executable, classifier, final_xlsx], check=False)
    else:
        print(f'\n(未找到 {classifier},跳过分类器步骤)')

    return 0


if __name__ == '__main__':
    sys.exit(main())
