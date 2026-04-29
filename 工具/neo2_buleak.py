#!/usr/bin/env python3
"""扫 neo2 微头条今日运行日志,精确统计每账号成功/失败,按"漏 N 补 N"写进账号配置.xlsx 的"待补漏" sheet。

精确算法 (双 pass):
  Pass 1: 扫日志,给每行打索引,记录所有 [剩余 X 篇] attempt 行的 (line_idx, account, doc)
  Pass 2: 每次 "已移至已发送" 事件 → 找之前最近的 attempt → ok_count[account] += 1
          每次 "X 发布失败" 事件 → 找之前最近的 attempt → fail_count[account] += 1

quota = 5 (微头条标准: 64 账号 × 5 = 320, 实际参与 54 账号 × 5 = 270 = 素材剩余)
"""
import os, re, sys
from datetime import datetime
from collections import defaultdict
import openpyxl
import bisect

LOG = os.path.expanduser("~/Desktop/微头条自动发布/运行报告/20260429/运行日志.txt")
XLSX = os.path.expanduser("~/Desktop/微头条自动发布/账号配置.xlsx")
QUOTA = 5

re_attempt = re.compile(r"\[剩余\s+\d+\s+篇\].*?\]\s+(?:\[补发\]\s+)?(\S+)\s+->\s+(\S+\.docx)")
re_ok      = re.compile(r"已移至已发送")
re_fail    = re.compile(r"X 发布失败")

# Pass 1: 收集所有 attempt 行 (line_idx, account, doc)
attempts = []  # list of (line_idx, account, doc)
ok_lines = []
fail_lines = []
all_lines = []

with open(LOG, encoding="utf-8") as f:
    for idx, line in enumerate(f):
        all_lines.append(line)
        if "[剩余" in line and "篇]" in line and "->" in line and ".docx" in line:
            m = re_attempt.search(line)
            if m:
                attempts.append((idx, m.group(1), m.group(2)))
        if re_ok.search(line):
            ok_lines.append(idx)
        if re_fail.search(line):
            fail_lines.append(idx)

print(f"\n=== 日志统计 ===")
print(f"attempt 行: {len(attempts)}")
print(f"已移至已发送 (=真成功): {len(ok_lines)}")
print(f"X 发布失败: {len(fail_lines)}")

# attempt_line_idxs 用于快速二分查找
attempt_idxs = [a[0] for a in attempts]

def nearest_attempt_before(line_idx):
    """找在 line_idx 之前最后一个 attempt 行,返回该 attempt 的 (account, doc),或 None"""
    pos = bisect.bisect_right(attempt_idxs, line_idx) - 1
    if pos < 0:
        return None
    return attempts[pos]

# Pass 2: 计数 (并去重: 同 attempt 不重复算)
ok_count = defaultdict(int)
fail_count = defaultdict(int)
attempted = set()  # (account, doc) 集合
seen_accounts = set()

# 每个 attempt 的归宿: 'ok' / 'fail' / 'unresolved'
# 用 attempt_idx 作 key 防同一 attempt 重复算
attempt_resolved = {}

# OK 事件: 找前一个 attempt,只算第一次
for ok_idx in ok_lines:
    a = nearest_attempt_before(ok_idx)
    if a is None:
        continue
    a_idx, acc, doc = a
    if attempt_resolved.get(a_idx):
        continue  # 该 attempt 已结案
    attempt_resolved[a_idx] = 'ok'
    ok_count[acc] += 1
    seen_accounts.add(acc)

# FAIL 事件: 找前一个 attempt,只算第一次,且不能跟 OK 撞
for fail_idx in fail_lines:
    a = nearest_attempt_before(fail_idx)
    if a is None:
        continue
    a_idx, acc, doc = a
    if attempt_resolved.get(a_idx):
        continue
    attempt_resolved[a_idx] = 'fail'
    fail_count[acc] += 1
    seen_accounts.add(acc)

# 把所有 attempt 涉及的账号都登记
for a_idx, acc, doc in attempts:
    seen_accounts.add(acc)

# 素材剩余 docx 数
MAT = os.path.expanduser("~/Desktop/微头条自动发布/素材")
mat_remaining = sum(1 for f in os.listdir(MAT) if f.endswith(".docx"))

# 加载 xlsx
wb = openpyxl.load_workbook(XLSX)

target_accounts = sorted(seen_accounts)

print(f"\n=== 今日总览 ===")
print(f"实际成功 (已移至已发送): {sum(ok_count.values())}")
print(f"实际失败 (X 发布失败): {sum(fail_count.values())}")
print(f"未结案 attempt (脚本中断/无后续): {len(attempts) - sum(1 for v in attempt_resolved.values())}")
print(f"参与账号: {len(seen_accounts)}")
print(f"素材剩余: {mat_remaining}")
print(f"理论值 64 账号 × 5 = 320; 实际参与 {len(seen_accounts)} × 5 = {len(seen_accounts)*5}")

# 算补漏 base
print(f"\n=== 每账号统计(base quota={QUOTA}) ===")
buleak_dict = {}  # account -> miss
for name in target_accounts:
    ok = ok_count.get(name, 0)
    fail = fail_count.get(name, 0)
    miss = QUOTA - ok
    if miss > 0:
        buleak_dict[name] = miss
    print(f"  {name}: ok={ok} fail={fail} base_miss={miss}")

base_total = sum(buleak_dict.values())
print(f"\nbase 补漏: {base_total} 篇,素材剩余 {mat_remaining} 篇")

# 让 sum = mat_remaining (一一对应素材,不多不少)
delta = mat_remaining - base_total
if delta > 0:
    # 素材有富余 → 分给 fail 多的账号 +1
    # 优先级: fail desc, ok asc
    candidates = sorted(buleak_dict.keys(),
                        key=lambda n: (-fail_count.get(n, 0), ok_count.get(n, 0)))
    print(f"\n=== 富余 {delta} 篇分给 fail 多的账号 +1 ===")
    given = 0
    for name in candidates:
        if given >= delta:
            break
        buleak_dict[name] += 1
        given += 1
        print(f"  +1 给 {name} (fail={fail_count.get(name,0)})")
elif delta < 0:
    # 素材不够 → 削减 ok 多的账号
    candidates = sorted(buleak_dict.keys(),
                        key=lambda n: (-ok_count.get(n, 0), -fail_count.get(n, 0)))
    print(f"\n=== 短缺 {-delta} 篇,从 ok 多的账号削 1 ===")
    cut = 0
    for name in candidates:
        if cut >= -delta:
            break
        if buleak_dict[name] > 0:
            buleak_dict[name] -= 1
            cut += 1
            print(f"  -1 从 {name}")
            if buleak_dict[name] == 0:
                del buleak_dict[name]

buleak = sorted(buleak_dict.items())
total_miss = sum(m for _,m in buleak)
print(f"\n=== 最终补漏汇总 ===")
print(f"  {len(buleak)} 个账号 / 共漏 {total_miss} 篇 (素材 {mat_remaining})")
assert total_miss == mat_remaining or (delta > 0 and total_miss <= mat_remaining), \
    f"sum {total_miss} != 素材 {mat_remaining}"

# 写"待补漏" sheet (清空 + 重写)
if "待补漏" not in wb.sheetnames:
    ws_b = wb.create_sheet("待补漏")
    ws_b.append(["账号名", "漏发数", "文稿名", "生成时间"])
else:
    ws_b = wb["待补漏"]
    if ws_b.max_row > 1:
        ws_b.delete_rows(2, ws_b.max_row - 1)

ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
for name, miss in buleak:
    ws_b.append([name, miss, "", ts])

# 清空"本轮已发"
if "本轮已发" in wb.sheetnames:
    ws_r = wb["本轮已发"]
    if ws_r.max_row > 1:
        ws_r.delete_rows(2, ws_r.max_row - 1)
    print(f"\n本轮已发 sheet 已清空 (供重新发文)")

wb.save(XLSX)
print(f"\nOK: 已写 {XLSX}")
print(f"  待补漏 sheet {len(buleak)} 行,共 {total_miss} 篇")
