#!/usr/bin/env python3
"""探 neo2 账号配置.xlsx 结构 + 当前数据"""
import openpyxl
import os

XLSX = os.path.expanduser("~/Desktop/微头条自动发布/账号配置.xlsx")
wb = openpyxl.load_workbook(XLSX, read_only=False, data_only=False)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== sheet: {sn}  dim={ws.dimensions}  rows={ws.max_row} cols={ws.max_column} ===")
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True)):
        print(f"  R{ri+1}:", list(row[:10]))
