"""
创作罐头定时发布脚本 - 图文文章版（Mac）
  Mac文章定时自动发布/
  ├── go.command          双击运行
  ├── gtg_timer.py
  ├── 定时发布.xlsx        配置：账号名 | 发布时间 | 文档文件名
  ├── 素材/               放 .docx 文件
  │   └── 已发送/         发完自动移入
  └── 运行报告/YYYYMMDD/

定时发布.xlsx 格式（第1行标题，从第2行起）：
  A列 账号名     B列 发布时间（2026-04-18 14:30）   C列 文档文件名（不含路径，可为空则随机取）
"""

import requests
import base64
import json
import websocket
import time
import os
import shutil
import glob
import sys
import random
import subprocess
import re
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== 配置 =====================
def _find_cdp_port():
    port_file = os.path.expanduser("~/Library/Application Support/创作罐头/DevToolsActivePort")
    try:
        with open(port_file) as f:
            return int(f.readline().strip())
    except Exception:
        return 9225

CDP_URL       = f"http://127.0.0.1:{_find_cdp_port()}"
ACCOUNT_CLASS = "account-RALrbJ"
WAIT_LOAD     = 4
EXCLUDE_ACCOUNTS = ["青春小馆"]
NOFIRST_ACCOUNTS = set()

NO_PROXY = {"http": "", "https": ""}
WS_OPTS  = {"suppress_origin": True}

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DOCS_FOLDER = os.path.join(BASE_DIR, "素材")
SENT_FOLDER = os.path.join(BASE_DIR, "素材", "已发送")
TIMER_EXCEL = os.path.join(BASE_DIR, "定时发布.xlsx")

RUN_REPORT_DIR = None
LOG_FILE       = None
FAIL_FILE      = None

os.environ["NO_PROXY"] = "127.0.0.1,localhost"
# ================================================

VIOLATION_KEYWORDS = {
    "违规/扣分": ["违规", "扣分", "处罚", "警告"],
    "禁言封号": ["禁言", "发言受限", "封禁", "封号"],
}


def _init_run_dir():
    global LOG_FILE, FAIL_FILE, RUN_REPORT_DIR
    ts = datetime.now().strftime("%Y%m%d")
    RUN_REPORT_DIR = os.path.join(BASE_DIR, "运行报告", ts)
    os.makedirs(RUN_REPORT_DIR, exist_ok=True)
    LOG_FILE  = os.path.join(RUN_REPORT_DIR, "运行日志.txt")
    FAIL_FILE = os.path.join(RUN_REPORT_DIR, "失败记录.xlsx")


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    sys.stdout.buffer.write((line + "\n").encode("utf-8", errors="replace"))
    sys.stdout.buffer.flush()
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def write_fail_excel(final_fails):
    if not final_fails:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "失败记录"
    for col, h in enumerate(["时间", "账号名", "定时时间", "文档", "失败原因"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    for row_data in final_fails:
        ws.append(list(row_data))
    try:
        wb.save(FAIL_FILE)
        log(f"失败记录已写入: {FAIL_FILE}")
    except Exception as e:
        log(f"  写入失败记录出错: {e}")


def _read_timer_excel():
    """读取定时发布.xlsx，返回任务列表 [(账号名, 发布时间str, 文档文件名或None), ...]"""
    tasks = []
    if not os.path.exists(TIMER_EXCEL):
        log(f"错误: 定时发布.xlsx 不存在: {TIMER_EXCEL}")
        return tasks
    try:
        wb = openpyxl.load_workbook(TIMER_EXCEL, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            name_v, time_v, doc_v = row[0], row[1], row[2] if len(row) > 2 else None
            if not name_v:
                continue
            name_v = str(name_v).strip()
            if not name_v or name_v.startswith('#'):
                continue
            if time_v is None:
                log(f"  跳过 {name_v}：发布时间为空")
                continue
            if hasattr(time_v, 'strftime'):
                t_str = time_v.strftime("%Y-%m-%d %H:%M")
            else:
                t_str = str(time_v).strip()[:16]
            doc_name = str(doc_v).strip() if doc_v else None
            tasks.append((name_v, t_str, doc_name))
        wb.close()
    except Exception as e:
        log(f"读取定时发布.xlsx失败: {e}")
    return tasks


def get_tabs():
    return requests.get(f"{CDP_URL}/json", timeout=5, proxies=NO_PROXY).json()


def get_main_ws_url():
    tabs = get_tabs()
    for t in tabs:
        if "czgts.cn" in t.get("url", "") and "webSocketDebuggerUrl" in t:
            return t["webSocketDebuggerUrl"]
    raise RuntimeError("找不到主窗口，请确认创作罐头已启动")


def ws_connect(url, timeout=10):
    return websocket.create_connection(url, timeout=timeout, **WS_OPTS)


def cdp(ws, method, params=None, mid=1):
    ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
    deadline = time.time() + 15
    while time.time() < deadline:
        try:
            d = json.loads(ws.recv())
            if d.get("id") == mid:
                return d.get("result", {})
        except:
            pass
    return {}


def js(ws, expr, mid=99):
    r = cdp(ws, "Runtime.evaluate", {"expression": expr}, mid)
    return r.get("result", {}).get("value")


def js_gesture(ws, expr, mid=99):
    """带用户手势的JS执行，用于触发原生文件对话框等需要user activation的操作"""
    r = cdp(ws, "Runtime.evaluate", {"expression": expr, "userGesture": True}, mid)
    return r.get("result", {}).get("value")


_gtg_minimize_recover_count = 0


def ensure_gtg_top():
    """每次cliclick前调用——最小化Code/Terminal等主要遮挡者，并把罐头置顶。
    若检测到创作罐头自身被最小化，主动取消最小化并重置窗口尺寸（自愈），并打日志方便排根因。
    """
    global _gtg_minimize_recover_count
    r = subprocess.run(["osascript", "-e", '''
tell application "Finder"
    set sb to bounds of window of desktop
    set screenW to item 3 of sb
    set screenH to item 4 of sb
end tell
tell application "System Events"
    repeat with pname in {"Code", "Google Chrome", "Safari", "Terminal", "Claude", "Feishu", "WeChat"}
        try
            tell process (pname as text)
                repeat with w in windows
                    try
                        set value of attribute "AXMinimized" of w to true
                    end try
                end repeat
            end tell
        end try
    end repeat
end tell
tell application "创作罐头" to activate
delay 0.2
tell application "System Events"
    tell process "创作罐头"
        try
            if (value of attribute "AXMinimized" of window 1) then
                set value of attribute "AXMinimized" of window 1 to false
                delay 0.5
                tell window 1
                    set position to {0, 25}
                    set size to {screenW, screenH - 25}
                end tell
                return "RECOVERED"
            end if
        end try
    end tell
end tell
return "OK"
'''], capture_output=True, text=True)
    if "RECOVERED" in (r.stdout or ""):
        _gtg_minimize_recover_count += 1
        log(f"  ⚠ 检测到创作罐头窗口被最小化，已自愈（累计 {_gtg_minimize_recover_count} 次）")
    time.sleep(0.3)


def click(ws, x, y, mid):
    p = {"button": "left", "clickCount": 1, "x": x, "y": y,
         "modifiers": 0, "timestamp": time.time() * 1000}
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mousePressed", **p}, mid)
    time.sleep(0.12)
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mouseReleased", **p}, mid + 1)


def scroll_find_account(main_ws, name):
    name_json = json.dumps(name)
    same_count = 0

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 10)
    time.sleep(0.5)

    for _ in range(500):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t === {name_json} || t.startsWith({name_json})){{
                    items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                    return 'found';
                }}
            }}
            return null;
        }})()
        """, 11)

        if v == 'found':
            time.sleep(0.3)
            pos = js(main_ws, f"""
            (function(){{
                var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    if(t === {name_json} || t.startsWith({name_json})){{
                        var r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                        r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                    }}
                }}
                return null;
            }})()
            """, 13)
            if pos:
                return json.loads(pos)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no-container';
            var before = c.scrollTop;
            c.scrollTop += 250;
            return before + '->' + c.scrollTop;
        })()
        """, 12)
        time.sleep(0.25)

        if result and result != "no-container":
            parts = result.split("->")
            if len(parts) == 2 and parts[0] == parts[1]:
                same_count += 1
                if same_count >= 5:
                    break
            else:
                same_count = 0

    return None


def _find_webview_once(main_ws, name):
    partition = js(main_ws, """
    (function(){
        var webviews = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<webviews.length;i++){
            var p = webviews[i].getAttribute('partition');
            if(!p || !p.startsWith('persist:')) continue;
            var r = webviews[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = p; }
        }
        return best;
    })()
    """, 15)

    if not partition:
        return None

    marker = f"_mk{random.randint(100000, 999999)}"
    r = cdp(main_ws, "Runtime.evaluate", {
        "expression": f"""
        new Promise(function(resolve){{
            var wv = document.querySelector('webview[partition="{partition}"]');
            if(!wv){{ resolve('no_wv'); return; }}
            wv.executeJavaScript('window.{marker}=1').then(function(){{
                resolve('ok');
            }}).catch(function(){{
                resolve('err');
            }});
        }})
        """,
        "awaitPromise": True,
        "timeout": 8000
    }, 200)

    if r.get("result", {}).get("value") != "ok":
        return None

    tabs = get_tabs()
    for t in tabs:
        if "webSocketDebuggerUrl" not in t:
            continue
        try:
            wsc = ws_connect(t["webSocketDebuggerUrl"], timeout=3)
            wsc.send(json.dumps({"id": 1, "method": "Runtime.evaluate",
                                 "params": {"expression": f"window.{marker}===1"}}))
            deadline = time.time() + 3
            found = False
            while time.time() < deadline:
                d = json.loads(wsc.recv())
                if d.get("id") == 1:
                    found = d.get("result", {}).get("result", {}).get("value") is True
                    break
            wsc.close()
            if found:
                return t["webSocketDebuggerUrl"]
        except:
            pass

    return None


def find_account_webview(main_ws, name):
    for retry in range(3):
        result = _find_webview_once(main_ws, name)
        if result:
            return result
        if retry < 2:
            log(f"  重试 webview ({retry+2}/3)...")
            time.sleep(2)
    return None


def detect_account_error(wsc):
    page_text = js(wsc, "document.body.innerText || ''", 70) or ""
    for reason, keywords in {
        "失登": ["请登录", "登录已失效", "账号已下线", "重新登录"],
        "封号": ["账号已被封禁", "账号异常", "账号被封"],
        "禁言": ["账号被禁言", "发言受限", "无法发布"],
        "限流": ["操作频繁", "请稍后再试"],
    }.items():
        for kw in keywords:
            if kw in page_text:
                return reason
    return None


def close_popup(ws):
    v = js(ws, """
    (function(){
        var b = document.querySelector('.close-btn,[class*="close-btn"]');
        if(b){var r=b.getBoundingClientRect();
        if(r.width>0) return JSON.stringify({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});}
        return null;
    })()
    """, 50)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], 51)
        time.sleep(0.5)


def close_current_tab(main_ws):
    v = js(main_ws, """
    (function(){
        var closes = document.querySelectorAll('.chrome-tab-close');
        if(closes.length === 0) return null;
        for(var i=0;i<closes.length;i++){
            var tab = closes[i].closest('.chrome-tab');
            if(tab && tab.classList.contains('active')){
                var r = closes[i].getBoundingClientRect();
                if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
        }
        var last = closes[closes.length-1];
        var r = last.getBoundingClientRect();
        return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
    })()
    """, 55)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 56)
        time.sleep(0.5)


# ========== 发布流程（定时发布版） ==========

def publish_article_timer(ws_url, doc_path, main_ws, account_name, timer_time):
    """定时发布一篇文章，timer_time格式: YYYY-MM-DD HH:MM"""
    try:
        wsc = ws_connect(ws_url, timeout=10)
    except Exception as e:
        return False, f"连接失败: {e}"

    close_popup(wsc)
    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    js(wsc, "location.href='https://mp.toutiao.com/profile_v4/graphic/publish'", 60)
    wsc.close()
    time.sleep(4)

    new_ws_url = None
    for _ in range(12):
        tabs = get_tabs()
        for t in tabs:
            if "graphic/publish" in t.get("url", "") and "webSocketDebuggerUrl" in t:
                new_ws_url = t["webSocketDebuggerUrl"]
                break
        if new_ws_url:
            break
        time.sleep(0.5)

    if not new_ws_url:
        return False, "导航到发文页失败"

    try:
        wsc = ws_connect(new_ws_url, timeout=10)
    except Exception as e:
        return False, f"重连失败: {e}"

    close_popup(wsc)
    time.sleep(1)

    current_url = js(wsc, "location.href", 59) or ""
    if "login" in current_url:
        wsc.close()
        return False, "失登"

    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    # 取 webview 屏幕坐标
    wv_s = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 61)

    if not wv_s:
        wsc.close()
        return False, "无法获取 webview 屏幕坐标"
    wv0 = json.loads(wv_s)

    def get_wv():
        r = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxArea = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var area = r.width * r.height;
                if(area > maxArea){ maxArea = area; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
        })()
        """, 62)
        return json.loads(r) if r else wv0

    # 关闭草稿提示条
    draft_close = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t === '\u7ee7\u7eed\u7f16\u8f91'){
                var bar = els[i].closest('[class]');
                while(bar){
                    var x = bar.querySelector('[class*="close"],[class*="Close"]');
                    if(x){ var r = x.getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}); }
                    bar = bar.parentElement && bar.parentElement.closest('[class]');
                    if(!bar) break;
                }
            }
        }
        return null;
    })()
    """, 58)
    if draft_close:
        dc = json.loads(draft_close)
        wv_t = get_wv()
        subprocess.run(["cliclick", f"c:{wv_t['sx']+dc['x']},{wv_t['sy']+dc['y']}"], capture_output=True)
        time.sleep(0.5)
    else:
        log("  无草稿提示条")

    # 激活窗口
    subprocess.run(["osascript", "-e", """
tell application "System Events"
    tell process "创作罐头"
        set frontmost to true
    end tell
end tell
"""], capture_output=True)
    time.sleep(0.6)

    # 点文档导入按钮（工具栏最后一个）
    v = None
    for _ in range(20):
        v = js(wsc, """
        (function(){
            var btns = Array.from(document.querySelectorAll('.syl-toolbar-button')).filter(function(b){
                return b.getBoundingClientRect().width > 0;
            });
            if(btns.length > 0){
                var last = btns[btns.length - 1];
                var r = last.getBoundingClientRect();
                return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
            return null;
        })()
        """, 63)
        if v: break
        time.sleep(0.5)
    if not v:
        wsc.close()
        return False, "找不到文档导入按钮"

    p = json.loads(v)
    wv_t = get_wv()
    import_x = wv_t['sx'] + p['x']
    import_y = wv_t['sy'] + p['y']
    title_x = wv_t['sx'] + 400
    title_y = wv_t['sy'] + 50
    ensure_gtg_top()
    subprocess.run(["cliclick", f"c:{title_x},{title_y}"], capture_output=True)
    time.sleep(0.5)
    log(f"  cliclick 点击文档导入 ({import_x},{import_y})")
    ensure_gtg_top()
    subprocess.run(["cliclick", f"m:{import_x},{import_y}"], capture_output=True)
    time.sleep(0.3)
    subprocess.run(["cliclick", f"c:{import_x},{import_y}"], capture_output=True)
    time.sleep(1.5)

    # 等"选择文档"按钮
    sel = None
    for _ in range(60):
        sel = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim() === '\u9009\u62e9\u6587\u6863'){
                    var r = btns[i].getBoundingClientRect();
                    if(r.width > 0) return JSON.stringify({bx: Math.round(r.left+r.width/2), by: Math.round(r.top+r.height/2)});
                }
            }
            return null;
        })()
        """, 65)
        if sel: break
        time.sleep(0.5)
    if not sel:
        wsc.close()
        return False, "文档导入弹窗未出现"

    sb = json.loads(sel)
    wv_t = get_wv()
    screen_x = wv_t['sx'] + sb['bx']
    screen_y = wv_t['sy'] + sb['by']

    doc_escaped = doc_path.replace("\\", "/")
    # AppleScript 字符串里的反斜杠和双引号要转义（路径里通常没有，但保险）
    safe_path = doc_escaped.replace("\\", "\\\\").replace('"', '\\"')
    result_holder = [None]

    def sheet_exists():
        r = subprocess.run([
            "osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of window 1)'
        ], capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def go_to_folder_sheet_exists():
        r = subprocess.run([
            "osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of sheet 1 of window 1)'
        ], capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def get_sheet_rect():
        """拿主对话框 position+size。NSOpenPanel 在新 macOS 里 entire contents 遍历超时，
        但 position/size 直接读很快。返回 (x,y,w,h) 或 None。"""
        r = subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        try
            set {px, py} to position of sheet 1 of window 1
            set {sw, sh} to size of sheet 1 of window 1
            return (px as string) & "|" & (py as string) & "|" & (sw as string) & "|" & (sh as string)
        on error
            return ""
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=5)
        try:
            nums = re.findall(r"\d+", r.stdout)
            if len(nums) >= 4:
                return tuple(int(v) for v in nums[:4])
        except Exception:
            pass
        return None

    def click_dialog_button(which):
        """which: 'open' 或 'cancel'。NSOpenPanel 右下角按钮坐标相对固定，
        用 cliclick 物理点击不经 keystroke，不受 frontmost 焦点影响。"""
        rect = get_sheet_rect()
        if not rect:
            return False
        x, y, w, h = rect
        if which == 'open':
            cx, cy = x + w - 55, y + h - 35
        else:
            cx, cy = x + w - 150, y + h - 35
        subprocess.run(["cliclick", f"m:{cx},{cy}"], capture_output=True)
        time.sleep(0.1)
        subprocess.run(["cliclick", f"c:{cx},{cy}"], capture_output=True)
        time.sleep(0.5)
        return True

    def press_esc(times=2):
        """兜底关对话框。优先 cliclick 点'取消'（物理层不受焦点影响），
        不行再 fallback key code 53。"""
        click_dialog_button('cancel')
        for _ in range(times):
            if not sheet_exists():
                return
            subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        key code 53
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.4)

    def fill_dialog():
        # 等文件对话框 sheet 出现
        deadline = time.time() + 15
        appeared = False
        while time.time() < deadline:
            if sheet_exists():
                appeared = True
                break
            time.sleep(0.3)
        if not appeared:
            result_holder[0] = False
            return
        time.sleep(0.6)
        # 方案A：直接给"前往文件夹"输入框赋值，绕过 clipboard。
        # clipboard 方案会被输入法/剪贴板管理器/循环间隔偶发污染，定时发布多账号连发时尤甚。
        # UI 路径由 probe_dialog.py 探测确认：text field 1 of sheet 1 of sheet 1 of window 1
        r = subprocess.run(["osascript", "-e", f'''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {{command down, shift down}}
        delay 1.5
        try
            set target to text field 1 of sheet 1 of sheet 1 of window 1
            set value of target to "{safe_path}"
            delay 0.3
            set rb to (value of target) as string
            if rb is "{safe_path}" then
                return "OK"
            else
                return "ERR:readback:" & rb
            end if
        on error errmsg
            return "ERR:" & errmsg
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=20)
        direct_ok = "OK" in (r.stdout or "")
        if not direct_ok:
            log(f"  直接赋值失败（{(r.stdout or '').strip()[:80]}），回退clipboard")
            # 关掉可能残留的"前往文件夹"小框
            press_esc(1)
            time.sleep(0.5)
            # 方案B：pbcopy + pbpaste 校验 + 最多 5 次重试，挡住 clipboard 被污染
            clipboard_ok = False
            for _ in range(5):
                subprocess.run(["pbcopy"], input=doc_escaped, text=True)
                time.sleep(0.15)
                rb = subprocess.run(["pbpaste"], capture_output=True, text=True).stdout
                if rb.strip() == doc_escaped.strip():
                    clipboard_ok = True
                    break
                time.sleep(0.2)
            if not clipboard_ok:
                log("  clipboard 校验 5 次失败")
                result_holder[0] = False
                return
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {command down, shift down}
        delay 1.5
        keystroke "a" using {command down}
        delay 0.4
        keystroke "v" using {command down}
        delay 1.2
    end tell
end tell
'''], capture_output=True)

        # Step 2: 按回车关"前往文件夹"小框，最多 5 次（keystroke 偶发焦点问题需检测+重试）
        go_closed = False
        for i in range(5):
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.1
tell application "System Events"
    tell process "创作罐头"
        keystroke return
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.8)
            if not go_to_folder_sheet_exists():
                go_closed = True
                if i > 0:
                    log(f"  前往文件夹 回车{i+1}次后关闭")
                break
        if not go_closed:
            log("  前往文件夹 5次回车未关 → cliclick 点取消")
            click_dialog_button('cancel')
            result_holder[0] = False
            return

        # Step 3: 等主对话框自动关闭（完整文件路径会被 NSOpenPanel 直接打开）
        for _ in range(12):  # 最多 6 秒
            time.sleep(0.5)
            if not sheet_exists():
                result_holder[0] = True
                return

        # Step 4: 主框没自动关 → cliclick 物理点"打开"按钮（最多 3 次）
        log("  主对话框未自动关闭 → cliclick 点打开按钮")
        for _ in range(3):
            if not click_dialog_button('open'):
                break
            for _ in range(4):  # 等 2 秒
                time.sleep(0.5)
                if not sheet_exists():
                    result_holder[0] = True
                    return

        # Step 5: 彻底卡死 → 点取消，外层 3 次重试会重开"选择文档"
        log("  对话框完全卡死 → cliclick 点取消")
        click_dialog_button('cancel')
        result_holder[0] = False

    # 重试最多 3 次，扛住偶发对话框 hang
    dialog_ok = False
    for dialog_attempt in range(3):
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        ensure_gtg_top()
        subprocess.run(["cliclick", f"m:{screen_x},{screen_y}"], capture_output=True)
        time.sleep(0.2)
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=30)
        if result_holder[0]:
            dialog_ok = True
            if dialog_attempt > 0:
                log(f"  文件对话框第{dialog_attempt+1}次成功")
            break
        # 失败前再保险按一次 ESC 确保对话框关闭，否则下次循环点"选择文档"会被堵
        press_esc(2)
        time.sleep(1)
        log(f"  第{dialog_attempt+1}次对话框处理失败，准备重试")

    if not dialog_ok:
        wsc.close()
        return False, "文件对话框反复卡住，3次重试均失败"

    time.sleep(5)
    char_count = 0
    for _ in range(15):
        v = js(wsc, """
        (function(){
            var el = document.querySelector('.ProseMirror');
            if(!el) return 0;
            return el.textContent.trim().length;
        })()
        """, 75)
        char_count = int(v) if v else 0
        if char_count >= 50:
            break
        time.sleep(0.8)

    log(f"  文章字数: {char_count}")
    if char_count < 50:
        wsc.close()
        return False, "文档导入后内容为空"

    js(wsc, """
    (function(){
        var el = document.querySelector('.ProseMirror') || document.body;
        el.scrollTop = el.scrollHeight;
        window.scrollTo(0, document.body.scrollHeight);
    })()
    """, 76)
    time.sleep(2)
    for _ in range(20):
        loading = js(wsc, """
        (function(){
            var imgs = document.querySelectorAll('img');
            for(var i=0;i<imgs.length;i++){
                if(!imgs[i].complete || imgs[i].naturalWidth === 0) return true;
            }
            return false;
        })()
        """, 77)
        if not loading:
            break
        time.sleep(0.5)

    # 读信用分 / 处理首发
    credit_raw = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t.indexOf('\u4fe1\u7528\u5206') !== -1 && t.length < 20){
                var m = t.match(/(\d{1,3})\u5206/);
                if(m){ var n = parseInt(m[1], 10);
                    if(n % 5 === 0 && n >= 5 && n <= 100) return n; }
            }
        }
        return null;
    })()
    """, 78)
    credit_score = int(credit_raw) if credit_raw is not None else None
    log(f"  信用分: {credit_score if credit_score is not None else '未读取到'}")
    should_first = (account_name not in NOFIRST_ACCOUNTS) and (credit_score is not None and credit_score >= 95)

    first_result = js(wsc, f"""
    (function(){{
        var shouldCheck = {'true' if should_first else 'false'};
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){{
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '\u5934\u6761\u9996\u53d1'){{
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){{
                    if(p.tagName === 'LABEL' && p.classList.contains('byte-checkbox')){{
                        var isChecked = p.classList.contains('byte-checkbox-checked');
                        if(isChecked === shouldCheck) return JSON.stringify({{already: true, checked: isChecked}});
                        var r = p.getBoundingClientRect();
                        if(r.width > 0 && r.height > 0)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        break;
                    }}
                    p = p.parentElement;
                }}
            }}
        }}
        return null;
    }})()
    """, 79)
    if first_result:
        fr = json.loads(first_result)
        if "already" in fr:
            log(f"  头条首发: 已是{'勾选' if fr['checked'] else '未勾选'}，无需操作")
        elif "x" in fr:
            wv_r = get_wv()
            subprocess.run(["cliclick", f"c:{wv_r['sx']+fr['x']},{wv_r['sy']+fr['y']}"], capture_output=True)
            time.sleep(0.3)
            log(f"  头条首发: {'勾选' if should_first else '取消勾选'}")
    else:
        log("  头条首发: 未找到复选框")

    # ---- 定时发布流程 ----
    wv_cur = get_wv()

    # cliclick 真实macOS鼠标点击（JS/CDP合成事件过不了React的 isTrusted 检查）
    # 关键：先scrollIntoView把按钮滚到视口中心，再换算成屏幕坐标
    popup_opened = False
    for attempt in range(5):
        pos = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim() === '\u5b9a\u65f6\u53d1\u5e03' && !btns[i].disabled){
                    var r = btns[i].getBoundingClientRect();
                    if(r.width > 0) return JSON.stringify({x: Math.round(r.left+r.width/2), y: Math.round(r.top+r.height/2)});
                }
            }
            return null;
        })()
        """, 149 + attempt * 3)
        if not pos:
            if attempt == 0:
                wsc.close()
                return False, "找不到定时发布按钮"
            time.sleep(1); continue
        p = json.loads(pos)
        # 每次重新取 wv 坐标（罐头layout可能变）
        wv_cur_r = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxA = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var a = r.width*r.height;
                if(a > maxA){ maxA = a; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX+best.left), sy: Math.round(window.screenY+best.top)});
        })()
        """, 150 + attempt * 3)
        wv_cur = json.loads(wv_cur_r) if wv_cur_r else get_wv()
        sx = wv_cur['sx'] + p['x']
        sy = wv_cur['sy'] + p['y']
        ensure_gtg_top()
        attempt_str = f" [第{attempt+1}次]" if attempt > 0 else ""
        log(f"  cliclick 点定时发布 ({sx},{sy}){attempt_str}")
        subprocess.run(["cliclick", f"m:{sx},{sy}"], capture_output=True)
        time.sleep(0.5)
        subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
        time.sleep(3)  # 罐头响应延迟
        # 等最多15秒看弹窗是否真打开
        for _ in range(30):
            v_cnt = js(wsc, """(function(){var bs=document.querySelectorAll('button');for(var i=0;i<bs.length;i++){var t=bs[i].textContent.trim();if(t==='\u9884\u89c8\u5e76\u5b9a\u65f6\u53d1\u5e03')return 1;}return 0;})()""", 151 + attempt * 3)
            if v_cnt and int(v_cnt) >= 1:
                popup_opened = True
                break
            time.sleep(0.5)
        if popup_opened:
            log(f"  弹窗已打开（第{attempt+1}次成功）")
            break
        log(f"  第{attempt+1}次点击后弹窗未开，重试")
    if not popup_opened:
        wsc.close()
        return False, "定时发布点击后弹窗始终未打开"
    p_t = {'y': -1}  # 占位

    # 等弹窗出现
    popup_ok = False
    for _ in range(20):
        v_popup = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim().indexOf('\u5b9a\u65f6') !== -1)
                    return 'found';
            }
            return null;
        })()
        """, 151)
        if v_popup == 'found':
            popup_ok = True
            break
        time.sleep(0.5)

    if not popup_ok:
        wsc.close()
        return False, "定时发布弹窗未出现"

    # 解析目标时间
    try:
        dt = datetime.strptime(timer_time, "%Y-%m-%d %H:%M")
    except Exception:
        wsc.close()
        return False, f"定时时间格式错误: {timer_time}"

    t_date1 = f"{dt.month}月{dt.day}日"
    t_date2 = f"{dt.month:02d}月{dt.day:02d}日"
    t_hour1 = str(dt.hour)
    t_hour2 = f"{dt.hour:02d}"
    t_min   = f"{dt.minute:02d}"

    # 弹窗DOM探测调试已移除（遍历全DOM会15s超时拖累流程）

    def click_select_option(select_idx, targets, mid_base):
        """点开第 select_idx 个(0-based)下拉，找 targets 中任一文字点击"""
        # 等到对应序号的 select-view 渲染出来（弹窗是渐进渲染的）
        slist = []
        for _ in range(30):
            sels_r = js(wsc, """
            (function(){
                var sels = document.querySelectorAll('[class*="arco-select-view-value"],[class*="arco-select-view-input"],[class*="select-view-value"]');
                var out = [];
                for(var i=0;i<sels.length;i++){
                    var r = sels[i].getBoundingClientRect();
                    if(r.width > 0 && r.height > 0)
                        out.push({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});
                }
                return JSON.stringify(out);
            })()
            """, mid_base)
            if sels_r:
                slist = json.loads(sels_r)
                if len(slist) > select_idx: break
            time.sleep(0.3)
        if select_idx >= len(slist):
            log(f"  select_idx={select_idx} 超出范围(共{len(slist)}个)")
            return False
        s = slist[select_idx]
        wv_now = get_wv()
        sx = wv_now['sx'] + s['x']
        sy = wv_now['sy'] + s['y']
        log(f"  cliclick 点开第{select_idx+1}个下拉 ({sx},{sy})")
        subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
        time.sleep(0.8)
        # 下拉打开后先滚到顶部，避免默认滚到当前时间导致目标选项在视野外
        js(wsc, """
        (function(){
            var lists = document.querySelectorAll('[class*="arco-select-popup"],[class*="arco-virtual-list"],[class*="select-popup"]');
            for(var i=0;i<lists.length;i++){
                var r = lists[i].getBoundingClientRect();
                if(r.width > 0 && r.height > 0){
                    lists[i].scrollTop = 0;
                    // 虚拟滚动内部容器也滚到顶
                    var inner = lists[i].querySelector('[class*="arco-virtual-list"]') || lists[i];
                    inner.scrollTop = 0;
                }
            }
        })()
        """, mid_base-1)
        time.sleep(0.3)
        for _ in range(20):  # 多给点循环，配合逐步滚动找到目标
            # 先尝试JS直接click（绕开坐标越界问题）
            js_clicked = js(wsc, f"""
            (function(){{
                var targets = {json.dumps(targets)};
                var items = document.querySelectorAll('[class*="arco-select-option"],[class*="select-option"],[class*="dropdown"] li,li[class*="option"]');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    for(var j=0;j<targets.length;j++){{
                        if(t === targets[j]){{
                            items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                            var ev = ['mousedown','mouseup','click'];
                            for(var k=0;k<ev.length;k++){{
                                items[i].dispatchEvent(new MouseEvent(ev[k],{{bubbles:true,cancelable:true,view:window}}));
                            }}
                            return 'clicked:' + targets[j];
                        }}
                    }}
                }}
                return null;
            }})()
            """, mid_base+1)
            if js_clicked:
                log(f"  JS点击 {js_clicked}")
                time.sleep(0.5)
                return True
            # 找不到则往下滚动下拉容器（已先滚到顶，逐步往下扫描所有选项）
            js(wsc, """
            (function(){
                var lists = document.querySelectorAll('[class*="arco-select-popup"],[class*="arco-virtual-list"],[class*="select-popup"]');
                for(var i=0;i<lists.length;i++){
                    var r = lists[i].getBoundingClientRect();
                    if(r.width > 0 && r.height > 0){
                        lists[i].scrollTop += 80;
                        var inner = lists[i].querySelector('[class*="arco-virtual-list"]') || lists[i];
                        inner.scrollTop += 80;
                    }
                }
            })()
            """, mid_base+2)
            time.sleep(0.2)
        log(f"  未找到目标选项: {targets}")
        return False

    # 等弹窗里的第一个下拉出现即可（click_select_option内部会自己等到3个都到位）
    time.sleep(1.2)

    if not click_select_option(0, [t_date1, t_date2], 160):
        wsc.close()
        return False, f"定时日期设置失败: {t_date1}"
    if not click_select_option(1, [t_hour1, t_hour2], 170):
        wsc.close()
        return False, f"定时小时设置失败: {t_hour1}"
    t_min_alt = str(dt.minute)
    if not click_select_option(2, [t_min, t_min_alt], 180):
        wsc.close()
        return False, f"定时分钟设置失败: {t_min}"

    log(f"  定时时间设置完成: {timer_time}")
    time.sleep(0.5)

    # JS 直接点"预览并定时发布"按钮
    prev_clicked = js(wsc, """
    (function(){
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            var t = btns[i].textContent.trim();
            if(t.indexOf('\u9884\u89c8') !== -1 && t.indexOf('\u5b9a\u65f6') !== -1 && !btns[i].disabled){
                btns[i].click();
                return 'clicked';
            }
        }
        return null;
    })()
    """, 190)
    if prev_clicked != 'clicked':
        wsc.close()
        return False, "找不到预览并定时发布按钮"
    log("  JS点击预览并定时发布")
    time.sleep(3)

    # 打印当前页面所有可见按钮（调试用）
    btns_now = js(wsc, """
    (function(){
        var out = [];
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            var r = btns[i].getBoundingClientRect();
            if(r.width > 0 && r.height > 0){
                out.push(btns[i].textContent.trim());
            }
        }
        return JSON.stringify(out);
    })()
    """, 192)
    log(f"  当前可见按钮: {btns_now}")

    # JS 直接点预览页的"定时发布"按钮（跟编辑页同名——预览层是后渲染的，取最后一个）
    # click 后做二次校验：按钮没消失就补点，避免click事件被React吞掉导致丢一篇
    click_expr = """
    (function(){
        var btns = document.querySelectorAll('button');
        var matched = [];
        for(var i=0;i<btns.length;i++){
            var t = btns[i].textContent.trim();
            if(t === '\u5b9a\u65f6\u53d1\u5e03' && !btns[i].disabled){
                var r = btns[i].getBoundingClientRect();
                if(r.width > 0 && r.height > 0) matched.push(btns[i]);
            }
        }
        if(matched.length === 0) return null;
        matched[matched.length - 1].click();
        return 'clicked:' + matched.length;
    })()
    """
    confirm_clicked = False
    for _ in range(60):
        time.sleep(0.5)
        v2 = js(wsc, click_expr, 191)
        if not v2:
            continue
        log(f"  JS点击预览页定时发布 ({v2})")
        confirm_clicked = True
        # 二次校验：等1.5秒看按钮是否消失/URL是否跳转，没动就补点
        for retry in range(2):
            time.sleep(1.5)
            still = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '\u5b9a\u65f6\u53d1\u5e03' && !btns[i].disabled){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0 && r.height > 0) return 1;
                    }
                }
                return 0;
            })()
            """, 193 + retry)
            if not still or int(still) == 0:
                break
            log(f"  按钮仍在，补点一次（retry={retry+1}）")
            js(wsc, click_expr, 195 + retry)
        break

    if not confirm_clicked:
        wsc.close()
        return False, "未出现预览页定时发布按钮"

    # 检测发布成功
    published = False
    for _ in range(20):
        time.sleep(0.5)
        t_txt = js(wsc, """
        (function(){
            var all = document.querySelectorAll('*');
            for(var i=0;i<all.length;i++){
                var t = all[i].textContent.trim();
                if((t==='发布成功！'||t==='发布成功'||t==='提交成功！'||t==='提交成功'||t==='定时发布成功')
                   && all[i].children.length<=1)
                    return t;
            }
            return null;
        })()
        """, 196)
        if t_txt:
            published = True
            break
        cur_url = js(wsc, "location.href", 197) or ""
        if "graphic" in cur_url and "publish" not in cur_url:
            published = True
            break

    if not published:
        err_after = detect_account_error(wsc)
        wsc.close()
        if err_after:
            return False, err_after
        return False, "未检测到发布成功"

    wsc.close()
    log(f"  OK 定时发布成功 → {timer_time}")
    return True, "成功"


def get_docs():
    docs = []
    for p in ["*.docx", "*.doc"]:
        docs.extend(glob.glob(os.path.join(DOCS_FOLDER, p)))
    return sorted([d for d in docs if "已发送" not in d])


def move_to_sent(doc_path):
    os.makedirs(SENT_FOLDER, exist_ok=True)
    dest = os.path.join(SENT_FOLDER, os.path.basename(doc_path))
    if os.path.exists(dest):
        name, ext = os.path.splitext(os.path.basename(doc_path))
        dest = os.path.join(SENT_FOLDER, f"{name}_{int(time.time())}{ext}")
    shutil.move(doc_path, dest)
    log(f"  已移至已发送: {os.path.basename(dest)}")


# ========== 主流程 ==========

def main():
    _init_run_dir()
    log("=" * 50)
    log("创作罐头图文文章定时发布 Mac版 启动")
    log(f"报告目录: {RUN_REPORT_DIR}")
    log("=" * 50)

    os.makedirs(DOCS_FOLDER, exist_ok=True)
    os.makedirs(SENT_FOLDER, exist_ok=True)

    # 读取任务配置
    tasks = _read_timer_excel()
    if not tasks:
        log("错误: 定时发布.xlsx 为空或不存在，请先填好配置")
        return

    log(f"共 {len(tasks)} 个定时发布任务:")
    for name, t_time, doc_name in tasks:
        log(f"  {name}  →  {t_time}  →  {doc_name or '(随机取素材)' }")

    # 准备文档池
    all_docs = get_docs()
    doc_map = {os.path.basename(d): d for d in all_docs}

    # 获取主窗口
    try:
        main_ws_url = get_main_ws_url()
    except RuntimeError as e:
        log(f"错误: {e}")
        return

    main_ws = ws_connect(main_ws_url, timeout=10)
    log("已连接主窗口")

    # 隐藏所有非罐头的可见应用 + 罐头按当前屏幕分辨率最大化（不写死像素，每台机子都适配）
    subprocess.run(["osascript", "-e", '''
tell application "Finder"
    set sb to bounds of window of desktop
    set screenW to item 3 of sb
    set screenH to item 4 of sb
end tell
tell application "System Events"
    repeat with p in (every process whose visible is true and background only is false)
        set pname to name of p
        if pname is not "创作罐头" and pname is not "罐头" and pname is not "Finder" then
            try
                set visible of p to false
            end try
        end if
    end repeat
end tell
tell application "创作罐头" to activate
delay 0.5
tell application "System Events"
    tell process "创作罐头"
        set frontmost to true
        tell window 1
            set position to {0, 25}
            set size to {screenW, screenH - 25}
        end tell
    end tell
end tell
'''], capture_output=True)
    time.sleep(1.0)
    log("已隐藏其他应用+罐头按屏幕分辨率最大化")

    fail_records = []
    success_count = 0

    for idx, (name, timer_time, doc_name) in enumerate(tasks):
        log(f"\n{'='*40}")
        log(f"任务 {idx+1}/{len(tasks)}: {name}  →  {timer_time}")

        # 确定文档路径
        if doc_name:
            doc_path = doc_map.get(doc_name)
            if not doc_path:
                log(f"  X 素材文件夹中找不到: {doc_name}")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, doc_name or "", "文档不存在"))
                continue
        else:
            # r[3] 存的是 basename，all_docs 是绝对路径，要按 basename 比对
            failed_basenames = {r[3] for r in fail_records if r[3]}
            remaining = [d for d in all_docs if os.path.basename(d) not in failed_basenames]
            if not remaining:
                remaining = all_docs
            if not remaining:
                log("  X 素材文件夹为空")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", "素材为空"))
                continue
            doc_path = random.choice(remaining)

        log(f"  文档: {os.path.basename(doc_path)}")

        # 找账号
        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未找到账号: {name}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), "侧边栏未找到"))
            continue

        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD)

        ws_url = find_account_webview(main_ws, name)
        if not ws_url:
            log("  X 找不到 webview")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), "webview匹配失败"))
            close_current_tab(main_ws)
            continue

        try:
            success, reason = publish_article_timer(ws_url, doc_path, main_ws, name, timer_time)
            if success:
                move_to_sent(doc_path)
                success_count += 1
                log(f"  ✓ 定时发布成功: {name}")
            else:
                log(f"  X 发布失败: {reason}")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), reason))
        except Exception as e:
            log(f"  X 异常: {e}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), f"异常: {e}"))
        finally:
            close_current_tab(main_ws)
            if idx < len(tasks) - 1:
                _d = random.randint(8, 20)
                log(f"  篇间等待 {_d} 秒...")
                time.sleep(_d)

    write_fail_excel(fail_records)

    main_ws.close()
    log(f"\n{'='*50}")
    log(f"完成! 成功:{success_count}  失败:{len(fail_records)}")
    log(f"{'='*50}")


if __name__ == "__main__":
    main()
