"""
[v1101] 2026-04-28 缺哥拍 — 大统一基线
v1101 改动:Step 3 6s 删 + ProseMirror 取最长 + 字数<50 重试 fill_dialog 1 次

创作罐头批量发布脚本 v3
项目目录结构:
  GTG_XXX/
  ├── 启动.bat
  ├── gtg_batch.py
  ├── 素材/           (放 .docx 文件)
  │   └── 已发送/     (发完自动移入)
  ├── gtg_log.txt     (运行日志)
  ├── 失败记录.xlsx   (发布失败的账号)
  ├── 高阅读量提醒.txt (阅读量超标时写入)
  └── 系统通知提醒.txt (检测到处罚消息时写入)
"""

import requests
import json
import websocket
import time
import os
import shutil
import glob
import sys
import random
import subprocess
import re
import threading
from datetime import datetime, timedelta
import docx as docxlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== 配置 =====================
def _find_cdp_port():
    port_file = os.path.expanduser("~/Library/Application Support/创作罐头/DevToolsActivePort")
    try:
        with open(port_file) as f:
            return int(f.readline().strip())
    except Exception:
        return 9223  # fallback

CDP_URL        = f"http://127.0.0.1:{_find_cdp_port()}"   # 自动检测罐头CDP端口
DEFAULT_TAG    = "原机构老号"               # 启动时未选择则使用此默认标签
ACCOUNT_CLASS  = "account-RALrbJ"
WAIT_LOAD      = 4
ALERT_THRESHOLD = 5000                     # 阅读量超过此值时写入提醒文件
EXCLUDE_ACCOUNTS = ["青春小馆"]            # 永不发文的账号（母账号等）
NOFIRST_ACCOUNTS = set()                   # 不选头条首发的账号（从nofirst.txt加载）

NO_PROXY       = {"http": "", "https": ""}
WS_OPTS        = {"suppress_origin": True}

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DOCS_FOLDER = os.path.join(BASE_DIR, "素材")
SENT_FOLDER = os.path.join(BASE_DIR, "素材", "已发送")
# 运行报告路径（main()开始时动态初始化）
RUN_REPORT_DIR   = None
LOG_FILE         = None
FAIL_FILE        = None
NOTICE_FILE      = None
ALERT_FILE       = None
VIOLATION_FILE   = None

VIOLATION_KEYWORDS = {
    "违规/扣分": ["违规", "扣分", "处罚", "警告"],
    "禁言封号": ["禁言", "发言受限", "封禁", "封号"],
    "原创侵权": ["原创违规", "侵权", "重复内容"],
}

# 绕过系统代理（避免 127.0.0.1 被代理拦截）
os.environ["NO_PROXY"] = "127.0.0.1,localhost"
# ================================================

# ========== 账号配置 Excel ==========
CONFIG_EXCEL  = os.path.join(BASE_DIR, "账号配置.xlsx")
SUMMARY_EXCEL = os.path.join(BASE_DIR, "发文汇总.xlsx")
_EXCEL_SHEETS = ["不首发", "永久跳过", "本轮已发", "白名单", "失败列表", "硬终止账号"]
_FAIL_HEADERS    = ["账号名", "失败原因", "文稿名", "失败时间", "轮次"]
_SUMMARY_HEADERS = ["账号名", "轮次", "发文时间", "失败时间", "补发成功时间"]
_HARD_TERMINATE_HEADERS = ["账号名", "终止原因", "终止时间", "本次已发篇数"]

# 4 类硬终止 reason — 命中即永久放弃,不再尝试
HARD_TERMINATE_REASONS = {"封号", "禁言", "侧边栏未找到", "信用分过低"}  # [v1101.1] 失登移出, 加信用分过低

def _ensure_config_excel():
    """如果账号配置.xlsx不存在则自动创建；已存在则确保关键 sheet 有表头"""
    if not os.path.exists(CONFIG_EXCEL):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sh in _EXCEL_SHEETS:
            ws_new = wb.create_sheet(sh)
            if sh == "失败列表":
                ws_new.append(_FAIL_HEADERS)
            elif sh == "硬终止账号":
                ws_new.append(_HARD_TERMINATE_HEADERS)
            elif sh == "白名单":
                ws_new.append(["账号名", "发文份数"])
            else:
                ws_new.append(["账号名"])
        wb.save(CONFIG_EXCEL)
        return
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        dirty = False
        if "失败列表" not in wb.sheetnames:
            ws_n = wb.create_sheet("失败列表")
            ws_n.append(_FAIL_HEADERS); dirty = True
        if "硬终止账号" not in wb.sheetnames:
            ws_n = wb.create_sheet("硬终止账号")
            ws_n.append(_HARD_TERMINATE_HEADERS); dirty = True
        if dirty:
            wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _ensure_summary_excel():
    if os.path.exists(SUMMARY_EXCEL):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发文汇总"
        ws.append(_SUMMARY_HEADERS)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass

def _read_excel_sheet(sheet_name):
    """读取账号配置.xlsx指定sheet的A列账号（跳过第1行标题，忽略空行和#开头）"""
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb[sheet_name]
        result = []
        for row in ws_r.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip() and not str(val).strip().startswith('#'):
                result.append(str(val).strip())
        wb.close()
        return result
    except Exception:
        return []

def _append_sent_excel(name):
    """往账号配置.xlsx的本轮已发sheet追加一行"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "本轮已发" not in wb.sheetnames:
            ws_s = wb.create_sheet("本轮已发")
            ws_s.append(["账号名"])
        else:
            ws_s = wb["本轮已发"]
        ws_s.append([name])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _append_fail_list(name, reason, doc_name, round_num):
    """失败列表追加一条（本轮内失败记录，供轮末补发）"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "失败列表" not in wb.sheetnames:
            ws_f = wb.create_sheet("失败列表")
            ws_f.append(_FAIL_HEADERS)
        else:
            ws_f = wb["失败列表"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_f.append([name, reason or "", doc_name or "", ts, round_num])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _read_fail_list():
    """读取失败列表 → [(name, reason, doc_name, ts, round_num), ...]"""
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "失败列表" not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb["失败列表"]
        result = []
        for row in ws_r.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            reason  = str(row[1]).strip() if row[1] is not None else ""
            docname = str(row[2]).strip() if row[2] is not None else ""
            ts      = str(row[3]).strip() if row[3] is not None else ""
            rnd     = int(row[4]) if (row[4] is not None and str(row[4]).strip().isdigit()) else 0
            result.append((name, reason, docname, ts, rnd))
        wb.close()
        return result
    except Exception:
        return []

def _remove_from_fail_list(name):
    """补发成功后从失败列表里删该账号所有记录"""
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "失败列表" not in wb.sheetnames:
            wb.close(); return
        ws_f = wb["失败列表"]
        rows_to_delete = []
        for idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] and str(row[0]).strip() == name:
                rows_to_delete.append(idx)
        for r_idx in reversed(rows_to_delete):
            ws_f.delete_rows(r_idx)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _append_summary(name, round_num, event_type, ts=None):
    """往独立 发文汇总.xlsx 追加/更新。event_type: '发文时间'/'失败时间'/'补发成功时间'"""
    if ts is None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    col_map = {"发文时间": 3, "失败时间": 4, "补发成功时间": 5}
    col = col_map.get(event_type)
    if col is None:
        return
    try:
        _ensure_summary_excel()
        wb = openpyxl.load_workbook(SUMMARY_EXCEL)
        ws_s = wb["发文汇总"] if "发文汇总" in wb.sheetnames else wb.create_sheet("发文汇总")
        if ws_s.max_row == 0:
            ws_s.append(_SUMMARY_HEADERS)
        target_row = None
        for r_idx in range(2, ws_s.max_row + 1):
            a = ws_s.cell(r_idx, 1).value
            b = ws_s.cell(r_idx, 2).value
            if a and str(a).strip() == name and b is not None and int(b) == round_num:
                target_row = r_idx
                break
        if target_row is None:
            ws_s.append([name, round_num, "", "", ""])
            target_row = ws_s.max_row
        ws_s.cell(target_row, col, ts)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass

def _append_hard_terminate(name, reason, count_so_far):
    """命中 4 类硬终止时,写入"硬终止账号"sheet。永久放弃,不再尝试。"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "硬终止账号" not in wb.sheetnames:
            ws_h = wb.create_sheet("硬终止账号")
            ws_h.append(_HARD_TERMINATE_HEADERS)
        else:
            ws_h = wb["硬终止账号"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_h.append([name, reason, ts, count_so_far])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _clear_round_sheets():
    """轮末齐活后清空 本轮已发 和 失败列表（保留表头）"""
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        for sn in ("本轮已发", "失败列表"):
            if sn in wb.sheetnames:
                ws_c = wb[sn]
                if ws_c.max_row > 1:
                    ws_c.delete_rows(2, ws_c.max_row - 1)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _sort_summary_by_account():
    """按账号分组排序 发文汇总.xlsx"""
    try:
        if not os.path.exists(SUMMARY_EXCEL):
            return
        wb = openpyxl.load_workbook(SUMMARY_EXCEL)
        if "发文汇总" not in wb.sheetnames:
            wb.close(); return
        ws_s = wb["发文汇总"]
        data = []
        for r_idx in range(2, ws_s.max_row + 1):
            row = [ws_s.cell(r_idx, c).value for c in range(1, 6)]
            if row[0]:
                data.append(row)
        data.sort(key=lambda r: (str(r[0]), int(r[1]) if r[1] is not None else 0))
        if ws_s.max_row > 1:
            ws_s.delete_rows(2, ws_s.max_row - 1)
        for row in data:
            ws_s.append(row)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass

# =====================================


def _init_run_dir():
    global LOG_FILE, FAIL_FILE, NOTICE_FILE, ALERT_FILE, VIOLATION_FILE, RUN_REPORT_DIR
    ts = datetime.now().strftime("%Y%m%d")
    RUN_REPORT_DIR = os.path.join(BASE_DIR, "运行报告", ts)
    os.makedirs(RUN_REPORT_DIR, exist_ok=True)
    LOG_FILE       = os.path.join(RUN_REPORT_DIR, "运行日志.txt")
    FAIL_FILE      = os.path.join(RUN_REPORT_DIR, "失败记录.xlsx")
    NOTICE_FILE    = os.path.join(RUN_REPORT_DIR, "系统通知.txt")
    ALERT_FILE     = os.path.join(RUN_REPORT_DIR, "高阅读提醒.txt")
    VIOLATION_FILE = os.path.join(RUN_REPORT_DIR, "违规提醒.txt")


# ========== 日志 ==========

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    sys.stdout.buffer.write((line + "\n").encode("utf-8", errors="replace"))
    sys.stdout.buffer.flush()
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ========== 状态文件 (实时刷新) ==========

STATUS_FILE = os.path.join(BASE_DIR, "当前状态.txt")

def write_status(big_round=None, sub_round=None, phase=None,
                 total_done=0, total_target=0, doc_pool_size=0,
                 sub_failed=0, dead_terminated=None, extra=""):
    """实时刷新状态文件,你随时打开就能看到当前进度"""
    try:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        lines = [f"[更新于 {ts}]"]
        if big_round is not None:
            lines.append(f"当前: 第 {big_round} 大循环 / 第 {sub_round} 小轮 / {phase}")
        lines.append(f"累计已发: {total_done} / {total_target} 篇")
        lines.append(f"文档池剩余: {doc_pool_size} 篇")
        if sub_failed:
            lines.append(f"本小轮失败队列: {sub_failed} 个账号")
        dead_terminated = dead_terminated or {}
        lines.append(f"硬终止: {len(dead_terminated)} 个" +
                     (f" {list(dead_terminated.keys())}" if dead_terminated else ""))
        if extra:
            lines.append(extra)
        with open(STATUS_FILE, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
    except Exception:
        pass


# ========== 失败记录 Excel ==========

def write_fail_excel(final_fails):
    """final_fails: [(ts, account_name, reason), ...] 只写最终失败账号"""
    if not final_fails:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "失败记录"
    for col, h in enumerate(["时间", "账号名", "失败原因"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    for ts_f, n_f, r_f in final_fails:
        ws.append([ts_f, n_f, r_f])
    try:
        wb.save(FAIL_FILE)
        log(f"最终失败记录已写入: {FAIL_FILE}")
    except Exception as e:
        log(f"  写入失败记录出错: {e}")


# ========== CDP 基础 ==========

def get_tabs():
    return requests.get(f"{CDP_URL}/json", timeout=5, proxies=NO_PROXY).json()


def get_main_ws_url():
    tabs = get_tabs()
    for t in tabs:
        if "czgts.cn" in t.get("url", "") and "webSocketDebuggerUrl" in t:
            return t["webSocketDebuggerUrl"]
    raise RuntimeError(f"找不到主窗口，请确认创作罐头已用 --remote-debugging-port={CDP_URL.split(':')[-1]} 启动")


def ws_connect(url, timeout=10):
    return websocket.create_connection(url, timeout=timeout, **WS_OPTS)


def cdp(ws, method, params=None, mid=1):
    ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
    deadline = time.time() + 15
    while time.time() < deadline:
        try:
            d = json.loads(ws.recv())
            if d.get("id") == mid:
                return d.get("result", {})
        except:
            pass
    return {}


def js(ws, expr, mid=99):
    r = cdp(ws, "Runtime.evaluate", {"expression": expr}, mid)
    return r.get("result", {}).get("value")


def click(ws, x, y, mid):
    p = {"button": "left", "clickCount": 1, "x": x, "y": y,
         "modifiers": 0, "timestamp": time.time() * 1000}
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mousePressed", **p}, mid)
    time.sleep(0.12)
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mouseReleased", **p}, mid + 1)


def click_text(ws, text, mid, area_top=None, area_bottom=None):
    text_json = json.dumps(text)
    area_check = ""
    if area_top is not None:
        area_check += f" && r.top >= {area_top}"
    if area_bottom is not None:
        area_check += f" && r.top < {area_bottom}"
    v = js(ws, f"""
    (function(){{
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){{
            var t = all[i].textContent.trim();
            if(t === {text_json} && all[i].children.length <= 2){{
                var r = all[i].getBoundingClientRect();
                if(r.width > 0{area_check})
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, mid)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], mid + 1)
        return True
    return False


# ========== 打开筛选面板 ==========

def _open_filter_panel(main_ws):
    """点击筛选图标，返回是否成功打开"""
    v = js(main_ws, """
    (function(){
        var candidates = document.querySelectorAll('[class*="filter"],[class*="Filter"],[class*="funnel"],[class*="screen"]');
        for(var i=0;i<candidates.length;i++){
            var r = candidates[i].getBoundingClientRect();
            if(r.width > 0 && r.width < 60 && r.top > 0 && r.top < 200)
                return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
        }
        var search = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(search){
            var r = search.getBoundingClientRect();
            return JSON.stringify({x:Math.round(r.right+20), y:Math.round(r.top+r.height/2)});
        }
        return null;
    })()
    """, 102)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 103)
        time.sleep(1.2)
        return True
    return False


def _close_filter_panel(main_ws):
    cdp(main_ws, "Input.dispatchKeyEvent", {
        "type": "keyDown", "key": "Escape", "code": "Escape", "windowsVirtualKeyCode": 27
    }, 109)
    time.sleep(0.5)


# ========== 交互式标签选择 ==========

def get_available_tags(main_ws):
    """
    打开筛选面板，读取所有可用标签列表。
    返回标签名列表，如读取失败返回空列表。
    """
    log("读取可用标签列表...")

    click_text(main_ws, "账号管理", 101, area_top=0, area_bottom=800)
    time.sleep(1.5)

    opened = _open_filter_panel(main_ws)
    if not opened:
        log("  警告: 未能打开筛选面板，跳过标签读取")
        return []

    tags_raw = js(main_ws, """
    (function(){
        var SKIP = ['全部','确定','取消','清除','清除筛选','确认','搜索','标签','分组',
                    '账号类型','平台','状态','按名称','排序','筛选'];
        // 尝试找筛选弹出容器
        var panel = document.querySelector(
            '[class*="popover"],[class*="Popover"],[class*="dropdown"],[class*="Dropdown"],' +
            '[class*="filter-panel"],[class*="filterPanel"],[class*="FilterPanel"]'
        );
        var root = panel || document.body;
        var items = root.querySelectorAll('span,label,li');
        var tags = [];
        for(var i=0;i<items.length;i++){
            var t = items[i].textContent.trim();
            if(!t || t.length < 1 || t.length > 20) continue;
            if(items[i].children.length > 0) continue;
            if(SKIP.indexOf(t) !== -1) continue;
            if(/^\d+$/.test(t)) continue;
            var r = items[i].getBoundingClientRect();
            if(r.width > 0 && r.top > 100) tags.push(t);
        }
        return JSON.stringify([...new Set(tags)]);
    })()
    """, 108)

    _close_filter_panel(main_ws)

    try:
        tags = json.loads(tags_raw) if tags_raw else []
    except:
        tags = []

    log(f"  读取到 {len(tags)} 个标签: {tags}")
    return tags


def select_tag_interactively(main_ws):
    """
    从创作罐头读取标签列表，让用户交互选择。
    返回选定的标签名。
    """
    tags = get_available_tags(main_ws)

    print("\n" + "=" * 45)
    if tags:
        print("检测到以下标签：")
        for i, tag in enumerate(tags, 1):
            print(f"  {i}. {tag}")
        print(f"\n默认标签: {DEFAULT_TAG}")
        print("请输入标签序号 (直接回车使用默认):")
        try:
            choice = input("> ").strip()
        except EOFError:
            choice = ""
        if choice.isdigit() and 1 <= int(choice) <= len(tags):
            selected = tags[int(choice) - 1]
        else:
            selected = DEFAULT_TAG
    else:
        print(f"未能读取标签列表，使用默认标签: {DEFAULT_TAG}")
        selected = DEFAULT_TAG

    print(f"已选标签: {selected}")
    print("=" * 45 + "\n")
    log(f"本次使用标签: {selected}")
    return selected


# ========== 标签筛选 ==========

def setup_tag_filter(main_ws, tag_name):
    log(f"设置标签筛选: {tag_name}")

    ok = click_text(main_ws, "账号管理", 101, area_top=0, area_bottom=800)
    if not ok:
        log("  警告: 未找到账号管理导航，尝试继续")
    time.sleep(1.5)

    _open_filter_panel(main_ws)

    ok = click_text(main_ws, "清除筛选", 104)
    if ok:
        time.sleep(0.8)
        log("  已清除筛选")

    tag_json = json.dumps(tag_name)
    v = js(main_ws, f"""
    (function(){{
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){{
            var t = all[i].textContent.trim();
            if(t === {tag_json}){{
                var r = all[i].getBoundingClientRect();
                if(r.width > 0)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 105)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 106)
        time.sleep(1)
        log(f"  已勾选标签: {tag_name}")
    else:
        log(f"  警告: 未找到标签 [{tag_name}]，将处理所有可见账号")
        _close_filter_panel(main_ws)
        return False

    _close_filter_panel(main_ws)
    return True


# ========== 收集账号列表 ==========

def collect_accounts(main_ws):
    log("开始收集账号列表...")

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 110)
    time.sleep(0.8)

    accounts = []
    seen = set()
    same_count = 0
    has_scrolled = False  # 曾经成功滚动过才开始计退出条件

    for _ in range(1000):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            var names = [];
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t) names.push(t);
            }}
            return JSON.stringify(names);
        }})()
        """, 111)

        if v:
            names = json.loads(v)
            for n in names:
                if n not in seen:
                    seen.add(n)
                    accounts.append(n)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no';
            var before = c.scrollTop;
            c.scrollTop += 200;
            return before + '->' + c.scrollTop;
        })()
        """, 112)
        time.sleep(1.0)

        if result and result != "no":
            parts = result.split("->")
            if len(parts) == 2:
                before_top = parts[0].strip()
                after_top = parts[1].strip()
                if after_top != '0' and after_top != before_top:
                    has_scrolled = True
                    same_count = 0
                elif before_top == after_top and has_scrolled:
                    same_count += 1
                    if same_count >= 6:
                        break

    # 兜底:虚拟滚动列表,最后几个账号常因 lazy render 漏读
    # 强制滚到容器最底部,等 1.5s DOM 稳定再扫一次
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = c.scrollHeight;
    })()
    """, 113)
    time.sleep(1.5)
    v = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        var names = [];
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t) names.push(t);
        }}
        return JSON.stringify(names);
    }})()
    """, 114)
    if v:
        names = json.loads(v)
        added = 0
        for n in names:
            if n not in seen:
                seen.add(n)
                accounts.append(n)
                added += 1
        if added:
            log(f"  兜底:滚到底再扫一次,补收 {added} 个账号(虚拟滚动 lazy render 漏发)")

    log(f"共收集到 {len(accounts)} 个账号")
    return accounts


# ========== 滚动查找并点击账号 ==========

def scroll_find_account(main_ws, name):
    name_json = json.dumps(name)
    same_count = 0

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 10)
    time.sleep(0.5)

    for _ in range(500):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t === {name_json} || t.startsWith({name_json})){{
                    items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                    return 'found';
                }}
            }}
            return null;
        }})()
        """, 11)

        if v == 'found':
            time.sleep(0.3)
            pos = js(main_ws, f"""
            (function(){{
                var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    if(t === {name_json} || t.startsWith({name_json})){{
                        // patch_lastacc: items[i] 是外层 account row (cliclick 不响应),
                        // 真可点击 row 是其子树中的 horizontalAccount-* (x≈83 中心).
                        var hor = items[i].querySelector('[class*="horizontalAccount"]');
                        var row = (hor && hor.getBoundingClientRect().width>=60) ? hor : items[i];
                        var r = row.getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        // 坐标不在视口内，再滚一次 (基于 row)
                        row.scrollIntoView({{block:'center', behavior:'instant'}});
                        r = row.getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                    }}
                }}
                return null;
            }})()
            """, 13)
            if pos:
                return json.loads(pos)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no-container';
            var before = c.scrollTop;
            c.scrollTop += 250;
            return before + '->' + c.scrollTop;
        })()
        """, 12)
        time.sleep(0.25)

        if result and result != "no-container":
            parts = result.split("->")
            if len(parts) == 2 and parts[0] == parts[1]:
                same_count += 1
                if same_count >= 5:
                    break
            else:
                same_count = 0

    # 滚动找不到,CDP JS 注入搜索框过滤(neo 同款 — 不需罐头前台,VSCode 占着不影响)
    # 直接调 React/Vue controlled input 内部 setter + dispatch input/change,绕开物理键盘
    log(f"  滚动未找到账号,CDP 注入搜索框过滤: {name}")
    filled = js(main_ws, f"""
    (function(){{
        var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(!s) return 'no_input';
        var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        setter.call(s, {name_json});
        s.dispatchEvent(new Event('input', {{bubbles:true}}));
        s.dispatchEvent(new Event('change', {{bubbles:true}}));
        return 'ok';
    }})()
    """, 14)
    if filled != 'ok':
        return None
    time.sleep(1.5)
    # 取过滤后的目标账号坐标
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 18)
    if pos:
        log(f"  搜索框过滤后找到账号: {name}")
        # CDP 注入清空搜索框,恢复完整列表
        js(main_ws, """
        (function(){
            var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
            if(!s) return 'no_input';
            var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(s, '');
            s.dispatchEvent(new Event('input', {bubbles:true}));
            s.dispatchEvent(new Event('change', {bubbles:true}));
            return 'ok';
        })()
        """, 19)
        time.sleep(0.5)
        return json.loads(pos)
    return None


# ========== webview 精确匹配（含3次重试） ==========

def _find_webview_once(main_ws, name):
    partition = js(main_ws, """
    (function(){
        var webviews = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<webviews.length;i++){
            var p = webviews[i].getAttribute('partition');
            if(!p || !p.startsWith('persist:')) continue;
            var r = webviews[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = p; }
        }
        return best;
    })()
    """, 15)

    if not partition:
        log("  警告: 未找到 partition")
        return None

    marker = f"_mk{random.randint(100000, 999999)}"
    r = cdp(main_ws, "Runtime.evaluate", {
        "expression": f"""
        new Promise(function(resolve){{
            var wv = document.querySelector('webview[partition="{partition}"]');
            if(!wv){{ resolve('no_wv'); return; }}
            wv.executeJavaScript('window.{marker}=1').then(function(){{
                resolve('ok');
            }}).catch(function(){{
                resolve('err');
            }});
        }})
        """,
        "awaitPromise": True,
        "timeout": 8000
    }, 200)

    inject_result = r.get("result", {}).get("value")
    if inject_result != "ok":
        log(f"  警告: 标记注入结果={inject_result}")
        return None

    tabs = get_tabs()
    for t in tabs:
        if "webSocketDebuggerUrl" not in t:
            continue
        try:
            wsc = ws_connect(t["webSocketDebuggerUrl"], timeout=3)
            wsc.send(json.dumps({"id": 1, "method": "Runtime.evaluate",
                                 "params": {"expression": f"window.{marker}===1"}}))
            deadline = time.time() + 3
            found = False
            while time.time() < deadline:
                d = json.loads(wsc.recv())
                if d.get("id") == 1:
                    found = d.get("result", {}).get("result", {}).get("value") is True
                    break
            wsc.close()
            if found:
                log("  webview 匹配成功")
                return t["webSocketDebuggerUrl"]
        except:
            pass

    return None


def find_account_webview(main_ws, name):
    for retry in range(3):
        result = _find_webview_once(main_ws, name)
        if result:
            return result
        if retry < 2:
            log(f"  重试 webview ({retry+2}/3)...")
            time.sleep(2)
    return None


def get_url_from_ws(ws_url):
    try:
        wsc = ws_connect(ws_url, timeout=4)
        v = js(wsc, "location.href", 1)
        wsc.close()
        return v or ""
    except:
        return ""


# ========== 系统通知检测 ==========

def check_system_notice(ws_url, account_name):
    """
    导航到消息中心 → 分别点击系统通知、审核通知频道
    → 只读取当天的消息原文写入 NOTICE_FILE
    """
    try:
        today = datetime.now()
        today_full = today.strftime("%Y-%m-%d")
        today_short = today.strftime("%m-%d")
        yesterday = today - timedelta(days=1)
        yesterday_full = yesterday.strftime("%Y-%m-%d")
        yesterday_short = yesterday.strftime("%m-%d")

        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/personal/message?type=message_letter'", 300)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        notices = []

        for channel in ["系统通知", "审核通知"]:
            channel_json = channel.replace('"', '\"')
            clicked = js(wsc, f"""
            (function(){{
                var spans = document.querySelectorAll('span.name');
                for(var i=0;i<spans.length;i++){{
                    if(spans[i].textContent.trim() === "{channel_json}"){{
                        var box = spans[i].closest('.conversation-box-primary') || spans[i].parentElement;
                        if(box){{ box.click(); return 'ok'; }}
                    }}
                }}
                return null;
            }})()
            """, 301)

            if not clicked:
                log(f"  未找到频道: {channel}")
                continue

            time.sleep(1.5)

            result = js(wsc, f"""
            (function(){{
                var todayFull = "{today_full}";
                var todayShort = "{today_short}";
                var yesterdayFull = "{yesterday_full}";
                var yesterdayShort = "{yesterday_short}";
                var list = document.querySelector('.chat-container-list');
                if(!list) return JSON.stringify([]);
                var items = list.children;
                var results = [];
                var isToday = false;
                for(var i=0;i<items.length;i++){{
                    var cls = items[i].className || '';
                    if(cls.indexOf('time-stamp') !== -1){{
                        var t = items[i].textContent.trim();
                        isToday = t.startsWith(todayFull) || t.startsWith(todayShort) || t.startsWith(yesterdayFull) || t.startsWith(yesterdayShort);
                    }} else if(isToday && cls.indexOf('chat-row') !== -1){{
                        var txt = items[i].textContent.trim();
                        if(txt) results.push(txt.substring(0, 300));
                        isToday = false;
                    }}
                }}
                return JSON.stringify(results);
            }})()
            """, 302)

            if result:
                try:
                    msgs = json.loads(result)
                    for msg in msgs:
                        notices.append(f"【{channel}】{msg}")
                except:
                    pass

        wsc.close()

        violation_count = 0
        if notices:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            # 分析违规类通知
            violations = []
            for n in notices:
                for cat, kws in VIOLATION_KEYWORDS.items():
                    for kw in kws:
                        if kw in n:
                            violations.append((cat, n))
                            break
            # 写系统通知
            content_str = f"[{ts}] 账号 {account_name} 当天通知:\n"
            for n in notices:
                content_str += f"  {n}\n"
            content_str += "\n"
            with open(NOTICE_FILE, "a", encoding="utf-8") as f:
                f.write(content_str)
            log(f"  ⚠ 当天通知 {len(notices)} 条")
            # 写违规提醒
            if violations:
                vcontent = f"[{ts}] 账号 {account_name} 违规/扣分提醒:\n"
                for cat, msg in violations:
                    vcontent += f"  [{cat}] {msg}\n"
                vcontent += "\n"
                with open(VIOLATION_FILE, "a", encoding="utf-8") as f:
                    f.write(vcontent)
                violation_count = len(violations)
                log(f"  ⚠ 违规/扣分 {violation_count} 条 → 违规提醒.txt")
        else:
            log("  系统/审核通知: 当天无新通知")
        return len(notices), violation_count
    except Exception as e:
        log(f"  系统通知检测出错: {e}")
        return 0, 0

# ========== 阅读量检测 ==========

def check_reading_stats(ws_url, account_name):
    """
    导航到数据概览页 → 读取三篇文章的阅读量。
    任一超过 ALERT_THRESHOLD 则写入 ALERT_FILE。
    """
    try:
        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/index'", 310)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        reads_raw = js(wsc, """
        (function(){
            var items = document.querySelectorAll('.recent-works-item');
            var results = [];
            for(var i=0;i<items.length;i++){
                var timeEl = items[i].querySelector('span.time');
                var timeText = timeEl ? timeEl.textContent.trim() : '未知时间';
                var labels = items[i].querySelectorAll('span.label');
                for(var j=0;j<labels.length;j++){
                    if(labels[j].textContent.trim() === '阅读量'){
                        var prev = labels[j].previousElementSibling;
                        if(prev){
                            var num = parseInt(prev.textContent.trim().replace(/,/g,''), 10);
                            if(!isNaN(num)) results.push({count: num, time: timeText});
                        }
                    }
                }
            }
            return JSON.stringify(results);
        })()
        """, 311)
        wsc.close()

        reads = json.loads(reads_raw) if reads_raw else []
        if not reads:
            log("  阅读量: 未读取到数据")
            return 0

        counts = [r['count'] for r in reads]
        log(f"  阅读量: {counts}")
        high = [r for r in reads if r['count'] >= ALERT_THRESHOLD]
        if high:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            lines = [f"[{ts}] 账号 {account_name} 高阅读量:"]
            for r in high:
                lines.append(f"  {r['time']} — {r['count']} 阅读量")
            msg = '\n'.join(lines) + '\n\n'
            with open(ALERT_FILE, "a", encoding="utf-8") as f:
                f.write(msg)
            log(f"  ★ 高阅读量提醒: {[r['count'] for r in high]} → 已写入 {ALERT_FILE}")
            return len(high)
        return 0
    except Exception as e:
        log(f"  阅读量检测出错: {e}")
        return 0


# ========== 账号状态检测 ==========

ERROR_KEYWORDS = {
    "失登": ["请登录", "登录已失效", "账号已下线", "重新登录"],
    "封号": ["账号已被封禁", "账号异常", "账号被封", "账号不可用"],
    "禁言": ["账号被禁言", "发言受限", "无法发布"],
    "限流": ["操作频繁", "请稍后再试", "访问受限"],
}


def detect_account_error(wsc):
    page_text = js(wsc, "document.body.innerText || ''", 70) or ""
    for reason, keywords in ERROR_KEYWORDS.items():
        for kw in keywords:
            if kw in page_text:
                return reason
    return None


# ========== 弹窗 & Tab ==========

def close_popup(ws):
    v = js(ws, """
    (function(){
        var b = document.querySelector('.close-btn,[class*="close-btn"]');
        if(b){var r=b.getBoundingClientRect();
        if(r.width>0) return JSON.stringify({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});}
        return null;
    })()
    """, 50)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], 51)
        time.sleep(0.5)


def close_current_tab(main_ws):
    v = js(main_ws, """
    (function(){
        var closes = document.querySelectorAll('.chrome-tab-close');
        if(closes.length === 0) return null;
        for(var i=0;i<closes.length;i++){
            var tab = closes[i].closest('.chrome-tab');
            if(tab && tab.classList.contains('active')){
                var r = closes[i].getBoundingClientRect();
                if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
        }
        var last = closes[closes.length-1];
        var r = last.getBoundingClientRect();
        return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
    })()
    """, 55)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 56)
        time.sleep(0.5)
        log("  已关闭tab")
    else:
        log("  未找到tab关闭按钮")


# ========== 读取文档 ==========

def read_docx_text(doc_path):
    doc = docxlib.Document(doc_path)
    lines = [p.text for p in doc.paragraphs if p.text.strip()]
    return '\n'.join(lines)


# ========== 发布流程 ==========

def publish_article(ws_url, doc_path, main_ws, account_name="", _credit_out=None):
    import subprocess, threading

    try:
        wsc = ws_connect(ws_url, timeout=10)
    except Exception as e:
        return False, f"连接失败: {e}"

    close_popup(wsc)
    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    js(wsc, "location.href='https://mp.toutiao.com/profile_v4/weitoutiao/publish'", 60)
    wsc.close()
    time.sleep(3)

    for _ in range(12):
        if "weitoutiao/publish" in get_url_from_ws(ws_url):
            break
        time.sleep(0.5)

    try:
        wsc = ws_connect(ws_url, timeout=10)
    except Exception as e:
        return False, f"重连失败: {e}"

    close_popup(wsc)
    time.sleep(0.5)

    current_url = js(wsc, "location.href", 59) or ""
    if "login" in current_url:
        wsc.close()
        return False, "失登"

    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    # 点底部"文档导入"按钮（轮询等待页面加载，最多等10秒）
    v = None
    for _ in range(20):
        v = js(wsc, """
        (function(){
            var els = document.querySelectorAll('*');
            for(var i=0;i<els.length;i++){
                var t = els[i].textContent.trim();
                if(t === '\u6587\u6863\u5bfc\u5165' && els[i].children.length <= 2){
                    var r = els[i].getBoundingClientRect();
                    if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                }
            }
            return null;
        })()
        """, 62)
        if v:
            break
        time.sleep(0.5)

    if not v:
        wsc.close()
        return False, "找不到文档导入按钮"

    p = json.loads(v)

    # 取 webview 真实屏幕坐标，用 cliclick 真实点击（CDP 合成点击无法触发弹窗）
    wv_s = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 61)

    if not wv_s:
        wsc.close()
        return False, "无法获取 webview 屏幕坐标"

    wv0 = json.loads(wv_s)
    import_x = wv0['sx'] + p['x']
    import_y = wv0['sy'] + p['y']
    log(f"  文档导入按钮真实屏幕坐标: ({import_x},{import_y})")

    import subprocess, threading
    # [v1101 P5] 强化 activate:unhide + AXRaise + verify frontmost + 重试 3 次
    subprocess.run(["osascript", "-e", """
tell application "创作罐头" to activate
delay 0.2
tell application "System Events"
    tell process "创作罐头"
        try
            if visible is false then set visible to true
        end try
        set frontmost to true
        try
            perform action "AXRaise" of window 1
        end try
        repeat 3 times
            if frontmost then exit repeat
            set frontmost to true
            delay 0.3
        end repeat
    end tell
end tell
"""], capture_output=True)
    time.sleep(0.6)

    # 激活后重新取坐标，防止窗口移动
    wv_s2 = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 62)
    if wv_s2:
        wv0b = json.loads(wv_s2)
        import_x = wv0b['sx'] + p['x']
        import_y = wv0b['sy'] + p['y']
        log(f"  激活后重新校正坐标: ({import_x},{import_y})")

    # 等"已恢复上次编辑"提示条消失（最多等8秒）
    for _ in range(16):
        has_dismiss = js(wsc, """
        (function(){
            var els = document.querySelectorAll('*');
            for(var i=0;i<els.length;i++){
                if(els[i].textContent.trim() === '\u649e\u9500'){
                    var r = els[i].getBoundingClientRect();
                    if(r.width > 0) return true;
                }
            }
            return false;
        })()
        """, 68)
        if not has_dismiss:
            break
        time.sleep(0.5)
    log("  准备点击文档导入")
    # 先点编辑区空白处让 webview 获得焦点
    editor_x = wv0b['sx'] + 400 if wv_s2 else wv0['sx'] + 400
    editor_y = wv0b['sy'] + 200 if wv_s2 else wv0['sy'] + 200
    subprocess.run(["cliclick", f"c:{editor_x},{editor_y}"], capture_output=True)
    time.sleep(0.5)
    # [v1101 P7] cliclick 文档导入 + 等弹窗,失败重试 3 次(每次重 activate + 重读坐标)
    sel = None
    for click_attempt in range(3):
        attempt_str = f" [第{click_attempt+1}次]" if click_attempt > 0 else ""
        log(f"  cliclick 点击文档导入 ({import_x},{import_y}){attempt_str}")
        subprocess.run(["cliclick", f"c:{import_x},{import_y}"], capture_output=True)
        time.sleep(1.5)

        for i in range(20):
            sel = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '选择文档'){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({bx: Math.round(r.left+r.width/2), by: Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 65)
            if sel:
                break
            if i == 4 and click_attempt == 0:
                all_btns = js(wsc, """
                (function(){
                    var btns = document.querySelectorAll('button');
                    var names = [];
                    for(var i=0;i<btns.length;i++){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) names.push(btns[i].textContent.trim());
                    }
                    return JSON.stringify(names);
                })()
                """, 67)
                log(f"  当前按钮: {all_btns}")
            time.sleep(0.5)

        if sel:
            if click_attempt > 0:
                log(f"  cliclick 第{click_attempt+1}次成功唤出弹窗")
            break

        # 失败 → 重 activate + 重读坐标准备下一次
        if click_attempt < 2:
            log(f"  弹窗未出,重 activate + 重读坐标后重试")
            subprocess.run(["osascript", "-e", """
tell application "创作罐头" to activate
delay 0.2
tell application "System Events"
    tell process "创作罐头"
        try
            if visible is false then set visible to true
        end try
        set frontmost to true
        try
            perform action "AXRaise" of window 1
        end try
    end tell
end tell
"""], capture_output=True)
            time.sleep(0.6)
            wv_s3 = js(main_ws, """
            (function(){
                var wvs = document.querySelectorAll('webview');
                var maxArea = 0, best = null;
                for(var i=0;i<wvs.length;i++){
                    var r = wvs[i].getBoundingClientRect();
                    var area = r.width * r.height;
                    if(area > maxArea){ maxArea = area; best = r; }
                }
                if(!best) return null;
                return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
            })()
            """, 62)
            if wv_s3:
                wv0c = json.loads(wv_s3)
                import_x = wv0c['sx'] + p['x']
                import_y = wv0c['sy'] + p['y']
                log(f"  重试前坐标更新: ({import_x},{import_y})")

    if not sel:
        wsc.close()
        return False, "文档导入弹窗未出现(3 次 cliclick 重试均失败)"

    # 从主窗口取 webview 真实屏幕坐标
    wv_screen = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 66)

    if not wv_screen:
        wsc.close()
        return False, "无法获取 webview 屏幕坐标"

    wv = json.loads(wv_screen)
    sb = json.loads(sel)
    screen_x = wv['sx'] + sb['bx']
    screen_y = wv['sy'] + sb['by']
    log(f"  webview屏幕原点:({wv['sx']},{wv['sy']}) 按钮CSS:({sb['bx']},{sb['by']}) => 真实屏幕:({screen_x},{screen_y})")

    doc_escaped = doc_path.replace("\\", "/")
    safe_path = doc_escaped.replace("\\", "\\\\").replace('"', '\\"')
    result_holder = [None]

    def sheet_exists():
        r = subprocess.run([
            "osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of window 1)'
        ], capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def go_to_folder_sheet_exists():
        r = subprocess.run([
            "osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of sheet 1 of window 1)'
        ], capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def get_sheet_rect():
        r = subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        try
            set {px, py} to position of sheet 1 of window 1
            set {sw, sh} to size of sheet 1 of window 1
            return (px as string) & "|" & (py as string) & "|" & (sw as string) & "|" & (sh as string)
        on error
            return ""
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=5)
        try:
            nums = re.findall(r"\d+", r.stdout)
            if len(nums) >= 4:
                return tuple(int(v) for v in nums[:4])
        except Exception:
            pass
        return None

    def click_dialog_button(which):
        rect = get_sheet_rect()
        if not rect:
            return False
        x, y, w, h = rect
        if which == 'open':
            cx, cy = x + w - 55, y + h - 35
        else:
            cx, cy = x + w - 150, y + h - 35
        subprocess.run(["cliclick", f"m:{cx},{cy}"], capture_output=True)
        time.sleep(0.1)
        subprocess.run(["cliclick", f"c:{cx},{cy}"], capture_output=True)
        time.sleep(0.5)
        return True

    def press_esc(times=2):
        """兜底关对话框：优先 cliclick 点取消（不受焦点影响），不行再 key code 53。"""
        click_dialog_button('cancel')
        for _ in range(times):
            if not sheet_exists():
                return
            subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        key code 53
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.4)

    def fill_dialog():
        # 等文件对话框 sheet 出现（最多 15s）
        deadline = time.time() + 15
        appeared = False
        while time.time() < deadline:
            if sheet_exists():
                appeared = True
                break
            time.sleep(0.3)
        if not appeared:
            result_holder[0] = False
            return
        time.sleep(0.6)

        # Step 1: Cmd+Shift+G 开"前往文件夹" + 直接赋值 text field（绕过 clipboard）
        r = subprocess.run(["osascript", "-e", f'''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {{command down, shift down}}
        delay 1.5
        try
            set target to text field 1 of sheet 1 of sheet 1 of window 1
            set value of target to "{safe_path}"
            delay 0.3
            set rb to (value of target) as string
            if rb is "{safe_path}" then
                return "OK"
            else
                return "ERR:readback:" & rb
            end if
        on error errmsg
            return "ERR:" & errmsg
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=20)
        direct_ok = "OK" in (r.stdout or "")

        if not direct_ok:
            log(f"  直接赋值失败（{(r.stdout or '').strip()[:80]}），回退clipboard")
            subprocess.run(["osascript", "-e",
                'tell application "System Events" to tell process "创作罐头" to key code 53'],
                capture_output=True)
            time.sleep(0.5)
            # pbcopy + pbpaste 校验重试
            clipboard_ok = False
            for _ in range(5):
                subprocess.run(["pbcopy"], input=doc_escaped, text=True)
                time.sleep(0.15)
                rb = subprocess.run(["pbpaste"], capture_output=True, text=True).stdout
                if rb.strip() == doc_escaped.strip():
                    clipboard_ok = True
                    break
                time.sleep(0.2)
            if not clipboard_ok:
                log("  clipboard 校验 5 次失败")
                click_dialog_button('cancel')
                result_holder[0] = False
                return
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {command down, shift down}
        delay 1.5
        keystroke "a" using {command down}
        delay 0.4
        keystroke "v" using {command down}
        delay 1.2
    end tell
end tell
'''], capture_output=True)

        # Step 2: 回车关"前往文件夹"小框，最多 5 次（应对 keystroke 焦点偶发丢失）
        go_closed = False
        for i in range(5):
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.1
tell application "System Events"
    tell process "创作罐头"
        keystroke return
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.8)
            if not go_to_folder_sheet_exists():
                go_closed = True
                if i > 0:
                    log(f"  前往文件夹 回车{i+1}次后关闭")
                break
        if not go_closed:
            log("  前往文件夹 5次回车未关 → cliclick 点取消")
            click_dialog_button('cancel')
            result_holder[0] = False
            return

        # [v1101] Step 3 跳过:macOS 26 NSOpenPanel 不再自动关,Step 4 兜底必须打,6s 硬等纯发呆

        # Step 4: 主对话框没自动关 → cliclick 物理点"打开"按钮
        log("  主对话框未自动关闭 → cliclick 点打开按钮")
        for _ in range(3):
            if not click_dialog_button('open'):
                break
            for _ in range(4):
                time.sleep(0.5)
                if not sheet_exists():
                    result_holder[0] = True
                    return

        # Step 5: 彻底卡死 → cliclick 点取消，外层 3 次重试重开
        log("  对话框完全卡死 → cliclick 点取消")
        click_dialog_button('cancel')
        result_holder[0] = False

    # 罐头置顶
    subprocess.run(["osascript", "-e", """
tell application "System Events"
    tell process "创作罐头"
        set frontmost to true
    end tell
end tell
"""], capture_output=True)
    time.sleep(0.3)

    # 重试最多 3 次扛偶发 hang
    dialog_ok = False
    for dialog_attempt in range(3):
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        log(f"  cliclick 点击 ({screen_x},{screen_y})")
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=30)
        if result_holder[0]:
            dialog_ok = True
            if dialog_attempt > 0:
                log(f"  文件对话框第{dialog_attempt+1}次成功")
            break
        press_esc(2)
        time.sleep(1)
        log(f"  第{dialog_attempt+1}次对话框处理失败，准备重试")

    if not dialog_ok:
        wsc.close()
        return False, "文件对话框反复卡住，3次重试均失败"

    # [v1101] 等内容加载;ProseMirror 取最长元素(避免命中标题 placeholder 5 字)
    def _read_char_count():
        cc = 0
        for _ in range(40):
            v = js(wsc, """
            (function(){
                var els = document.querySelectorAll('.ProseMirror');
                var max = 0;
                for (var i = 0; i < els.length; i++) {
                    var l = els[i].textContent.trim().length;
                    if (l > max) max = l;
                }
                return max;
            })()
            """, 75)
            cc = int(v) if v else 0
            if cc >= 50:
                break
            time.sleep(1.0)
        return cc

    time.sleep(5)
    char_count = _read_char_count()
    log(f"  文章字数: {char_count}")

    # [v1101] 字数 < 50 → 重试 fill_dialog 一次(借鉴 neo retry 机制)
    if char_count < 50:
        log(f"  对话框已关但字数仅 {char_count}（文档未真导入），重试 fill_dialog")
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=30)
        time.sleep(5)
        char_count = _read_char_count()
        log(f"  重试后字数: {char_count}")
        if char_count < 50:
            wsc.close()
            return False, "文档导入后内容为空(重试1次仍空)"

    # 读取信用分，决定是否勾选头条首发
    credit_raw = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t.indexOf('\u4fe1\u7528\u5206') !== -1 && t.length < 20){
                var m = t.match(/(\d{1,3})\u5206/);
                if(m){
                    var n = parseInt(m[1], 10);
                    if(n % 5 === 0 && n >= 5 && n <= 100) return n;
                }
            }
        }
        return null;
    })()
    """, 78)
    credit_score = int(credit_raw) if credit_raw is not None else None
    log(f"  信用分: {credit_score if credit_score is not None else '未读取到'}")
    # [v1101.1] 信用分 < 60 → 硬终止
    if credit_score is not None and credit_score < 60:
        log(f"  ★ 信用分 {credit_score} < 60,硬终止")
        wsc.close()
        return False, "信用分过低"
    if _credit_out is not None:
        _credit_out.append(credit_score)
    should_first = (account_name not in NOFIRST_ACCOUNTS) and (credit_score is not None and credit_score >= 95)

    # 根据信用分设置头条首发复选框
    first_result = js(wsc, f"""
    (function(){{
        var shouldCheck = {'true' if should_first else 'false'};
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){{
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '\u5934\u6761\u9996\u53d1'){{
                // 向上找 LABEL.byte-checkbox，用 byte-checkbox-checked 判断勾选状态
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){{
                    if(p.tagName === 'LABEL' && p.classList.contains('byte-checkbox')){{
                        var isChecked = p.classList.contains('byte-checkbox-checked');
                        if(isChecked === shouldCheck){{
                            return JSON.stringify({{already: true, checked: isChecked}});
                        }}
                        var r = p.getBoundingClientRect();
                        if(r.width > 0 && r.height > 0){{
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        }}
                        break;
                    }}
                    p = p.parentElement;
                }}
            }}
        }}
        return null;
    }})()
    """, 79)
    if first_result:
        fr = json.loads(first_result)
        if "already" in fr:
            log(f"  头条首发: 已是{'勾选' if fr['checked'] else '未勾选'}，无需操作")
        elif "x" in fr:
            # cliclick真实点击（CDP虚拟事件对自定义复选框无效）
            wv_r = js(main_ws, """
            (function(){
                var wvs = document.querySelectorAll('webview');
                var maxArea = 0, best = null;
                for(var i=0;i<wvs.length;i++){
                    var r = wvs[i].getBoundingClientRect();
                    var area = r.width * r.height;
                    if(area > maxArea){ maxArea = area; best = r; }
                }
                if(!best) return null;
                return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
            })()
            """, 80)
            wv = json.loads(wv_r) if wv_r else None
            if wv:
                sx = wv['sx'] + fr['x']
                sy = wv['sy'] + fr['y']
                subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
                log(f"  头条首发 cliclick ({sx},{sy})")
            else:
                log("  头条首发: webview坐标获取失败，跳过点击")
            time.sleep(0.3)
            log(f"  头条首发: {'勾选' if should_first else '取消勾选'}")
    else:
        log("  头条首发: 未找到复选框")

    # 点"发布"
    v = js(wsc, """
    (function(){
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            if(btns[i].textContent.trim()==='发布' && !btns[i].disabled){
                var r = btns[i].getBoundingClientRect();
                if(r.width>0) return JSON.stringify({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});
            }
        }
        return null;
    })()
    """, 90)

    if not v:
        wsc.close()
        return False, "找不到发布按钮"

    p = json.loads(v)
    click(wsc, p["x"], p["y"], 91)
    log("  已点击发布")
    time.sleep(1)

    published = False
    for _ in range(20):
        time.sleep(0.5)
        t2 = js(wsc, """
        (function(){
            var all = document.querySelectorAll('*');
            for(var i=0;i<all.length;i++){
                var t = all[i].textContent.trim();
                if((t==='发布成功！'||t==='发布成功'||t==='提交成功！'||t==='提交成功') && all[i].children.length<=1)
                    return t;
            }
            return null;
        })()
        """, 96)
        if t2:
            published = True
            break
        cur_url = js(wsc, "location.href", 97) or ""
        if "weitoutiao" in cur_url and "publish" not in cur_url:
            published = True
            break

    if not published:
        err_after = detect_account_error(wsc)
        wsc.close()
        if err_after:
            return False, err_after
        return False, "未检测到发布成功"

    wsc.close()
    log("  OK 发布成功")
    return True, "成功"


# ========== 文档管理 ==========

def get_docs():
    docs = []
    for p in ["*.docx", "*.doc"]:
        docs.extend(glob.glob(os.path.join(DOCS_FOLDER, p)))
    return sorted([d for d in docs if "已发送" not in d])


def move_to_sent(doc_path):
    os.makedirs(SENT_FOLDER, exist_ok=True)
    dest = os.path.join(SENT_FOLDER, os.path.basename(doc_path))
    if os.path.exists(dest):
        name, ext = os.path.splitext(os.path.basename(doc_path))
        dest = os.path.join(SENT_FOLDER, f"{name}_{int(time.time())}{ext}")
    shutil.move(doc_path, dest)
    log(f"  已移至已发送: {os.path.basename(dest)}")


# ========== 死磕主循环 (公共函数) ==========

def run_death_grip(
    accounts,
    per_account_quota,
    doc_pool,
    main_ws,
    sub_rounds=5,              # 微头条 5,文章/定时 3
    max_fail_per_sub=3,
    sent_accounts_set=None,
    credit_records=None,
    fail_records=None,         # Mac 风格: list of (ts, name, reason),由 main 持有
    success_accounts=None,     # Mac 风格: set
    initial_acc_count=None,
    initial_dead=None,
    log_label="",
):
    """死磕循环:大循环 N 小轮 + 外层无限磕。
    - 4 类硬终止(失登/封号/禁言/侧边栏未找到) → 即刻永久放弃,写"硬终止账号"sheet
    - 其他失败 → 本小轮内累计 max_fail_per_sub 次跳过本小轮,下小轮恢复
    - N 小轮跑完 = 1 大循环;还有 quota 没满且 doc 没空 → 继续下一大循环(无限磕)
    返回: dict { acc_count, dead_terminated, doc_pool, ok_count, fail_count, total_notices, total_violations, total_alerts }
    """
    sent_accounts_set = sent_accounts_set if sent_accounts_set is not None else set()
    credit_records   = credit_records if credit_records is not None else {}
    fail_records     = fail_records if fail_records is not None else []
    success_accounts = success_accounts if success_accounts is not None else set()
    acc_count        = dict(initial_acc_count) if initial_acc_count else {}
    for a in accounts:
        acc_count.setdefault(a, 0)
    dead_terminated  = dict(initial_dead) if initial_dead else {}

    ok_count = fail_count = 0
    total_notices = total_violations = total_alerts = 0
    big_round = 0

    def _do_publish(name, doc, round_label, is_retry=False):
        nonlocal ok_count, fail_count, total_notices, total_violations, total_alerts
        log(f"\n  {log_label}[剩余 {len(doc_pool)} 篇] {round_label} {'[补发]' if is_retry else ''} {name}  ->  {os.path.basename(doc)}")

        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未在侧边栏找到账号: {name}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, "侧边栏未找到"))
            fail_count += 1
            return False, "侧边栏未找到"

        log(f"  点击坐标({pos['x']},{pos['y']})")
        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD)

        ws_url = find_account_webview(main_ws, name)
        if not ws_url:
            log("  X 找不到 webview")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, "webview匹配失败"))
            fail_count += 1
            close_current_tab(main_ws)
            return False, "webview匹配失败"

        page_url = get_url_from_ws(ws_url)
        if "login" in page_url:
            log("  X 账号失登")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, "失登"))
            fail_count += 1
            close_current_tab(main_ws)
            return False, "失登"

        nc, vc = check_system_notice(ws_url, name)
        total_notices += nc
        total_violations += vc
        time.sleep(2)

        _d_wait = random.randint(8, 20)
        try:
            credit_buf = []
            success, reason = publish_article(ws_url, doc, main_ws, name, credit_buf)
            if credit_buf and credit_buf[0] is not None:
                credit_records[name] = credit_buf[0]
            if success:
                move_to_sent(doc)
                if doc in doc_pool:
                    doc_pool.remove(doc)
                acc_count[name] = acc_count.get(name, 0) + 1
                success_accounts.add(name)
                ok_count += 1
                total_alerts += check_reading_stats(ws_url, name)
                return True, ""
            else:
                log(f"  X 发布失败: {reason}")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, reason))
                fail_count += 1
                return False, reason
        except Exception as e:
            log(f"  X 异常: {e}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, f"异常: {e}"))
            fail_count += 1
            return False, f"异常: {e}"
        finally:
            close_current_tab(main_ws)
            log(f"  篇间等待 {_d_wait} 秒...")
            time.sleep(_d_wait)

    def _is_eligible(name, sub_skipped):
        return (name not in dead_terminated and
                name not in sub_skipped and
                name not in sent_accounts_set and
                acc_count.get(name, 0) < per_account_quota.get(name, 0))

    def _handle_failure(name, reason, sub_fail_count, sub_skipped):
        if reason in HARD_TERMINATE_REASONS:
            cnt_so_far = acc_count.get(name, 0)
            dead_terminated[name] = (reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cnt_so_far)
            _append_hard_terminate(name, reason, cnt_so_far)
            log(f"  ★ 4 类硬终止: {name} -> {reason} (本次已发 {cnt_so_far} 篇,永久放弃)")
            return True
        sub_fail_count[name] = sub_fail_count.get(name, 0) + 1
        if sub_fail_count[name] >= max_fail_per_sub:
            sub_skipped.add(name)
            log(f"  - {name} 本小轮 {max_fail_per_sub} 次失败,跳过本小轮(下小轮恢复)")
        return False

    def _total_target():
        return sum(per_account_quota.get(a, 0) for a in accounts if a not in dead_terminated)

    def _total_done():
        return sum(acc_count.get(a, 0) for a in accounts)

    while True:
        if not doc_pool:
            log(f"\n{log_label}文档池已空,死磕结束")
            break
        active = [a for a in accounts
                  if a not in dead_terminated
                  and acc_count.get(a, 0) < per_account_quota.get(a, 0)]
        if not active:
            log(f"\n{log_label}所有账号都已满 quota,死磕结束")
            break

        big_round += 1
        log(f"\n{'='*20} {log_label}第 {big_round} 大循环 开始 (active {len(active)} 账号,文档池 {len(doc_pool)} 篇) {'='*20}")

        for sub_idx in range(1, sub_rounds + 1):
            log(f"\n----- {log_label}第 {big_round} 大循环 / 第 {sub_idx} 小轮 (Phase A) -----")
            sub_fail_count = {}
            sub_skipped = set()
            sub_round_id = big_round * 100 + sub_idx

            for name in list(accounts):
                if not doc_pool:
                    log(f"  Phase A 中文档池已空,提前结束")
                    break
                if not _is_eligible(name, sub_skipped):
                    continue
                doc = random.choice(doc_pool)
                round_label = f"[大{big_round}/小{sub_idx}/A]"
                write_status(big_round, sub_idx, "Phase A",
                             total_done=_total_done(), total_target=_total_target(),
                             doc_pool_size=len(doc_pool),
                             sub_failed=len(sub_skipped), dead_terminated=dead_terminated)
                success, reason = _do_publish(name, doc, round_label, is_retry=False)
                if success:
                    sent_accounts_set.add(name)
                    _append_sent_excel(name)
                    _append_summary(name, sub_round_id, "发文时间")
                else:
                    _append_fail_list(name, reason, os.path.basename(doc), sub_round_id)
                    _append_summary(name, sub_round_id, "失败时间")
                    _handle_failure(name, reason, sub_fail_count, sub_skipped)

            phase_b_round = 0
            while phase_b_round < max_fail_per_sub and doc_pool:
                phase_b_round += 1
                pending = [f for f in _read_fail_list()
                           if f[4] == sub_round_id
                           and f[0] not in sent_accounts_set
                           and f[0] not in dead_terminated
                           and f[0] not in sub_skipped]
                if not pending:
                    log(f"  Phase B: 失败列表已清空,本小轮齐活")
                    break
                log(f"\n  {log_label}Phase B 第 {phase_b_round}/{max_fail_per_sub} 次, 待补 {len(pending)} 个")
                for f_name, f_reason, f_docname, f_ts, f_rnd in pending:
                    if not _is_eligible(f_name, sub_skipped):
                        continue
                    doc_path = None
                    for d in doc_pool:
                        if os.path.basename(d) == f_docname:
                            doc_path = d
                            break
                    if doc_path is None:
                        if not doc_pool:
                            break
                        doc_path = random.choice(doc_pool)
                    round_label = f"[大{big_round}/小{sub_idx}/B-{phase_b_round}]"
                    write_status(big_round, sub_idx, f"Phase B-{phase_b_round}",
                                 total_done=_total_done(), total_target=_total_target(),
                                 doc_pool_size=len(doc_pool),
                                 sub_failed=len(sub_skipped), dead_terminated=dead_terminated)
                    success, reason = _do_publish(f_name, doc_path, round_label, is_retry=True)
                    if success:
                        sent_accounts_set.add(f_name)
                        _append_sent_excel(f_name)
                        _remove_from_fail_list(f_name)
                        _append_summary(f_name, sub_round_id, "补发成功时间")
                        log(f"  ✓ 补发成功: {f_name}")
                    else:
                        _handle_failure(f_name, reason, sub_fail_count, sub_skipped)

            log(f"\n{log_label}第 {big_round} 大循环 / 第 {sub_idx} 小轮 结束。本小轮跳过 {len(sub_skipped)} 个(下小轮恢复)。硬终止累计 {len(dead_terminated)} 个")
            sent_accounts_set.clear()
            _clear_round_sheets()

        log(f"\n{'='*20} {log_label}第 {big_round} 大循环 结束 {'='*20}")

    write_status(big_round, sub_rounds, "结束",
                 total_done=_total_done(), total_target=_total_target(),
                 doc_pool_size=len(doc_pool),
                 sub_failed=0, dead_terminated=dead_terminated,
                 extra=f"{log_label}{big_round} 大循环跑完。最终硬终止 {len(dead_terminated)} 个")

    return {
        "acc_count": acc_count,
        "dead_terminated": dead_terminated,
        "doc_pool": doc_pool,
        "ok_count": ok_count,
        "fail_count": fail_count,
        "total_notices": total_notices,
        "total_violations": total_violations,
        "total_alerts": total_alerts,
        "big_rounds": big_round,
    }


# ========== 主流程 ==========

def main():
    _init_run_dir()
    log("=" * 50)
    log("创作罐头批量发布 v3 启动")
    log(f"报告目录: {RUN_REPORT_DIR}")
    log("=" * 50)

    os.makedirs(DOCS_FOLDER, exist_ok=True)
    os.makedirs(SENT_FOLDER, exist_ok=True)

    fail_records = []         # [(ts, name, reason)] 记录所有失败尝试
    success_accounts = set()  # 记录最终成功发布的账号
    credit_records = {}       # {账号名: 信用分}
    total_notices = 0
    total_violations = 0
    total_alerts = 0

    # 读取账号配置.xlsx - 不首发sheet
    _ensure_config_excel()
    try:
        _nofirst_list = _read_excel_sheet("不首发")
        NOFIRST_ACCOUNTS.update(_nofirst_list)
        if _nofirst_list:
            log(f"账号配置.xlsx[不首发]已加载，不首发账号: {len(NOFIRST_ACCOUNTS)} 个")
    except Exception as _ne:
        log(f"读取账号配置.xlsx[不首发]失败: {_ne}")

    # 读取账号配置.xlsx - 永久跳过sheet
    try:
        _skip_list = _read_excel_sheet("永久跳过")
        for _sv in _skip_list:
            if _sv not in EXCLUDE_ACCOUNTS:
                EXCLUDE_ACCOUNTS.append(_sv)
        if _skip_list:
            log(f"账号配置.xlsx[永久跳过]已加载，永久跳过: {EXCLUDE_ACCOUNTS}")
    except Exception as _se:
        log(f"读取账号配置.xlsx[永久跳过]失败: {_se}")

    # 读取账号配置.xlsx - 本轮已发sheet（中断恢复时跳过已发账号）
    sent_accounts_set = set()
    try:
        _sent_list = _read_excel_sheet("本轮已发")
        for _sv in _sent_list:
            sent_accounts_set.add(_sv)
        if sent_accounts_set:
            log(f"账号配置.xlsx[本轮已发]已加载，本轮已发: {len(sent_accounts_set)} 个账号（本轮将跳过）")
    except Exception as _se:
        log(f"读取账号配置.xlsx[本轮已发]失败: {_se}")

    # 读取账号配置.xlsx - 失败列表sheet（中断恢复时继续补发）
    try:
        _fail_pre = _read_fail_list()
        if _fail_pre:
            log(f"账号配置.xlsx[失败列表]已加载 {len(_fail_pre)} 条（本轮末将补发）")
    except Exception as _fe:
        log(f"读取账号配置.xlsx[失败列表]失败: {_fe}")

    docs = get_docs()
    if not docs:
        log("错误: 素材文件夹中没有 docx 文件，请放入文档后重试")
        input("按回车退出...")
        return

    log(f"待发文档: {len(docs)} 份")

    try:
        main_ws_url = get_main_ws_url()
    except RuntimeError as e:
        log(f"错误: {e}")
        input("按回车退出...")
        return

    main_ws = ws_connect(main_ws_url, timeout=10)
    log("已连接主窗口")


    # 不使用标签筛选，发全部账号（由 EXCLUDE_ACCOUNTS 排除不需要发的账号）

    # 收集账号列表
    accounts = collect_accounts(main_ws)
    if EXCLUDE_ACCOUNTS:
        before = len(accounts)
        accounts = [a for a in accounts if not any(ex in a or a in ex for ex in EXCLUDE_ACCOUNTS)]
        skipped = before - len(accounts)
        if skipped:
            log(f"已排除 {skipped} 个永不发文账号: {EXCLUDE_ACCOUNTS}")

    # 读取账号配置.xlsx - 白名单sheet（有内容则只跑白名单内的账号）
    try:
        _include = _read_excel_sheet("白名单")
        if _include:
            accounts = [a for a in accounts if any(inc in a or a in inc for inc in _include)]
            log(f"账号配置.xlsx[白名单]已加载，白名单 {len(_include)} 个，过滤后剩 {len(accounts)} 个账号")
    except Exception as _e:
        log(f"读取账号配置.xlsx[白名单]失败: {_e}")

    if not accounts:
        log("错误: 未找到任何账号，请检查创作罐头是否正常登录")
        main_ws.close()
        input("按回车退出...")
        return

    # 配额:总篇数÷账号数(死磕模式不再支持补漏专属 quota,统一按平均配额)
    quota = len(docs) // len(accounts) if len(accounts) > 0 else 1
    quota = max(quota, 1)
    quota_map = {a: quota for a in accounts}
    log(f"本次发布: {len(accounts)} 个账号，{len(docs)} 篇文档，每账号配额 {quota} 篇")

    doc_pool = list(docs)

    # 调用死磕主循环 (微头条 5 小轮 / 大循环 + 外层无限磕)
    result = run_death_grip(
        accounts=accounts,
        per_account_quota=quota_map,
        doc_pool=doc_pool,
        main_ws=main_ws,
        sub_rounds=5,
        max_fail_per_sub=3,
        sent_accounts_set=sent_accounts_set,
        credit_records=credit_records,
        fail_records=fail_records,
        success_accounts=success_accounts,
    )

    acc_count        = result["acc_count"]
    dead_terminated  = result["dead_terminated"]
    doc_pool         = result["doc_pool"]
    ok_count         = result["ok_count"]
    fail_count       = result["fail_count"]
    total_notices    = result["total_notices"]
    total_violations = result["total_violations"]
    total_alerts     = result["total_alerts"]

    if doc_pool:
        log(f"\n! 文档池还剩 {len(doc_pool)} 篇未发完(死磕已无可用账号)")
        for d in doc_pool:
            log(f"  未发: {os.path.basename(d)}")

    # 写最终失败记录(放弃账号 + 本次从未成功的账号)
    final_fails = [(ts, n, r) for ts, n, r in fail_records
                   if (n not in success_accounts) or (n in dead_terminated)]
    write_fail_excel(final_fails)

    # ========== 收尾:硬终止账号汇报 ==========
    if dead_terminated:
        log(f"\n★ 4 类硬终止账号 {len(dead_terminated)} 个 (已写入'硬终止账号'sheet,需人工处理):")
        for name, (reason, ts, cnt) in dead_terminated.items():
            log(f"  - {name}\t{reason}\t本次已发 {cnt} 篇\t{ts}")
    else:
        log("\n★ 无硬终止账号,本次全部账号正常完成或仍在 active")

    # 排序"发文汇总"：同账号的多轮放一起，按轮次升序
    _sort_summary_by_account()

    # 写信用分记录
    if credit_records:
        credit_file = os.path.join(RUN_REPORT_DIR, "信用分记录.txt")
        lines = ["信用分记录\n", "=" * 40 + "\n"]
        for acc, score in sorted(credit_records.items(), key=lambda x: x[1]):
            flag = " ← 危险" if score < 60 else (" ← 警告" if score < 80 else "")
            lines.append(f"{score}分  {acc}{flag}\n")
        with open(credit_file, "a", encoding="utf-8") as f:
            f.writelines(lines)

    # 生成汇总报告
    run_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    final_fail_names = sorted(set(n for _, n, _ in final_fails))
    summary_lines = [
        "=" * 60,
        f"运行汇总 - {run_time_str}",
        "=" * 60,
        f"类型    : 微头条发布(死磕模式)",
        f"报告目录: {RUN_REPORT_DIR}",
        "",
        f"成功发布: {ok_count} 篇",
        f"中途失败: {fail_count} 次(已重试到底)",
        f"硬终止号: {len(dead_terminated)} 个 (失登/封号/禁言/找不到)",
        f"文档剩余: {len(doc_pool)} 篇",
    ]
    if dead_terminated:
        summary_lines.append("")
        summary_lines.append("硬终止账号 (需人工处理):")
        for name, (reason, ts, cnt) in dead_terminated.items():
            summary_lines.append(f"  - {name}\t{reason}\t本次已发 {cnt} 篇")
    summary_lines.extend([
        "",
        f"系统通知: {total_notices} 条  → 系统通知.txt",
        f"违规提醒: {total_violations} 条  → 违规提醒.txt",
        f"高阅读提醒: {total_alerts} 条  → 高阅读提醒.txt",
        "=" * 60,
    ])
    summary_text = "\n".join(summary_lines)
    summary_file = os.path.join(RUN_REPORT_DIR, "汇总报告.txt")
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write(summary_text)

    main_ws.close()
    log(f"\n{'='*50}")
    log(f"完成! 成功:{ok_count}  最终失败:{len(final_fail_names)}  硬终止:{len(dead_terminated)}")
    log("\n" + summary_text)
    log(f"{'='*50}")


if __name__ == "__main__":
    main()
