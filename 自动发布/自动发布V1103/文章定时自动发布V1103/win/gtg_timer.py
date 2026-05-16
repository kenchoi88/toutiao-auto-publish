"""
创作罐头定时发布脚本 - 图文文章版（Win台机版）
  GTG_青春小馆定时自动发布/
  ├── go.bat                双击运行
  ├── debug_launch.bat      先用这个启动罐头到 CDP 9223
  ├── gtg_timer.py          ← 本文件
  ├── gtg_batch.py          发文核心（共享，已加 schedule_time 参数）
  ├── 定时配置.xlsx          B1 = 发布日期（每次只改这一格）
  ├── 账号配置.xlsx          白名单/永久跳过/失败列表（与 gtg_batch.py 共用）
  ├── 素材/                 .docx 文稿
  │   └── 已发送/
  └── 运行报告/YYYYMMDD/

定时配置.xlsx 结构：
  A1: "发布日期"   B1: 2026-04-26   ← 每次只改这一格
  A3: "账号名"     B3: "发文数"     ← 表头
  A4起 数据行：    A=账号名         B=发文数（1/2/3）

自动排程规则（与 mini 一致）：
  - 早窗 08:01 起：发文数≥1 的账号各 1 篇
  - 中窗 12:01 起：发文数≥2 的账号各 1 篇
  - 晚窗 19:01 起：发文数=3 的账号各 1 篇
  - 间隔 1 分钟，按 Excel 顺序发；超时自然延后不截断
"""

import os
import sys
import time
import shutil
import glob
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 复用 gtg_batch 里所有 CDP / 账号操作 / publish_article（含定时分支）+ 死磕主循环
import gtg_batch as gb
from gtg_batch import (
    CDP_URL, EXCLUDE_ACCOUNTS,
    get_tabs, get_main_ws_url, ws_connect, click,
    scroll_find_account, collect_accounts, find_account_webview,
    close_current_tab, publish_article,
    HARD_TERMINATE_REASONS, _append_hard_terminate, run_death_grip,
)

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DOCS_FOLDER  = os.path.join(BASE_DIR, "素材")
SENT_FOLDER  = os.path.join(BASE_DIR, "素材", "已发送")
TIMER_EXCEL  = os.path.join(BASE_DIR, "定时配置.xlsx")
CONFIG_EXCEL = os.path.join(BASE_DIR, "账号配置.xlsx")

WAIT_LOAD = 4

RUN_REPORT_DIR = None
LOG_FILE       = None
FAIL_FILE      = None


def _init_run_dir():
    global LOG_FILE, FAIL_FILE, RUN_REPORT_DIR
    ts = datetime.now().strftime("%Y%m%d")
    RUN_REPORT_DIR = os.path.join(BASE_DIR, "运行报告", ts)
    os.makedirs(RUN_REPORT_DIR, exist_ok=True)
    LOG_FILE  = os.path.join(RUN_REPORT_DIR, "运行日志.txt")
    FAIL_FILE = os.path.join(RUN_REPORT_DIR, "定时失败记录.xlsx")
    # 让 gtg_batch 的 log() 也写到同一份运行报告
    gb._init_run_dir()


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    sys.stdout.buffer.write((line + "\n").encode("utf-8", errors="replace"))
    sys.stdout.buffer.flush()
    if LOG_FILE:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")


def write_fail_excel(final_fails):
    if not final_fails:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "定时失败记录"
    headers = ["时间", "账号名", "定时时间", "文档", "失败原因"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = Alignment(horizontal="center")
    for w, c in zip([16, 25, 18, 30, 30], "ABCDE"):
        ws.column_dimensions[c].width = w
    for row in final_fails:
        ws.append(list(row))
    try:
        wb.save(FAIL_FILE)
        log(f"失败记录已写入: {FAIL_FILE}")
    except Exception as e:
        log(f"  写入失败记录出错: {e}")


def _read_publish_date():
    if not os.path.exists(TIMER_EXCEL):
        log(f"错误: 定时配置.xlsx 不存在: {TIMER_EXCEL}")
        return None
    try:
        wb = openpyxl.load_workbook(TIMER_EXCEL, read_only=True, data_only=True)
        ws = wb.active
        v = ws.cell(row=1, column=2).value
        wb.close()
        if not v:
            log("错误: 定时配置.xlsx B1 未填发布日期")
            return None
        return v.strftime("%Y-%m-%d") if hasattr(v, 'strftime') else str(v).strip()[:10]
    except Exception as e:
        log(f"读取定时配置.xlsx失败: {e}")
        return None


def _read_skip_set():
    skip = set()
    if not os.path.exists(CONFIG_EXCEL):
        return skip
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "永久跳过" in wb.sheetnames:
            for row in wb["永久跳过"].iter_rows(min_row=2, max_col=1, values_only=True):
                v = row[0]
                if v:
                    s = str(v).strip()
                    if s and not s.startswith("#"):
                        skip.add(s)
        wb.close()
    except Exception as e:
        log(f"读取永久跳过失败: {e}")
    return skip


def _read_whitelist():
    """白名单 sheet：[(账号名, 发文份数1-3)]，非空则只发这些账号"""
    wl = []
    if not os.path.exists(CONFIG_EXCEL):
        return wl
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "白名单" in wb.sheetnames:
            for row in wb["白名单"].iter_rows(min_row=2, max_col=2, values_only=True):
                name_v, q_v = row[0], row[1]
                if not name_v:
                    continue
                name = str(name_v).strip()
                if not name or name.startswith("#"):
                    continue
                try:
                    q = int(q_v) if q_v is not None else 3
                except (ValueError, TypeError):
                    q = 3
                q = max(1, min(3, q))
                wl.append((name, q))
        wb.close()
    except Exception as e:
        log(f"读取白名单失败: {e}")
    return wl


def _read_sent_with_count():
    """[v1102] 读「本轮已发」 sheet → {账号: 已发次数}"""
    if not os.path.exists(CONFIG_EXCEL): return {}
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "本轮已发" not in wb.sheetnames: wb.close(); return {}
        ws_r = wb["本轮已发"]
        result = {}
        for row in ws_r.iter_rows(min_row=2, max_col=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            cnt = 1
            if len(row) > 1 and row[1] is not None:
                try: cnt = int(row[1])
                except: cnt = 1
            result[name] = cnt
        wb.close(); return result
    except Exception:
        return {}


def _append_sent_excel(name):
    """[v1102] 写「本轮已发」 sheet:行存在 count+1,不存在 append (账号, 1)"""
    if not os.path.exists(CONFIG_EXCEL): return
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "本轮已发" not in wb.sheetnames:
            ws_s = wb.create_sheet("本轮已发")
            ws_s.append(["账号名", "已发次数"])
            ws_s.append([name, 1])
        else:
            ws_s = wb["本轮已发"]
            found_row = None
            for row_idx, row in enumerate(ws_s.iter_rows(min_row=2, max_col=2, values_only=False), start=2):
                if row[0].value and str(row[0].value).strip() == name:
                    found_row = row_idx
                    break
            if found_row:
                cur = ws_s.cell(row=found_row, column=2).value or 0
                try: cur = int(cur)
                except: cur = 0
                ws_s.cell(row=found_row, column=2).value = cur + 1
            else:
                ws_s.append([name, 1])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass


def _expand_tasks(accounts_quota, date_str):
    """按 早/中/晚 三窗 + 1 分钟间隔 展开任务"""
    GAP_MIN = 1
    windows = [
        # [v1102] 缺 N 篇 → 排前 N 个窗(剩 1=早 / 剩 2=早+中 / 剩 3=早+中+晚)
        ("早", "08:01", [n for n, q in accounts_quota if q >= 1]),
        ("中", "12:01", [n for n, q in accounts_quota if q >= 2]),
        ("晚", "19:01", [n for n, q in accounts_quota if q >= 3]),
    ]
    tasks = []
    for label, start_str, names in windows:
        if not names:
            continue
        base = datetime.strptime(f"{date_str} {start_str}", "%Y-%m-%d %H:%M")
        for i, name in enumerate(names):
            t = base + timedelta(minutes=GAP_MIN * i)
            tasks.append((name, t.strftime("%Y-%m-%d %H:%M")))
        last = base + timedelta(minutes=GAP_MIN * (len(names) - 1))
        log(f"  {label}窗 {start_str} 起 {len(names)} 个账号（末个 {last.strftime('%H:%M')}）")
    tasks.sort(key=lambda x: x[1])
    return tasks


def get_docs():
    docs = []
    for p in ["*.docx", "*.doc"]:
        docs.extend(glob.glob(os.path.join(DOCS_FOLDER, p)))
    return sorted([d for d in docs if "已发送" not in d])


# [v1101.4] doc_pool 顺序取 + 校验,救"分发完源必删"导致罐头找不到文件
def _pop_doc(doc_pool):
    """从 doc_pool 顺序取一篇实存的 docx,失效引用就地剔除。返回 None 表示池已空。"""
    while doc_pool:
        doc = doc_pool.pop(0)
        if os.path.exists(doc):
            return doc
        log(f"  ! 源已删除(可能被外部分发),跳过: {os.path.basename(doc)}")
    return None


def move_to_sent(doc_path):
    os.makedirs(SENT_FOLDER, exist_ok=True)
    dest = os.path.join(SENT_FOLDER, os.path.basename(doc_path))
    if os.path.exists(dest):
        name, ext = os.path.splitext(os.path.basename(doc_path))
        dest = os.path.join(SENT_FOLDER, f"{name}_{int(time.time())}{ext}")
    shutil.move(doc_path, dest)
    log(f"  已移至已发送: {os.path.basename(dest)}")


def _finalize_config(accounts_quota, success_by_acct, fail_docs_by_acct=None):
    """收尾：清空白名单；把提交失败账号写入待补漏"""
    if not os.path.exists(CONFIG_EXCEL):
        log("警告: 账号配置.xlsx 不存在，跳过收尾")
        return
    fail_docs_by_acct = fail_docs_by_acct or {}
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb = openpyxl.load_workbook(CONFIG_EXCEL)

        if "白名单" in wb.sheetnames:
            wl = wb["白名单"]
            if wl.max_row >= 2:
                wl.delete_rows(2, wl.max_row - 1)

        pending_rows = []
        for name, quota in accounts_quota:
            miss = quota - success_by_acct.get(name, 0)
            if miss > 0:
                docs = fail_docs_by_acct.get(name, [])
                doc_names = " | ".join(d for d in docs if d)
                pending_rows.append((name, miss, doc_names, now))

        if "待补漏" in wb.sheetnames:
            pen = wb["待补漏"]
            if pen.max_row >= 2:
                pen.delete_rows(2, pen.max_row - 1)
            for i, row in enumerate(pending_rows, start=2):
                for j, v in enumerate(row, start=1):
                    pen.cell(row=i, column=j, value=v)

        wb.save(CONFIG_EXCEL)
        log("\n收尾：白名单已清空")
        if pending_rows:
            log(f"收尾：待补漏写入 {len(pending_rows)} 行（总漏发 {sum(r[1] for r in pending_rows)} 篇）")
            for r in pending_rows:
                log(f"  - {r[0]} 漏 {r[1]} 篇")
        else:
            log("收尾：本次无提交失败，待补漏留空")
    except Exception as e:
        log(f"警告: 收尾写账号配置失败: {e}")


# ========== 主流程 ==========

def main():
    _init_run_dir()
    log("=" * 50)
    log("创作罐头图文文章定时发布 Win台机版 启动")
    log(f"报告目录: {RUN_REPORT_DIR}")
    log("=" * 50)

    os.makedirs(DOCS_FOLDER, exist_ok=True)
    os.makedirs(SENT_FOLDER, exist_ok=True)

    date_str = _read_publish_date()
    if not date_str:
        return
    log(f"发布日期: {date_str}")

    try:
        main_ws_url = get_main_ws_url()
    except RuntimeError as e:
        log(f"错误: {e}（先跑 debug_launch.bat 把罐头开起来到 CDP 9223）")
        return

    main_ws = ws_connect(main_ws_url, timeout=10)
    log("已连接主窗口")

    skip = _read_skip_set()
    if skip:
        log(f"「永久跳过」：{sorted(skip)}")

    # [v1102.9 缺哥拍 2026-05-16] 总是先 collect 全员, 白名单只优先排序不剔除, B 列 quota 忽略
    all_names = collect_accounts(main_ws)
    all_accounts_full = [n for n in all_names if n not in skip and n not in EXCLUDE_ACCOUNTS]
    wl = _read_whitelist()
    if wl:
        log(f"白名单 {len(wl)} 个优先排序, 全员 {len(all_accounts_full)} 个账号都发")
        _wl_names = [n for n, _ in wl if n not in skip and n not in EXCLUDE_ACCOUNTS]
        _seen = set()
        all_accounts = []
        for wn in _wl_names:
            for a in all_accounts_full:
                if a not in _seen and (wn in a or a in wn):
                    all_accounts.append(a)
                    _seen.add(a)
                    break
        for a in all_accounts_full:
            if a not in _seen:
                all_accounts.append(a)
                _seen.add(a)
    else:
        log(f"白名单为空, 全员 {len(all_accounts_full)} 个账号")
        all_accounts = all_accounts_full
    wl_quota_map = {}  # [v1102.9] 白名单 B 列 quota 忽略, 统一用 quota_total
    _full_accounts_count = len(all_accounts_full)

    if not all_accounts:
        log("错误: 没有可发文账号")
        main_ws.close()
        return

    doc_pool = list(get_docs())
    docs_count = len(doc_pool)

    sent_count_map = _read_sent_with_count()
    sent_total = sum(sent_count_map.values())
    # [v1102.9 缺哥拍 2026-05-16] quota 算法取消 cap 3, 分母用 _full_accounts_count(全员), 缺 N 补 N
    # 物理上限由 _expand_tasks 3 时段(早/中/晚)自然限制每号 ≤ 3 篇 — 算法层去 cap, 剩余素材留池
    _total = docs_count + sent_total
    _n = _full_accounts_count
    base_q = max(1, _total // _n) if _n else 1
    extra = (_total - base_q * _n) if _n else 0
    quota_per_account = {n: base_q + (1 if i < extra else 0) for i, n in enumerate(all_accounts)}
    log(f"本次发布: {len(all_accounts)} 个账号(白名单/全员={_full_accounts_count}), {docs_count} 篇文档, 每号 quota={base_q}{'+1' if extra else ''} 篇 (素材{docs_count}+已发{sent_total}=总{_total}/{_n} 全员={base_q} 余 {extra} 篇分前 {extra} 号 +1, 无 cap, 缺 N 补 N — _expand_tasks 物理 3 时段限制每号实发 ≤ 3 篇)")

    accounts_quota = []
    skipped_full = 0
    reduced = 0
    for n in all_accounts:
        already = sent_count_map.get(n, 0)
        wl_q = wl_quota_map.get(n)
        my_q = quota_per_account[n]
        target_q = min(wl_q, my_q) if wl_q else my_q
        miss = target_q - already
        if miss <= 0:
            skipped_full += 1
            continue
        if already > 0:
            reduced += 1
        accounts_quota.append((n, miss))

    if sent_count_map:
        current_round = sent_total // len(all_accounts) + 1
        done_in_round = sent_total % len(all_accounts)
        remaining_in_round = len(all_accounts) - done_in_round
        log(f"中断恢复:第 {current_round} 小轮第 {done_in_round + 1} 账号断点(已发 {sent_total} 篇,quota 满跳 {skipped_full} 个 / 部分扣 {reduced} 个)")
        log(f"    本小轮({current_round})剩 {remaining_in_round} 账号 + 后续 {quota_total - current_round} 小轮各 {len(all_accounts)} 账号")
        log(f"    共 {sum(q for _, q in accounts_quota)} 篇未发")

    if not accounts_quota:
        log("✓ 所有账号本大循环已齐活,无需排程")
        main_ws.close()
        return

    tasks = _expand_tasks(accounts_quota, date_str)
    log(f"\n共 {len(tasks)} 个定时发布任务")
    log(f"素材池共 {docs_count} 份文稿")

    doc_pool = list(get_docs())
    log(f"素材池共 {len(doc_pool)} 份文稿")
    if len(doc_pool) < len(tasks):
        log(f"警告：素材（{len(doc_pool)}）少于任务（{len(tasks)}），后面任务将记为素材不足")

    fail_records = []
    success_count = 0
    success_by_acct = {}
    fail_docs_by_acct = {}
    dead_terminated = {}   # name -> (reason, ts, count_so_far) — 4 类硬终止,不再 requeue

    doc_retry_set = set()
    def requeue_doc(p):
        if not p or p in doc_retry_set:
            return False
        doc_retry_set.add(p)
        doc_pool.append(p)
        log(f"  → 文档已回池尾待后续重试: {os.path.basename(p)}")
        return True

    def _terminate(name, reason, doc_path):
        """命中 4 类硬终止 → 写'硬终止账号' sheet,不 requeue"""
        cnt = success_by_acct.get(name, 0)
        dead_terminated[name] = (reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cnt)
        _append_hard_terminate(name, reason, cnt)
        log(f"  ★ 4 类硬终止: {name} -> {reason} (本次已发 {cnt} 篇,不再尝试)")
        # 文档不 requeue,但也不丢:回池给 Stage 2 用
        if doc_path and doc_path not in doc_pool:
            doc_pool.append(doc_path)

    consecutive_fail = 0
    MAX_CONSECUTIVE_FAIL = 6
    COOLDOWN_AT = 3
    COOLDOWN_SEC = 60

    for idx, (name, timer_time) in enumerate(tasks):
        log(f"\n{'='*40}")
        log(f"任务 {idx+1}/{len(tasks)}: {name}  →  {timer_time}")

        if name in dead_terminated:
            log(f"  ★ 该账号已硬终止 ({dead_terminated[name][0]}),跳过本任务,文档留给 Stage 2")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", f"已硬终止/{dead_terminated[name][0]}"))
            continue

        if not doc_pool:
            log("  X 素材池已空")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", "素材不足"))
            fail_docs_by_acct.setdefault(name, []).append("")
            continue
        # [v1101.4] _pop_doc 替代 doc_pool.pop(0): 校验存在 + 失效就地剔除
        doc_path = _pop_doc(doc_pool)
        if doc_path is None:
            log("  X 素材池已空(全失效)")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", "素材不足"))
            fail_docs_by_acct.setdefault(name, []).append("")
            continue
        log(f"  文档: {os.path.basename(doc_path)}")

        this_task_failed = False
        this_task_fail_reason = ""

        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未找到账号: {name}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), "侧边栏未找到"))
            fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
            _terminate(name, "侧边栏未找到", doc_path)  # 4 类硬终止之一
            this_task_failed = True
            this_task_fail_reason = "侧边栏未找到"
        else:
            click(main_ws, pos["x"], pos["y"], 20)
            time.sleep(WAIT_LOAD)

            ws_url = find_account_webview(main_ws, name)
            if not ws_url:
                log("  X 找不到 webview")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), "webview匹配失败"))
                fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
                requeue_doc(doc_path)  # webview匹配失败不在 4 类,正常 requeue
                close_current_tab(main_ws)
                this_task_failed = True
                this_task_fail_reason = "webview匹配失败"
            else:
                try:
                    success, reason = publish_article(ws_url, doc_path, main_ws, name=name, schedule_time=timer_time)
                    if success:
                        move_to_sent(doc_path)
                        success_count += 1
                        success_by_acct[name] = success_by_acct.get(name, 0) + 1
                        log(f"  ✓ 定时发布成功: {name}")
                        _append_sent_excel(name)  # [v1102] 写「本轮已发」count +1
                    else:
                        log(f"  X 发布失败: {reason}")
                        fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), reason))
                        fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
                        if reason in HARD_TERMINATE_REASONS:
                            _terminate(name, reason, doc_path)
                        else:
                            requeue_doc(doc_path)
                        this_task_failed = True
                        this_task_fail_reason = reason
                except Exception as e:
                    log(f"  X 异常: {e}")
                    fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), f"异常: {e}"))
                    fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
                    requeue_doc(doc_path)
                    this_task_failed = True
                    this_task_fail_reason = f"异常: {e}"
                finally:
                    close_current_tab(main_ws)

        if this_task_failed:
            consecutive_fail += 1
            if consecutive_fail >= MAX_CONSECUTIVE_FAIL:
                log(f"\n!! 连续 {consecutive_fail} 次失败触发熔断（最近: {this_task_fail_reason}），Stage 1 提前结束,Stage 2 接手")
                break
            if consecutive_fail >= COOLDOWN_AT:
                log(f"  连败 {consecutive_fail} 次，先 sleep {COOLDOWN_SEC}s 等罐头缓过来再继续")
                time.sleep(COOLDOWN_SEC)
        else:
            consecutive_fail = 0

        if idx < len(tasks) - 1:
            _d = random.randint(8, 20)
            log(f"  篇间等待 {_d} 秒...")
            time.sleep(_d)

    write_fail_excel(fail_records)
    log(f"\n{'='*50}")
    log(f"Stage 1 (定时排程) 完成! 成功:{success_count}  失败:{len(fail_records)}  硬终止:{len(dead_terminated)}")
    log(f"{'='*50}")

    # ========== Stage 2:死磕补尾(自动衔接) ==========
    # 漏发账号 = 配置的 quota - 实际成功数。已硬终止的不参与
    remaining_quota = {}
    remaining_accounts = []
    for name, q in accounts_quota:
        if name in dead_terminated:
            continue
        sent = success_by_acct.get(name, 0)
        if sent < q:
            remaining_quota[name] = q - sent
            remaining_accounts.append(name)

    if remaining_accounts and doc_pool:
        log(f"\n{'#'*60}")
        log(f"# Stage 2 启动:死磕补尾,{len(remaining_accounts)} 个账号待补,文档池剩 {len(doc_pool)} 篇")
        log(f"# 不再带定时,立即发布,直到全部完成 / 4 类硬终止 / Ctrl+C")
        log(f"{'#'*60}")
        try:
            stage2 = run_death_grip(
                accounts=remaining_accounts,
                per_account_quota=remaining_quota,
                doc_pool=doc_pool,
                main_ws=main_ws,
                sub_rounds=3,
                max_fail_per_sub=3,
                initial_dead=dead_terminated,
                log_label="[Stage 2] ",
            )
            stage2_dead = stage2.get("dead_terminated", {})
            new_dead = {n: v for n, v in stage2_dead.items() if n not in dead_terminated}
            dead_terminated.update(stage2_dead)
            for n in stage2.get("acc_count", {}):
                if n in stage2["acc_count"]:
                    success_by_acct[n] = success_by_acct.get(n, 0) + stage2["acc_count"][n]
            log(f"\nStage 2 完成: 新增成功 {stage2.get('ok_count', 0)} 篇,新增硬终止 {len(new_dead)} 个")
        except KeyboardInterrupt:
            log(f"\n!! Stage 2 被人工 Ctrl+C 中断")
    elif not remaining_accounts:
        log("\n★ Stage 1 全部账号已达 quota 或硬终止,无需 Stage 2")
    else:
        log("\n★ Stage 1 结束时文档池已空,Stage 2 无文档可发")

    _finalize_config(accounts_quota, success_by_acct, fail_docs_by_acct)

    main_ws.close()

    if dead_terminated:
        log(f"\n★★ 4 类硬终止账号 {len(dead_terminated)} 个 (需人工处理):")
        for name, (reason, ts, cnt) in dead_terminated.items():
            log(f"  - {name}\t{reason}\t本次已发 {cnt} 篇\t{ts}")

    log(f"\n{'='*50}")
    log(f"全流程完成! 总成功:{sum(success_by_acct.values())}  Stage1失败记录:{len(fail_records)}  硬终止:{len(dead_terminated)}")
    log(f"{'='*50}")
    os.system("pause")


if __name__ == "__main__":
    main()
