"""
创作罐头定时发布脚本 - 图文文章版（Mac）
  Mac文章定时自动发布/
  ├── go.command          双击运行
  ├── gtg_timer.py
  ├── 定时发布.xlsx        配置：B1=日期；A4起 账号名 | 发文数(1/2/3)
  ├── 素材/               放 .docx 文件
  │   └── 已发送/         发完自动移入
  └── 运行报告/YYYYMMDD/

定时发布.xlsx 结构：
  A1: "发布日期"       B1: 2026-04-23         ← 每次只改这一格
  A3: "账号名"         B3: "发文数"           ← 表头
  A4起 数据行：        A=账号名               B=发文数（1/2/3）

自动排程规则：
  - 早窗 08:01 起：发文数≥1 的账号各发 1 篇
  - 中窗 12:01 起：发文数≥2 的账号各发 1 篇
  - 晚窗 19:01 起：发文数=3 的账号各发 1 篇
  - 每窗内按 Excel 顺序，相邻账号间隔 1 分钟（超出自然延后，不截断）
"""

import requests
import base64
import json
import websocket
import time
import os
import shutil
import glob
import sys
import random
import subprocess
import re
import threading
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== 配置 =====================
def _find_cdp_port():
    port_file = os.path.expanduser("~/Library/Application Support/创作罐头/DevToolsActivePort")
    try:
        with open(port_file) as f:
            return int(f.readline().strip())
    except Exception:
        return 9225

CDP_URL       = f"http://127.0.0.1:{_find_cdp_port()}"
ACCOUNT_CLASS = "account-RALrbJ"
WAIT_LOAD     = 4
EXCLUDE_ACCOUNTS = ["青春小馆"]
NOFIRST_ACCOUNTS = set()

NO_PROXY = {"http": "", "https": ""}
WS_OPTS  = {"suppress_origin": True}

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DOCS_FOLDER  = os.path.join(BASE_DIR, "素材")
SENT_FOLDER  = os.path.join(BASE_DIR, "素材", "已发送")
TIMER_EXCEL  = os.path.join(BASE_DIR, "定时发布.xlsx")      # 只存 B1 发布日期
CONFIG_EXCEL = os.path.join(BASE_DIR, "账号配置.xlsx")       # 与自动发那边同结构可互换

RUN_REPORT_DIR = None
# [v1101 P9] 阅读量回查告警阈值
ALERT_THRESHOLD = 5000
ALERT_FILE = None  # 在 _init_run_dir 里赋值
LOG_FILE       = None
FAIL_FILE      = None
# [v1102 NOTICE 重建] timer.py 补
NOTICE_FILE         = None  # 系统通知.txt
NOTICE_CHECKED_FILE = None  # notice_checked.txt 持久化已检查账号
LAST_PUBLISHED_FILE = None  # [v1102] 持久化最近 publish 成功账号
VIOLATION_FILE      = None  # 违规提醒.txt
notice_checked_set  = set()  # 全局: 当天已检查 NOTICE 账号

os.environ["NO_PROXY"] = "127.0.0.1,localhost"
# ================================================

VIOLATION_KEYWORDS = {
    "违规/扣分": ["违规", "扣分", "处罚", "警告"],
    "禁言封号": ["禁言", "发言受限", "封禁", "封号"],
}


def _init_run_dir():
    global LOG_FILE, FAIL_FILE, RUN_REPORT_DIR, NOTICE_FILE, NOTICE_CHECKED_FILE, LAST_PUBLISHED_FILE, VIOLATION_FILE
    ts = datetime.now().strftime("%Y%m%d")
    RUN_REPORT_DIR = os.path.join(BASE_DIR, "运行报告", ts)
    os.makedirs(RUN_REPORT_DIR, exist_ok=True)
    LOG_FILE  = os.path.join(RUN_REPORT_DIR, "运行日志.txt")
    FAIL_FILE = os.path.join(RUN_REPORT_DIR, "失败记录.xlsx")
    global ALERT_FILE
    ALERT_FILE = os.path.join(RUN_REPORT_DIR, "高阅读量提醒.txt")
    NOTICE_FILE         = os.path.join(RUN_REPORT_DIR, "系统通知.txt")
    NOTICE_CHECKED_FILE = os.path.join(RUN_REPORT_DIR, "notice_checked.txt")
    LAST_PUBLISHED_FILE = os.path.join(RUN_REPORT_DIR, "last_published.txt")
    VIOLATION_FILE      = os.path.join(RUN_REPORT_DIR, "违规提醒.txt")


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    sys.stdout.buffer.write((line + "\n").encode("utf-8", errors="replace"))
    sys.stdout.buffer.flush()
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def write_fail_excel(final_fails):
    if not final_fails:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "失败记录"
    for col, h in enumerate(["时间", "账号名", "定时时间", "文档", "失败原因"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    for row_data in final_fails:
        ws.append(list(row_data))
    try:
        wb.save(FAIL_FILE)
        log(f"失败记录已写入: {FAIL_FILE}")
    except Exception as e:
        log(f"  写入失败记录出错: {e}")


def _read_publish_date():
    """从 定时发布.xlsx B1 读取发布日期字符串 yyyy-mm-dd"""
    if not os.path.exists(TIMER_EXCEL):
        log(f"错误: 定时发布.xlsx 不存在: {TIMER_EXCEL}")
        return None
    try:
        wb = openpyxl.load_workbook(TIMER_EXCEL, read_only=True, data_only=True)
        ws = wb.active
        v = ws.cell(row=1, column=2).value
        wb.close()
        if not v:
            log("错误: 定时发布.xlsx B1 未填发布日期")
            return None
        return v.strftime("%Y-%m-%d") if hasattr(v, 'strftime') else str(v).strip()[:10]
    except Exception as e:
        log(f"读取定时发布.xlsx失败: {e}")
        return None


def _read_skip_set():
    """读账号配置.xlsx「永久跳过」sheet，返回要剔除的账号集合"""
    skip = set()
    if not os.path.exists(CONFIG_EXCEL):
        return skip
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "永久跳过" in wb.sheetnames:
            for row in wb["永久跳过"].iter_rows(min_row=2, max_col=1, values_only=True):
                v = row[0]
                if v:
                    s = str(v).strip()
                    if s and not s.startswith("#"):
                        skip.add(s)
        wb.close()
    except Exception as e:
        log(f"读取永久跳过失败: {e}")
    return skip


def _read_whitelist():
    """读账号配置.xlsx「白名单」sheet：返回 [(账号名, 发文份数)]。
    白名单非空时代表缺哥手动指定「只发这些账号」，是第一优先级。
    """
    wl = []
    if not os.path.exists(CONFIG_EXCEL):
        return wl
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "白名单" in wb.sheetnames:
            for row in wb["白名单"].iter_rows(min_row=2, max_col=2, values_only=True):
                name_v, q_v = row[0], row[1]
                if not name_v:
                    continue
                name = str(name_v).strip()
                if not name or name.startswith("#"):
                    continue
                try:
                    q = int(q_v) if q_v is not None else 3
                except (ValueError, TypeError):
                    q = 3
                if q < 1:
                    q = 1
                if q > 3:
                    q = 3
                wl.append((name, q))
        wb.close()
    except Exception as e:
        log(f"读取白名单失败: {e}")
    return wl


def _read_excel_sheet(sheet_name):
    """[v1101.5] 读账号配置.xlsx 指定 sheet 的 A 列账号(跳第 1 行标题,忽略空行和 # 开头)"""
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb[sheet_name]
        result = []
        for row in ws_r.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip() and not str(val).strip().startswith('#'):
                result.append(str(val).strip())
        wb.close()
        return result
    except Exception:
        return []


def _append_sent_excel(name):
    """[v1101.5] 往账号配置.xlsx 的本轮已发 sheet 写一行 [name, count];若 name 已存在则 count+1"""
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "本轮已发" not in wb.sheetnames:
            ws_s = wb.create_sheet("本轮已发")
            ws_s.append(["账号名", "已发次数"])
            ws_s.append([name, 1])
        else:
            ws_s = wb["本轮已发"]
            # 找现有行
            found_row = None
            for row_idx, row in enumerate(ws_s.iter_rows(min_row=2, max_col=2, values_only=False), start=2):
                if row[0].value and str(row[0].value).strip() == name:
                    found_row = row_idx
                    break
            if found_row:
                cur = ws_s.cell(row=found_row, column=2).value or 0
                try: cur = int(cur)
                except: cur = 0
                ws_s.cell(row=found_row, column=2).value = cur + 1
            else:
                ws_s.append([name, 1])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass


def _read_sent_with_count():
    """[v1101.5] 读「本轮已发」sheet → {账号名: 已发次数}"""
    if not os.path.exists(CONFIG_EXCEL):
        return {}
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "本轮已发" not in wb.sheetnames:
            wb.close(); return {}
        ws_r = wb["本轮已发"]
        result = {}
        for row in ws_r.iter_rows(min_row=2, max_col=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            cnt = 1
            if len(row) > 1 and row[1] is not None:
                try: cnt = int(row[1])
                except: cnt = 1
            result[name] = cnt
        wb.close()
        return result
    except Exception:
        return {}


def _clear_round_sheets():
    """[v1101.5] 收尾时清空本轮已发 sheet(保留表头)"""
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "本轮已发" in wb.sheetnames:
            ws_c = wb["本轮已发"]
            if ws_c.max_row > 1:
                ws_c.delete_rows(2, ws_c.max_row - 1)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass



def _expand_tasks(accounts_quota, date_str):
    """按 早/中/晚 窗 + 8 分钟间隔 展开任务：
      - 早窗 07:00 起：发文数≥1 的账号各 1 篇
      - 中窗 12:00 起：发文数≥2 的账号各 1 篇
      - 晚窗 17:00 起：发文数=3 的账号各 1 篇
    间隔 8 分钟；若账号较多导致时间溢出窗口上限，自然延后，不截断。
    """
    GAP_MIN = 1
    # [v1101.5] 缺 N 篇 → 排最后 N 个窗(中断恢复:跑完早窗=q剩2 → 排中+晚,跑完早+中=q剩1 → 排晚)
    # q=3 没发过 → 早+中+晚;q=2 已发早 → 中+晚;q=1 已发早+中 → 晚
    windows = [
        ("早", "08:01", [n for n, q in accounts_quota if q >= 3]),
        ("中", "12:01", [n for n, q in accounts_quota if q >= 2]),
        ("晚", "19:01", [n for n, q in accounts_quota if q >= 1]),
    ]
    tasks = []
    for label, start_str, names in windows:
        if not names:
            continue
        base = datetime.strptime(f"{date_str} {start_str}", "%Y-%m-%d %H:%M")
        for i, name in enumerate(names):
            t = base + timedelta(minutes=GAP_MIN * i)
            tasks.append((name, t.strftime("%Y-%m-%d %H:%M")))
        last = base + timedelta(minutes=GAP_MIN * (len(names) - 1))
        log(f"  {label}窗 {start_str} 起 {len(names)} 个账号（末个 {last.strftime('%H:%M')}）")
    tasks.sort(key=lambda x: x[1])
    return tasks


def get_tabs():
    return requests.get(f"{CDP_URL}/json", timeout=5, proxies=NO_PROXY).json()


def get_main_ws_url():
    tabs = get_tabs()
    for t in tabs:
        if "czgts.cn" in t.get("url", "") and "webSocketDebuggerUrl" in t:
            return t["webSocketDebuggerUrl"]
    raise RuntimeError("找不到主窗口，请确认创作罐头已启动")


def ws_connect(url, timeout=10):
    return websocket.create_connection(url, timeout=timeout, **WS_OPTS)


def cdp(ws, method, params=None, mid=1):
    ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
    deadline = time.time() + 15
    while time.time() < deadline:
        try:
            d = json.loads(ws.recv())
            if d.get("id") == mid:
                return d.get("result", {})
        except:
            pass
    return {}


def js(ws, expr, mid=99):
    r = cdp(ws, "Runtime.evaluate", {"expression": expr}, mid)
    return r.get("result", {}).get("value")


def js_gesture(ws, expr, mid=99):
    """带用户手势的JS执行，用于触发原生文件对话框等需要user activation的操作"""
    r = cdp(ws, "Runtime.evaluate", {"expression": expr, "userGesture": True}, mid)
    return r.get("result", {}).get("value")


_gtg_minimize_recover_count = 0


def ensure_gtg_top():
    """[v1101 P5] 强化:除了原来的最小化别人 + 取消最小化罐头,
    再加 set visible to true(处理 Cmd-H Hidden) + AXRaise(强制 reorder) + verify frontmost(失败重试 3 次)。
    """
    global _gtg_minimize_recover_count
    r = subprocess.run(["osascript", "-e", '''
tell application "System Events"
    repeat with pname in {"Google Chrome", "Safari", "Claude", "Feishu", "WeChat"}
        try
            tell process (pname as text)
                repeat with w in windows
                    try
                        set value of attribute "AXMinimized" of w to true
                    end try
                end repeat
            end tell
        end try
    end repeat
end tell
tell application "创作罐头" to activate
delay 0.2
tell application "System Events"
    tell process "创作罐头"
        try
            if (value of attribute "AXMinimized" of window 1) then
                set value of attribute "AXMinimized" of window 1 to false
                delay 0.5
                return "RECOVERED"
            end if
        end try
    end tell
end tell
return "OK"
'''], capture_output=True, text=True)
    if "RECOVERED" in (r.stdout or ""):
        _gtg_minimize_recover_count += 1
        log(f"  ⚠ 检测到创作罐头窗口被最小化，已自愈（累计 {_gtg_minimize_recover_count} 次）")
    # [v1101 P5] unhide + AXRaise + verify frontmost 重试 3 次
    subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.1
tell application "System Events"
    tell process "创作罐头"
        try
            if visible is false then set visible to true
        end try
        try
            perform action "AXRaise" of window 1
        end try
        repeat 3 times
            if frontmost then exit repeat
            set frontmost to true
            delay 0.3
        end repeat
    end tell
end tell
'''], capture_output=True)
    time.sleep(0.3)


def click(ws, x, y, mid):
    p = {"button": "left", "clickCount": 1, "x": x, "y": y,
         "modifiers": 0, "timestamp": time.time() * 1000}
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mousePressed", **p}, mid)
    time.sleep(0.12)
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mouseReleased", **p}, mid + 1)


def scroll_find_account(main_ws, name):
    name_json = json.dumps(name)
    same_count = 0

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 10)
    time.sleep(0.5)

    for _ in range(500):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t === {name_json} || t.startsWith({name_json})){{
                    items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                    return 'found';
                }}
            }}
            return null;
        }})()
        """, 11)

        if v == 'found':
            time.sleep(0.3)
            pos = js(main_ws, f"""
            (function(){{
                var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    if(t === {name_json} || t.startsWith({name_json})){{
                        var r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        // 坐标不在视口内，再滚一次
                        items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                        r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                    }}
                }}
                return null;
            }})()
            """, 13)
            if pos:
                return json.loads(pos)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no-container';
            var before = c.scrollTop;
            c.scrollTop += 250;
            return before + '->' + c.scrollTop;
        })()
        """, 12)
        time.sleep(0.25)

        if result and result != "no-container":
            parts = result.split("->")
            if len(parts) == 2 and parts[0] == parts[1]:
                same_count += 1
                if same_count >= 5:
                    break
            else:
                same_count = 0

    # 滚动找不到，尝试搜索框输入账号名过滤
    log(f"  滚动未找到账号，尝试搜索框输入: {name}")
    search_pos = js(main_ws, """
    (function(){
        var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(!s) return null;
        var r = s.getBoundingClientRect();
        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
        return null;
    })()
    """, 14)
    if not search_pos:
        return None
    # v1101.2: CDP nativeInputValueSetter 注入,绕过键盘+罐头前台依赖(实战 air 295→300)
    js(main_ws, f"""
    (function(){{
        var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(!s) return 'no_input';
        var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        setter.call(s, {name_json});
        s.dispatchEvent(new Event('input',  {{bubbles:true}}));
        s.dispatchEvent(new Event('change', {{bubbles:true}}));
        return 'ok';
    }})()
    """, 15)
    time.sleep(1.5)
    # 再取坐标
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 18)
    if pos:
        log(f"  搜索框过滤后找到账号: {name}")
        # v1101.2: CDP 清空搜索框,恢复完整列表
        js(main_ws, """
        (function(){
            var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
            if(!s) return;
            var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(s, '');
            s.dispatchEvent(new Event('input',  {bubbles:true}));
            s.dispatchEvent(new Event('change', {bubbles:true}));
        })()
        """, 19)
        time.sleep(0.5)
        return json.loads(pos)
    return None


# ========== webview 精确匹配（含3次重试） ==========


def collect_accounts(main_ws):
    """从罐头左侧栏读取所有账号（处理虚拟滚动列表）"""
    log("开始收集账号列表...")
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 110)
    time.sleep(0.8)

    accounts = []
    seen = set()
    same_count = 0
    has_scrolled = False

    for _ in range(1000):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            var names = [];
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t) names.push(t);
            }}
            return JSON.stringify(names);
        }})()
        """, 111)
        if v:
            for n in json.loads(v):
                if n not in seen:
                    seen.add(n)
                    accounts.append(n)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no';
            var before = c.scrollTop;
            c.scrollTop += 200;
            return before + '->' + c.scrollTop;
        })()
        """, 112)
        time.sleep(0.3)

        if result and result != "no":
            parts = result.split("->")
            if len(parts) == 2:
                before_top, after_top = parts[0].strip(), parts[1].strip()
                if after_top != '0' and after_top != before_top:
                    has_scrolled = True
                    same_count = 0
                elif before_top == after_top and has_scrolled:
                    same_count += 1
                    if same_count >= 4:
                        break

    # 兜底:虚拟滚动列表,最后几个账号常因 lazy render 漏读
    # 强制滚到容器最底部,等 1.5s DOM 稳定再扫一次
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = c.scrollHeight;
    })()
    """, 113)
    time.sleep(1.5)
    v = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        var names = [];
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t) names.push(t);
        }}
        return JSON.stringify(names);
    }})()
    """, 114)
    if v:
        names = json.loads(v)
        added = 0
        for n in names:
            if n not in seen:
                seen.add(n)
                accounts.append(n)
                added += 1
        if added:
            log(f"  兜底:滚到底再扫一次,补收 {added} 个账号(虚拟滚动 lazy render 漏发)")

    log(f"共收集到 {len(accounts)} 个账号")
    return accounts


def _find_webview_once(main_ws, name):
    partition = js(main_ws, """
    (function(){
        var webviews = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<webviews.length;i++){
            var p = webviews[i].getAttribute('partition');
            if(!p || !p.startsWith('persist:')) continue;
            var r = webviews[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = p; }
        }
        return best;
    })()
    """, 15)

    if not partition:
        return None

    marker = f"_mk{random.randint(100000, 999999)}"
    r = cdp(main_ws, "Runtime.evaluate", {
        "expression": f"""
        new Promise(function(resolve){{
            var wv = document.querySelector('webview[partition="{partition}"]');
            if(!wv){{ resolve('no_wv'); return; }}
            wv.executeJavaScript('window.{marker}=1').then(function(){{
                resolve('ok');
            }}).catch(function(){{
                resolve('err');
            }});
        }})
        """,
        "awaitPromise": True,
        "timeout": 8000
    }, 200)

    if r.get("result", {}).get("value") != "ok":
        return None

    tabs = get_tabs()
    for t in tabs:
        if "webSocketDebuggerUrl" not in t:
            continue
        try:
            wsc = ws_connect(t["webSocketDebuggerUrl"], timeout=3)
            wsc.send(json.dumps({"id": 1, "method": "Runtime.evaluate",
                                 "params": {"expression": f"window.{marker}===1"}}))
            deadline = time.time() + 3
            found = False
            while time.time() < deadline:
                d = json.loads(wsc.recv())
                if d.get("id") == 1:
                    found = d.get("result", {}).get("result", {}).get("value") is True
                    break
            wsc.close()
            if found:
                return t["webSocketDebuggerUrl"]
        except:
            pass

    return None


def find_account_webview(main_ws, name):
    for retry in range(3):
        result = _find_webview_once(main_ws, name)
        if result:
            return result
        if retry < 2:
            log(f"  重试 webview ({retry+2}/3)...")
            time.sleep(2)
    return None


def _search_box_set(main_ws, value):
    """侧边栏搜索框设置值(空字符串=清空)。React-friendly: 用原型 setter 触发 input/change 事件。"""
    val_json = json.dumps(value)
    return js(main_ws, f"""
    (function(){{
        var inputs = document.querySelectorAll('input');
        for(var i=0;i<inputs.length;i++){{
            var ph = inputs[i].getAttribute('placeholder') || '';
            if(ph.indexOf('账号') !== -1 || ph.indexOf('手机号') !== -1){{
                var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                setter.call(inputs[i], '');
                inputs[i].dispatchEvent(new Event('input', {{bubbles:true}}));
                if({val_json}.length > 0){{
                    setter.call(inputs[i], {val_json});
                    inputs[i].dispatchEvent(new Event('input', {{bubbles:true}}));
                    inputs[i].dispatchEvent(new Event('change', {{bubbles:true}}));
                }}
                return 'ok';
            }}
        }}
        return 'no_input';
    }})()
    """, 50)


def _locate_filtered_account(main_ws, name):
    """搜索框过滤后,直接抓第一个匹配账号的中心坐标(viewport 相对)。"""
    name_json = json.dumps(name)
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 51)
    if not pos:
        return None
    return json.loads(pos)


def find_or_reopen_webview(main_ws, name, reopen_attempts=2):
    """虚拟滚动底部账号 webview partition 不渲染时,通过侧边栏搜索框定位账号让 webview 重建。

    流程:正常 find_account_webview 失败 → 关 tab → 搜索框输入账号名 → 等过滤 →
    抓过滤后的账号坐标 → click → 等更长时间 → 再尝试 find_account_webview。
    搜索框过滤后只剩匹配项渲染在 DOM 顶部,不再受虚拟滚动影响。
    """
    ws_url = find_account_webview(main_ws, name)
    if ws_url:
        return ws_url

    for attempt in range(reopen_attempts):
        log(f"  webview partition 失败,搜索框重建 {attempt+1}/{reopen_attempts}: 输入 \"{name}\"")
        try:
            close_current_tab(main_ws)
        except Exception as e:
            log(f"  关 tab 异常(忽略): {e}")
        time.sleep(1.0)

        # 用搜索框过滤
        rs = _search_box_set(main_ws, name)
        if rs != 'ok':
            log("  搜索框定位失败,降级用 scroll_find_account")
            pos = scroll_find_account(main_ws, name)
        else:
            time.sleep(1.5)  # 等过滤结果 React 渲染完
            pos = _locate_filtered_account(main_ws, name)
            if not pos:
                log("  搜索过滤后仍找不到,降级用 scroll_find_account")
                _search_box_set(main_ws, "")  # 清空恢复列表
                time.sleep(0.5)
                pos = scroll_find_account(main_ws, name)

        if not pos:
            log(f"  搜索/滚动都找不到 {name},重建中止")
            _search_box_set(main_ws, "")
            return None

        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD + 2)  # 比首次多等 2 秒,给虚拟滚动 lazy render 留余地

        ws_url = find_account_webview(main_ws, name)

        # 不论成败,清空搜索框还原列表(避免影响后续账号)
        _search_box_set(main_ws, "")
        time.sleep(0.3)

        if ws_url:
            log(f"  webview 重建成功(尝试 {attempt+1})")
            return ws_url

    return None


def detect_account_error(wsc):
    page_text = js(wsc, "document.body.innerText || ''", 70) or ""
    for reason, keywords in {
        "失登": ["请登录", "登录已失效", "账号已下线", "重新登录"],
        "封号": ["账号已被封禁", "账号异常", "账号被封"],
        "禁言": ["账号被禁言", "发言受限", "无法发布"],
        "限流": ["操作频繁", "请稍后再试"],
    }.items():
        for kw in keywords:
            if kw in page_text:
                return reason
    return None


def close_popup(ws):
    v = js(ws, """
    (function(){
        var b = document.querySelector('.close-btn,[class*="close-btn"]');
        if(b){var r=b.getBoundingClientRect();
        if(r.width>0) return JSON.stringify({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});}
        return null;
    })()
    """, 50)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], 51)
        time.sleep(0.5)



# [v1101 P9] 阅读量回查 — 抄自 mini 微头条
def check_reading_stats(ws_url, account_name):
    """导航到数据概览页 → 读取三篇文章的阅读量。任一超过 ALERT_THRESHOLD 则写入 ALERT_FILE。"""
    try:
        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/index'", 310)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        reads_raw = js(wsc, """
        (function(){
            var items = document.querySelectorAll('.recent-works-item');
            var results = [];
            for(var i=0;i<items.length;i++){
                var timeEl = items[i].querySelector('span.time');
                var timeText = timeEl ? timeEl.textContent.trim() : '未知时间';
                var labels = items[i].querySelectorAll('span.label');
                for(var j=0;j<labels.length;j++){
                    if(labels[j].textContent.trim() === '阅读量'){
                        var prev = labels[j].previousElementSibling;
                        if(prev){
                            var num = parseInt(prev.textContent.trim().replace(/,/g,''), 10);
                            if(!isNaN(num)) results.push({count: num, time: timeText});
                        }
                    }
                }
            }
            return JSON.stringify(results);
        })()
        """, 311)
        wsc.close()

        reads = json.loads(reads_raw) if reads_raw else []
        if not reads:
            log("  阅读量: 未读取到数据")
            return 0

        counts = [r['count'] for r in reads]
        log(f"  阅读量: {counts}")
        high = [r for r in reads if r['count'] >= ALERT_THRESHOLD]
        if high and ALERT_FILE:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            with open(ALERT_FILE, "a", encoding="utf-8") as f:
                f.write(f"[{ts}] 账号 {account_name} 高阅读量:")
                f.write(chr(10))
                for r in high:
                    f.write(f"  {r['time']} — {r['count']} 阅读量")
                    f.write(chr(10))
                f.write(chr(10))
            log(f"  ★ 高阅读量提醒: {[r['count'] for r in high]} → 已写入 {ALERT_FILE}")
            return len(high)
        return 0
    except Exception as e:
        log(f"  阅读量检测出错: {e}")
        return 0


def close_current_tab(main_ws):
    v = js(main_ws, """
    (function(){
        var closes = document.querySelectorAll('.chrome-tab-close');
        if(closes.length === 0) return null;
        for(var i=0;i<closes.length;i++){
            var tab = closes[i].closest('.chrome-tab');
            if(tab && tab.classList.contains('active')){
                var r = closes[i].getBoundingClientRect();
                if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
        }
        var last = closes[closes.length-1];
        var r = last.getBoundingClientRect();
        return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
    })()
    """, 55)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 56)
        time.sleep(0.5)


# ========== 发布流程（定时发布版） ==========

def check_system_notice(ws_url, account_name):
    """
    [v1102] 导航到消息中心 → 点击 系统通知 + 审核通知 频道
    → 读取 2 天内(今天+昨天)的完整消息原文写入 NOTICE_FILE
    新 selector: .conversation-box.notify-im-user-item (替代旧 span.name)
    新提取: body.innerText 按日期行切分(MM-DD HH:MM / YYYY-MM-DD / 昨日/今日 HH:MM)
    """
    try:
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        today_short     = today.strftime("%m-%d")
        yesterday_short = yesterday.strftime("%m-%d")
        today_full      = today.strftime("%Y-%m-%d")
        yesterday_full  = yesterday.strftime("%Y-%m-%d")

        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/personal/message?type=message_letter'", 300)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        time.sleep(2.5)
        notices = []

        for channel in ["系统通知", "审核通知"]:
            channel_json = channel.replace('"', '\\"')
            clicked = js(wsc, f"""
            (function(){{
                var items = document.querySelectorAll('.conversation-box.notify-im-user-item');
                for(var i=0; i<items.length; i++){{
                    var t = (items[i].innerText || '').trim();
                    if(t.indexOf("{channel_json}") === 0){{
                        items[i].click();
                        return 'ok';
                    }}
                }}
                return null;
            }})()
            """, 301)

            if not clicked:
                log(f"  未找到频道: {channel}")
                continue

            time.sleep(2.5)

            result = js(wsc, f"""
            (function() {{
                var todayShort = "{today_short}";
                var yesterdayShort = "{yesterday_short}";
                var todayFull = "{today_full}";
                var yesterdayFull = "{yesterday_full}";
                var lines = (document.body.innerText || '').split(/\\r?\\n/);
                var results = [];
                var current = '';
                var currentDate = '';
                var inWindow = false;
                function dateInfo(line) {{
                    var m = line.match(/^(\\d{{2}}-\\d{{2}})\\s+\\d{{2}}:\\d{{2}}$/);
                    if (m) return m[1] === todayShort || m[1] === yesterdayShort;
                    m = line.match(/^(\\d{{4}}-\\d{{2}}-\\d{{2}})\\s+\\d{{2}}:\\d{{2}}$/);
                    if (m) return m[1] === todayFull || m[1] === yesterdayFull;
                    if (/^昨日\\s+\\d{{2}}:\\d{{2}}$/.test(line)) return true;
                    if (/^今日\\s+\\d{{2}}:\\d{{2}}$/.test(line)) return true;
                    return null;
                }}
                function isDateLine(line) {{
                    return /^(\\d{{2}}-\\d{{2}}|\\d{{4}}-\\d{{2}}-\\d{{2}}|昨日|今日)\\s+\\d{{2}}:\\d{{2}}$/.test(line);
                }}
                for (var i = 0; i < lines.length; i++) {{
                    var line = lines[i].trim();
                    if (!line) continue;
                    if (isDateLine(line)) {{
                        if (inWindow && current.trim()) {{
                            results.push(currentDate + '\\n' + current.trim());
                        }}
                        currentDate = line;
                        current = '';
                        inWindow = (dateInfo(line) === true);
                    }} else if (inWindow) {{
                        current += line + '\\n';
                    }}
                }}
                if (inWindow && current.trim()) {{
                    results.push(currentDate + '\\n' + current.trim());
                }}
                // 去重(预览和详情可能重复)
                var seen = {{}};
                var dedup = [];
                for (var k = 0; k < results.length; k++) {{
                    var key = results[k].substring(0, 80);
                    if (!seen[key]) {{ seen[key] = true; dedup.push(results[k]); }}
                }}
                return JSON.stringify(dedup);
            }})()
            """, 302)

            if result:
                try:
                    msgs = json.loads(result)
                    for msg in msgs:
                        notices.append(f"【{channel}】\n{msg}")
                except:
                    pass

        wsc.close()

        violation_count = 0
        if notices:
            ts_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            # 分析违规类通知
            violations = []
            for n in notices:
                for cat, kws in VIOLATION_KEYWORDS.items():
                    for kw in kws:
                        if kw in n:
                            violations.append((cat, n))
                            break
            # 写系统通知 — 完整内容,不截断
            content_str = f"\n[{ts_str}] 账号 {account_name} 2 天内通知 ({len(notices)} 条):\n"
            for n in notices:
                content_str += f"\n--- 通知 ---\n{n}\n"
            content_str += "\n" + "=" * 60 + "\n"
            with open(NOTICE_FILE, "a", encoding="utf-8") as f:
                f.write(content_str)
            log(f"  ⚠ 2 天内通知 {len(notices)} 条 → 系统通知.txt")
            # 写违规提醒
            if violations:
                vcontent = f"[{ts_str}] 账号 {account_name} 违规/扣分提醒:\n"
                for cat, msg in violations:
                    vcontent += f"  [{cat}] {msg[:300]}...\n"
                vcontent += "\n"
                with open(VIOLATION_FILE, "a", encoding="utf-8") as f:
                    f.write(vcontent)
                violation_count = len(violations)
                log(f"  ⚠ 违规/扣分 {violation_count} 条 → 违规提醒.txt")
        else:
            log("  系统/审核通知: 2 天内无新通知")
        return len(notices), violation_count
    except Exception as e:
        log(f"  系统通知检测出错: {e}")
        return 0, 0

def _check_notice_once(ws_url, account_name):
    """[v1102] publish 调用前 wrapper: 当天读 1 次 NOTICE,持久化 set"""
    global notice_checked_set
    if account_name in notice_checked_set:
        log(f"  系统/审核通知:{account_name} 当天已读过,跳过")
        return
    check_system_notice(ws_url, account_name)
    notice_checked_set.add(account_name)
    try:
        with open(NOTICE_CHECKED_FILE, "a", encoding="utf-8") as _ncf:
            _ncf.write(f"{account_name}|{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    except Exception as _e:
        log(f"  [v1102] 写 notice_checked.txt 失败: {_e}")


def publish_article_timer(ws_url, doc_path, main_ws, account_name, timer_time=None):
    """定时发布一篇文章，timer_time格式: YYYY-MM-DD HH:MM"""
    # [v1102] publish 前检 NOTICE(每号当天 1 次,持久化)
    _check_notice_once(ws_url, account_name)
    try:
        wsc = ws_connect(ws_url, timeout=10)
    except Exception as e:
        return False, f"连接失败: {e}"

    close_popup(wsc)
    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    js(wsc, "location.href='https://mp.toutiao.com/profile_v4/graphic/publish'", 60)
    wsc.close()
    time.sleep(4)

    new_ws_url = None
    for _ in range(12):
        tabs = get_tabs()
        for t in tabs:
            if "graphic/publish" in t.get("url", "") and "webSocketDebuggerUrl" in t:
                new_ws_url = t["webSocketDebuggerUrl"]
                break
        if new_ws_url:
            break
        time.sleep(0.5)

    if not new_ws_url:
        return False, "导航到发文页失败"

    try:
        wsc = ws_connect(new_ws_url, timeout=10)
    except Exception as e:
        return False, f"重连失败: {e}"

    close_popup(wsc)
    time.sleep(1)

    current_url = js(wsc, "location.href", 59) or ""
    if "login" in current_url:
        wsc.close()
        return False, "失登"

    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    # 取 webview 屏幕坐标
    wv_s = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 61)

    if not wv_s:
        wsc.close()
        return False, "无法获取 webview 屏幕坐标"
    wv0 = json.loads(wv_s)

    def get_wv():
        r = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxArea = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var area = r.width * r.height;
                if(area > maxArea){ maxArea = area; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
        })()
        """, 62)
        return json.loads(r) if r else wv0

    # 关闭草稿提示条
    draft_close = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t === '\u7ee7\u7eed\u7f16\u8f91'){
                var bar = els[i].closest('[class]');
                while(bar){
                    var x = bar.querySelector('[class*="close"],[class*="Close"]');
                    if(x){ var r = x.getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}); }
                    bar = bar.parentElement && bar.parentElement.closest('[class]');
                    if(!bar) break;
                }
            }
        }
        return null;
    })()
    """, 58)
    if draft_close:
        dc = json.loads(draft_close)
        wv_t = get_wv()
        subprocess.run(["cliclick", f"c:{wv_t['sx']+dc['x']},{wv_t['sy']+dc['y']}"], capture_output=True)
        time.sleep(0.5)
    else:
        log("  无草稿提示条")

    # 激活窗口
    subprocess.run(["osascript", "-e", """
tell application "System Events"
    tell process "创作罐头"
        set frontmost to true
    end tell
end tell
"""], capture_output=True)
    time.sleep(0.6)

    # 点文档导入按钮（工具栏最后一个）
    v = None
    for _ in range(20):
        v = js(wsc, """
        (function(){
            var btns = Array.from(document.querySelectorAll('.syl-toolbar-button')).filter(function(b){
                return b.getBoundingClientRect().width > 0;
            });
            if(btns.length > 0){
                var last = btns[btns.length - 1];
                var r = last.getBoundingClientRect();
                return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
            return null;
        })()
        """, 63)
        if v: break
        time.sleep(0.5)
    if not v:
        wsc.close()
        return False, "找不到文档导入按钮"

    p = json.loads(v)
    wv_t = get_wv()
    import_x = wv_t['sx'] + p['x']
    import_y = wv_t['sy'] + p['y']
    title_x = wv_t['sx'] + 400
    title_y = wv_t['sy'] + 50
    ensure_gtg_top()
    subprocess.run(["cliclick", f"c:{title_x},{title_y}"], capture_output=True)
    time.sleep(0.5)
    # [v1101 P7] cliclick 文档导入 + 等弹窗,失败重试 3 次(每次重 activate + 重读坐标)
    sel = None
    for click_attempt in range(3):
        attempt_str = f" [第{click_attempt+1}次]" if click_attempt > 0 else ""
        log(f"  cliclick 点击文档导入 ({import_x},{import_y}){attempt_str}")
        ensure_gtg_top()
        subprocess.run(["cliclick", f"m:{import_x},{import_y}"], capture_output=True)
        time.sleep(0.3)
        subprocess.run(["cliclick", f"c:{import_x},{import_y}"], capture_output=True)
        time.sleep(1.5)

        for _ in range(60):
            sel = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '选择文档'){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({bx: Math.round(r.left+r.width/2), by: Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 65)
            if sel: break
            time.sleep(0.5)

        if sel:
            if click_attempt > 0:
                log(f"  cliclick 第{click_attempt+1}次成功唤出弹窗")
            break

        if click_attempt < 2:
            log(f"  弹窗未出,重 activate + 重读坐标后重试")
            ensure_gtg_top()
            wv_re = get_wv()
            import_x = wv_re['sx'] + p['x']
            import_y = wv_re['sy'] + p['y']
            log(f"  重试前坐标更新: ({import_x},{import_y})")

    if not sel:
        wsc.close()
        return False, "文档导入弹窗未出现(3 次 cliclick 重试均失败)"

    sb = json.loads(sel)
    wv_t = get_wv()
    screen_x = wv_t['sx'] + sb['bx']
    screen_y = wv_t['sy'] + sb['by']

    doc_escaped = doc_path.replace("\\", "/")
    # AppleScript 字符串里的反斜杠和双引号要转义（路径里通常没有，但保险）
    safe_path = doc_escaped.replace("\\", "\\\\").replace('"', '\\"')
    result_holder = [None]

    def sheet_exists():
        r = subprocess.run([
            "osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of window 1)'
        ], capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def go_to_folder_sheet_exists():
        r = subprocess.run([
            "osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of sheet 1 of window 1)'
        ], capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def get_sheet_rect():
        """拿主对话框 position+size。NSOpenPanel 在新 macOS 里 entire contents 遍历超时，
        但 position/size 直接读很快。返回 (x,y,w,h) 或 None。"""
        r = subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        try
            set {px, py} to position of sheet 1 of window 1
            set {sw, sh} to size of sheet 1 of window 1
            return (px as string) & "|" & (py as string) & "|" & (sw as string) & "|" & (sh as string)
        on error
            return ""
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=5)
        try:
            nums = re.findall(r"\d+", r.stdout)
            if len(nums) >= 4:
                return tuple(int(v) for v in nums[:4])
        except Exception:
            pass
        return None

    def click_dialog_button(which):
        """which: 'open' 或 'cancel'。NSOpenPanel 右下角按钮坐标相对固定，
        用 cliclick 物理点击不经 keystroke，不受 frontmost 焦点影响。"""
        rect = get_sheet_rect()
        if not rect:
            return False
        x, y, w, h = rect
        if which == 'open':
            cx, cy = x + w - 55, y + h - 35
        else:
            cx, cy = x + w - 150, y + h - 35
        subprocess.run(["cliclick", f"m:{cx},{cy}"], capture_output=True)
        time.sleep(0.1)
        subprocess.run(["cliclick", f"c:{cx},{cy}"], capture_output=True)
        time.sleep(0.5)
        return True

    def press_esc(times=2):
        """兜底关对话框。优先 cliclick 点'取消'（物理层不受焦点影响），
        不行再 fallback key code 53。"""
        click_dialog_button('cancel')
        for _ in range(times):
            if not sheet_exists():
                return
            subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        key code 53
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.4)

    def fill_dialog():
        # 等文件对话框 sheet 出现
        deadline = time.time() + 15
        appeared = False
        while time.time() < deadline:
            if sheet_exists():
                appeared = True
                break
            time.sleep(0.3)
        if not appeared:
            result_holder[0] = False
            return
        time.sleep(0.6)
        # 方案A：直接给"前往文件夹"输入框赋值，绕过 clipboard。
        # clipboard 方案会被输入法/剪贴板管理器/循环间隔偶发污染，定时发布多账号连发时尤甚。
        # UI 路径由 probe_dialog.py 探测确认：text field 1 of sheet 1 of sheet 1 of window 1
        r = subprocess.run(["osascript", "-e", f'''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {{command down, shift down}}
        delay 1.5
        try
            set target to text field 1 of sheet 1 of sheet 1 of window 1
            set value of target to "{safe_path}"
            delay 0.3
            set rb to (value of target) as string
            if rb is "{safe_path}" then
                return "OK"
            else
                return "ERR:readback:" & rb
            end if
        on error errmsg
            return "ERR:" & errmsg
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=20)
        direct_ok = "OK" in (r.stdout or "")
        if not direct_ok:
            log(f"  直接赋值失败（{(r.stdout or '').strip()[:80]}），回退clipboard")
            # 关掉可能残留的"前往文件夹"小框
            press_esc(1)
            time.sleep(0.5)
            # 方案B：pbcopy + pbpaste 校验 + 最多 5 次重试，挡住 clipboard 被污染
            clipboard_ok = False
            for _ in range(5):
                subprocess.run(["pbcopy"], input=doc_escaped, text=True)
                time.sleep(0.15)
                rb = subprocess.run(["pbpaste"], capture_output=True, text=True).stdout
                if rb.strip() == doc_escaped.strip():
                    clipboard_ok = True
                    break
                time.sleep(0.2)
            if not clipboard_ok:
                log("  clipboard 校验 5 次失败")
                result_holder[0] = False
                return
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {command down, shift down}
        delay 1.5
        keystroke "a" using {command down}
        delay 0.4
        keystroke "v" using {command down}
        delay 1.2
    end tell
end tell
'''], capture_output=True)

        # Step 2: 按回车关"前往文件夹"小框，最多 5 次（keystroke 偶发焦点问题需检测+重试）
        go_closed = False
        for i in range(5):
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.1
tell application "System Events"
    tell process "创作罐头"
        keystroke return
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.8)
            if not go_to_folder_sheet_exists():
                go_closed = True
                if i > 0:
                    log(f"  前往文件夹 回车{i+1}次后关闭")
                break
        if not go_closed:
            log("  前往文件夹 5次回车未关 → cliclick 点取消")
            click_dialog_button('cancel')
            result_holder[0] = False
            return

        # [v1101 P1] Step 3 跳过:macOS 26 NSOpenPanel 不再自动关,Step 4 必须打,6s 硬等纯发呆

        # Step 4: 主框没自动关 → cliclick 物理点"打开"按钮（最多 3 次）
        log("  主对话框未自动关闭 → cliclick 点打开按钮")
        for _ in range(3):
            if not click_dialog_button('open'):
                break
            for _ in range(4):  # 等 2 秒
                time.sleep(0.5)
                if not sheet_exists():
                    result_holder[0] = True
                    return

        # Step 5: 彻底卡死 → 点取消，外层 3 次重试会重开"选择文档"
        log("  对话框完全卡死 → cliclick 点取消")
        click_dialog_button('cancel')
        result_holder[0] = False

    # 重试最多 3 次，扛住偶发对话框 hang
    dialog_ok = False
    for dialog_attempt in range(3):
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        ensure_gtg_top()
        subprocess.run(["cliclick", f"m:{screen_x},{screen_y}"], capture_output=True)
        time.sleep(0.2)
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=30)
        if result_holder[0]:
            dialog_ok = True
            if dialog_attempt > 0:
                log(f"  文件对话框第{dialog_attempt+1}次成功")
            break
        # 失败前再保险按一次 ESC 确保对话框关闭，否则下次循环点"选择文档"会被堵
        press_esc(2)
        time.sleep(1)
        log(f"  第{dialog_attempt+1}次对话框处理失败，准备重试")

    if not dialog_ok:
        wsc.close()
        return False, "文件对话框反复卡住，3次重试均失败"

    time.sleep(5)
    char_count = 0
    for _ in range(15):
        v = js(wsc, """
        (function(){
            // [v1101 P3] 取最长 ProseMirror 元素(避免命中标题 placeholder 5 字)
            var els = document.querySelectorAll('.ProseMirror');
            var max = 0;
            for (var i = 0; i < els.length; i++) {
                var l = els[i].textContent.trim().length;
                if (l > max) max = l;
            }
            return max;
        })()
        """, 75)
        char_count = int(v) if v else 0
        if char_count >= 50:
            break
        time.sleep(0.8)

    log(f"  文章字数: {char_count}")
    # [v1101 P2] 字数<50 重试 fill_dialog 一次(借鉴 neo retry,救夜间冤杀)
    if char_count < 50:
        log(f"  对话框已关但字数仅 {char_count}（文档未真导入），重试 fill_dialog")
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        ensure_gtg_top()
        subprocess.run(["cliclick", f"m:{screen_x},{screen_y}"], capture_output=True)
        time.sleep(0.2)
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=30)
        time.sleep(5)
        char_count = 0
        for _ in range(15):
            v = js(wsc, """
            (function(){
                var els = document.querySelectorAll('.ProseMirror');
                var max = 0;
                for (var i = 0; i < els.length; i++) {
                    var l = els[i].textContent.trim().length;
                    if (l > max) max = l;
                }
                return max;
            })()
            """, 75)
            char_count = int(v) if v else 0
            if char_count >= 50:
                break
            time.sleep(0.8)
        log(f"  重试后字数: {char_count}")
        if char_count < 50:
            wsc.close()
            return False, "文档导入后内容为空(重试1次仍空)"

    js(wsc, """
    (function(){
        var el = document.querySelector('.ProseMirror') || document.body;
        el.scrollTop = el.scrollHeight;
        window.scrollTo(0, document.body.scrollHeight);
    })()
    """, 76)
    time.sleep(2)
    for _ in range(20):
        loading = js(wsc, """
        (function(){
            var imgs = document.querySelectorAll('img');
            for(var i=0;i<imgs.length;i++){
                if(!imgs[i].complete || imgs[i].naturalWidth === 0) return true;
            }
            return false;
        })()
        """, 77)
        if not loading:
            break
        time.sleep(0.5)

    # 读信用分 / 处理首发
    credit_raw = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t.indexOf('\u4fe1\u7528\u5206') !== -1 && t.length < 20){
                var m = t.match(/(\d{1,3})\u5206/);
                if(m){ var n = parseInt(m[1], 10);
                    if(n % 5 === 0 && n >= 5 && n <= 100) return n; }
            }
        }
        return null;
    })()
    """, 78)
    credit_score = int(credit_raw) if credit_raw is not None else None
    log(f"  信用分: {credit_score if credit_score is not None else '未读取到'}")
    # [v1101.1] 信用分 < 60 → 硬终止
    if credit_score is not None and credit_score < 60:
        log(f"  ★ 信用分 {credit_score} < 60,硬终止")
        wsc.close()
        return False, "信用分过低"
    should_first = (account_name not in NOFIRST_ACCOUNTS) and (credit_score is not None and credit_score >= 95)

    # [v1101.3] 头条首发复选框: 探测 + cliclick 点击 + 回读校验 + 三轮兜底 + 硬保护
    _PROBE_JS = r"""
    (function(){
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '头条首发'){
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){
                    if(p.tagName === 'LABEL' && p.classList.contains('byte-checkbox')){
                        var isChecked = p.classList.contains('byte-checkbox-checked');
                        var r = p.getBoundingClientRect();
                        var px = Math.round(r.left + r.width/2);
                        var py = Math.round(r.top + r.height/2);
                        return JSON.stringify({found:true, checked:isChecked, cb_x:px, cb_y:py});
                    }
                    p = p.parentElement;
                }
            }
        }
        return null;
    })()
    """
    _CLICK_JS = r"""
    (function(){
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '头条首发'){
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){
                    if(p.tagName === 'LABEL' && p.classList.contains('byte-checkbox')){
                        try{ p.click(); }catch(e){}
                        var inp = p.querySelector('input[type="checkbox"]');
                        if(inp){
                            try{ inp.dispatchEvent(new Event('change',{bubbles:true})); }catch(e){}
                        }
                        return 'label';
                    }
                    p = p.parentElement;
                }
            }
        }
        return null;
    })()
    """
    _WV_JS = r"""
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """

    def _probe_first():
        raw = js(wsc, _PROBE_JS, 79)
        if not raw:
            return None
        try:
            return json.loads(raw)
        except Exception:
            return None

    def _wv_origin():
        raw = js(main_ws, _WV_JS, 80)
        if not raw:
            return None
        try:
            return json.loads(raw)
        except Exception:
            return None

    fr = _probe_first()
    if fr is None:
        log("  头条首发: 未找到复选框")
    elif bool(fr.get('checked')) == should_first:
        log(f"  头条首发: 已是{'勾选' if fr['checked'] else '未勾选'}，无需操作")
    else:
        attempts = []
        verified = False
        for attempt in range(1, 4):
            if attempt < 3 and fr.get('found'):
                wv = _wv_origin()
                if wv:
                    sx = wv['sx'] + fr['cb_x']
                    sy = wv['sy'] + fr['cb_y']
                    subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
                    attempts.append(f"cliclick#{attempt}@({sx},{sy})")
                    time.sleep(0.4)
                else:
                    attempts.append(f"cliclick#{attempt}=no_wv")
                    time.sleep(0.2)
            else:
                rc = js(wsc, _CLICK_JS, 30)
                time.sleep(0.4)
                attempts.append(f"js#{attempt}={rc}")
            fr2 = _probe_first()
            if fr2 and bool(fr2.get('checked')) == should_first:
                log(f"  头条首发: {'勾选' if should_first else '取消勾选'} 已校验 [{'/'.join(attempts)}]")
                verified = True
                break
            fr = fr2 if fr2 else fr
        if not verified:
            actual = fr.get('checked') if fr else None
            log(f"  ✗ 头条首发校准失败: 目标={should_first} 实际={actual} 尝试=[{'/'.join(attempts)}]")
            if (not should_first) and actual is True:
                log(f"  ★ 硬保护: 应取消首发但仍勾选,跳过该篇避免扣 5 分")
                wsc.close()
                return False, "首发取消失败(硬保护)"

    # ---- 立即/定时分支 fork ----
    if timer_time is None:
        # 点"预览并发布" + 等"确认发布"，最多重试5次（封面图加载慢时第一次点可能无效）
        confirm_clicked = False
        for attempt in range(5):
            # 每次重新取 webview 屏幕坐标
            wv_cur_r = js(main_ws, """
            (function(){
                var wvs = document.querySelectorAll('webview');
                var maxArea = 0, best = null;
                for(var i=0;i<wvs.length;i++){
                    var r = wvs[i].getBoundingClientRect();
                    var area = r.width * r.height;
                    if(area > maxArea){ maxArea = area; best = r; }
                }
                if(!best) return null;
                return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
            })()
            """, 79)
            wv_cur = json.loads(wv_cur_r) if wv_cur_r else wv0
    
            # 找"预览并发布"按钮
            v = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '\u9884\u89c8\u5e76\u53d1\u5e03' && !btns[i].disabled){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 80)
    
            if not v:
                if attempt == 0:
                    wsc.close()
                    return False, "找不到预览并发布按钮"
            else:
                p = json.loads(v)
                preview_x = wv_cur['sx'] + p['x']
                preview_y = wv_cur['sy'] + p['y']
                attempt_str = f" [第{attempt+1}次]" if attempt > 0 else ""
                log(f"  cliclick 点击预览并发布 ({preview_x},{preview_y}){attempt_str}")
                subprocess.run(["cliclick", f"m:{preview_x},{preview_y}"], capture_output=True)
                time.sleep(0.5)
                subprocess.run(["cliclick", f"c:{preview_x},{preview_y}"], capture_output=True)
                log("  已点击预览并发布")
                time.sleep(4)
    
            # 等"确认发布"按钮
            for i in range(60):
                time.sleep(0.5)
                if i == 10 and not confirm_clicked:
                    log(f"  [V1102.3] 5s 未见确认发布,补点预览并发布 ({preview_x},{preview_y})")
                    subprocess.run(["cliclick", f"c:{preview_x},{preview_y}"], capture_output=True)
                v2 = js(wsc, """
                (function(){
                    var btns = document.querySelectorAll('button');
                    for(var i=0;i<btns.length;i++){
                        var t = btns[i].textContent.trim();
                        if(t === '\u786e\u8ba4\u53d1\u5e03' || t.indexOf('\u786e\u8ba4\u53d1\u5e03') !== -1){
                            var r = btns[i].getBoundingClientRect();
                            if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                        }
                    }
                    return null;
                })()
                """, 83)
                if v2:
                    p2 = json.loads(v2)
                    confirm_x = wv_cur['sx'] + p2['x']
                    confirm_y = wv_cur['sy'] + p2['y']
                    log(f"  cliclick 点击确认发布 ({confirm_x},{confirm_y})")
                    subprocess.run(["cliclick", f"c:{confirm_x},{confirm_y}"], capture_output=True)
                    confirm_clicked = True
                    break
    
            if confirm_clicked:
                break
            if attempt < 4:
                # diag: 抓 DOM 写诊断,定位 JS 选择器为啥漏匹配 "确认发布"
                try:
                    diag = js(wsc, """
                    (function(){
                        var out = [];
                        var btns = document.querySelectorAll('button');
                        out.push('btn-tags='+btns.length);
                        for(var i=0;i<btns.length;i++){
                            var r = btns[i].getBoundingClientRect();
                            if(r.width>0 && r.height>0){
                                out.push('btn['+i+']="'+btns[i].textContent.trim().slice(0,30)+'" disabled='+btns[i].disabled);
                            }
                        }
                        out.push('---含发布字非button可见叶子---');
                        var all = document.querySelectorAll('*');
                        for(var i=0;i<all.length;i++){
                            if(all[i].tagName==='BUTTON') continue;
                            if(all[i].childElementCount!==0) continue;
                            var t = all[i].textContent.trim();
                            if(t.length<25 && t.indexOf('发布')!==-1){
                                var r = all[i].getBoundingClientRect();
                                if(r.width>0) out.push(all[i].tagName+'="'+t+'"');
                            }
                        }
                        return out.join(' | ');
                    })()
                    """, 84)
                    log(f"  [DIAG] {diag}")
                except Exception as e:
                    log(f"  [DIAG] dump失败: {e}")
                log(f"  确认发布未出现，封面图可能未加载完，准备第{attempt+2}次点击预览并发布...")
    
        if not confirm_clicked:
            wsc.close()
            return False, "未出现确认发布按钮"
    
        # 检测发布成功
        published = False
        for _ in range(20):
            time.sleep(0.5)
            t = js(wsc, """
            (function(){
                var all = document.querySelectorAll('*');
                for(var i=0;i<all.length;i++){
                    var t = all[i].textContent.trim();
                    if((t==='发布成功！'||t==='发布成功'||t==='提交成功！'||t==='提交成功')
                       && all[i].children.length<=1)
                        return t;
                }
                return null;
            })()
            """, 96)
            if t:
                published = True
                break
            cur_url = js(wsc, "location.href", 97) or ""
            if "graphic" in cur_url and "publish" not in cur_url:
                published = True
                break
    
        if not published:
            err_after = detect_account_error(wsc)
            wsc.close()
            if err_after:
                return False, err_after
            return False, "未检测到发布成功"
    
        wsc.close()
        log("  OK 立即发布成功")
        return True, "成功"

    # ---- 定时发布流程 ----
    wv_cur = get_wv()

    # cliclick 真实macOS鼠标点击（JS/CDP合成事件过不了React的 isTrusted 检查）
    # 关键：先scrollIntoView把按钮滚到视口中心，再换算成屏幕坐标
    popup_opened = False
    for attempt in range(5):
        pos = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim() === '\u5b9a\u65f6\u53d1\u5e03' && !btns[i].disabled){
                    var r = btns[i].getBoundingClientRect();
                    if(r.width > 0) return JSON.stringify({x: Math.round(r.left+r.width/2), y: Math.round(r.top+r.height/2)});
                }
            }
            return null;
        })()
        """, 149 + attempt * 3)
        if not pos:
            if attempt == 0:
                wsc.close()
                return False, "找不到定时发布按钮"
            time.sleep(1); continue
        p = json.loads(pos)
        # 每次重新取 wv 坐标（罐头layout可能变）
        wv_cur_r = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxA = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var a = r.width*r.height;
                if(a > maxA){ maxA = a; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX+best.left), sy: Math.round(window.screenY+best.top)});
        })()
        """, 150 + attempt * 3)
        wv_cur = json.loads(wv_cur_r) if wv_cur_r else get_wv()
        sx = wv_cur['sx'] + p['x']
        sy = wv_cur['sy'] + p['y']
        ensure_gtg_top()
        attempt_str = f" [第{attempt+1}次]" if attempt > 0 else ""
        log(f"  cliclick 点定时发布 ({sx},{sy}){attempt_str}")
        subprocess.run(["cliclick", f"m:{sx},{sy}"], capture_output=True)
        time.sleep(0.5)
        subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
        time.sleep(3)  # 罐头响应延迟
        # 等最多15秒看弹窗是否真打开
        for _ in range(30):
            v_cnt = js(wsc, """(function(){var bs=document.querySelectorAll('button');for(var i=0;i<bs.length;i++){var t=bs[i].textContent.trim();if(t==='\u9884\u89c8\u5e76\u5b9a\u65f6\u53d1\u5e03')return 1;}return 0;})()""", 151 + attempt * 3)
            if v_cnt and int(v_cnt) >= 1:
                popup_opened = True
                break
            time.sleep(0.5)
        if popup_opened:
            log(f"  弹窗已打开（第{attempt+1}次成功）")
            break
        log(f"  第{attempt+1}次点击后弹窗未开，重试")
    if not popup_opened:
        wsc.close()
        return False, "定时发布点击后弹窗始终未打开"
    p_t = {'y': -1}  # 占位

    # 等弹窗出现
    popup_ok = False
    for _ in range(20):
        v_popup = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim().indexOf('\u5b9a\u65f6') !== -1)
                    return 'found';
            }
            return null;
        })()
        """, 151)
        if v_popup == 'found':
            popup_ok = True
            break
        time.sleep(0.5)

    if not popup_ok:
        wsc.close()
        return False, "定时发布弹窗未出现"

    # 解析目标时间
    try:
        dt = datetime.strptime(timer_time, "%Y-%m-%d %H:%M")
    except Exception:
        wsc.close()
        return False, f"定时时间格式错误: {timer_time}"

    t_date1 = f"{dt.month}月{dt.day}日"
    t_date2 = f"{dt.month:02d}月{dt.day:02d}日"
    t_hour1 = str(dt.hour)
    t_hour2 = f"{dt.hour:02d}"
    t_min   = f"{dt.minute:02d}"

    # 弹窗DOM探测调试已移除（遍历全DOM会15s超时拖累流程）

    def click_select_option(select_idx, targets, mid_base):
        """点开第 select_idx 个(0-based)下拉，找 targets 中任一文字点击"""
        # 等到对应序号的 select-view 渲染出来（弹窗是渐进渲染的）
        slist = []
        for _ in range(30):
            sels_r = js(wsc, """
            (function(){
                var sels = document.querySelectorAll('[class*="arco-select-view-value"],[class*="arco-select-view-input"],[class*="select-view-value"]');
                var out = [];
                for(var i=0;i<sels.length;i++){
                    var r = sels[i].getBoundingClientRect();
                    if(r.width > 0 && r.height > 0)
                        out.push({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});
                }
                return JSON.stringify(out);
            })()
            """, mid_base)
            if sels_r:
                slist = json.loads(sels_r)
                if len(slist) > select_idx: break
            time.sleep(0.3)
        if select_idx >= len(slist):
            log(f"  select_idx={select_idx} 超出范围(共{len(slist)}个)")
            return False
        s = slist[select_idx]
        wv_now = get_wv()
        sx = wv_now['sx'] + s['x']
        sy = wv_now['sy'] + s['y']
        log(f"  cliclick 点开第{select_idx+1}个下拉 ({sx},{sy})")
        subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
        time.sleep(0.8)
        # 下拉打开后先滚到顶部，避免默认滚到当前时间导致目标选项在视野外
        js(wsc, """
        (function(){
            var lists = document.querySelectorAll('[class*="arco-select-popup"],[class*="arco-virtual-list"],[class*="select-popup"]');
            for(var i=0;i<lists.length;i++){
                var r = lists[i].getBoundingClientRect();
                if(r.width > 0 && r.height > 0){
                    lists[i].scrollTop = 0;
                    // 虚拟滚动内部容器也滚到顶
                    var inner = lists[i].querySelector('[class*="arco-virtual-list"]') || lists[i];
                    inner.scrollTop = 0;
                }
            }
        })()
        """, mid_base-1)
        time.sleep(0.3)
        for _ in range(20):  # 多给点循环，配合逐步滚动找到目标
            # 先尝试JS直接click（绕开坐标越界问题）
            js_clicked = js(wsc, f"""
            (function(){{
                var targets = {json.dumps(targets)};
                var items = document.querySelectorAll('[class*="arco-select-option"],[class*="select-option"],[class*="dropdown"] li,li[class*="option"]');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    for(var j=0;j<targets.length;j++){{
                        if(t === targets[j]){{
                            items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                            var ev = ['mousedown','mouseup','click'];
                            for(var k=0;k<ev.length;k++){{
                                items[i].dispatchEvent(new MouseEvent(ev[k],{{bubbles:true,cancelable:true,view:window}}));
                            }}
                            return 'clicked:' + targets[j];
                        }}
                    }}
                }}
                return null;
            }})()
            """, mid_base+1)
            if js_clicked:
                log(f"  JS点击 {js_clicked}")
                time.sleep(0.5)
                return True
            # 找不到则往下滚动下拉容器（已先滚到顶，逐步往下扫描所有选项）
            js(wsc, """
            (function(){
                var lists = document.querySelectorAll('[class*="arco-select-popup"],[class*="arco-virtual-list"],[class*="select-popup"]');
                for(var i=0;i<lists.length;i++){
                    var r = lists[i].getBoundingClientRect();
                    if(r.width > 0 && r.height > 0){
                        lists[i].scrollTop += 80;
                        var inner = lists[i].querySelector('[class*="arco-virtual-list"]') || lists[i];
                        inner.scrollTop += 80;
                    }
                }
            })()
            """, mid_base+2)
            time.sleep(0.2)
        log(f"  未找到目标选项: {targets}")
        return False

    # 等弹窗里的第一个下拉出现即可（click_select_option内部会自己等到3个都到位）
    time.sleep(1.2)

    if not click_select_option(0, [t_date1, t_date2], 160):
        wsc.close()
        return False, f"定时日期设置失败: {t_date1}"
    if not click_select_option(1, [t_hour1, t_hour2], 170):
        wsc.close()
        return False, f"定时小时设置失败: {t_hour1}"
    t_min_alt = str(dt.minute)
    if not click_select_option(2, [t_min, t_min_alt], 180):
        wsc.close()
        return False, f"定时分钟设置失败: {t_min}"

    log(f"  定时时间设置完成: {timer_time}")
    time.sleep(0.5)

    # JS 直接点"预览并定时发布"按钮
    prev_clicked = js(wsc, """
    (function(){
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            var t = btns[i].textContent.trim();
            if(t.indexOf('\u9884\u89c8') !== -1 && t.indexOf('\u5b9a\u65f6') !== -1 && !btns[i].disabled){
                btns[i].click();
                return 'clicked';
            }
        }
        return null;
    })()
    """, 190)
    if prev_clicked != 'clicked':
        wsc.close()
        return False, "找不到预览并定时发布按钮"
    log("  JS点击预览并定时发布")
    time.sleep(3)

    # 打印当前页面所有可见按钮（调试用）
    btns_now = js(wsc, """
    (function(){
        var out = [];
        var btns = document.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            var r = btns[i].getBoundingClientRect();
            if(r.width > 0 && r.height > 0){
                out.push(btns[i].textContent.trim());
            }
        }
        return JSON.stringify(out);
    })()
    """, 192)
    log(f"  当前可见按钮: {btns_now}")

    # JS 直接点预览页的"定时发布"按钮（跟编辑页同名——预览层是后渲染的，取最后一个）
    # click 后做二次校验：按钮没消失就补点，避免click事件被React吞掉导致丢一篇
    click_expr = """
    (function(){
        var btns = document.querySelectorAll('button');
        var matched = [];
        for(var i=0;i<btns.length;i++){
            var t = btns[i].textContent.trim();
            if(t === '\u5b9a\u65f6\u53d1\u5e03' && !btns[i].disabled){
                var r = btns[i].getBoundingClientRect();
                if(r.width > 0 && r.height > 0) matched.push(btns[i]);
            }
        }
        if(matched.length === 0) return null;
        matched[matched.length - 1].click();
        return 'clicked:' + matched.length;
    })()
    """
    confirm_clicked = False
    for _ in range(60):
        time.sleep(0.5)
        v2 = js(wsc, click_expr, 191)
        if not v2:
            continue
        log(f"  JS点击预览页定时发布 ({v2})")
        confirm_clicked = True
        # 二次校验：等1.5秒看按钮是否消失/URL是否跳转，没动就补点
        for retry in range(2):
            time.sleep(1.5)
            still = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '\u5b9a\u65f6\u53d1\u5e03' && !btns[i].disabled){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0 && r.height > 0) return 1;
                    }
                }
                return 0;
            })()
            """, 193 + retry)
            if not still or int(still) == 0:
                break
            log(f"  按钮仍在，补点一次（retry={retry+1}）")
            js(wsc, click_expr, 195 + retry)
        break

    if not confirm_clicked:
        wsc.close()
        return False, "未出现预览页定时发布按钮"

    # 检测发布成功
    published = False
    for _ in range(20):
        time.sleep(0.5)
        t_txt = js(wsc, """
        (function(){
            var all = document.querySelectorAll('*');
            for(var i=0;i<all.length;i++){
                var t = all[i].textContent.trim();
                if((t==='发布成功！'||t==='发布成功'||t==='提交成功！'||t==='提交成功'||t==='定时发布成功')
                   && all[i].children.length<=1)
                    return t;
            }
            return null;
        })()
        """, 196)
        if t_txt:
            published = True
            break
        cur_url = js(wsc, "location.href", 197) or ""
        if "graphic" in cur_url and "publish" not in cur_url:
            published = True
            break

    if not published:
        err_after = detect_account_error(wsc)
        wsc.close()
        if err_after:
            return False, err_after
        return False, "未检测到发布成功"

    wsc.close()
    log(f"  OK 定时发布成功 → {timer_time}")
    # [v1102] 持久化最近 publish 成功账号
    try:
        with open(LAST_PUBLISHED_FILE, "a", encoding="utf-8") as _lpf:
            _lpf.write(f"{account_name}|{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    except Exception as _e:
        log(f"  [v1102] 写 last_published.txt 失败: {_e}")
    return True, "成功"


def get_docs():
    docs = []
    for p in ["*.docx", "*.doc"]:
        docs.extend(glob.glob(os.path.join(DOCS_FOLDER, p)))
    return sorted([d for d in docs if "已发送" not in d])


# [v1101.4] doc_pool 顺序取 + 校验,救"分发完源必删"导致罐头找不到文件
def _pop_doc(doc_pool):
    """从 doc_pool 顺序取一篇实存的 docx,失效引用就地剔除。返回 None 表示池已空。"""
    while doc_pool:
        doc = doc_pool.pop(0)
        if os.path.exists(doc):
            return doc
        log(f"  ! 源已删除(可能被外部分发),跳过: {os.path.basename(doc)}")
    return None


# [v1101.4] Stage 2 死磕用: 随机抽存在的 doc + 大循环重扫池
def _pick_doc(doc_pool):
    """从 doc_pool 抽一篇实存的 docx,失效引用就地清理。返回 None 表示池已空。"""
    import random
    while doc_pool:
        doc = random.choice(doc_pool)
        if os.path.exists(doc):
            return doc
        log(f"  ! 源已删除(可能被外部分发),从池剔除: {os.path.basename(doc)}")
        doc_pool.remove(doc)
    return None


def _resync_pool(doc_pool):
    """大循环开始前重扫素材池:剔除幽灵引用 + 加入新到的素材。返回 (剔除数, 新增数)。"""
    cur = set(get_docs())
    before = len(doc_pool)
    doc_pool[:] = [d for d in doc_pool if d in cur]
    removed = before - len(doc_pool)
    pool_set = set(doc_pool)
    new_docs = [d for d in cur if d not in pool_set]
    if new_docs:
        doc_pool.extend(new_docs)
    return removed, len(new_docs)


def _finalize_config(accounts_quota, success_by_acct, fail_docs_by_acct=None):
    """收尾：清空账号配置.xlsx「白名单」；把提交失败的账号写入「待补漏」。
    漏发数 = quota - 本次提交成功数（只考虑提交阶段，不管次日罐头实际发布）
    fail_docs_by_acct: {账号: [失败的文档basename...]}，用于填写「待补漏」的文稿名列
    """
    if not os.path.exists(CONFIG_EXCEL):
        log(f"警告: 账号配置.xlsx 不存在，跳过收尾")
        return
    fail_docs_by_acct = fail_docs_by_acct or {}
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb = openpyxl.load_workbook(CONFIG_EXCEL)

        # 1. 清空白名单
        if "白名单" in wb.sheetnames:
            wl = wb["白名单"]
            if wl.max_row >= 2:
                wl.delete_rows(2, wl.max_row - 1)

        # 2. 写待补漏（先清空旧数据再写新的）
        pending_rows = []
        for name, quota in accounts_quota:
            miss = quota - success_by_acct.get(name, 0)
            if miss > 0:
                docs = fail_docs_by_acct.get(name, [])
                # 文稿名列：该账号失败过的文档名拼起来（用 | 分隔），多于 miss 个也都记下便于追溯
                doc_names = " | ".join(d for d in docs if d)
                pending_rows.append((name, miss, doc_names, now))

        if "待补漏" in wb.sheetnames:
            pen = wb["待补漏"]
            if pen.max_row >= 2:
                pen.delete_rows(2, pen.max_row - 1)
            for i, row in enumerate(pending_rows, start=2):
                for j, v in enumerate(row, start=1):
                    pen.cell(row=i, column=j, value=v)

        # [v1102] 全员齐活 → 清空「本轮已发」 sheet(大循环末才 clear)
        全员齐活 = not pending_rows
        if 全员齐活 and "本轮已发" in wb.sheetnames:
            ws_sent = wb["本轮已发"]
            if ws_sent.max_row > 1:
                ws_sent.delete_rows(2, ws_sent.max_row - 1)

        wb.save(CONFIG_EXCEL)
        log(f"\n收尾：白名单已清空")
        if pending_rows:
            log(f"收尾：待补漏写入 {len(pending_rows)} 行（总漏发 {sum(r[1] for r in pending_rows)} 篇）")
            for r in pending_rows:
                log(f"  - {r[0]} 漏 {r[1]} 篇")
            log(f"  [v1102] 本次未齐活,「本轮已发」 sheet 保留(下次启动接着跑)")
        else:
            log(f"收尾：本次无提交失败，待补漏留空")
            log(f"  [v1102] 大循环全员齐活 → 「本轮已发」 sheet 已清空")
    except Exception as e:
        log(f"警告: 收尾写账号配置失败: {e}")


def move_to_sent(doc_path):
    os.makedirs(SENT_FOLDER, exist_ok=True)
    dest = os.path.join(SENT_FOLDER, os.path.basename(doc_path))
    if os.path.exists(dest):
        name, ext = os.path.splitext(os.path.basename(doc_path))
        dest = os.path.join(SENT_FOLDER, f"{name}_{int(time.time())}{ext}")
    shutil.move(doc_path, dest)
    log(f"  已移至已发送: {os.path.basename(dest)}")


# ========== 主流程 ==========

# ========== Stage 2 死磕主循环 (定时模式补尾) ==========
# 漏发账号在 Stage 1 跑完后,这里把它们在"now + 30min"重新排程,直到全部发完或命中 4 类硬终止
# 4 类硬终止 = {失登, 封号, 禁言, 侧边栏未找到}

def run_death_grip_timer(
    accounts,
    per_account_quota,       # dict[name -> 还需补几篇]
    doc_pool,
    main_ws,
    sub_rounds=3,
    max_fail_per_sub=3,
    initial_dead=None,
):
    """
    Stage 2: 用 publish_article_timer 立即模式 (timer_time=None) 补尾漏发账号——直接发，不再走定时排程。
    - 4 类硬终止 → 立即写硬终止 sheet 永久放弃
    - 其他失败 → 本小轮 max_fail_per_sub 次失败 → 跳本小轮(下小轮恢复)
    - 大循环 N 小轮 + 外层无限磕,直到全发完 / 4 类硬终止 / Ctrl+C
    """
    HARD_TERMINATE_REASONS = {"封号", "禁言", "侧边栏未找到", "信用分过低"}  # [v1101.1] 失登移除(改软重试),加信用分过低
    _HARD_TERMINATE_HEADERS = ["账号名", "终止原因", "终止时间", "本次已发篇数"]
    dead_terminated = dict(initial_dead) if initial_dead else {}
    acc_count = {a: 0 for a in accounts}

    def _append_hard_terminate(name, reason, count_so_far):
        try:
            wb = openpyxl.load_workbook(CONFIG_EXCEL)
            if "硬终止账号" not in wb.sheetnames:
                ws_h = wb.create_sheet("硬终止账号")
                ws_h.append(_HARD_TERMINATE_HEADERS)
            else:
                ws_h = wb["硬终止账号"]
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_h.append([name, reason, ts, count_so_far])
            wb.save(CONFIG_EXCEL)
        except Exception:
            pass

    def _do_publish(name, doc):
        log(f"\n  [Stage2 剩余 {len(doc_pool)} 篇] {name}  ->  {os.path.basename(doc)}")
        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未在侧边栏找到账号: {name}")
            return False, "侧边栏未找到"
        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD)

        ws_url = find_or_reopen_webview(main_ws, name)
        if not ws_url:
            log("  X 找不到 webview")
            close_current_tab(main_ws)
            return False, "webview匹配失败"

        try:
            success, reason = publish_article_timer(ws_url, doc, main_ws, name)  # timer_time=None → 立即发布
            if success:
                if doc in doc_pool:
                    doc_pool.remove(doc)
                acc_count[name] = acc_count.get(name, 0) + 1
                try:
                    move_to_sent(doc)
                except Exception:
                    pass
                # [v1101 P9] 发文成功后回查阅读量(高阅读触发告警)
                try:
                    check_reading_stats(ws_url, name)
                except Exception as _e:
                    log(f"  阅读量回查异常: {_e}")
                return True, ""
            else:
                log(f"  X 发布失败: {reason}")
                return False, reason
        except Exception as e:
            log(f"  X 异常: {e}")
            return False, f"异常: {e}"
        finally:
            close_current_tab(main_ws)
            _d = random.randint(8, 20)
            log(f"  篇间等待 {_d} 秒...")
            time.sleep(_d)

    def _is_eligible(name, sub_skipped):
        return (name not in dead_terminated and
                name not in sub_skipped and
                acc_count.get(name, 0) < per_account_quota.get(name, 0))

    def _handle_fail(name, reason, sub_fail_count, sub_skipped):
        if reason in HARD_TERMINATE_REASONS:
            cnt = acc_count.get(name, 0)
            dead_terminated[name] = (reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cnt)
            _append_hard_terminate(name, reason, cnt)
            log(f"  ★ 4 类硬终止: {name} -> {reason}")
            return True
        sub_fail_count[name] = sub_fail_count.get(name, 0) + 1
        if sub_fail_count[name] >= max_fail_per_sub:
            sub_skipped.add(name)
            log(f"  - {name} 本小轮 {max_fail_per_sub} 次失败,跳本小轮")
        return False

    # [v1101 P10] Stage 2 死磕熔断
    stage2_consecutive_fail = 0
    STAGE2_MAX_FAIL = 6
    big_round = 0
    while True:
        if not doc_pool:
            log("\n[Stage2] 文档池已空,死磕结束")
            break
        active = [a for a in accounts
                  if a not in dead_terminated
                  and acc_count.get(a, 0) < per_account_quota.get(a, 0)]
        if not active:
            log("\n[Stage2] 所有账号已满 quota,死磕结束")
            break
        # [v1101.4] 大循环开头重扫池, 实时同步外部 mutate
        _removed, _added = _resync_pool(doc_pool)
        if _removed or _added:
            log(f"  [v1101.4] doc_pool 重扫: 剔除 {_removed} 条幽灵, 新增 {_added} 篇")
        if not doc_pool:
            log("\n[Stage2] 重扫后文档池已空,死磕结束")
            break
        big_round += 1
        log(f"\n{'='*20} [Stage2] 第 {big_round} 大循环 (active {len(active)},文档池 {len(doc_pool)}) {'='*20}")

        for sub_idx in range(1, sub_rounds + 1):
            log(f"\n----- [Stage2] 大{big_round}/小{sub_idx} -----")
            sub_fail_count = {}
            sub_skipped = set()
            for name in list(accounts):
                if not doc_pool:
                    break
                if not _is_eligible(name, sub_skipped):
                    continue
                # [v1101.4] _pick_doc 替代 random.choice: 校验存在 + 失效就地剔除
                doc = _pick_doc(doc_pool)
                if doc is None:
                    log("[Stage2] 文档池被 _pick_doc 清空(全失效), 提前结束")
                    break
                ok, reason = _do_publish(name, doc)
                if not ok:
                    _handle_fail(name, reason, sub_fail_count, sub_skipped)
                    # [v1101 P10] Stage 2 熔断
                    # [v1101.1] 取消 Stage 2 熔断,改用账号级 layered retry,死磕到 quota 满
            log(f"\n[Stage2] 大{big_round}/小{sub_idx} 结束。本小轮跳过 {len(sub_skipped)},硬终止累计 {len(dead_terminated)}")

    return {
        "acc_count": acc_count,
        "dead_terminated": dead_terminated,
        "doc_pool": doc_pool,
        "big_rounds": big_round,
    }


def main():
    _init_run_dir()
    # [v1102] 从 notice_checked.txt 恢复 notice_checked_set
    global notice_checked_set
    notice_checked_set = set()
    if NOTICE_CHECKED_FILE and os.path.exists(NOTICE_CHECKED_FILE):
        try:
            with open(NOTICE_CHECKED_FILE, encoding='utf-8') as _ncf:
                for _line in _ncf:
                    _name = _line.strip().split('|')[0]
                    if _name:
                        notice_checked_set.add(_name)
        except Exception as _e:
            pass
    log("=" * 50)
    log("创作罐头图文文章定时发布 Mac版 启动")
    log(f"报告目录: {RUN_REPORT_DIR}")
    if notice_checked_set:
        log(f"  [v1102] 从 notice_checked.txt 恢复 {len(notice_checked_set)} 个已检查账号(中断恢复)")
    log("=" * 50)

    os.makedirs(DOCS_FOLDER, exist_ok=True)
    os.makedirs(SENT_FOLDER, exist_ok=True)

    # 读发布日期（定时发布.xlsx B1）
    date_str = _read_publish_date()
    if not date_str:
        return
    log(f"发布日期: {date_str}")

    # 获取主窗口
    try:
        main_ws_url = get_main_ws_url()
    except RuntimeError as e:
        log(f"错误: {e}")
        return

    main_ws = ws_connect(main_ws_url, timeout=10)
    log("已连接主窗口")


    # 账号来源：白名单第一优先，为空则从罐头左侧栏动态读
    skip = _read_skip_set()
    if skip:
        log(f"「永久跳过」：{sorted(skip)}")

    # [v1102] 主线内化中断恢复 — quota 动态算 = (已发总+池剩) // 账号数
    wl = _read_whitelist()
    if wl:
        log(f"检测到「白名单」非空 {len(wl)} 个账号 —— 本次只发白名单指定的账号")
        all_accounts = [n for n, _ in wl if n not in skip and n not in EXCLUDE_ACCOUNTS]
        wl_quota_map = {n: q for n, q in wl}
        if len(all_accounts) != len(wl):
            log(f"  有 {len(wl) - len(all_accounts)} 个白名单账号被「永久跳过/排除」过滤")
    else:
        log("白名单为空 —— 从罐头左侧栏动态读取所有账号")
        all_names = collect_accounts(main_ws)
        all_accounts = [n for n in all_names if n not in skip and n not in EXCLUDE_ACCOUNTS]
        wl_quota_map = {}
        log(f"过滤后发文账号: {len(all_accounts)}")

    if not all_accounts:
        log("错误: 没有可发文账号")
        main_ws.close()
        return

    # [v1102] 主线主控 v2:读 last_published.txt 拿最近 publish 账号 → 找 idx → 环形重排让下一位置首
    _last_published_acc = None
    if LAST_PUBLISHED_FILE and os.path.exists(LAST_PUBLISHED_FILE):
        try:
            with open(LAST_PUBLISHED_FILE, encoding='utf-8') as _lpf:
                _lines = [_l.strip() for _l in _lpf if _l.strip()]
                if _lines:
                    _last_published_acc = _lines[-1].split('|')[0].strip()
        except Exception as _e:
            log(f"  [v1102] last_published.txt 读取失败: {_e}")
    if _last_published_acc and all_accounts:
        _last_idx = -1
        for _i, _a in enumerate(all_accounts):
            if _last_published_acc in _a or _a in _last_published_acc:
                _last_idx = _i
                break
        if _last_idx >= 0:
            _next = (_last_idx + 1) % len(all_accounts)
            all_accounts = all_accounts[_next:] + all_accounts[:_next]
            log(f"  [v1102] 中断处自动接续:最近 publish「{_last_published_acc}」(idx={_last_idx}) → 从下一位「{all_accounts[0]}」起跑")

    # 准备文档池(提前到 quota 算之前)
    doc_pool = list(get_docs())
    docs_count = len(doc_pool)

    # [v1102] 读「本轮已发」 → 算 quota_total = (已发总篇数 + 池剩) // 账号数
    sent_count_map = _read_sent_with_count()
    sent_total = sum(sent_count_map.values())
    quota_total = max(1, min(3, (docs_count + sent_total) // len(all_accounts)))  # timer 3 窗上限
    log(f"  [v1102] 素材池剩 {docs_count} 篇 + 已发累计 {sent_total} 篇 = 总 {docs_count + sent_total} 篇 / {len(all_accounts)} 账号 = **每号 quota={quota_total}**(本大循环 {quota_total} 小轮)")

    # 每号缺 = max(0, quota_total - 已发次数);白名单 B 列指定 quota 时取 min
    accounts_quota = []
    skipped_full = 0
    reduced = 0
    for n in all_accounts:
        already = sent_count_map.get(n, 0)
        wl_q = wl_quota_map.get(n)
        target_q = min(wl_q, quota_total) if wl_q else quota_total
        缺 = target_q - already
        if 缺 <= 0:
            skipped_full += 1
            continue
        if already > 0:
            reduced += 1
        accounts_quota.append((n, 缺))

    if sent_count_map:
        # 当前在第几小轮 = 已发总 // 账号数 + 1
        current_round = sent_total // len(all_accounts) + 1
        done_in_round = sent_total % len(all_accounts)
        remaining_in_round = len(all_accounts) - done_in_round
        log(f"  [v1102] 中断恢复:第 {current_round} 小轮第 {done_in_round + 1} 账号断点(已发 {sent_total} 篇,quota 满跳 {skipped_full} 个 / 部分扣 {reduced} 个)")
        log(f"    本小轮({current_round})剩 {remaining_in_round} 账号 + 后续 {quota_total - current_round} 小轮各 {len(all_accounts)} 账号")
        log(f"    共 {sum(q for _, q in accounts_quota)} 篇未发")

    if not accounts_quota:
        log("✓ 所有账号本大循环已齐活,无需排程")
        main_ws.close()
        return

    tasks = _expand_tasks(accounts_quota, date_str)
    log(f"\n共 {len(tasks)} 个定时发布任务")
    log(f"素材池共 {docs_count} 份文稿")
    if docs_count < len(tasks):
        log(f"警告:素材({docs_count})少于任务({len(tasks)}),后面任务将记为素材不足")

    fail_records = []
    success_count = 0
    success_by_acct = {}  # 每账号提交成功篇数，收尾算漏发
    fail_docs_by_acct = {}  # 每账号失败过的文档名列表，给待补漏用

    doc_retry_set = set()  # 已回池过的文档路径（每篇最多回一次，防死循环）
    def requeue_doc(p):
        if not p or p in doc_retry_set:
            return False
        doc_retry_set.add(p)
        doc_pool.append(p)  # 回池尾，让别的好素材先消耗
        log(f"  → 文档已回池尾待后续重试: {os.path.basename(p)}")
        return True

    # 4 类硬终止集合 + 硬终止账号字典(本次累计) + 写"硬终止账号"sheet
    HARD_TERMINATE_REASONS = {"封号", "禁言", "侧边栏未找到", "信用分过低"}  # [v1101.1] 失登移除(改软重试),加信用分过低
    dead_terminated = {}  # name -> (reason, ts, count_so_far)
    _HARD_TERMINATE_HEADERS = ["账号名", "终止原因", "终止时间", "本次已发篇数"]

    def _append_hard_terminate(name, reason, count_so_far):
        try:
            wb = openpyxl.load_workbook(CONFIG_EXCEL)
            if "硬终止账号" not in wb.sheetnames:
                ws_h = wb.create_sheet("硬终止账号")
                ws_h.append(_HARD_TERMINATE_HEADERS)
            else:
                ws_h = wb["硬终止账号"]
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_h.append([name, reason, ts, count_so_far])
            wb.save(CONFIG_EXCEL)
        except Exception:
            pass

    def _terminate(name, reason, doc_path):
        cnt = success_by_acct.get(name, 0)
        dead_terminated[name] = (reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cnt)
        _append_hard_terminate(name, reason, cnt)
        log(f"  ★ 4 类硬终止: {name} -> {reason} (本次已发 {cnt} 篇,永久放弃)")
        # 文档不 requeue,留池子让别的账号用
        if doc_path and doc_path not in doc_pool:
            doc_pool.append(doc_path)

    consecutive_fail = 0
    MAX_CONSECUTIVE_FAIL = 6  # 连败熔断阈值
    COOLDOWN_AT = 3           # 连败到此值时,先 sleep 一段让罐头缓口气
    COOLDOWN_SEC = 1

    for idx, (name, timer_time) in enumerate(tasks):
        log(f"\n{'='*40}")
        log(f"任务 {idx+1}/{len(tasks)}: {name}  →  {timer_time}")

        if name in dead_terminated:
            log(f"  ★ 该账号已硬终止 ({dead_terminated[name][0]}),跳过本任务")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", f"已硬终止/{dead_terminated[name][0]}"))
            continue

        # 从素材池顺序取一篇(非随机)
        if not doc_pool:
            log("  X 素材池已空")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", "素材不足"))
            fail_docs_by_acct.setdefault(name, []).append("")
            continue
        # [v1101.4] _pop_doc 替代 doc_pool.pop(0): 校验存在 + 失效就地剔除
        doc_path = _pop_doc(doc_pool)
        if doc_path is None:
            log("  X 素材池已空(全失效)")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, "", "素材不足"))
            fail_docs_by_acct.setdefault(name, []).append("")
            continue

        log(f"  文档: {os.path.basename(doc_path)}")

        this_task_failed = False
        this_task_fail_reason = ""

        # 找账号
        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未找到账号: {name}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), "侧边栏未找到"))
            fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
            _terminate(name, "侧边栏未找到", doc_path)  # 4 类硬终止,不 requeue
            this_task_failed = True
            this_task_fail_reason = "侧边栏未找到"
        else:
            click(main_ws, pos["x"], pos["y"], 20)
            time.sleep(WAIT_LOAD)

            ws_url = find_or_reopen_webview(main_ws, name)
            if not ws_url:
                log("  X 找不到 webview")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), "webview匹配失败"))
                fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
                requeue_doc(doc_path)  # webview 不在 4 类,正常 requeue
                close_current_tab(main_ws)
                this_task_failed = True
                this_task_fail_reason = "webview匹配失败"
            else:
                try:
                    success, reason = publish_article_timer(ws_url, doc_path, main_ws, name, timer_time)
                    if success:
                        move_to_sent(doc_path)
                        success_count += 1
                        success_by_acct[name] = success_by_acct.get(name, 0) + 1
                        log(f"  ✓ 定时发布成功: {name}")
                        _append_sent_excel(name)  # [v1101.5] 写「本轮已发」sheet,中断恢复用
                    else:
                        log(f"  X 发布失败: {reason}")
                        fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), reason))
                        fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
                        if reason in HARD_TERMINATE_REASONS:
                            _terminate(name, reason, doc_path)
                        else:
                            requeue_doc(doc_path)
                        this_task_failed = True
                        this_task_fail_reason = reason
                except Exception as e:
                    log(f"  X 异常: {e}")
                    fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, timer_time, os.path.basename(doc_path), f"异常: {e}"))
                    fail_docs_by_acct.setdefault(name, []).append(os.path.basename(doc_path))
                    requeue_doc(doc_path)
                    this_task_failed = True
                    this_task_fail_reason = f"异常: {e}"
                finally:
                    close_current_tab(main_ws)

        # 连败熔断 + 系统通知
        if this_task_failed:
            consecutive_fail += 1
            # [v1101.1] 取消 Stage 1 6 连败熔断,改用账号级 layered retry
            if consecutive_fail >= COOLDOWN_AT:
                log(f"  连败 {consecutive_fail} 次，先 sleep {COOLDOWN_SEC}s 等罐头缓过来再继续")
                time.sleep(COOLDOWN_SEC)
        else:
            consecutive_fail = 0

        if idx < len(tasks) - 1:
            _d = random.randint(8, 20)
            log(f"  篇间等待 {_d} 秒...")
            time.sleep(_d)

    write_fail_excel(fail_records)
    log(f"\n{'='*50}")
    log(f"Stage 1 (定时排程) 完成! 成功:{success_count}  失败:{len(fail_records)}  硬终止:{len(dead_terminated)}")
    log(f"{'='*50}")

    # ========== Stage 2:死磕补尾 (自动衔接) ==========
    # 漏发账号(quota - 已发) 通过 publish_article_timer 排程到 now+30min,串行死磕
    remaining_quota = {}
    remaining_accounts = []
    for name, q in accounts_quota:
        if name in dead_terminated:
            continue
        sent = success_by_acct.get(name, 0)
        if sent < q:
            remaining_quota[name] = q - sent
            remaining_accounts.append(name)

    if remaining_accounts and doc_pool:
        log(f"\n{'#'*60}")
        log(f"# Stage 2 启动:死磕补尾,{len(remaining_accounts)} 个账号待补,文档池剩 {len(doc_pool)} 篇")
        log(f"# 漏发账号立即发布（不走定时），死磕直到全部成功 / 4 类硬终止 / Ctrl+C")
        log(f"# 直到全部排程成功 / 命中 4 类硬终止 / 你 Ctrl+C")
        log(f"{'#'*60}")
        try:
            stage2 = run_death_grip_timer(
                accounts=remaining_accounts,
                per_account_quota=remaining_quota,
                doc_pool=doc_pool,
                main_ws=main_ws,
                sub_rounds=3,
                max_fail_per_sub=3,
                initial_dead=dead_terminated,
            )
            stage2_dead = stage2.get("dead_terminated", {})
            new_dead = {n: v for n, v in stage2_dead.items() if n not in dead_terminated}
            dead_terminated.update(stage2_dead)
            stage2_added = sum(stage2.get("acc_count", {}).values())
            for n, v in stage2.get("acc_count", {}).items():
                if v > 0:
                    success_by_acct[n] = success_by_acct.get(n, 0) + v
            log(f"\nStage 2 完成: 新增成功 {stage2_added} 篇,新增硬终止 {len(new_dead)} 个")
        except KeyboardInterrupt:
            log(f"\n!! Stage 2 被人工 Ctrl+C 中断")
    elif not remaining_accounts:
        log("\n★ Stage 1 全部账号已达 quota 或硬终止,无需 Stage 2")
    else:
        log("\n★ Stage 1 结束时文档池已空,Stage 2 无文档可发")

    _finalize_config(accounts_quota, success_by_acct, fail_docs_by_acct)
    main_ws.close()

    if dead_terminated:
        log(f"\n★★ 4 类硬终止账号 {len(dead_terminated)} 个 (已写入'硬终止账号'sheet,需人工处理):")
        for d_name, (d_reason, d_ts, d_cnt) in dead_terminated.items():
            log(f"  - {d_name}\t{d_reason}\t本次已发 {d_cnt} 篇\t{d_ts}")

    log(f"\n{'='*50}")
    log(f"全流程完成! 总成功:{sum(success_by_acct.values())}  Stage1失败:{len(fail_records)}  硬终止:{len(dead_terminated)}")
    log(f"{'='*50}")


if __name__ == "__main__":
    main()
