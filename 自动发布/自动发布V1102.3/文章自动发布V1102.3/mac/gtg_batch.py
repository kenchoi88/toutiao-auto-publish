"""
创作罐头批量发布脚本 - 图文文章版（Mac）
  Mac文章自动发布/
  ├── go.command      双击运行
  ├── gtg_batch.py
  ├── 素材/           放 .docx 文件
  │   └── 已发送/     发完自动移入
  ├── gtg_log.txt
  ├── 失败记录.xlsx
  └── 系统通知提醒.txt
"""

import requests
import json
import websocket
import time
import os
import shutil
import glob
import sys
import random
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== 配置 =====================
def _find_cdp_port():
    port_file = os.path.expanduser("~/Library/Application Support/创作罐头/DevToolsActivePort")
    try:
        with open(port_file) as f:
            return int(f.readline().strip())
    except Exception:
        return 9225

CDP_URL          = f"http://127.0.0.1:{_find_cdp_port()}"
ACCOUNT_CLASS    = "account-RALrbJ"
WAIT_LOAD        = 4
EXCLUDE_ACCOUNTS = ["青春小馆"]
NOFIRST_ACCOUNTS = set()                   # 不选头条首发的账号（从nofirst.txt加载）

NO_PROXY = {"http": "", "https": ""}
WS_OPTS  = {"suppress_origin": True}

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DOCS_FOLDER = os.path.join(BASE_DIR, "素材")
SENT_FOLDER = os.path.join(BASE_DIR, "素材", "已发送")
ALERT_THRESHOLD  = 5000

# 运行报告路径（main()开始时动态初始化）
RUN_REPORT_DIR   = None
LOG_FILE         = None
FAIL_FILE        = None
NOTICE_FILE      = None
NOTICE_CHECKED_FILE = None  # [v1102] 持久化已检查账号集合(中断恢复后仍只读 1 次)
LAST_PUBLISHED_FILE = None  # [v1102.2] 持久化最近 publish 成功账号(中断恢复后从此账号下一位起跑)
ALERT_FILE       = None
VIOLATION_FILE   = None

os.environ["NO_PROXY"] = "127.0.0.1,localhost"
# ================================================


VIOLATION_KEYWORDS = {
    "违规/扣分": ["违规", "扣分", "处罚", "警告"],
    "禁言封号": ["禁言", "发言受限", "封禁", "封号"],
    "原创侵权": ["原创违规", "侵权", "重复内容"],
}

# ========== 账号配置 Excel ==========
CONFIG_EXCEL  = os.path.join(BASE_DIR, "账号配置.xlsx")
SUMMARY_EXCEL = os.path.join(BASE_DIR, "发文汇总.xlsx")
_EXCEL_SHEETS = ["不首发", "永久跳过", "本轮已发", "白名单", "失败列表", "硬终止账号"]
_FAIL_HEADERS    = ["账号名", "失败原因", "文稿名", "失败时间", "轮次"]
_SUMMARY_HEADERS = ["账号名", "轮次", "发文时间", "失败时间", "补发成功时间"]
_HARD_TERMINATE_HEADERS = ["账号名", "终止原因", "终止时间", "本次已发篇数"]

# 4 类硬终止 reason — 命中即永久放弃,不再尝试
HARD_TERMINATE_REASONS = {"封号", "禁言", "侧边栏未找到", "信用分过低"}  # [v1101.1] 失登 移除(改软重试),加 信用分过低

def _ensure_config_excel():
    """如果账号配置.xlsx不存在则自动创建;已存在则确保关键 sheet 有表头"""
    if not os.path.exists(CONFIG_EXCEL):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sh in _EXCEL_SHEETS:
            ws_new = wb.create_sheet(sh)
            if sh == "失败列表":
                ws_new.append(_FAIL_HEADERS)
            elif sh == "硬终止账号":
                ws_new.append(_HARD_TERMINATE_HEADERS)
            elif sh == "白名单":
                ws_new.append(["账号名", "发文份数"])
            else:
                ws_new.append(["账号名"])
        wb.save(CONFIG_EXCEL)
        return
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        dirty = False
        if "失败列表" not in wb.sheetnames:
            ws_n = wb.create_sheet("失败列表")
            ws_n.append(_FAIL_HEADERS); dirty = True
        if "硬终止账号" not in wb.sheetnames:
            ws_n = wb.create_sheet("硬终止账号")
            ws_n.append(_HARD_TERMINATE_HEADERS); dirty = True
        if dirty:
            wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _ensure_summary_excel():
    if os.path.exists(SUMMARY_EXCEL):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发文汇总"
        ws.append(_SUMMARY_HEADERS)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass


def _read_whitelist_with_quota():
    """读白名单sheet：A列账号 + B列配额（空/非数字视为1）。返回 {账号: 配额}"""
    if not os.path.exists(CONFIG_EXCEL):
        return {}
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "白名单" not in wb.sheetnames:
            wb.close()
            return {}
        ws_r = wb["白名单"]
        result = {}
        for row in ws_r.iter_rows(min_row=2, max_col=2, values_only=True):
            name = row[0]
            if not name or not str(name).strip() or str(name).strip().startswith('#'):
                continue
            name = str(name).strip()
            q_raw = row[1] if len(row) > 1 else None
            try:
                q = int(q_raw) if q_raw is not None and str(q_raw).strip() != "" else 1
            except Exception:
                q = 1
            if q < 1:
                q = 1
            result[name] = q
        wb.close()
        return result
    except Exception:
        return {}


def _read_excel_sheet(sheet_name):
    """读取账号配置.xlsx指定sheet的A列账号（跳过第1行标题，忽略空行和#开头）"""
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb[sheet_name]
        result = []
        for row in ws_r.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip() and not str(val).strip().startswith('#'):
                result.append(str(val).strip())
        wb.close()
        return result
    except Exception:
        return []

def _append_sent_excel(name):
    """[v1102] 写「本轮已发」 sheet:行存在 count+1,不存在 append (账号, 1)"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "本轮已发" not in wb.sheetnames:
            ws_s = wb.create_sheet("本轮已发")
            ws_s.append(["账号名", "已发次数"])
            ws_s.append([name, 1])
        else:
            ws_s = wb["本轮已发"]
            found_row = None
            for row_idx, row in enumerate(ws_s.iter_rows(min_row=2, max_col=2, values_only=False), start=2):
                if row[0].value and str(row[0].value).strip() == name:
                    found_row = row_idx
                    break
            if found_row:
                cur = ws_s.cell(row=found_row, column=2).value or 0
                try: cur = int(cur)
                except: cur = 0
                ws_s.cell(row=found_row, column=2).value = cur + 1
            else:
                ws_s.append([name, 1])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass


def _read_sent_with_count():
    """[v1102] 读「本轮已发」 sheet → {账号: 已发次数}"""
    if not os.path.exists(CONFIG_EXCEL): return {}
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "本轮已发" not in wb.sheetnames: wb.close(); return {}
        ws_r = wb["本轮已发"]
        result = {}
        for row in ws_r.iter_rows(min_row=2, max_col=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            cnt = 1
            if len(row) > 1 and row[1] is not None:
                try: cnt = int(row[1])
                except: cnt = 1
            result[name] = cnt
        wb.close(); return result
    except Exception:
        return {}

def _append_fail_list(name, reason, doc_name, round_num):
    """失败列表追加一条（本轮内失败记录，供轮末补发）"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "失败列表" not in wb.sheetnames:
            ws_f = wb.create_sheet("失败列表")
            ws_f.append(_FAIL_HEADERS)
        else:
            ws_f = wb["失败列表"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_f.append([name, reason or "", doc_name or "", ts, round_num])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _read_fail_list():
    """读取失败列表 → [(name, reason, doc_name, ts, round_num), ...]"""
    if not os.path.exists(CONFIG_EXCEL):
        return []
    try:
        wb = openpyxl.load_workbook(CONFIG_EXCEL, read_only=True, data_only=True)
        if "失败列表" not in wb.sheetnames:
            wb.close()
            return []
        ws_r = wb["失败列表"]
        result = []
        for row in ws_r.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if not name or name.startswith('#'): continue
            reason  = str(row[1]).strip() if row[1] is not None else ""
            docname = str(row[2]).strip() if row[2] is not None else ""
            ts      = str(row[3]).strip() if row[3] is not None else ""
            rnd     = int(row[4]) if (row[4] is not None and str(row[4]).strip().isdigit()) else 0
            result.append((name, reason, docname, ts, rnd))
        wb.close()
        return result
    except Exception:
        return []

def _remove_from_fail_list(name):
    """补发成功后，从失败列表里删除该账号的所有记录"""
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "失败列表" not in wb.sheetnames:
            wb.close(); return
        ws_f = wb["失败列表"]
        rows_to_delete = []
        for idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] and str(row[0]).strip() == name:
                rows_to_delete.append(idx)
        for r_idx in reversed(rows_to_delete):
            ws_f.delete_rows(r_idx)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _append_summary(name, round_num, event_type, ts=None):
    """往独立 发文汇总.xlsx 追加/更新。event_type: 发文时间/失败时间/补发成功时间"""
    if ts is None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    col_map = {"发文时间": 3, "失败时间": 4, "补发成功时间": 5}
    col = col_map.get(event_type)
    if col is None:
        return
    try:
        _ensure_summary_excel()
        wb = openpyxl.load_workbook(SUMMARY_EXCEL)
        ws_s = wb["发文汇总"] if "发文汇总" in wb.sheetnames else wb.create_sheet("发文汇总")
        if ws_s.max_row == 0:
            ws_s.append(_SUMMARY_HEADERS)
        target_row = None
        for r_idx in range(2, ws_s.max_row + 1):
            a = ws_s.cell(r_idx, 1).value
            b = ws_s.cell(r_idx, 2).value
            if a and str(a).strip() == name and b is not None and int(b) == round_num:
                target_row = r_idx
                break
        if target_row is None:
            ws_s.append([name, round_num, "", "", ""])
            target_row = ws_s.max_row
        ws_s.cell(target_row, col, ts)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass

def _append_hard_terminate(name, reason, count_so_far):
    """命中 4 类硬终止时,写入"硬终止账号"sheet。永久放弃,不再尝试。"""
    try:
        _ensure_config_excel()
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        if "硬终止账号" not in wb.sheetnames:
            ws_h = wb.create_sheet("硬终止账号")
            ws_h.append(_HARD_TERMINATE_HEADERS)
        else:
            ws_h = wb["硬终止账号"]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_h.append([name, reason, ts, count_so_far])
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass






def _clear_round_sheets():
    """轮末齐活后清空 本轮已发 和 失败列表（保留表头）"""
    try:
        if not os.path.exists(CONFIG_EXCEL):
            return
        wb = openpyxl.load_workbook(CONFIG_EXCEL)
        for sn in ("本轮已发", "失败列表"):
            if sn in wb.sheetnames:
                ws_c = wb[sn]
                if ws_c.max_row > 1:
                    ws_c.delete_rows(2, ws_c.max_row - 1)
        wb.save(CONFIG_EXCEL)
    except Exception:
        pass

def _sort_summary_by_account():
    """按账号分组排序 发文汇总.xlsx"""
    try:
        if not os.path.exists(SUMMARY_EXCEL):
            return
        wb = openpyxl.load_workbook(SUMMARY_EXCEL)
        if "发文汇总" not in wb.sheetnames:
            wb.close(); return
        ws_s = wb["发文汇总"]
        data = []
        for r_idx in range(2, ws_s.max_row + 1):
            row = [ws_s.cell(r_idx, c).value for c in range(1, 6)]
            if row[0]:
                data.append(row)
        data.sort(key=lambda r: (str(r[0]), int(r[1]) if r[1] is not None else 0))
        if ws_s.max_row > 1:
            ws_s.delete_rows(2, ws_s.max_row - 1)
        for row in data:
            ws_s.append(row)
        wb.save(SUMMARY_EXCEL)
    except Exception:
        pass


# =====================================


def _init_run_dir():
    global LOG_FILE, FAIL_FILE, NOTICE_FILE, NOTICE_CHECKED_FILE, LAST_PUBLISHED_FILE, ALERT_FILE, VIOLATION_FILE, RUN_REPORT_DIR
    ts = datetime.now().strftime("%Y%m%d")
    RUN_REPORT_DIR = os.path.join(BASE_DIR, "运行报告", ts)
    os.makedirs(RUN_REPORT_DIR, exist_ok=True)
    LOG_FILE       = os.path.join(RUN_REPORT_DIR, "运行日志.txt")
    FAIL_FILE      = os.path.join(RUN_REPORT_DIR, "失败记录.xlsx")
    NOTICE_FILE    = os.path.join(RUN_REPORT_DIR, "系统通知.txt")
    NOTICE_CHECKED_FILE = os.path.join(RUN_REPORT_DIR, "notice_checked.txt")  # [v1102] 已检查账号持久化
    LAST_PUBLISHED_FILE = os.path.join(RUN_REPORT_DIR, "last_published.txt")  # [v1102.2] 最近 publish 成功账号持久化
    ALERT_FILE     = os.path.join(RUN_REPORT_DIR, "高阅读提醒.txt")
    VIOLATION_FILE = os.path.join(RUN_REPORT_DIR, "违规提醒.txt")


def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    sys.stdout.buffer.write((line + "\n").encode("utf-8", errors="replace"))
    sys.stdout.buffer.flush()
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def write_status(*args, **kwargs):
    pass


def write_fail_excel(final_fails):
    """final_fails: [(ts, account_name, reason), ...] 只写最终失败账号"""
    if not final_fails:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "失败记录"
    for col, h in enumerate(["时间", "账号名", "失败原因"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    for ts_f, n_f, r_f in final_fails:
        ws.append([ts_f, n_f, r_f])
    try:
        wb.save(FAIL_FILE)
        log(f"最终失败记录已写入: {FAIL_FILE}")
    except Exception as e:
        log(f"  写入失败记录出错: {e}")


def get_tabs():
    return requests.get(f"{CDP_URL}/json", timeout=5, proxies=NO_PROXY).json()


def get_main_ws_url():
    tabs = get_tabs()
    for t in tabs:
        if "czgts.cn" in t.get("url", "") and "webSocketDebuggerUrl" in t:
            return t["webSocketDebuggerUrl"]
    raise RuntimeError("找不到主窗口，请确认创作罐头已启动")


def ws_connect(url, timeout=10):
    return websocket.create_connection(url, timeout=timeout, **WS_OPTS)


def cdp(ws, method, params=None, mid=1):
    ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
    deadline = time.time() + 15
    while time.time() < deadline:
        try:
            d = json.loads(ws.recv())
            if d.get("id") == mid:
                return d.get("result", {})
        except:
            pass
    return {}


def js(ws, expr, mid=99):
    r = cdp(ws, "Runtime.evaluate", {"expression": expr}, mid)
    return r.get("result", {}).get("value")


def click(ws, x, y, mid):
    p = {"button": "left", "clickCount": 1, "x": x, "y": y,
         "modifiers": 0, "timestamp": time.time() * 1000}
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mousePressed", **p}, mid)
    time.sleep(0.12)
    cdp(ws, "Input.dispatchMouseEvent", {"type": "mouseReleased", **p}, mid + 1)


def collect_accounts(main_ws):
    log("开始收集账号列表...")
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 110)
    time.sleep(0.8)

    accounts = []
    seen = set()
    same_count = 0
    has_scrolled = False

    for _ in range(1000):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            var names = [];
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t) names.push(t);
            }}
            return JSON.stringify(names);
        }})()
        """, 111)
        if v:
            for n in json.loads(v):
                if n not in seen:
                    seen.add(n)
                    accounts.append(n)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no';
            var before = c.scrollTop;
            c.scrollTop += 200;
            return before + '->' + c.scrollTop;
        })()
        """, 112)
        time.sleep(0.3)

        if result and result != "no":
            parts = result.split("->")
            if len(parts) == 2:
                before_top, after_top = parts[0].strip(), parts[1].strip()
                if after_top != '0' and after_top != before_top:
                    has_scrolled = True
                    same_count = 0
                elif before_top == after_top and has_scrolled:
                    same_count += 1
                    if same_count >= 4:
                        break

    # 兜底:虚拟滚动列表,最后几个账号常因 lazy render 漏读
    # 强制滚到容器最底部,等 1.5s DOM 稳定再扫一次
    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = c.scrollHeight;
    })()
    """, 113)
    time.sleep(1.5)
    v = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        var names = [];
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t) names.push(t);
        }}
        return JSON.stringify(names);
    }})()
    """, 114)
    if v:
        names = json.loads(v)
        added = 0
        for n in names:
            if n not in seen:
                seen.add(n)
                accounts.append(n)
                added += 1
        if added:
            log(f"  兜底:滚到底再扫一次,补收 {added} 个账号(虚拟滚动 lazy render 漏发)")

    log(f"共收集到 {len(accounts)} 个账号")
    return accounts


def scroll_find_account(main_ws, name):
    name_json = json.dumps(name)
    same_count = 0

    js(main_ws, """
    (function(){
        var c = document.querySelector('[class*="menuMainWarpper"]');
        if(c) c.scrollTop = 0;
    })()
    """, 10)
    time.sleep(0.5)

    for _ in range(500):
        v = js(main_ws, f"""
        (function(){{
            var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
            for(var i=0;i<items.length;i++){{
                var t = items[i].textContent.trim();
                if(t === {name_json} || t.startsWith({name_json})){{
                    items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                    return 'found';
                }}
            }}
            return null;
        }})()
        """, 11)

        if v == 'found':
            time.sleep(0.3)
            pos = js(main_ws, f"""
            (function(){{
                var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
                for(var i=0;i<items.length;i++){{
                    var t = items[i].textContent.trim();
                    if(t === {name_json} || t.startsWith({name_json})){{
                        var r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                        // 坐标不在视口内，再滚一次
                        items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                        r = items[i].getBoundingClientRect();
                        if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                            return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
                    }}
                }}
                return null;
            }})()
            """, 13)
            if pos:
                return json.loads(pos)

        result = js(main_ws, """
        (function(){
            var c = document.querySelector('[class*="menuMainWarpper"]');
            if(!c) return 'no-container';
            var before = c.scrollTop;
            c.scrollTop += 250;
            return before + '->' + c.scrollTop;
        })()
        """, 12)
        time.sleep(0.25)

        if result and result != "no-container":
            parts = result.split("->")
            if len(parts) == 2 and parts[0] == parts[1]:
                same_count += 1
                if same_count >= 5:
                    break
            else:
                same_count = 0

    # 滚动找不到，尝试搜索框输入账号名过滤
    log(f"  滚动未找到账号，尝试搜索框输入: {name}")
    search_pos = js(main_ws, """
    (function(){
        var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(!s) return null;
        var r = s.getBoundingClientRect();
        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
        return null;
    })()
    """, 14)
    if not search_pos:
        return None
    # v1101.2: CDP nativeInputValueSetter 注入,绕过键盘+罐头前台依赖(实战 air 295→300)
    js(main_ws, f"""
    (function(){{
        var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
        if(!s) return 'no_input';
        var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
        setter.call(s, {name_json});
        s.dispatchEvent(new Event('input',  {{bubbles:true}}));
        s.dispatchEvent(new Event('change', {{bubbles:true}}));
        return 'ok';
    }})()
    """, 15)
    time.sleep(1.5)
    # 再取坐标
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0 && r.top >= 0 && r.top <= window.innerHeight)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 18)
    if pos:
        log(f"  搜索框过滤后找到账号: {name}")
        # v1101.2: CDP 清空搜索框,恢复完整列表
        js(main_ws, """
        (function(){
            var s = document.querySelector('input[placeholder*="账号"],input[placeholder*="手机"]');
            if(!s) return;
            var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(s, '');
            s.dispatchEvent(new Event('input',  {bubbles:true}));
            s.dispatchEvent(new Event('change', {bubbles:true}));
        })()
        """, 19)
        time.sleep(0.5)
        return json.loads(pos)
    return None


def _find_webview_once(main_ws, name):
    partition = js(main_ws, """
    (function(){
        var webviews = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<webviews.length;i++){
            var p = webviews[i].getAttribute('partition');
            if(!p || !p.startsWith('persist:')) continue;
            var r = webviews[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = p; }
        }
        return best;
    })()
    """, 15)

    if not partition:
        return None

    marker = f"_mk{random.randint(100000, 999999)}"
    r = cdp(main_ws, "Runtime.evaluate", {
        "expression": f"""
        new Promise(function(resolve){{
            var wv = document.querySelector('webview[partition="{partition}"]');
            if(!wv){{ resolve('no_wv'); return; }}
            wv.executeJavaScript('window.{marker}=1').then(function(){{
                resolve('ok');
            }}).catch(function(){{
                resolve('err');
            }});
        }})
        """,
        "awaitPromise": True,
        "timeout": 8000
    }, 200)

    if r.get("result", {}).get("value") != "ok":
        return None

    tabs = get_tabs()
    for t in tabs:
        if "webSocketDebuggerUrl" not in t:
            continue
        try:
            wsc = ws_connect(t["webSocketDebuggerUrl"], timeout=3)
            wsc.send(json.dumps({"id": 1, "method": "Runtime.evaluate",
                                 "params": {"expression": f"window.{marker}===1"}}))
            deadline = time.time() + 3
            found = False
            while time.time() < deadline:
                d = json.loads(wsc.recv())
                if d.get("id") == 1:
                    found = d.get("result", {}).get("result", {}).get("value") is True
                    break
            wsc.close()
            if found:
                log("  webview 匹配成功")
                return t["webSocketDebuggerUrl"]
        except:
            pass

    return None


def find_account_webview(main_ws, name):
    for retry in range(3):
        result = _find_webview_once(main_ws, name)
        if result:
            return result
        if retry < 2:
            log(f"  重试 webview ({retry+2}/3)...")
            time.sleep(2)
    return None


def _search_box_set(main_ws, value):
    """侧边栏搜索框设置值(空字符串=清空)。React-friendly: 用原型 setter 触发 input/change 事件。"""
    val_json = json.dumps(value)
    return js(main_ws, f"""
    (function(){{
        var inputs = document.querySelectorAll('input');
        for(var i=0;i<inputs.length;i++){{
            var ph = inputs[i].getAttribute('placeholder') || '';
            if(ph.indexOf('账号') !== -1 || ph.indexOf('手机号') !== -1){{
                var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                setter.call(inputs[i], '');
                inputs[i].dispatchEvent(new Event('input', {{bubbles:true}}));
                if({val_json}.length > 0){{
                    setter.call(inputs[i], {val_json});
                    inputs[i].dispatchEvent(new Event('input', {{bubbles:true}}));
                    inputs[i].dispatchEvent(new Event('change', {{bubbles:true}}));
                }}
                return 'ok';
            }}
        }}
        return 'no_input';
    }})()
    """, 50)


def _locate_filtered_account(main_ws, name):
    """搜索框过滤后,直接抓第一个匹配账号的中心坐标(viewport 相对)。"""
    name_json = json.dumps(name)
    pos = js(main_ws, f"""
    (function(){{
        var items = document.querySelectorAll('.{ACCOUNT_CLASS}');
        for(var i=0;i<items.length;i++){{
            var t = items[i].textContent.trim();
            if(t === {name_json} || t.startsWith({name_json})){{
                items[i].scrollIntoView({{block:'center', behavior:'instant'}});
                var r = items[i].getBoundingClientRect();
                if(r.width > 0)
                    return JSON.stringify({{x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)}});
            }}
        }}
        return null;
    }})()
    """, 51)
    if not pos:
        return None
    return json.loads(pos)


def find_or_reopen_webview(main_ws, name, reopen_attempts=2):
    """虚拟滚动底部账号 webview partition 不渲染时,通过侧边栏搜索框定位账号让 webview 重建。

    流程:正常 find_account_webview 失败 → 关 tab → 搜索框输入账号名 → 等过滤 →
    抓过滤后的账号坐标 → click → 等更长时间 → 再尝试 find_account_webview。
    搜索框过滤后只剩匹配项渲染在 DOM 顶部,不再受虚拟滚动影响。
    """
    ws_url = find_account_webview(main_ws, name)
    if ws_url:
        return ws_url

    for attempt in range(reopen_attempts):
        log(f"  webview partition 失败,搜索框重建 {attempt+1}/{reopen_attempts}: 输入 \"{name}\"")
        try:
            close_current_tab(main_ws)
        except Exception as e:
            log(f"  关 tab 异常(忽略): {e}")
        time.sleep(1.0)

        # 用搜索框过滤
        rs = _search_box_set(main_ws, name)
        if rs != 'ok':
            log("  搜索框定位失败,降级用 scroll_find_account")
            pos = scroll_find_account(main_ws, name)
        else:
            time.sleep(1.5)  # 等过滤结果 React 渲染完
            pos = _locate_filtered_account(main_ws, name)
            if not pos:
                log("  搜索过滤后仍找不到,降级用 scroll_find_account")
                _search_box_set(main_ws, "")  # 清空恢复列表
                time.sleep(0.5)
                pos = scroll_find_account(main_ws, name)

        if not pos:
            log(f"  搜索/滚动都找不到 {name},重建中止")
            _search_box_set(main_ws, "")
            return None

        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD + 2)  # 比首次多等 2 秒,给虚拟滚动 lazy render 留余地

        ws_url = find_account_webview(main_ws, name)

        # 不论成败,清空搜索框还原列表(避免影响后续账号)
        _search_box_set(main_ws, "")
        time.sleep(0.3)

        if ws_url:
            log(f"  webview 重建成功(尝试 {attempt+1})")
            return ws_url

    return None


def get_url_from_ws(ws_url):
    try:
        wsc = ws_connect(ws_url, timeout=4)
        v = js(wsc, "location.href", 1)
        wsc.close()
        return v or ""
    except:
        return ""


def check_system_notice(ws_url, account_name):
    """
    [v1102] 导航到消息中心 → 点击 系统通知 + 审核通知 频道
    → 读取 2 天内(今天+昨天)的完整消息原文写入 NOTICE_FILE
    新 selector: .conversation-box.notify-im-user-item (替代旧 span.name)
    新提取: body.innerText 按日期行切分(MM-DD HH:MM / YYYY-MM-DD / 昨日/今日 HH:MM)
    """
    try:
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        today_short     = today.strftime("%m-%d")
        yesterday_short = yesterday.strftime("%m-%d")
        today_full      = today.strftime("%Y-%m-%d")
        yesterday_full  = yesterday.strftime("%Y-%m-%d")

        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/personal/message?type=message_letter'", 300)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        time.sleep(2.5)
        notices = []

        for channel in ["系统通知", "审核通知"]:
            channel_json = channel.replace('"', '\\"')
            clicked = js(wsc, f"""
            (function(){{
                var items = document.querySelectorAll('.conversation-box.notify-im-user-item');
                for(var i=0; i<items.length; i++){{
                    var t = (items[i].innerText || '').trim();
                    if(t.indexOf("{channel_json}") === 0){{
                        items[i].click();
                        return 'ok';
                    }}
                }}
                return null;
            }})()
            """, 301)

            if not clicked:
                log(f"  未找到频道: {channel}")
                continue

            time.sleep(2.5)

            result = js(wsc, f"""
            (function() {{
                var todayShort = "{today_short}";
                var yesterdayShort = "{yesterday_short}";
                var todayFull = "{today_full}";
                var yesterdayFull = "{yesterday_full}";
                var lines = (document.body.innerText || '').split(/\\r?\\n/);
                var results = [];
                var current = '';
                var currentDate = '';
                var inWindow = false;
                function dateInfo(line) {{
                    var m = line.match(/^(\\d{{2}}-\\d{{2}})\\s+\\d{{2}}:\\d{{2}}$/);
                    if (m) return m[1] === todayShort || m[1] === yesterdayShort;
                    m = line.match(/^(\\d{{4}}-\\d{{2}}-\\d{{2}})\\s+\\d{{2}}:\\d{{2}}$/);
                    if (m) return m[1] === todayFull || m[1] === yesterdayFull;
                    if (/^昨日\\s+\\d{{2}}:\\d{{2}}$/.test(line)) return true;
                    if (/^今日\\s+\\d{{2}}:\\d{{2}}$/.test(line)) return true;
                    return null;
                }}
                function isDateLine(line) {{
                    return /^(\\d{{2}}-\\d{{2}}|\\d{{4}}-\\d{{2}}-\\d{{2}}|昨日|今日)\\s+\\d{{2}}:\\d{{2}}$/.test(line);
                }}
                for (var i = 0; i < lines.length; i++) {{
                    var line = lines[i].trim();
                    if (!line) continue;
                    if (isDateLine(line)) {{
                        if (inWindow && current.trim()) {{
                            results.push(currentDate + '\\n' + current.trim());
                        }}
                        currentDate = line;
                        current = '';
                        inWindow = (dateInfo(line) === true);
                    }} else if (inWindow) {{
                        current += line + '\\n';
                    }}
                }}
                if (inWindow && current.trim()) {{
                    results.push(currentDate + '\\n' + current.trim());
                }}
                // 去重(预览和详情可能重复)
                var seen = {{}};
                var dedup = [];
                for (var k = 0; k < results.length; k++) {{
                    var key = results[k].substring(0, 80);
                    if (!seen[key]) {{ seen[key] = true; dedup.push(results[k]); }}
                }}
                return JSON.stringify(dedup);
            }})()
            """, 302)

            if result:
                try:
                    for msg in json.loads(result):
                        notices.append(f"【{channel}】\n{msg}")
                except:
                    pass

        wsc.close()

        violation_count = 0
        if notices:
            ts_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            # 分析违规类通知
            violations = []
            for n in notices:
                for cat, kws in VIOLATION_KEYWORDS.items():
                    for kw in kws:
                        if kw in n:
                            violations.append((cat, n))
                            break
            # 写系统通知 — 完整内容,不截断
            content = f"\n[{ts_str}] 账号 {account_name} 2 天内通知 ({len(notices)} 条):\n"
            for n in notices:
                content += f"\n--- 通知 ---\n{n}\n"
            content += "\n" + "=" * 60 + "\n"
            with open(NOTICE_FILE, "a", encoding="utf-8") as f:
                f.write(content)
            log(f"  ⚠ 2 天内通知 {len(notices)} 条 → 系统通知.txt")
            # 写违规提醒
            if violations:
                vcontent = f"[{ts_str}] 账号 {account_name} 违规/扣分提醒:\n"
                for cat, msg in violations:
                    vcontent += f"  [{cat}] {msg[:300]}...\n"
                vcontent += "\n"
                with open(VIOLATION_FILE, "a", encoding="utf-8") as f:
                    f.write(vcontent)
                violation_count = len(violations)
                log(f"  ⚠ 违规/扣分 {violation_count} 条 → 违规提醒.txt")
        else:
            log("  系统/审核通知: 2 天内无新通知")
        return len(notices), violation_count
    except Exception as e:
        log(f"  系统通知检测出错: {e}")
        return 0, 0


def check_reading_stats(ws_url, account_name):
    """导航到主页 → 读取最近文章阅读量，超过 ALERT_THRESHOLD 则写入 ALERT_FILE。"""
    try:
        wsc = ws_connect(ws_url, timeout=8)
        js(wsc, "location.href='https://mp.toutiao.com/profile_v4/index'", 310)
        wsc.close()
        time.sleep(3)

        wsc = ws_connect(ws_url, timeout=8)
        reads_raw = js(wsc, """
        (function(){
            var items = document.querySelectorAll('.recent-works-item');
            var results = [];
            for(var i=0;i<items.length;i++){
                var timeEl = items[i].querySelector('span.time');
                var timeText = timeEl ? timeEl.textContent.trim() : '未知时间';
                var labels = items[i].querySelectorAll('span.label');
                for(var j=0;j<labels.length;j++){
                    if(labels[j].textContent.trim() === '阅读量'){
                        var prev = labels[j].previousElementSibling;
                        if(prev){
                            var num = parseInt(prev.textContent.trim().replace(/,/g,''), 10);
                            if(!isNaN(num)) results.push({count: num, time: timeText});
                        }
                    }
                }
            }
            return JSON.stringify(results);
        })()
        """, 311)
        wsc.close()

        reads = json.loads(reads_raw) if reads_raw else []
        if not reads:
            log("  阅读量: 未读取到数据")
            return

        counts = [r['count'] for r in reads]
        log(f"  阅读量: {counts}")
        high = [r for r in reads if r['count'] >= ALERT_THRESHOLD]
        if high:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M")
            lines = [f"[{ts}] 账号 {account_name} 高阅读量:"]
            for r in high:
                lines.append(f"  {r['time']} — {r['count']} 阅读量")
            msg = '\n'.join(lines) + '\n\n'
            with open(ALERT_FILE, "a", encoding="utf-8") as f:
                f.write(msg)
            log(f"  ★ 高阅读量提醒: {[r['count'] for r in high]} → 已写入 {ALERT_FILE}")
            return len(high)
        return 0
    except Exception as e:
        log(f"  阅读量检测出错: {e}")
        return 0


def detect_account_error(wsc):
    page_text = js(wsc, "document.body.innerText || ''", 70) or ""
    for reason, keywords in {
        "失登": ["请登录", "登录已失效", "账号已下线", "重新登录"],
        "封号": ["账号已被封禁", "账号异常", "账号被封"],
        "禁言": ["账号被禁言", "发言受限", "无法发布"],
    }.items():
        for kw in keywords:
            if kw in page_text:
                return reason
    return None


def close_popup(ws):
    v = js(ws, """
    (function(){
        var b = document.querySelector('.close-btn,[class*="close-btn"]');
        if(b){var r=b.getBoundingClientRect();
        if(r.width>0) return JSON.stringify({x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)});}
        return null;
    })()
    """, 50)
    if v:
        p = json.loads(v)
        click(ws, p["x"], p["y"], 51)
        time.sleep(0.5)


def close_current_tab(main_ws):
    v = js(main_ws, """
    (function(){
        var closes = document.querySelectorAll('.chrome-tab-close');
        if(closes.length === 0) return null;
        for(var i=0;i<closes.length;i++){
            var tab = closes[i].closest('.chrome-tab');
            if(tab && tab.classList.contains('active')){
                var r = closes[i].getBoundingClientRect();
                if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
        }
        var last = closes[closes.length-1];
        var r = last.getBoundingClientRect();
        return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
    })()
    """, 55)
    if v:
        p = json.loads(v)
        click(main_ws, p["x"], p["y"], 56)
        time.sleep(0.5)
        log("  已关闭tab")
    else:
        log("  未找到tab关闭按钮")


# ========== 文件选择对话框（Mac） ==========

def _fill_mac_file_dialog(doc_path, result_holder):
    """等 Mac 文件选择对话框出现，用 osascript 填入路径并确认"""
    deadline = time.time() + 20
    while time.time() < deadline:
        # 检测文件选择对话框是否在前台
        check = subprocess.run(
            ["osascript", "-e",
             'tell application "System Events" to get name of first process whose frontmost is true'],
            capture_output=True, text=True
        )
        front = check.stdout.strip()
        if front in ("electron", "创作罐头", "Electron"):
            # 尝试发 Cmd+Shift+G 打开路径输入框（Mac 文件对话框快捷键）
            subprocess.run(["osascript", "-e", """
tell application "System Events"
    keystroke "g" using {command down, shift down}
end tell
"""], capture_output=True)
            time.sleep(0.5)
            # 输入路径并回车确认
            subprocess.run(["osascript", "-e", f"""
tell application "System Events"
    keystroke "{doc_path}"
    key code 36
    delay 0.5
    key code 36
end tell
"""], capture_output=True)
            result_holder[0] = True
            return
        time.sleep(0.5)
    result_holder[0] = False


# ========== 发布流程（图文文章） ==========

def publish_article(ws_url, doc_path, main_ws, account_name="", _credit_out=None):
    try:
        wsc = ws_connect(ws_url, timeout=10)
    except Exception as e:
        return False, f"连接失败: {e}"

    close_popup(wsc)
    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    # 导航到图文发布页
    js(wsc, "location.href='https://mp.toutiao.com/profile_v4/graphic/publish'", 60)
    wsc.close()
    time.sleep(4)

    # 导航后 target id 会变，从 tabs 里重新找 graphic/publish 页面
    new_ws_url = None
    for _ in range(12):
        tabs = get_tabs()
        for t in tabs:
            if "graphic/publish" in t.get("url","") and "webSocketDebuggerUrl" in t:
                new_ws_url = t["webSocketDebuggerUrl"]
                break
        if new_ws_url:
            break
        time.sleep(0.5)

    if not new_ws_url:
        return False, "导航到发文页失败"

    try:
        wsc = ws_connect(new_ws_url, timeout=10)
    except Exception as e:
        return False, f"重连失败: {e}"

    close_popup(wsc)
    time.sleep(1)

    current_url = js(wsc, "location.href", 59) or ""
    if "login" in current_url:
        wsc.close()
        return False, "失登"

    err = detect_account_error(wsc)
    if err:
        wsc.close()
        return False, err

    import subprocess, threading

    # 取 webview 屏幕坐标（草稿条关闭用）
    wv_early = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 57)
    wv_early = json.loads(wv_early) if wv_early else None

    # 关闭草稿提示条（cliclick 点 × 按钮）
    draft_close = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t === '\u7ee7\u7eed\u7f16\u8f91'){
                var bar = els[i].closest('[class]');
                while(bar){
                    var x = bar.querySelector('[class*="close"],[class*="Close"]');
                    if(x){
                        var r = x.getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                    }
                    bar = bar.parentElement && bar.parentElement.closest('[class]');
                    if(!bar) break;
                }
            }
        }
        return null;
    })()
    """, 58)

    if draft_close and wv_early:
        dc = json.loads(draft_close)
        dc_x = wv_early['sx'] + dc['x']
        dc_y = wv_early['sy'] + dc['y']
        log(f"  关闭草稿条: cliclick ({dc_x},{dc_y})")
        subprocess.run(["cliclick", f"c:{dc_x},{dc_y}"], capture_output=True)
        time.sleep(0.5)
    else:
        log("  无草稿提示条")

    # 取 webview 真实屏幕坐标
    wv_s = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 61)

    if not wv_s:
        wsc.close()
        return False, "无法获取 webview 屏幕坐标"

    wv0 = json.loads(wv_s)

    # 点工具栏最后一个按钮（文档导入图标，轮询等待页面加载，最多等10秒）
    v = None
    for _ in range(20):
        v = js(wsc, """
        (function(){
            var btns = Array.from(document.querySelectorAll('.syl-toolbar-button')).filter(function(b){
                return b.getBoundingClientRect().width > 0;
            });
            if(btns.length > 0){
                var last = btns[btns.length - 1];
                var r = last.getBoundingClientRect();
                return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
            }
            return null;
        })()
        """, 62)
        if v:
            break
        time.sleep(0.5)

    if not v:
        wsc.close()
        return False, "找不到文档导入按钮"

    p = json.loads(v)

    # [v1101 P5] 激活强化:unhide + AXRaise + verify frontmost + 重试 3 次
    subprocess.run(["osascript", "-e", """
tell application "创作罐头" to activate
delay 0.2
tell application "System Events"
    tell process "创作罐头"
        try
            if visible is false then set visible to true
        end try
        set frontmost to true
        try
            perform action "AXRaise" of window 1
        end try
        repeat 3 times
            if frontmost then exit repeat
            set frontmost to true
            delay 0.3
        end repeat
    end tell
end tell
"""], capture_output=True)
    time.sleep(0.6)

    wv_s2 = js(main_ws, """
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """, 63)
    if wv_s2:
        wv0 = json.loads(wv_s2)

    import_x = wv0['sx'] + p['x']
    import_y = wv0['sy'] + p['y']
    log(f"  文档导入按钮CSS坐标:({p['x']},{p['y']}) webview原点:({wv0['sx']},{wv0['sy']}) => 屏幕:({import_x},{import_y})")

    # 先点标题栏空白处让 webview 获得焦点（标题栏在顶部，不会点到草稿内容）
    title_x = wv0['sx'] + 400
    title_y = wv0['sy'] + 50
    subprocess.run(["cliclick", f"c:{title_x},{title_y}"], capture_output=True)
    time.sleep(0.5)
    # [v1101 P7] cliclick 文档导入 + 等弹窗,失败重试 3 次
    sel = None
    for click_attempt in range(3):
        attempt_str = f" [第{click_attempt+1}次]" if click_attempt > 0 else ""
        log(f"  cliclick 点击文档导入 ({import_x},{import_y}){attempt_str}")
        subprocess.run(["cliclick", f"c:{import_x},{import_y}"], capture_output=True)
        time.sleep(1.5)

        for i in range(20):
            sel = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim() === '选择文档'){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({bx: Math.round(r.left+r.width/2), by: Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 65)
            if sel:
                break
            if i == 4 and click_attempt == 0:
                all_btns = js(wsc, """
                (function(){
                    var btns = document.querySelectorAll('button');
                    var names = [];
                    for(var i=0;i<btns.length;i++){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) names.push(btns[i].textContent.trim());
                    }
                    return JSON.stringify(names);
                })()
                """, 67)
                log(f"  当前按钮: {all_btns}")
            time.sleep(0.5)

        if sel:
            if click_attempt > 0:
                log(f"  cliclick 第{click_attempt+1}次成功唤出弹窗")
            break

        if click_attempt < 2:
            log(f"  弹窗未出,重 activate + 重读坐标后重试")
            subprocess.run(["osascript", "-e", """
tell application "创作罐头" to activate
delay 0.2
tell application "System Events"
    tell process "创作罐头"
        try
            if visible is false then set visible to true
        end try
        set frontmost to true
        try
            perform action "AXRaise" of window 1
        end try
    end tell
end tell
"""], capture_output=True)
            time.sleep(0.6)
            wv_s3 = js(main_ws, """
            (function(){
                var wvs = document.querySelectorAll('webview');
                var maxArea = 0, best = null;
                for(var i=0;i<wvs.length;i++){
                    var r = wvs[i].getBoundingClientRect();
                    var area = r.width * r.height;
                    if(area > maxArea){ maxArea = area; best = r; }
                }
                if(!best) return null;
                return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
            })()
            """, 62)
            if wv_s3:
                wv0c = json.loads(wv_s3)
                import_x = wv0c['sx'] + p['x']
                import_y = wv0c['sy'] + p['y']
                log(f"  重试前坐标更新: ({import_x},{import_y})")

    if not sel:
        wsc.close()
        return False, "文档导入弹窗未出现(3 次 cliclick 重试均失败)"

    sb = json.loads(sel)
    screen_x = wv0['sx'] + sb['bx']
    screen_y = wv0['sy'] + sb['by']
    log(f"  选择文档屏幕坐标: ({screen_x},{screen_y})")

    doc_escaped = doc_path.replace("\\", "/")
    safe_path = doc_escaped.replace("\\", "\\\\").replace('"', '\\"')
    result_holder = [None]

    def sheet_exists():
        r = subprocess.run(["osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of window 1)'],
            capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def go_to_folder_sheet_exists():
        r = subprocess.run(["osascript", "-e",
            'tell application "System Events" to tell process "创作罐头" to return (exists sheet 1 of sheet 1 of window 1)'],
            capture_output=True, text=True)
        return 'true' in r.stdout.lower()

    def get_sheet_rect():
        """拿主对话框 position+size。entire contents 会超时，但 position/size 直接读很快。"""
        r = subprocess.run(["osascript", "-e", '''
tell application "System Events"
    tell process "创作罐头"
        try
            set {px, py} to position of sheet 1 of window 1
            set {sw, sh} to size of sheet 1 of window 1
            return (px as string) & "|" & (py as string) & "|" & (sw as string) & "|" & (sh as string)
        on error
            return ""
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=5)
        try:
            nums = re.findall(r"\d+", r.stdout)
            if len(nums) >= 4:
                return tuple(int(v) for v in nums[:4])
        except Exception:
            pass
        return None

    def click_dialog_button(which):
        """which: 'open' 或 'cancel'。cliclick 物理点击，不经 keystroke，不受 frontmost 焦点影响。"""
        rect = get_sheet_rect()
        if not rect:
            return False
        x, y, w, h = rect
        if which == 'open':
            cx, cy = x + w - 55, y + h - 35
        else:
            cx, cy = x + w - 150, y + h - 35
        subprocess.run(["cliclick", f"m:{cx},{cy}"], capture_output=True)
        time.sleep(0.1)
        subprocess.run(["cliclick", f"c:{cx},{cy}"], capture_output=True)
        time.sleep(0.5)
        return True

    def fill_dialog():
        # 等对话框出现
        deadline = time.time() + 15
        appeared = False
        while time.time() < deadline:
            if sheet_exists():
                appeared = True
                break
            time.sleep(0.3)
        if not appeared:
            result_holder[0] = False
            return
        time.sleep(0.6)

        # Step 1: Cmd+Shift+G 开"前往文件夹" + 直接赋值 text field（绕过 clipboard）
        r = subprocess.run(["osascript", "-e", f'''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {{command down, shift down}}
        delay 1.5
        try
            set target to text field 1 of sheet 1 of sheet 1 of window 1
            set value of target to "{safe_path}"
            delay 0.3
            set rb to (value of target) as string
            if rb is "{safe_path}" then
                return "OK"
            else
                return "ERR:readback:" & rb
            end if
        on error errmsg
            return "ERR:" & errmsg
        end try
    end tell
end tell
'''], capture_output=True, text=True, timeout=20)
        direct_ok = "OK" in (r.stdout or "")

        if not direct_ok:
            log(f"  直接赋值失败（{(r.stdout or '').strip()[:80]}），回退clipboard")
            subprocess.run(["osascript", "-e",
                'tell application "System Events" to tell process "创作罐头" to key code 53'],
                capture_output=True)
            time.sleep(0.5)
            clipboard_ok = False
            for _ in range(5):
                subprocess.run(["pbcopy"], input=doc_escaped.encode("utf-8"))
                time.sleep(0.15)
                rb = subprocess.run(["pbpaste"], capture_output=True).stdout.decode("utf-8", errors="replace")
                if rb.strip() == doc_escaped.strip():
                    clipboard_ok = True
                    break
                time.sleep(0.2)
            if not clipboard_ok:
                log("  clipboard 校验 5 次失败")
                click_dialog_button('cancel')
                result_holder[0] = False
                return
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.3
tell application "System Events"
    tell process "创作罐头"
        keystroke "g" using {command down, shift down}
        delay 1.5
        keystroke "a" using {command down}
        delay 0.4
        keystroke "v" using {command down}
        delay 1.2
    end tell
end tell
'''], capture_output=True)

        # Step 2: 回车关"前往文件夹"小框，最多 5 次
        go_closed = False
        for i in range(5):
            subprocess.run(["osascript", "-e", '''
tell application "创作罐头" to activate
delay 0.1
tell application "System Events"
    tell process "创作罐头"
        keystroke return
    end tell
end tell
'''], capture_output=True)
            time.sleep(0.8)
            if not go_to_folder_sheet_exists():
                go_closed = True
                if i > 0:
                    log(f"  前往文件夹 回车{i+1}次后关闭")
                break
        if not go_closed:
            log("  前往文件夹 5次回车未关 → cliclick 点取消")
            click_dialog_button('cancel')
            result_holder[0] = False
            return

        # [v1101 P1] Step 3 跳过:macOS 26 NSOpenPanel 不再自动关,Step 4 必打,6s 硬等纯发呆

        # Step 4: 主框没关 → cliclick 物理点"打开"按钮
        log("  主对话框未自动关闭 → cliclick 点打开按钮")
        for _ in range(3):
            if not click_dialog_button('open'):
                break
            for _ in range(4):
                time.sleep(0.5)
                if not sheet_exists():
                    result_holder[0] = True
                    return

        # Step 5: 彻底卡死 → 点取消
        log("  对话框完全卡死 → cliclick 点取消")
        click_dialog_button('cancel')
        result_holder[0] = False

    # 外层最多 3 次重试，扛住偶发对话框 hang（原本只试 1 次，现加上）
    dialog_ok = False
    for dialog_attempt in range(3):
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=45)
        if result_holder[0]:
            dialog_ok = True
            if dialog_attempt > 0:
                log(f"  文件对话框第{dialog_attempt+1}次成功")
            break
        # 保险：失败前再点一次取消确保对话框真关了
        click_dialog_button('cancel')
        time.sleep(1)
        log(f"  第{dialog_attempt+1}次对话框处理失败，准备重试")

    if not dialog_ok:
        wsc.close()
        return False, "文件选择对话框反复卡住，3次重试均失败"

    # 等待内容加载
    time.sleep(5)
    char_count = 0
    for _ in range(15):
        v = js(wsc, """
        (function(){
            // [v1101 P3] 取最长 ProseMirror(避免命中标题 placeholder 5 字)
            var els = document.querySelectorAll('.ProseMirror');
            var max = 0;
            for (var i = 0; i < els.length; i++) {
                var l = els[i].textContent.trim().length;
                if (l > max) max = l;
            }
            return max;
        })()
        """, 75)
        char_count = int(v) if v else 0
        if char_count >= 50:
            break
        time.sleep(0.8)

    log(f"  文章字数: {char_count}")
    # [v1101 P2] 字数<50 重试 fill_dialog 一次
    if char_count < 50:
        log(f"  对话框已关但字数仅 {char_count}（文档未真导入），重试 fill_dialog")
        result_holder[0] = None
        t = threading.Thread(target=fill_dialog, daemon=True)
        t.start()
        time.sleep(0.2)
        subprocess.run(["cliclick", f"c:{screen_x},{screen_y}"], capture_output=True)
        t.join(timeout=30)
        time.sleep(5)
        char_count = 0
        for _ in range(15):
            v = js(wsc, """
            (function(){
                var els = document.querySelectorAll('.ProseMirror');
                var max = 0;
                for (var i = 0; i < els.length; i++) {
                    var l = els[i].textContent.trim().length;
                    if (l > max) max = l;
                }
                return max;
            })()
            """, 75)
            char_count = int(v) if v else 0
            if char_count >= 50:
                break
            time.sleep(0.8)
        log(f"  重试后字数: {char_count}")
        if char_count < 50:
            wsc.close()
            return False, "文档导入后内容为空(重试1次仍空)"

    # 滚动到页面最底部，等待图片全部加载
    js(wsc, """
    (function(){
        var el = document.querySelector('.ProseMirror') || document.body;
        el.scrollTop = el.scrollHeight;
        window.scrollTo(0, document.body.scrollHeight);
    })()
    """, 76)
    time.sleep(2)
    # 等图片全部加载完
    for _ in range(20):
        loading = js(wsc, """
        (function(){
            var imgs = document.querySelectorAll('img');
            for(var i=0;i<imgs.length;i++){
                if(!imgs[i].complete || imgs[i].naturalWidth === 0) return true;
            }
            return false;
        })()
        """, 77)
        if not loading:
            break
        time.sleep(0.5)
    log("  图片加载完成")

    # 读取信用分，决定是否勾选头条首发
    credit_raw = js(wsc, """
    (function(){
        var els = document.querySelectorAll('*');
        for(var i=0;i<els.length;i++){
            var t = els[i].textContent.trim();
            if(t.indexOf('\u4fe1\u7528\u5206') !== -1 && t.length < 20){
                var m = t.match(/(\d{1,3})\u5206/);
                if(m){
                    var n = parseInt(m[1], 10);
                    if(n % 5 === 0 && n >= 5 && n <= 100) return n;
                }
            }
        }
        return null;
    })()
    """, 78)
    credit_score = int(credit_raw) if credit_raw is not None else None
    log(f"  信用分: {credit_score if credit_score is not None else '未读取到'}")
    if _credit_out is not None:
        _credit_out.append(credit_score)
    # [v1101.1] 信用分 < 60 → 硬终止,跟禁言/封号同等放弃
    if credit_score is not None and credit_score < 60:
        log(f"  ★ 信用分 {credit_score} < 60,硬终止")
        wsc.close()
        return False, "信用分过低"
    should_first = (account_name not in NOFIRST_ACCOUNTS) and (credit_score is not None and credit_score >= 95)

    # [v1101.3] 头条首发复选框: 探测 + cliclick 点击 + 回读校验 + 三轮兜底 + 硬保护
    _PROBE_JS = r"""
    (function(){
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '头条首发'){
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){
                    if(p.tagName === 'LABEL' && p.classList.contains('byte-checkbox')){
                        var isChecked = p.classList.contains('byte-checkbox-checked');
                        var r = p.getBoundingClientRect();
                        var px = Math.round(r.left + r.width/2);
                        var py = Math.round(r.top + r.height/2);
                        return JSON.stringify({found:true, checked:isChecked, cb_x:px, cb_y:py});
                    }
                    p = p.parentElement;
                }
            }
        }
        return null;
    })()
    """
    _CLICK_JS = r"""
    (function(){
        var all = document.querySelectorAll('*');
        for(var i=0;i<all.length;i++){
            if(all[i].childElementCount === 0 && all[i].textContent.trim() === '头条首发'){
                var p = all[i].parentElement;
                while(p && p.tagName !== 'BODY'){
                    if(p.tagName === 'LABEL' && p.classList.contains('byte-checkbox')){
                        try{ p.click(); }catch(e){}
                        var inp = p.querySelector('input[type="checkbox"]');
                        if(inp){
                            try{ inp.dispatchEvent(new Event('change',{bubbles:true})); }catch(e){}
                        }
                        return 'label';
                    }
                    p = p.parentElement;
                }
            }
        }
        return null;
    })()
    """
    _WV_JS = r"""
    (function(){
        var wvs = document.querySelectorAll('webview');
        var maxArea = 0, best = null;
        for(var i=0;i<wvs.length;i++){
            var r = wvs[i].getBoundingClientRect();
            var area = r.width * r.height;
            if(area > maxArea){ maxArea = area; best = r; }
        }
        if(!best) return null;
        return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
    })()
    """

    def _probe_first():
        raw = js(wsc, _PROBE_JS, 79)
        if not raw:
            return None
        try:
            return json.loads(raw)
        except Exception:
            return None

    def _wv_origin():
        raw = js(main_ws, _WV_JS, 80)
        if not raw:
            return None
        try:
            return json.loads(raw)
        except Exception:
            return None

    fr = _probe_first()
    if fr is None:
        log("  头条首发: 未找到复选框")
    elif bool(fr.get('checked')) == should_first:
        log(f"  头条首发: 已是{'勾选' if fr['checked'] else '未勾选'}，无需操作")
    else:
        attempts = []
        verified = False
        for attempt in range(1, 4):
            if attempt < 3 and fr.get('found'):
                wv = _wv_origin()
                if wv:
                    sx = wv['sx'] + fr['cb_x']
                    sy = wv['sy'] + fr['cb_y']
                    subprocess.run(["cliclick", f"c:{sx},{sy}"], capture_output=True)
                    attempts.append(f"cliclick#{attempt}@({sx},{sy})")
                    time.sleep(0.4)
                else:
                    attempts.append(f"cliclick#{attempt}=no_wv")
                    time.sleep(0.2)
            else:
                rc = js(wsc, _CLICK_JS, 30)
                time.sleep(0.4)
                attempts.append(f"js#{attempt}={rc}")
            fr2 = _probe_first()
            if fr2 and bool(fr2.get('checked')) == should_first:
                log(f"  头条首发: {'勾选' if should_first else '取消勾选'} 已校验 [{'/'.join(attempts)}]")
                verified = True
                break
            fr = fr2 if fr2 else fr
        if not verified:
            actual = fr.get('checked') if fr else None
            log(f"  ✗ 头条首发校准失败: 目标={should_first} 实际={actual} 尝试=[{'/'.join(attempts)}]")
            if (not should_first) and actual is True:
                log(f"  ★ 硬保护: 应取消首发但仍勾选,跳过该篇避免扣 5 分")
                wsc.close()
                return False, "首发取消失败(硬保护)"

    # 点"预览并发布" + 等"确认发布"，最多重试5次（封面图加载慢时第一次点可能无效）
    confirm_clicked = False
    for attempt in range(5):
        # 每次重新取 webview 屏幕坐标
        wv_cur_r = js(main_ws, """
        (function(){
            var wvs = document.querySelectorAll('webview');
            var maxArea = 0, best = null;
            for(var i=0;i<wvs.length;i++){
                var r = wvs[i].getBoundingClientRect();
                var area = r.width * r.height;
                if(area > maxArea){ maxArea = area; best = r; }
            }
            if(!best) return null;
            return JSON.stringify({sx: Math.round(window.screenX + best.left), sy: Math.round(window.screenY + best.top)});
        })()
        """, 79)
        wv_cur = json.loads(wv_cur_r) if wv_cur_r else wv0

        # 找"预览并发布"按钮
        v = js(wsc, """
        (function(){
            var btns = document.querySelectorAll('button');
            for(var i=0;i<btns.length;i++){
                if(btns[i].textContent.trim() === '\u9884\u89c8\u5e76\u53d1\u5e03' && !btns[i].disabled){
                    var r = btns[i].getBoundingClientRect();
                    if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                }
            }
            return null;
        })()
        """, 80)

        if not v:
            if attempt == 0:
                wsc.close()
                return False, "找不到预览并发布按钮"
        else:
            p = json.loads(v)
            preview_x = wv_cur['sx'] + p['x']
            preview_y = wv_cur['sy'] + p['y']
            attempt_str = f" [第{attempt+1}次]" if attempt > 0 else ""
            log(f"  cliclick 点击预览并发布 ({preview_x},{preview_y}){attempt_str}")
            subprocess.run(["cliclick", f"m:{preview_x},{preview_y}"], capture_output=True)
            time.sleep(0.5)
            subprocess.run(["cliclick", f"c:{preview_x},{preview_y}"], capture_output=True)
            log("  已点击预览并发布")
            time.sleep(4)

        # 等"确认发布"按钮
        for i in range(60):
            time.sleep(0.5)
            if i == 10 and not confirm_clicked:
                log(f"  [V1102.3] 5s 未见确认发布,补点预览并发布 ({preview_x},{preview_y})")
                subprocess.run(["cliclick", f"c:{preview_x},{preview_y}"], capture_output=True)
            v2 = js(wsc, """
            (function(){
                var btns = document.querySelectorAll('button');
                for(var i=0;i<btns.length;i++){
                    var t = btns[i].textContent.trim();
                    if(t === '\u786e\u8ba4\u53d1\u5e03' || t.indexOf('\u786e\u8ba4\u53d1\u5e03') !== -1){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width > 0) return JSON.stringify({x:Math.round(r.left+r.width/2), y:Math.round(r.top+r.height/2)});
                    }
                }
                return null;
            })()
            """, 83)
            if v2:
                p2 = json.loads(v2)
                confirm_x = wv_cur['sx'] + p2['x']
                confirm_y = wv_cur['sy'] + p2['y']
                log(f"  cliclick 点击确认发布 ({confirm_x},{confirm_y})")
                subprocess.run(["cliclick", f"c:{confirm_x},{confirm_y}"], capture_output=True)
                confirm_clicked = True
                break

        if confirm_clicked:
            break
        if attempt < 4:
            # [v1101 P4] 抓 DOM 写诊断
            try:
                diag = js(wsc, """
                (function(){
                    var out = [];
                    var btns = document.querySelectorAll('button');
                    out.push('btn-tags='+btns.length);
                    for(var i=0;i<btns.length;i++){
                        var r = btns[i].getBoundingClientRect();
                        if(r.width>0 && r.height>0){
                            out.push('btn['+i+']="'+btns[i].textContent.trim().slice(0,30)+'" disabled='+btns[i].disabled);
                        }
                    }
                    out.push('---含发布字非button可见叶子---');
                    var all = document.querySelectorAll('*');
                    for(var i=0;i<all.length;i++){
                        if(all[i].tagName==='BUTTON') continue;
                        if(all[i].childElementCount!==0) continue;
                        var t = all[i].textContent.trim();
                        if(t.length<25 && t.indexOf('发布')!==-1){
                            var r = all[i].getBoundingClientRect();
                            if(r.width>0) out.push(all[i].tagName+'="'+t+'"');
                        }
                    }
                    return out.join(' | ');
                })()
                """, 84)
                log(f"  [DIAG] {diag}")
            except Exception as e:
                log(f"  [DIAG] dump失败: {e}")
            log(f"  确认发布未出现，封面图可能未加载完，准备第{attempt+2}次点击预览并发布...")

    if not confirm_clicked:
        wsc.close()
        return False, "未出现确认发布按钮"

    # 检测发布成功
    published = False
    for _ in range(20):
        time.sleep(0.5)
        t = js(wsc, """
        (function(){
            var all = document.querySelectorAll('*');
            for(var i=0;i<all.length;i++){
                var t = all[i].textContent.trim();
                if((t==='发布成功！'||t==='发布成功'||t==='提交成功！'||t==='提交成功')
                   && all[i].children.length<=1)
                    return t;
            }
            return null;
        })()
        """, 96)
        if t:
            published = True
            break
        cur_url = js(wsc, "location.href", 97) or ""
        if "graphic" in cur_url and "publish" not in cur_url:
            published = True
            break

    if not published:
        err_after = detect_account_error(wsc)
        wsc.close()
        if err_after:
            return False, err_after
        return False, "未检测到发布成功"

    wsc.close()
    log("  OK 文章发布成功")
    return True, "成功"


# ========== 文档管理 ==========

def get_docs():
    docs = []
    for p in ["*.docx", "*.doc"]:
        docs.extend(glob.glob(os.path.join(DOCS_FOLDER, p)))
    return sorted([d for d in docs if "已发送" not in d])


# [v1101.4] doc_pool 实时校验 + 重扫工具,救"分发完源必删"导致罐头找不到文件
def _pick_doc(doc_pool):
    """从 doc_pool 抽一篇实存的 docx,失效引用就地清理。返回 None 表示池已空。"""
    while doc_pool:
        doc = random.choice(doc_pool)
        if os.path.exists(doc):
            return doc
        log(f"  ! 源已删除(可能被外部分发),从池剔除: {os.path.basename(doc)}")
        doc_pool.remove(doc)
    return None


def _resync_pool(doc_pool):
    """大循环开始前重扫素材池:剔除幽灵引用 + 加入新到的素材。返回 (剔除数, 新增数)。"""
    cur = set(get_docs())
    before = len(doc_pool)
    doc_pool[:] = [d for d in doc_pool if d in cur]
    removed = before - len(doc_pool)
    pool_set = set(doc_pool)
    new_docs = [d for d in cur if d not in pool_set]
    if new_docs:
        doc_pool.extend(new_docs)
    return removed, len(new_docs)


def move_to_sent(doc_path):
    os.makedirs(SENT_FOLDER, exist_ok=True)
    dest = os.path.join(SENT_FOLDER, os.path.basename(doc_path))
    if os.path.exists(dest):
        name, ext = os.path.splitext(os.path.basename(doc_path))
        dest = os.path.join(SENT_FOLDER, f"{name}_{int(time.time())}{ext}")
    shutil.move(doc_path, dest)
    log(f"  已移至已发送: {os.path.basename(dest)}")


# ========== 死磕主循环 (公共函数) ==========

def run_death_grip(
    accounts,
    per_account_quota,
    doc_pool,
    main_ws,
    sub_rounds=3,
    max_fail_per_sub=3,
    sent_accounts_set=None,
    credit_records=None,
    fail_records=None,
    success_accounts=None,
    initial_acc_count=None,
    initial_dead=None,
    log_label="",
):
    """死磕循环:大循环 N 小轮 + 外层无限磕。
    - 4 类硬终止 → 即刻永久放弃
    - 其他失败 → 本小轮内累计 max_fail_per_sub 次跳过本小轮,下小轮恢复
    - 大循环跑完 + 还有 quota → 继续下一大循环(无限磕)
    """
    sent_accounts_set = sent_accounts_set if sent_accounts_set is not None else set()
    credit_records   = credit_records if credit_records is not None else {}
    fail_records     = fail_records if fail_records is not None else []
    success_accounts = success_accounts if success_accounts is not None else set()
    acc_count        = dict(initial_acc_count) if initial_acc_count else {}
    for a in accounts:
        acc_count.setdefault(a, 0)
    dead_terminated  = dict(initial_dead) if initial_dead else {}

    ok_count = fail_count = 0
    total_notices = total_violations = total_alerts = 0
    big_round = 0
    # [v1102] 每账号每天只读 1 次,持久化到 NOTICE_CHECKED_FILE,中断恢复后不重读
    notice_checked_set = set()
    if NOTICE_CHECKED_FILE and os.path.exists(NOTICE_CHECKED_FILE):
        try:
            with open(NOTICE_CHECKED_FILE, encoding='utf-8') as _ncf:
                for _line in _ncf:
                    _name = _line.strip().split('|')[0]
                    if _name:
                        notice_checked_set.add(_name)
            if notice_checked_set:
                log(f"  [v1102] 从 notice_checked.txt 恢复 {len(notice_checked_set)} 个已检查账号(中断恢复)")
        except Exception as _e:
            log(f"  [v1102] notice_checked.txt 读取失败: {_e}")

    def _do_publish(name, doc, round_label, is_retry=False):
        nonlocal ok_count, fail_count, total_notices, total_violations, total_alerts
        log(f"\n  {log_label}[剩余 {len(doc_pool)} 篇] {round_label} {'[补发]' if is_retry else ''} {name}  ->  {os.path.basename(doc)}")

        pos = scroll_find_account(main_ws, name)
        if not pos:
            log(f"  X 未在侧边栏找到账号: {name}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, "侧边栏未找到"))
            fail_count += 1
            return False, "侧边栏未找到"

        log(f"  点击坐标({pos['x']},{pos['y']})")
        click(main_ws, pos["x"], pos["y"], 20)
        time.sleep(WAIT_LOAD)

        ws_url = find_or_reopen_webview(main_ws, name)
        if not ws_url:
            log("  X 找不到 webview")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, "webview匹配失败"))
            fail_count += 1
            close_current_tab(main_ws)
            return False, "webview匹配失败"

        page_url = get_url_from_ws(ws_url)
        if "login" in page_url:
            log("  X 账号失登")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, "失登"))
            fail_count += 1
            close_current_tab(main_ws)
            return False, "失登"

        # [v1102] 每账号每天只读 1 次审核/系统通知 — 持久化,中断恢复后仍跳过
        if name not in notice_checked_set:
            nc, vc = check_system_notice(ws_url, name)
            total_notices += nc
            total_violations += vc
            notice_checked_set.add(name)
            # 同步写持久化文件,防中断后重读
            try:
                with open(NOTICE_CHECKED_FILE, "a", encoding="utf-8") as _ncf:
                    _ncf.write(f"{name}|{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            except Exception as _e:
                log(f"  [v1102] 写 notice_checked.txt 失败: {_e}")
            time.sleep(2)
        else:
            log(f"  系统/审核通知:{name} 当天已读过,跳过")

        _d_wait = random.randint(8, 20)
        try:
            credit_buf = []
            success, reason = publish_article(ws_url, doc, main_ws, name, credit_buf)
            if credit_buf and credit_buf[0] is not None:
                credit_records[name] = credit_buf[0]
            if success:
                move_to_sent(doc)
                if doc in doc_pool:
                    doc_pool.remove(doc)
                acc_count[name] = acc_count.get(name, 0) + 1
                success_accounts.add(name)
                ok_count += 1
                # [v1102.2] 持久化最近 publish 成功账号
                try:
                    with open(LAST_PUBLISHED_FILE, "a", encoding="utf-8") as _lpf:
                        _lpf.write(f"{name}|{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                except Exception as _e:
                    log(f"  [v1102.2] 写 last_published.txt 失败: {_e}")
                total_alerts += check_reading_stats(ws_url, name)
                return True, ""
            else:
                log(f"  X 发布失败: {reason}")
                fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, reason))
                fail_count += 1
                return False, reason
        except Exception as e:
            log(f"  X 异常: {e}")
            fail_records.append((datetime.now().strftime("%Y-%m-%d %H:%M"), name, f"异常: {e}"))
            fail_count += 1
            return False, f"异常: {e}"
        finally:
            close_current_tab(main_ws)
            log(f"  篇间等待 {_d_wait} 秒...")
            time.sleep(_d_wait)

    def _is_eligible(name, sub_skipped):
        return (name not in dead_terminated and
                name not in sub_skipped and
                name not in sent_accounts_set and
                acc_count.get(name, 0) < per_account_quota.get(name, 0))

    def _handle_failure(name, reason, sub_fail_count, sub_skipped):
        if reason in HARD_TERMINATE_REASONS:
            cnt_so_far = acc_count.get(name, 0)
            dead_terminated[name] = (reason, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), cnt_so_far)
            _append_hard_terminate(name, reason, cnt_so_far)
            log(f"  ★ 4 类硬终止: {name} -> {reason} (本次已发 {cnt_so_far} 篇,永久放弃)")
            return True
        sub_fail_count[name] = sub_fail_count.get(name, 0) + 1
        if sub_fail_count[name] >= max_fail_per_sub:
            sub_skipped.add(name)
            log(f"  - {name} 本小轮 {max_fail_per_sub} 次失败,跳过本小轮(下小轮恢复)")
        return False

    def _total_target():
        return sum(per_account_quota.get(a, 0) for a in accounts if a not in dead_terminated)

    def _total_done():
        return sum(acc_count.get(a, 0) for a in accounts)

    while True:
        if not doc_pool:
            log(f"\n{log_label}文档池已空,死磕结束")
            break
        active = [a for a in accounts
                  if a not in dead_terminated
                  and acc_count.get(a, 0) < per_account_quota.get(a, 0)]
        if not active:
            log(f"\n{log_label}所有账号都已满 quota,死磕结束")
            break

        # [v1101.4] 大循环开头重扫池, 实时同步外部 mutate (scp+rm 等)
        _removed, _added = _resync_pool(doc_pool)
        if _removed or _added:
            log(f"  [v1101.4] doc_pool 重扫: 剔除 {_removed} 条幽灵引用, 新增 {_added} 篇素材")
        if not doc_pool:
            log(f"\n{log_label}重扫后文档池已空,死磕结束")
            break

        big_round += 1
        log(f"\n{'='*20} {log_label}第 {big_round} 大循环 开始 (active {len(active)} 账号,文档池 {len(doc_pool)} 篇) {'='*20}")

        for sub_idx in range(1, sub_rounds + 1):
            log(f"\n----- {log_label}第 {big_round} 大循环 / 第 {sub_idx} 小轮 (Phase A) -----")
            sub_fail_count = {}
            sub_skipped = set()
            sub_round_id = big_round * 100 + sub_idx

            for name in list(accounts):
                if not doc_pool:
                    log(f"  Phase A 中文档池已空,提前结束")
                    break
                if not _is_eligible(name, sub_skipped):
                    continue
                # [v1101.4] _pick_doc 替代 random.choice: 校验存在 + 失效就地剔除
                doc = _pick_doc(doc_pool)
                if doc is None:
                    log(f"  Phase A 中文档池被 _pick_doc 清空(全失效),提前结束")
                    break
                round_label = f"[大{big_round}/小{sub_idx}/A]"
                write_status(big_round, sub_idx, "Phase A",
                             total_done=_total_done(), total_target=_total_target(),
                             doc_pool_size=len(doc_pool),
                             sub_failed=len(sub_skipped), dead_terminated=dead_terminated)
                success, reason = _do_publish(name, doc, round_label, is_retry=False)
                if success:
                    sent_accounts_set.add(name)
                    _append_sent_excel(name)
                    _append_summary(name, sub_round_id, "发文时间")
                else:
                    _append_fail_list(name, reason, os.path.basename(doc), sub_round_id)
                    _append_summary(name, sub_round_id, "失败时间")
                    _handle_failure(name, reason, sub_fail_count, sub_skipped)

            phase_b_round = 0
            while phase_b_round < max_fail_per_sub and doc_pool:
                phase_b_round += 1
                pending = [f for f in _read_fail_list()
                           if f[4] == sub_round_id
                           and f[0] not in sent_accounts_set
                           and f[0] not in dead_terminated
                           and f[0] not in sub_skipped]
                if not pending:
                    log(f"  Phase B: 失败列表已清空,本小轮齐活")
                    break
                log(f"\n  {log_label}Phase B 第 {phase_b_round}/{max_fail_per_sub} 次, 待补 {len(pending)} 个")
                for f_name, f_reason, f_docname, f_ts, f_rnd in pending:
                    if not _is_eligible(f_name, sub_skipped):
                        continue
                    # [v1101.4] 先找原失败 doc, 失效就剔除; 找不到 fallback 用 _pick_doc
                    doc_path = None
                    for d in list(doc_pool):
                        if os.path.basename(d) == f_docname:
                            if os.path.exists(d):
                                doc_path = d
                            else:
                                log(f"  ! 原失败 doc 已删除: {f_docname}, 改抽随机")
                                doc_pool.remove(d)
                            break
                    if doc_path is None:
                        doc_path = _pick_doc(doc_pool)
                        if doc_path is None:
                            log(f"  Phase B 文档池清空, 提前结束")
                            break
                    round_label = f"[大{big_round}/小{sub_idx}/B-{phase_b_round}]"
                    write_status(big_round, sub_idx, f"Phase B-{phase_b_round}",
                                 total_done=_total_done(), total_target=_total_target(),
                                 doc_pool_size=len(doc_pool),
                                 sub_failed=len(sub_skipped), dead_terminated=dead_terminated)
                    success, reason = _do_publish(f_name, doc_path, round_label, is_retry=True)
                    if success:
                        sent_accounts_set.add(f_name)
                        _append_sent_excel(f_name)
                        _remove_from_fail_list(f_name)
                        _append_summary(f_name, sub_round_id, "补发成功时间")
                        log(f"  ✓ 补发成功: {f_name}")
                    else:
                        _handle_failure(f_name, reason, sub_fail_count, sub_skipped)

            log(f"\n{log_label}第 {big_round} 大循环 / 第 {sub_idx} 小轮 结束。本小轮跳过 {len(sub_skipped)} 个(下小轮恢复)。硬终止累计 {len(dead_terminated)} 个")
            sent_accounts_set.clear()
            # [v1102] sheet 不再小轮末 clear,累积到大循环末才 clear

        log(f"\n{'='*20} {log_label}第 {big_round} 大循环 结束 {'='*20}")
        # [v1102] 全员齐活才 clear 「本轮已发」 sheet
        active_left = [a for a in accounts if a not in dead_terminated and acc_count.get(a, 0) < per_account_quota.get(a, 0)]
        if not active_left:
            _clear_round_sheets()
            log(f"  [v1102] 大循环全员齐活 → 「本轮已发」 sheet 已清空")

    write_status(big_round, sub_rounds, "结束",
                 total_done=_total_done(), total_target=_total_target(),
                 doc_pool_size=len(doc_pool),
                 sub_failed=0, dead_terminated=dead_terminated,
                 extra=f"{log_label}{big_round} 大循环跑完。最终硬终止 {len(dead_terminated)} 个")

    return {
        "acc_count": acc_count,
        "dead_terminated": dead_terminated,
        "doc_pool": doc_pool,
        "ok_count": ok_count,
        "fail_count": fail_count,
        "total_notices": total_notices,
        "total_violations": total_violations,
        "total_alerts": total_alerts,
        "big_rounds": big_round,
    }


# ========== 主流程 ==========

def main():
    _init_run_dir()
    log("=" * 50)
    log("创作罐头图文文章批量发布 Mac版 启动")
    log(f"报告目录: {RUN_REPORT_DIR}")
    log("=" * 50)

    os.makedirs(DOCS_FOLDER, exist_ok=True)
    os.makedirs(SENT_FOLDER, exist_ok=True)

    fail_records = []         # [(ts, name, reason)] 记录所有失败尝试
    success_accounts = set()  # 记录最终成功发布的账号
    credit_records = {}       # {账号名: 信用分}
    total_notices = 0
    total_violations = 0
    total_alerts = 0

    # 读取账号配置.xlsx - 不首发sheet
    _ensure_config_excel()
    try:
        _nofirst_list = _read_excel_sheet("不首发")
        NOFIRST_ACCOUNTS.update(_nofirst_list)
        if _nofirst_list:
            log(f"账号配置.xlsx[不首发]已加载，不首发账号: {len(NOFIRST_ACCOUNTS)} 个")
    except Exception as _ne:
        log(f"读取账号配置.xlsx[不首发]失败: {_ne}")

    # 读取账号配置.xlsx - 永久跳过sheet
    try:
        _skip_list = _read_excel_sheet("永久跳过")
        for _sv in _skip_list:
            if _sv not in EXCLUDE_ACCOUNTS:
                EXCLUDE_ACCOUNTS.append(_sv)
        if _skip_list:
            log(f"账号配置.xlsx[永久跳过]已加载，永久跳过: {EXCLUDE_ACCOUNTS}")
    except Exception as _se:
        log(f"读取账号配置.xlsx[永久跳过]失败: {_se}")

    # [v1102] 读「本轮已发」sheet → {账号: 已发次数} + 注入 sent_accounts_set
    sent_accounts_set = set()
    sent_count_map = {}
    try:
        sent_count_map = _read_sent_with_count()
        for _sv in sent_count_map:
            sent_accounts_set.add(_sv)
        if sent_count_map:
            log(f"账号配置.xlsx[本轮已发]已加载,已发累计 {sum(sent_count_map.values())} 篇 / {len(sent_count_map)} 个账号")
    except Exception as _se:
        log(f"读取账号配置.xlsx[本轮已发]失败: {_se}")
    sent_total = sum(sent_count_map.values())

    # 读取账号配置.xlsx - 失败列表sheet（中断恢复时继续补发）
    try:
        _fail_pre = _read_fail_list()
        if _fail_pre:
            log(f"账号配置.xlsx[失败列表]已加载 {len(_fail_pre)} 条（本轮末将补发）")
    except Exception as _fe:
        log(f"读取账号配置.xlsx[失败列表]失败: {_fe}")

    docs = get_docs()
    if not docs:
        log("错误: 素材文件夹中没有 docx 文件")
        return

    log(f"待发文档: {len(docs)} 份")

    try:
        main_ws_url = get_main_ws_url()
    except RuntimeError as e:
        log(f"错误: {e}")
        return

    main_ws = ws_connect(main_ws_url, timeout=10)
    log("已连接主窗口")

    accounts = collect_accounts(main_ws)
    if EXCLUDE_ACCOUNTS:
        before = len(accounts)
        accounts = [a for a in accounts if not any(ex in a or a in ex for ex in EXCLUDE_ACCOUNTS)]
        skipped = before - len(accounts)
        if skipped:
            log(f"已排除 {skipped} 个永不发文账号: {EXCLUDE_ACCOUNTS}")

    # 读取账号配置.xlsx - 白名单 sheet (B 列可指定账号独立配额)
    quota_map = {}
    try:
        _wl_map = _read_whitelist_with_quota()
        if _wl_map:
            # [v1102.1] 按白名单 dict 顺序重排 accounts(catchup 写白名单按断点环形排)
            _orig_accounts = list(accounts)
            _seen = set()
            _new_accounts = []
            for inc in _wl_map.keys():
                for a in _orig_accounts:
                    if a not in _seen and (inc in a or a in inc):
                        _new_accounts.append(a)
                        _seen.add(a)
                        break
            accounts = _new_accounts
            # [v1102.2] 主线主控 v2:读 last_published.txt 拿最近 publish 账号 → 找 idx → 环形重排让下一位置首
            _last_published_acc = None
            if LAST_PUBLISHED_FILE and os.path.exists(LAST_PUBLISHED_FILE):
                try:
                    with open(LAST_PUBLISHED_FILE, encoding='utf-8') as _lpf:
                        _lines = [_l.strip() for _l in _lpf if _l.strip()]
                        if _lines:
                            _last_published_acc = _lines[-1].split('|')[0].strip()
                except Exception as _e:
                    log(f"  [v1102.2] last_published.txt 读取失败: {_e}")
            if _last_published_acc and accounts:
                _last_idx = -1
                for _i, _a in enumerate(accounts):
                    if _last_published_acc in _a or _a in _last_published_acc:
                        _last_idx = _i
                        break
                if _last_idx >= 0:
                    _next = (_last_idx + 1) % len(accounts)
                    accounts = accounts[_next:] + accounts[:_next]
                    log(f"  [v1102.2] 中断处自动接续:最近 publish「{_last_published_acc}」(idx={_last_idx}) → 从下一位「{accounts[0]}」起跑")
            for a in accounts:
                for inc, q in _wl_map.items():
                    if inc in a or a in inc:
                        quota_map[a] = q
                        break
            log(f"账号配置.xlsx[白名单]已加载,白名单 {len(_wl_map)} 个,过滤+重排后剩 {len(accounts)} 个账号(首位={accounts[0] if accounts else '空'})")
            log(f"白名单配额: {quota_map}")
    except Exception as _e:
        log(f"读取账号配置.xlsx[白名单]失败: {_e}")

    if not accounts:
        log("错误: 未找到任何账号")
        main_ws.close()
        return

    # 没有白名单独立配额 → 全局 = 总篇数 ÷ 账号数
    if not quota_map:
        quota = (len(docs) + sent_total) // len(accounts) if len(accounts) > 0 else 1  # [v1102] 加已发累计
        quota = max(quota, 1)
        quota_map = {a: quota for a in accounts}
        log(f"本次发布: {len(accounts)} 个账号,{len(docs)} 篇文档,每账号配额 {quota} 篇")
    else:
        total_quota = sum(quota_map.get(a, 1) for a in accounts)
        log(f"本次发布: {len(accounts)} 个账号,{len(docs)} 篇文档,按独立配额共 {total_quota} 篇")

    doc_pool = list(docs)

    # 调用死磕主循环 (文章 3 小轮 / 大循环 + 外层无限磕)
    result = run_death_grip(
        accounts=accounts,
        per_account_quota=quota_map,
        doc_pool=doc_pool,
        main_ws=main_ws,
        sub_rounds=3,
        max_fail_per_sub=3,
        sent_accounts_set=sent_accounts_set,
        credit_records=credit_records,
        fail_records=fail_records,
        success_accounts=success_accounts,
        initial_acc_count=sent_count_map,  # [v1102] 传入已发次数,主循环 acc_count<quota 自然停
    )

    acc_count        = result["acc_count"]
    dead_terminated  = result["dead_terminated"]
    doc_pool         = result["doc_pool"]
    ok_count         = result["ok_count"]
    fail_count       = result["fail_count"]
    total_notices    = result["total_notices"]
    total_violations = result["total_violations"]
    total_alerts     = result["total_alerts"]

    if doc_pool:
        log(f"\n! 文档池还剩 {len(doc_pool)} 篇未发完(死磕已无可用账号)")
        for d in doc_pool:
            log(f"  未发: {os.path.basename(d)}")

    # 写最终失败记录
    final_fails = [(ts, n, r) for ts, n, r in fail_records
                   if (n not in success_accounts) or (n in dead_terminated)]
    write_fail_excel(final_fails)

    # ========== 收尾:硬终止账号汇报 ==========
    if dead_terminated:
        log(f"\n★ 4 类硬终止账号 {len(dead_terminated)} 个 (已写入'硬终止账号'sheet,需人工处理):")
        for name, (reason, ts, cnt) in dead_terminated.items():
            log(f"  - {name}\t{reason}\t本次已发 {cnt} 篇\t{ts}")
    else:
        log("\n★ 无硬终止账号,本次全部账号正常完成或仍在 active")

    # 排序"发文汇总"：同账号的多轮放一起，按轮次升序
    _sort_summary_by_account()

    # 写信用分记录
    if credit_records:
        credit_file = os.path.join(RUN_REPORT_DIR, "信用分记录.txt")
        lines = ["信用分记录\n", "=" * 40 + "\n"]
        for acc, score in sorted(credit_records.items(), key=lambda x: x[1]):
            flag = " ← 危险" if score < 60 else (" ← 警告" if score < 80 else "")
            lines.append(f"{score}分  {acc}{flag}\n")
        with open(credit_file, "a", encoding="utf-8") as f:
            f.writelines(lines)

    # 生成汇总报告
    run_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    final_fail_names = sorted(set(n for _, n, _ in final_fails))
    summary_lines = [
        "=" * 60,
        f"运行汇总 - {run_time_str}",
        "=" * 60,
        f"类型    : 文章发布(死磕模式)",
        f"报告目录: {RUN_REPORT_DIR}",
        "",
        f"成功发布: {ok_count} 篇",
        f"中途失败: {fail_count} 次(已重试到底)",
        f"硬终止号: {len(dead_terminated)} 个 (失登/封号/禁言/找不到)",
        f"文档剩余: {len(doc_pool)} 篇",
    ]
    if dead_terminated:
        summary_lines.append("")
        summary_lines.append("硬终止账号 (需人工处理):")
        for name, (reason, ts, cnt) in dead_terminated.items():
            summary_lines.append(f"  - {name}	{reason}	本次已发 {cnt} 篇")
    summary_lines.extend([
        "",
        f"系统通知: {total_notices} 条  → 系统通知.txt",
        f"违规提醒: {total_violations} 条  → 违规提醒.txt",
        f"高阅读提醒: {total_alerts} 条  → 高阅读提醒.txt",
        "=" * 60,
    ])
    summary_text = "\n".join(summary_lines)
    summary_file = os.path.join(RUN_REPORT_DIR, "汇总报告.txt")
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write(summary_text)

    main_ws.close()
    log(f"\n{'='*50}")
    log(f"完成! 成功:{ok_count}  最终失败:{len(final_fail_names)}  硬终止:{len(dead_terminated)}")
    log("\n" + summary_text)
    log(f"{'='*50}")


if __name__ == "__main__":
    main()
