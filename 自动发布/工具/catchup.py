#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
catchup.py — 中断恢复一键工具(三大件通用)
v1101.5 · 2026-05-02 · 全员独立脚本控制,主线零内部控制

用法:
    cd ~/Desktop/<件>自动发布 && python3 catchup.py

逻辑(自适应件类型):
    1. detect_kind() → 微头条 (QUOTA=5) / 文章 (QUOTA=3)
    2. 数据源优先级:
       优先 1: 文章定时件「待补漏」sheet(若存在 + 当前是文章件 → 文章定时件中断救场)
       优先 2: 自己今日 log 反推(quota = QUOTA_TARGET - 已成功)
    3. 找断点账号(自己 log 最后一篇成功的下一个)
    4. 环形重排:断点为首 → 末 → 绕回开头
    5. 写本件「白名单」+ 备份

部署:
    - 5 机 ~/Desktop/{微头条,文章}自动发布/catchup.py
    - Win 台机 ~/Desktop/台机专用自动发布/{微头条,文章}自动发布/catchup.py
    - 仓库主版: 自动发布/工具/catchup.py
"""
import os, sys, glob, shutil, time, re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺 openpyxl: pip install openpyxl"); sys.exit(1)

HERE = Path(__file__).parent.resolve()
HOME = Path.home()


def detect_kind():
    """返回 (件名, QUOTA_TARGET)"""
    n = HERE.name
    if "微头条" in n: return ("微头条", 5)
    if "文章" in n:   return ("文章", 3)
    print(f"✗ 未识别件类型: {n}"); sys.exit(1)


def find_timer_dir():
    """找文章定时件目录(各机命名差异)"""
    cands = [
        HOME / "Desktop/文章定时自动发布",
        HOME / "Desktop/Mac文章定时自动发布",
        HOME / "code/头条自动发布/文章定时自动发布",
        HOME / "code/头条自动发布/Mac文章定时自动发布",
        HOME / "Desktop/台机专用自动发布/文章定时自动发布",
    ]
    for d in cands:
        if d.exists() and (d / "账号配置.xlsx").exists():
            return d
    return None


def read_待补漏(timer_dir):
    """读 文章定时件 账号配置.xlsx「待补漏」sheet → [(账号, 漏数), ...]"""
    p = timer_dir / "账号配置.xlsx"
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        if "待补漏" not in wb.sheetnames:
            wb.close(); return []
        ws = wb["待补漏"]
        items = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            miss = ws.cell(r, 2).value
            if name and miss:
                try:
                    m = int(miss)
                    if m > 0:
                        items.append((str(name).strip(), m))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return items
    except Exception as e:
        print(f"  ! 读「待补漏」失败: {e}")
        return []


def find_self_log():
    """找当前件最新 log"""
    log_root = HERE / "运行报告"
    if not log_root.exists(): return None
    logs = sorted(log_root.glob("*/运行日志.txt"), key=lambda p: p.stat().st_mtime, reverse=True)
    return logs[0] if logs else None


def reverse_quota_from_log(log_file, accounts_known, quota_target):
    """从自己 log 反推每号已发数 → quota = QUOTA_TARGET - 已发
    返回 [(账号, 漏数), ...] 按 log 出现顺序(早到晚)"""
    if not log_file or not log_file.exists():
        return []
    # 微头条/文章 gtg_batch log 格式: [HH:MM:SS] [剩余 N 篇] [大1/小1/A] 账号名 -> docx
    # OK 行: "OK 发布成功" / "✓ 发文成功"
    task_pat = re.compile(r"\[剩余\s*\d+\s*篇\]\s*\[\S+?\]\s*(?:\[补发\]\s*)?(\S+?)\s*->")
    ok_pat = re.compile(r"OK 发布成功|✓ 发文成功|发布成功")

    sent_count = {}
    cur_acc = None
    appear_order = []
    try:
        with open(log_file, encoding="utf-8") as f:
            for line in f:
                m = task_pat.search(line)
                if m:
                    cur_acc = m.group(1).strip()
                    if cur_acc not in appear_order:
                        appear_order.append(cur_acc)
                    continue
                if ok_pat.search(line) and cur_acc:
                    sent_count[cur_acc] = sent_count.get(cur_acc, 0) + 1
                    cur_acc = None
    except Exception as e:
        print(f"  ! 读 log 失败: {e}")
        return []

    items = []
    # 按出现顺序排,每号 quota = max(0, target - sent)
    for acc in appear_order:
        q = max(0, quota_target - sent_count.get(acc, 0))
        if q > 0:
            items.append((acc, q))
    return items, sent_count


def find_breakpoint(items, log_file):
    """从 log 找最后一篇 ✓ 成功的下一个 → items 中的 index"""
    if not items or not log_file or not log_file.exists():
        return 0
    last_account = None
    pat = re.compile(r"OK 发布成功|✓ 发文成功|✓ 定时发布成功")
    task_pat = re.compile(r"\[剩余\s*\d+\s*篇\]\s*\[\S+?\]\s*(?:\[补发\]\s*)?(\S+?)\s*->")
    cur_acc = None
    try:
        with open(log_file, encoding="utf-8") as f:
            for line in f:
                tm = task_pat.search(line)
                if tm:
                    cur_acc = tm.group(1).strip()
                    continue
                if pat.search(line):
                    if cur_acc:
                        last_account = cur_acc
                    else:
                        m = re.search(r"成功:\s*(\S+)", line)
                        if m: last_account = m.group(1).strip()
    except Exception:
        return 0

    if not last_account: return 0
    names = [n for n, _ in items]
    if last_account not in names:
        return 0
    idx = names.index(last_account)
    return (idx + 1) % len(items)


def write_whitelist(items, start_idx, kind, quota_total):
    p = HERE / "账号配置.xlsx"
    if not p.exists():
        print(f"✗ 当前件无 账号配置.xlsx: {p}"); sys.exit(1)
    ts = time.strftime("%Y%m%d_%H%M%S")
    bak = str(p) + f".bak_catchup_{ts}"
    shutil.copy2(p, bak)
    print(f"备份: {os.path.basename(bak)}")

    ordered = items[start_idx:] + items[:start_idx]

    wb = load_workbook(p)
    if "白名单" not in wb.sheetnames:
        ws = wb.create_sheet("白名单")
        ws.cell(1, 1, "账号名"); ws.cell(1, 2, "发文份数")
    else:
        ws = wb["白名单"]
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
    r = 2
    # [2026-05-16 缺哥拍] B 列写 QUOTA_TOTAL 不写漏数,跟 gtg_batch acc_count<quota 语义对齐
    # 修前: 漏数 (QUOTA-已发) → 接续点累计已发 > 漏数被立刻跳; 修后: QUOTA_TOTAL 让累计上限自然停
    for name, miss in ordered:
        ws.cell(r, 1, name); ws.cell(r, 2, quota_total); r += 1
    wb.save(p)

    total = sum(m for _, m in ordered)
    print(f"✓ 白名单 {r-2} 行 / 共 {total} 篇待补 (xlsx B 列写 QUOTA={quota_total})")
    print(f"  首位(断点): {ordered[0][0]}  漏 {ordered[0][1]} 篇")
    print(f"  末位:         {ordered[-1][0]}  漏 {ordered[-1][1]} 篇")


def main():
    kind, quota_target = detect_kind()
    print(f"=== catchup.py · {kind}自动发布 (QUOTA={quota_target}) ===")
    print(f"    HERE: {HERE}\n")

    items = []
    log_file_for_breakpoint = None
    source = ""

    # 数据源优先级 1: 文章定时件「待补漏」sheet (仅文章件适用)
    if kind == "文章":
        timer_dir = find_timer_dir()
        if timer_dir:
            待补漏_items = read_待补漏(timer_dir)
            if 待补漏_items:
                items = 待补漏_items
                source = f"文章定时件「待补漏」({timer_dir})"
                # 断点用 文章定时件 log
                log_root = timer_dir / "运行报告"
                logs = sorted(log_root.glob("*/运行日志.txt"), key=lambda p: p.stat().st_mtime, reverse=True)
                log_file_for_breakpoint = logs[0] if logs else None
                print(f"  ✓ 数据源: {source}")
                print(f"     {len(items)} 账号 / {sum(m for _, m in items)} 篇待补")

    # 数据源优先级 2: 自己今日 log 反推
    if not items:
        log_file_for_breakpoint = find_self_log()
        if log_file_for_breakpoint:
            result = reverse_quota_from_log(log_file_for_breakpoint, [], quota_target)
            if isinstance(result, tuple):
                items, sent_count = result
                if items:
                    source = f"自己今日 log 反推 ({log_file_for_breakpoint.parent.name})"
                    print(f"  ✓ 数据源: {source}")
                    print(f"     {len(items)} 账号 / {sum(m for _, m in items)} 篇待补 (已发{sum(sent_count.values())} 篇)")

    if not items:
        print("✗ 无数据 — 没有「待补漏」数据,自己 log 也没成功记录")
        print("  人工写白名单或确认是否本日已跑过")
        sys.exit(1)

    # 找断点
    start_idx = find_breakpoint(items, log_file_for_breakpoint)
    if start_idx == 0:
        print(f"  断点: items[0] (无法定位 / 最后成功账号已发完)")
    else:
        print(f"  断点: items[{start_idx}] = {items[start_idx][0]}\n")

    write_whitelist(items, start_idx, kind, quota_target)
    print()
    print("下一步: 双击 go.command (或 go.bat) 启动跑")


if __name__ == "__main__":
    main()
