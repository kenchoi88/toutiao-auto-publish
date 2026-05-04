#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""5 机 × 3 件 账号配置.xlsx 结构审计 — sheet 名 / 列名 / 行数,不动数据"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
fp = sys.argv[1] if len(sys.argv) > 1 else None
if not fp or not os.path.exists(fp):
    print(f"FILE_NOT_FOUND: {fp}"); sys.exit(1)

wb = openpyxl.load_workbook(fp, read_only=True)
print(f"=== {fp} ===")
print(f"  sheet 数: {len(wb.sheetnames)}: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"    [{sn}] 空"); continue
    # 列头行
    headers = rows[0] if rows else []
    data_rows = [r for r in rows[1:] if any(c not in (None, '', '-') for c in r)]
    print(f"    [{sn}] 列名: {list(headers)} / 数据行: {len(data_rows)}")
