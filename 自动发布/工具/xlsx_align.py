#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""5 机 × 3 件 账号配置.xlsx 结构对齐
   - 加缺失标准 sheet
   - 统一列名(只改 header 行)
   - 删 v1101 过时 sheet(待补漏/补漏历史)
   - 非标准非过时 sheet 仅警告保留(可能是用户自定义)
   - idempotent: 无改动时自动删备份
"""
import sys, os, shutil
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

STANDARD = [
    ('不首发',     ['账号名']),
    ('永久跳过',   ['账号名']),
    ('本轮已发',   ['账号名', '已发次数']),
    ('白名单',     ['账号名', '发文份数']),
    ('失败列表',   ['账号名', '失败原因', '文稿名', '失败时间', '轮次']),
    ('硬终止账号', ['账号名', '终止原因', '终止时间', '本次已发篇数']),
    ('发文汇总',   ['账号名', '轮次', '发文时间', '失败时间', '补发成功时间']),
]
OBSOLETE = ['待补漏', '补漏历史']
STANDARD_NAMES = {n for n, _ in STANDARD}

if len(sys.argv) < 2:
    print("用法: python xlsx_align.py <账号配置.xlsx>"); sys.exit(1)
fp = sys.argv[1]
if not os.path.exists(fp):
    print(f"FILE_NOT_FOUND: {fp}"); sys.exit(1)

ts = datetime.now().strftime("%Y%m%d_%H%M%S")
bak = f"{fp}.bak_pre_align_{ts}"
shutil.copy(fp, bak)

try:
    wb = openpyxl.load_workbook(fp)
except Exception as e:
    print(f"LOAD_FAIL: {e}"); os.remove(bak); sys.exit(2)

changes = []

for sn in OBSOLETE:
    if sn in wb.sheetnames:
        del wb[sn]
        changes.append(f"- 删过时 sheet [{sn}]")

for sheet_name, cols in STANDARD:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for c, name in enumerate(cols, 1):
            ws.cell(row=1, column=c, value=name)
        changes.append(f"+ 加 sheet [{sheet_name}]")
    else:
        ws = wb[sheet_name]
        for c, name in enumerate(cols, 1):
            cur = ws.cell(row=1, column=c).value
            if cur != name:
                ws.cell(row=1, column=c, value=name)
                changes.append(f"  [{sheet_name}] 列{c}: {repr(cur)} → {repr(name)}")

unknown = [sn for sn in wb.sheetnames if sn not in STANDARD_NAMES]
if unknown:
    print(f"  [{fp}] ⚠ 非标准 sheet 保留(未识别): {unknown}")

if changes:
    try:
        wb.save(fp)
    except Exception as e:
        print(f"SAVE_FAIL: {e}  (备份已保留: {bak})"); sys.exit(3)
    print(f"  [{fp}]")
    for c in changes: print(f"    {c}")
    print(f"    备份: {bak}")
else:
    print(f"  [{fp}] 已对齐,无改动")
    os.remove(bak)
